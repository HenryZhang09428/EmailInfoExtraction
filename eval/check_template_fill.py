import json
import os
from pathlib import Path
from core.template.parser import parse_template_xlsx
from core.template.writer import apply_fill_plan

def main():
    eval_dir = Path(__file__).parent
    fixtures_dir = eval_dir / "fixtures"
    
    template_path = fixtures_dir / "template.xlsx"
    extracted_json_path = fixtures_dir / "extracted_sample.json"
    fill_plan_path = fixtures_dir / "fill_plan.json"
    output_path = fixtures_dir / "filled_template.xlsx"
    
    if not template_path.exists():
        print(f"Template not found: {template_path}")
        print("Creating a minimal template for testing...")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "姓名"
        ws["B1"] = "邮箱"
        ws["C1"] = "电话"
        ws["A2"] = "张三"
        ws["B2"] = "test@example.com"
        ws["C2"] = "13800138000"
        wb.save(template_path)
        print(f"Created template: {template_path}")
    
    if not extracted_json_path.exists():
        print(f"Extracted JSON not found: {extracted_json_path}")
        print("Creating sample extracted JSON...")
        extracted_data = {
            "sources": [
                {
                    "filename": "test.xlsx",
                    "source_type": "excel",
                    "extracted": {
                        "姓名": "李四",
                        "邮箱": "lisi@example.com",
                        "电话": "13900139000"
                    }
                }
            ]
        }
        with open(extracted_json_path, 'w', encoding='utf-8') as f:
            json.dump(extracted_data, f, ensure_ascii=False, indent=2)
        print(f"Created extracted JSON: {extracted_json_path}")
    
    print("Parsing template...")
    template_schema = parse_template_xlsx(str(template_path))
    print(f"Found {len(template_schema.sheet_schemas)} sheet(s)")
    for sheet_schema in template_schema.sheet_schemas:
        print(f"  Sheet: {sheet_schema.sheet}, Regions: {len(sheet_schema.regions)}")
    
    if fill_plan_path.exists():
        print(f"Loading fill plan from: {fill_plan_path}")
        with open(fill_plan_path, 'r', encoding='utf-8') as f:
            fill_plan = json.load(f)
    else:
        print("No fill_plan.json found. Creating a minimal fill plan...")
        with open(extracted_json_path, 'r', encoding='utf-8') as f:
            extracted_data = json.load(f)
        
        sheet_name = template_schema.sheet_schemas[0].sheet if template_schema.sheet_schemas else "Sheet"
        region_id = template_schema.sheet_schemas[0].regions[0].region_id if template_schema.sheet_schemas and template_schema.sheet_schemas[0].regions else "region_1"
        
        extracted = extracted_data["sources"][0]["extracted"] if extracted_data.get("sources") else {}
        
        fill_plan = {
            "target": {
                "sheet": sheet_name,
                "region_id": region_id,
                "layout_type": "table",
                "clear_policy": "clear_values_keep_format"
            },
            "clear_ranges": ["A2:C10"],
            "row_writes": [
                {
                    "start_cell": "A2",
                    "rows": [extracted],
                    "column_mapping": {
                        "姓名": "A",
                        "邮箱": "B",
                        "电话": "C"
                    }
                }
            ],
            "writes": [],
            "warnings": []
        }
        
        with open(fill_plan_path, 'w', encoding='utf-8') as f:
            json.dump(fill_plan, f, ensure_ascii=False, indent=2)
        print(f"Created fill plan: {fill_plan_path}")
    
    print("Applying fill plan...")
    apply_fill_plan(str(template_path), fill_plan, str(output_path))
    
    if output_path.exists():
        print(f"✅ OK: Filled template saved to {output_path}")
        
        from openpyxl import load_workbook
        wb = load_workbook(output_path, data_only=True)
        ws = wb.active
        print(f"Cell A2: {ws['A2'].value}")
        print(f"Cell B2: {ws['B2'].value}")
        print(f"Cell C2: {ws['C2'].value}")
        wb.close()
    else:
        print(f"❌ ERROR: Output file not created")

if __name__ == "__main__":
    main()
