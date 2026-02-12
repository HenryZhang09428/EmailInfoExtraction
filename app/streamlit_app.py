from dotenv import load_dotenv
load_dotenv()

import atexit
import shutil
import streamlit as st
import streamlit.components.v1 as components
import json
import inspect
import tempfile
import os
from pathlib import Path
import sys
import hashlib
from datetime import datetime, date

# Ensure repo root is on sys.path for "core" imports
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


# ============================================================================
# Cache Cleanup
# ============================================================================
def cleanup_cache() -> None:
    """
    Clean up the .cache/eml_parts directory.
    
    This function is registered with atexit to ensure cleanup on application exit,
    and can also be called at startup to clean up leftovers from crashed sessions.
    """
    cache_dir = Path(__file__).parent.parent / ".cache" / "eml_parts"
    if cache_dir.exists():
        try:
            shutil.rmtree(cache_dir)
        except Exception:
            pass  # Silently ignore cleanup errors


# Register cleanup function for application exit
atexit.register(cleanup_cache)
from typing import List, Optional, Any
from core.pipeline import run_pipeline, update_ir_scores, build_stable_ir_signature
try:
    from core.pipeline import run_extract, fill_template
except ImportError as e:
    import traceback
    print(f"Warning: Failed to import template functions: {e}")
    traceback.print_exc()
    run_extract = None
    fill_template = None
from core.mapping.attribute_set import build_attribute_set
from core.mapping.mapper import map_to_schema
from core.config import get_settings
from core.llm import get_llm_client
from core.prompts_loader import get_prompts
from core.logger import get_logger
import pandas as pd

logger = get_logger(__name__)

# ============================================================================
# Constants
# ============================================================================
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB per file
ALLOWED_EXTENSIONS = ['.xlsx', '.xls', '.docx', '.txt', '.eml', '.png', '.jpg', '.jpeg', '.gif', '.csv']


# ============================================================================
# Helper Functions (DRY Principle)
# ============================================================================

def validate_file_size(uploaded_file) -> bool:
    """
    Validate that the uploaded file doesn't exceed the maximum size limit.
    
    Returns:
        True if file size is valid, False otherwise.
    """
    if uploaded_file.size > MAX_FILE_SIZE:
        size_mb = uploaded_file.size / (1024 * 1024)
        max_mb = MAX_FILE_SIZE / (1024 * 1024)
        st.error(f"❌ 文件 '{uploaded_file.name}' 太大 ({size_mb:.1f}MB)。最大允许大小为 {max_mb:.0f}MB。")
        logger.warning("File %s rejected: size %d bytes exceeds limit %d", 
                      uploaded_file.name, uploaded_file.size, MAX_FILE_SIZE)
        return False
    return True


def handle_download(file_path: str, label: str, file_name: Optional[str] = None,
                   mime_type: str = "application/octet-stream", disabled: bool = False) -> None:
    """
    Create a download button for a file.
    
    Args:
        file_path: Path to the file to download.
        label: Button label text.
        file_name: Name for the downloaded file. Defaults to original filename.
        mime_type: MIME type of the file.
    """
    if file_name is None:
        file_name = Path(file_path).name
    
    if disabled:
        st.download_button(
            label=label,
            data=b"",
            file_name=file_name,
            mime=mime_type,
            disabled=True
        )
        return
    
    if not os.path.exists(file_path):
        st.error(f"文件不存在: {file_path}")
        return
    
    try:
        with open(file_path, "rb") as f:
            file_data = f.read()
        
        st.download_button(
            label=label,
            data=file_data,
            file_name=file_name,
            mime=mime_type
        )
    except Exception as e:
        st.error(f"无法读取文件: {e}")
        logger.error("Failed to read file for download: %s", e)


def display_extraction_results(source_docs: List[Any], current_source_id: Optional[str] = None) -> None:
    """
    Display extraction results for source documents.
    
    Args:
        source_docs: List of SourceDoc objects.
        current_source_id: Currently selected source ID.
    """
    if not source_docs:
        st.info("没有可显示的源文档。请先运行提取。")
        return
    
    # Source selector if multiple sources
    if len(source_docs) > 1:
        source_options = {f"{s.filename} ({s.source_type})": s.source_id for s in source_docs}
        
        # Find current index
        current_index = 0
        if current_source_id:
            source_ids = list(source_options.values())
            if current_source_id in source_ids:
                current_index = source_ids.index(current_source_id)
        
        selected_source_name = st.selectbox(
            "选择要查看的数据源",
            options=list(source_options.keys()),
            index=current_index,
            key="extracted_json_source_selector"
        )
        selected_source_id = source_options[selected_source_name]
        
        # Find the selected source
        display_source = next((s for s in source_docs if s.source_id == selected_source_id), None)
    else:
        display_source = source_docs[0] if source_docs else None
    
    if display_source:
        st.subheader(f"提取结果: {display_source.filename}")
        st.write(f"**类型**: {display_source.source_type}")
        st.write(f"**Source ID**: `{display_source.source_id}`")
        
        if display_source.extracted:
            st.json(display_source.extracted)
            if st.session_state.get("show_debug") and isinstance(display_source.extracted, dict):
                metadata = display_source.extracted.get("metadata", {}) if isinstance(display_source.extracted, dict) else {}
                warnings = display_source.extracted.get("warnings", [])
                coverage = metadata.get("coverage")
                fallback_used = metadata.get("fallback_used")
                semantic_keys_count = metadata.get("semantic_keys_count")
                semantic_key_by_header = metadata.get("semantic_key_by_header") or {}

                st.divider()
                st.subheader("调试信息（抽取）")
                if coverage is not None:
                    st.write(f"**coverage**: {coverage}")
                if fallback_used is not None:
                    st.write(f"**fallback_used**: {fallback_used}")
                if semantic_keys_count is not None:
                    st.write(f"**semantic_keys_count**: {semantic_keys_count}")
                if warnings:
                    st.write(f"**warnings**: {warnings}")

                if isinstance(semantic_key_by_header, dict) and semantic_key_by_header:
                    preview_items = list(semantic_key_by_header.items())[:20]
                    preview_dict = {k: v for k, v in preview_items}
                    st.write("**semantic_key_by_header (前20项)**")
                    st.json(preview_dict)
        else:
            st.info("没有可用的提取数据。")
    else:
        st.info("未选择数据源。")


def display_excel_sheet(ws, sheet_name: str, template_schema: Any = None, max_rows: int = 1000) -> None:
    """
    Display an Excel worksheet as a DataFrame.
    
    Args:
        ws: openpyxl Worksheet object.
        sheet_name: Name of the sheet.
        template_schema: Optional TemplateSchema for header mapping.
        max_rows: Maximum rows to read.
    """
    from openpyxl.utils import get_column_letter
    
    # Read sheet to DataFrame
    data = []
    max_col = ws.max_column
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_col, values_only=True), start=1):
        row_data = [value if value is not None else "" for value in row]
        data.append(row_data)
    
    if not data:
        st.info("表格为空，没有填充数据。")
        return
    
    # Normalize row lengths
    max_cols = max(len(row) for row in data) if data else 0
    for row in data:
        while len(row) < max_cols:
            row.append("")
    
    df_raw = pd.DataFrame(data)
    
    if df_raw.empty:
        st.info("表格为空，没有填充数据。")
        return
    
    st.divider()
    st.subheader(f"Sheet: {sheet_name}")
    
    # Detect header row
    header_row = None
    for idx in range(min(5, len(df_raw))):
        non_null_count = df_raw.iloc[idx].notna().sum()
        if non_null_count >= 2:
            header_row = idx
            break
    
    if header_row is None:
        st.info("无法检测表头行，显示原始数据（所有列）：")
        st.dataframe(_coerce_df_for_display(df_raw.head(20)), use_container_width=True, hide_index=True)
        return
    
    df = df_raw.copy()
    
    # Build header values
    header_values = _build_header_values(df_raw, header_row, sheet_name, template_schema)
    
    # Ensure enough header values
    if len(header_values) < len(df.columns):
        for i in range(len(header_values), len(df.columns)):
            header_values.append(f"Col_{get_column_letter(i+1)}")
    
    df.columns = header_values[:len(df.columns)]
    
    # Determine data start row: use template schema if available
    data_start_idx = header_row + 1
    if template_schema:
        sheet_schema = next((ss for ss in template_schema.sheet_schemas if ss.sheet == sheet_name), None)
        if sheet_schema and sheet_schema.regions:
            region = sheet_schema.regions[0]
            if region.header_rows:
                # header_rows is 1-indexed (Excel row numbers)
                # Convert to 0-indexed DataFrame index: Excel row N -> index N-1
                # Data starts after max header row
                max_header_excel_row = max(region.header_rows)
                data_start_idx = max_header_excel_row  # Excel row N+1 -> index N
    
    df = df.iloc[data_start_idx:].reset_index(drop=True)
    df = df.dropna(how='all')
    
    if not df.empty:
        st.dataframe(_coerce_df_for_display(df), use_container_width=True, hide_index=True)
    else:
        st.info("处理表头后未找到数据行。")
        st.write("原始数据预览（前10行）：")
        st.dataframe(_coerce_df_for_display(df_raw.head(10)), use_container_width=True, hide_index=True)


def _build_header_values(df_raw: pd.DataFrame, header_row: int, sheet_name: str, 
                        template_schema: Any) -> List[str]:
    """Build header values from template schema or raw data."""
    from openpyxl.utils import get_column_letter
    
    header_values = []
    
    if template_schema:
        sheet_schema = next((ss for ss in template_schema.sheet_schemas if ss.sheet == sheet_name), None)
        
        if sheet_schema and sheet_schema.regions:
            region = sheet_schema.regions[0]
            if region.table and region.table.header:
                header_map = {h.col_letter: h.header_path for h in region.table.header}
                
                seen_names = {}
                for i in range(len(df_raw.columns)):
                    col_letter = get_column_letter(i + 1)
                    if col_letter in header_map:
                        header_name = header_map[col_letter]
                        if header_name:
                            if header_name in seen_names:
                                seen_names[header_name] += 1
                                header_values.append(f"{header_name}_{seen_names[header_name]}")
                            else:
                                seen_names[header_name] = 0
                                header_values.append(header_name)
                        else:
                            cell_value = df_raw.iloc[header_row, i]
                            header_values.append(str(cell_value) if pd.notna(cell_value) and str(cell_value).strip() else f"Col_{col_letter}")
                    else:
                        cell_value = df_raw.iloc[header_row, i]
                        header_values.append(str(cell_value) if pd.notna(cell_value) and str(cell_value).strip() else f"Col_{col_letter}")
                return header_values
    
    # Fallback: use raw values
    for i, x in enumerate(df_raw.iloc[header_row]):
        if pd.notna(x) and str(x).strip():
            header_values.append(str(x))
        else:
            header_values.append(f"Col_{get_column_letter(i+1)}")
    
    return header_values


def save_uploaded_file(uploaded_file, temp_dir: str) -> Optional[str]:
    """
    Save an uploaded file to a temporary directory.
    
    Args:
        uploaded_file: Streamlit UploadedFile object.
        temp_dir: Path to temporary directory.
    
    Returns:
        Path to saved file, or None if failed.
    """
    if not validate_file_size(uploaded_file):
        return None
    
    try:
        file_content = uploaded_file.read()
        temp_path = os.path.join(temp_dir, uploaded_file.name)
        
        with open(temp_path, "wb") as f:
            f.write(file_content)
        
        logger.debug("Saved uploaded file: %s (%d bytes)", temp_path, len(file_content))
        return temp_path
    except Exception as e:
        st.error(f"保存文件 '{uploaded_file.name}' 失败: {e}")
        logger.error("Failed to save uploaded file %s: %s", uploaded_file.name, e)
        return None


def _safe_stringify_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, pd.Timestamp)):
        try:
            return value.isoformat(sep=" ", timespec="seconds")
        except Exception:
            return str(value)
    return value


def _coerce_df_for_display(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df_disp = df.copy()
    for col in df_disp.columns:
        if pd.api.types.is_datetime64_any_dtype(df_disp[col]):
            df_disp[col] = df_disp[col].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
        else:
            df_disp[col] = df_disp[col].apply(_safe_stringify_cell)
    return df_disp


def get_safe_settings_display() -> dict:
    """
    Get settings for display, with sensitive values masked.
    
    Returns:
        Dict with safe settings (API keys masked).
    """
    settings = get_settings()
    return {
        "OPENAI_MODEL": settings.OPENAI_MODEL,
        "OPENAI_VISION_MODEL": settings.OPENAI_VISION_MODEL,
        "OPENAI_BASE_URL": settings.OPENAI_BASE_URL,
        "OPENAI_API_KEY": "***已隐藏***" if settings.OPENAI_API_KEY else "未设置",
        "TEMPERATURE": settings.TEMPERATURE,
        "REQUEST_TIMEOUT": settings.REQUEST_TIMEOUT,
    }


def _build_fill_cache_key(template_bytes: Optional[bytes], ir: Any) -> Optional[str]:
    if not template_bytes or ir is None:
        return None
    ir_signature = build_stable_ir_signature(ir)
    digest = hashlib.sha256(template_bytes + ir_signature.encode("utf-8")).hexdigest()
    return digest

# ============================================================================
# Session State Initialization
# ============================================================================

def init_session_state():
    """Initialize all session state variables."""
    defaults = {
        "uploaded_files_dict": {},
        "current_source_id": None,
        "pinned_source_id": None,
        "ir": None,
        "final_json": None,
        "template_path": None,
        "template_bytes": None,
        "filled_template_path": None,
        "template_schema": None,
        "fill_plan": None,
        "record_routing_debug": None,
        "fill_cache": {},
        "temp_dir": None,  # Track temporary directory for cleanup
        "cache_cleaned": False,
    }
    for key, default_value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default_value

init_session_state()

# One-time cache cleanup per session to avoid deleting active attachments
if not st.session_state.cache_cleaned:
    cleanup_cache()
    st.session_state.cache_cleaned = True

# Create or get persistent temporary directory for this session
if st.session_state.temp_dir is None or not os.path.exists(st.session_state.temp_dir):
    st.session_state.temp_dir = tempfile.mkdtemp(prefix="email_extraction_")

st.title("Multi-Source Input Pipeline")

# 添加状态指示器
if st.session_state.uploaded_files_dict:
    st.info(f"📊 当前已加载 {len(st.session_state.uploaded_files_dict)} 个文件")

with st.sidebar:
    st.header("File Management")
    
    if st.button("Clear cache", use_container_width=True, help="清理 .cache/eml_parts 缓存目录"):
        cleanup_cache()
        st.session_state.cache_cleaned = True
        st.success("缓存已清理")
    
    # 调试信息开关
    debug_mode = st.checkbox("显示调试信息", key="show_debug")
    
    if debug_mode:
        st.write(f"已上传文件数: {len(st.session_state.uploaded_files_dict)}")
        st.write(f"Session state keys: {list(st.session_state.keys())}")
        llm_client = get_llm_client()
        last_call = llm_client.get_last_call_info() if llm_client else None
        if last_call:
            with st.expander("最近一次 LLM 调用", expanded=False):
                start_ts = last_call.get("start_time")
                end_ts = last_call.get("end_time")
                start_str = datetime.fromtimestamp(start_ts).strftime("%H:%M:%S") if start_ts else "-"
                end_str = datetime.fromtimestamp(end_ts).strftime("%H:%M:%S") if end_ts else "-"
                st.write(f"step: {last_call.get('step')}")
                st.write(f"method: {last_call.get('method')}")
                st.write(f"status: {last_call.get('status')}")
                st.write(f"model: {last_call.get('model')}")
                st.write(f"timeout: {last_call.get('timeout')}")
                st.write(f"prompt_chars: {last_call.get('prompt_chars')}")
                st.write(f"retries: {last_call.get('retries')}")
                st.write(f"start: {start_str}")
                st.write(f"end: {end_str}")
                st.write(f"elapsed_ms: {last_call.get('elapsed_ms')}")
    
    uploaded_files = st.file_uploader(
        "上传多个文件",
        accept_multiple_files=True,
        key="file_uploader",
        help="支持 .xlsx, .xls, .docx, .txt, .eml, .png, .jpg, .jpeg 等格式"
    )
    
    # 调试信息
    if debug_mode:
        st.write(f"uploaded_files 类型: {type(uploaded_files)}")
        st.write(f"uploaded_files 值: {uploaded_files}")
        if uploaded_files is not None:
            st.write(f"uploaded_files 长度: {len(uploaded_files)}")
            for i, f in enumerate(uploaded_files):
                st.write(f"文件 {i}: {f.name}, 大小: {f.size}")
    
    # 处理上传的文件
    # 注意：uploaded_files 可能是 None（未选择）或空列表（已选择但被清空）或文件列表
    if uploaded_files is not None and len(uploaded_files) > 0:
        new_files = []
        rejected_files = []
        
        for uploaded_file in uploaded_files:
            # 检查文件是否已经上传过
            if uploaded_file.name not in st.session_state.uploaded_files_dict:
                # 文件大小检查
                if not validate_file_size(uploaded_file):
                    rejected_files.append(uploaded_file.name)
                    continue
                
                try:
                    if debug_mode:
                        st.write(f"正在处理文件: {uploaded_file.name} ({uploaded_file.size / 1024:.1f}KB)")
                    
                    # 保存到会话临时目录
                    temp_path = save_uploaded_file(uploaded_file, st.session_state.temp_dir)
                    
                    if temp_path:
                        st.session_state.uploaded_files_dict[uploaded_file.name] = temp_path
                        new_files.append(uploaded_file.name)
                        
                        if debug_mode:
                            st.write(f"文件已保存到: {temp_path}")
                    
                except Exception as e:
                    error_msg = f"上传文件 {uploaded_file.name} 失败: {str(e)}"
                    st.error(error_msg)
                    logger.error("File upload failed: %s", e)
                    if debug_mode:
                        import traceback
                        st.code(traceback.format_exc())
        
        # 显示上传结果
        if new_files:
            st.success(f"✅ 成功上传 {len(new_files)} 个文件: {', '.join(new_files)}")
        
        if rejected_files:
            st.warning(f"⚠️ {len(rejected_files)} 个文件因超出大小限制被拒绝")
    
    # 显示已上传的文件列表
    if st.session_state.uploaded_files_dict:
        st.divider()
        st.subheader(f"已上传的文件 ({len(st.session_state.uploaded_files_dict)})")
        for filename in list(st.session_state.uploaded_files_dict.keys()):
            col1, col2 = st.columns([4, 1])
            with col1:
                st.text(filename)
            with col2:
                if st.button("🗑️", key=f"delete_{filename}", use_container_width=True, help="删除文件"):
                    del st.session_state.uploaded_files_dict[filename]
                    st.rerun()
    else:
        st.info("📁 还没有上传任何文件")
    
    st.divider()
    st.subheader("Source List")
    
    if st.session_state.ir and st.session_state.ir.sources:
        for source in st.session_state.ir.sources:
            is_pinned = st.session_state.pinned_source_id == source.source_id
            is_current = st.session_state.current_source_id == source.source_id
            
            col1, col2 = st.columns([3, 1])
            with col1:
                prefix = ""
                if source.parent_source_id:
                    prefix = "📎 "
                label = f"{'📌 ' if is_pinned else ''}{prefix}{source.filename}"
                if st.button(label, key=f"select_{source.source_id}", use_container_width=True):
                    st.session_state.current_source_id = source.source_id
                    st.rerun()
            with col2:
                if is_pinned:
                    if st.button("🔓", key=f"unpin_{source.source_id}"):
                        st.session_state.pinned_source_id = None
                        st.rerun()
                else:
                    if st.button("📌", key=f"pin_{source.source_id}"):
                        st.session_state.pinned_source_id = source.source_id
                        st.session_state.current_source_id = source.source_id
                        st.rerun()
    else:
        st.info("No sources available. Run pipeline first.")

st.divider()
st.subheader("Processing")

col1, col2 = st.columns(2)

with col1:
    if st.button("Extract", type="primary", use_container_width=True):
        if not st.session_state.uploaded_files_dict:
            st.error("Please upload source files first.")
        elif run_extract is None:
            st.error("run_extract function is not available. Please check template module imports.")
        else:
            with st.spinner("Extracting..."):
                try:
                    file_paths = list(st.session_state.uploaded_files_dict.values())
                    st.session_state.ir = run_extract(file_paths)
                    st.session_state.final_json = None
                    st.session_state.filled_template_path = None
                    st.session_state.template_schema = None
                    st.session_state.fill_plan = None
                    st.session_state.record_routing_debug = None
                    
                    if st.session_state.ir.sources:
                        pinned_exists = False
                        if st.session_state.pinned_source_id:
                            for source in st.session_state.ir.sources:
                                if source.source_id == st.session_state.pinned_source_id:
                                    st.session_state.current_source_id = st.session_state.pinned_source_id
                                    pinned_exists = True
                                    break
                        if not pinned_exists:
                            st.session_state.current_source_id = st.session_state.ir.sources[0].source_id
                            st.session_state.pinned_source_id = None
                    
                    st.success("Extraction completed!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Extraction failed: {e}")
                    import traceback
                    st.code(traceback.format_exc())

with col2:
    template_file = st.file_uploader(
        "上传模板 Excel",
        type=["xlsx"],
        key="template_uploader",
        help="上传 Excel 模板文件 (.xlsx)，最大 50MB"
    )
    
    if template_file:
        # 文件大小检查
        if not validate_file_size(template_file):
            st.session_state.template_path = None
        else:
            try:
                st.session_state.template_bytes = template_file.getvalue()
                temp_path = save_uploaded_file(template_file, st.session_state.temp_dir)
                if temp_path:
                    st.session_state.template_path = temp_path
                    st.success(f"✅ 模板已上传: {template_file.name}")
            except Exception as e:
                st.error(f"上传模板失败: {e}")
                logger.error("Template upload failed: %s", e)

if st.session_state.ir and st.session_state.template_path:
    if fill_template is None:
        st.warning("fill_template function is not available. Please check template module imports.")
    else:
        btn_col1, btn_col2 = st.columns(2)
        fill_clicked = btn_col1.button("Fill Template", type="primary", use_container_width=True)
        regen_clicked = btn_col2.button("Regenerate (LLM)", use_container_width=True)
        
        if fill_clicked or regen_clicked:
            with st.spinner("Filling template..."):
                try:
                    template_bytes = st.session_state.template_bytes
                    if template_bytes is None and st.session_state.template_path:
                        with open(st.session_state.template_path, "rb") as f:
                            template_bytes = f.read()
                    cache_key = _build_fill_cache_key(template_bytes, st.session_state.ir)
                    
                    bypass_cache = regen_clicked
                    cache_hit = cache_key and cache_key in st.session_state.fill_cache and not bypass_cache
                    
                    if cache_hit:
                        cached = st.session_state.fill_cache[cache_key]
                        filled_path = cached["filled_path"]
                        template_schema = cached["template_schema"]
                        fill_plan = cached["fill_plan"]
                        st.info("✅ 命中缓存结果（未重新调用 LLM）")
                    else:
                        supports_require_llm = "require_llm" in inspect.signature(fill_template).parameters
                        if regen_clicked and not supports_require_llm:
                            st.error("当前运行的后端版本不支持 require_llm，请重启 Streamlit 以加载最新代码。")
                            st.stop()
                        if supports_require_llm:
                            filled_path, template_schema, fill_plan = fill_template(
                                st.session_state.ir,
                                st.session_state.template_path,
                                require_llm=regen_clicked
                            )
                        else:
                            filled_path, template_schema, fill_plan = fill_template(
                                st.session_state.ir,
                                st.session_state.template_path
                            )
                        if cache_key:
                            st.session_state.fill_cache[cache_key] = {
                                "filled_path": filled_path,
                                "template_schema": template_schema,
                                "fill_plan": fill_plan
                            }
                    
                    st.session_state.filled_template_path = filled_path
                    st.session_state.template_schema = template_schema
                    st.session_state.fill_plan = fill_plan
                    
                    # Show success message with LLM usage indicator
                    llm_used = fill_plan.get("llm_used", False)
                    inferred_count = fill_plan.get("constant_values_count", 0)
                    debug_info = fill_plan.get("debug", {}) if isinstance(fill_plan, dict) else {}
                    llm_status = debug_info.get("llm_status")
                    cells_written = debug_info.get("cells_written")
                    warnings = fill_plan.get("warnings", []) if isinstance(fill_plan, dict) else []
                    fill_failed = (cells_written == 0) or any("0 cells written" in w for w in warnings)
                    
                    if fill_failed:
                        st.error("❌ 模板填充失败（未写入任何单元格）")
                    elif llm_used:
                        msg = "🤖 模板填充成功！（使用 LLM 智能映射）"
                        if inferred_count > 0:
                            msg += f"\n📝 从上下文推断了 {inferred_count} 个字段值"
                        st.success(msg)
                    else:
                        reason = f"（原因: {llm_status}）" if llm_status and llm_status != "ok" else ""
                        st.warning(f"⚠️ 模板填充完成（使用规则匹配，未启用 LLM）{reason}")
                    
                    st.session_state.record_routing_debug = debug_info.get("record_routing_debug")
                    
                    st.rerun()
                except Exception as e:
                    st.error(f"Fill template failed: {e}")
                    import traceback
                    st.code(traceback.format_exc())

if st.session_state.ir and st.session_state.current_source_id:
    current_source = None
    for source in st.session_state.ir.sources:
        if source.source_id == st.session_state.current_source_id:
            current_source = source
            break
    
    if current_source:
        tab1, tab2 = st.tabs(["Extracted JSON", "Final Output"])
        
        with tab1:
            display_extraction_results(
                st.session_state.ir.sources,
                current_source.source_id
            )
        
        with tab2:
            if st.session_state.filled_template_path:
                st.subheader("已填充的模板")
                
                # Show LLM usage indicator
                if st.session_state.fill_plan:
                    llm_used = st.session_state.fill_plan.get("llm_used", False)
                    inferred_count = st.session_state.fill_plan.get("constant_values_count", 0)
                    debug_info = st.session_state.fill_plan.get("debug", {})
                    llm_status = debug_info.get("llm_status")
                    cells_written = debug_info.get("cells_written")
                    warnings = st.session_state.fill_plan.get("warnings", [])
                    fill_failed = (cells_written == 0) or any("0 cells written" in w for w in warnings)
                    
                    if fill_failed:
                        st.error("❌ 填充失败（未写入任何单元格）")
                    elif llm_used:
                        indicator_msg = "🤖 使用 LLM 智能映射"
                        if inferred_count > 0:
                            indicator_msg += f" | 📝 从上下文推断了 {inferred_count} 个字段"
                        st.info(indicator_msg)
                    else:
                        reason = f"（原因: {llm_status}）" if llm_status and llm_status != "ok" else ""
                        st.warning(f"⚠️ 使用规则匹配（未启用 LLM）{reason}")
                
                if os.path.exists(st.session_state.filled_template_path):
                    debug_info = st.session_state.fill_plan.get("debug", {}) if st.session_state.fill_plan else {}
                    warnings = st.session_state.fill_plan.get("warnings", []) if st.session_state.fill_plan else []
                    cells_written = debug_info.get("cells_written")
                    fill_failed = (cells_written == 0) or any("0 cells written" in w for w in warnings)
                    
                    # Download button using helper function
                    handle_download(
                        st.session_state.filled_template_path,
                        "📥 下载已填充的模板",
                        "filled_template.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        disabled=fill_failed
                    )
                    
                    # Display Excel content
                    try:
                        from openpyxl import load_workbook
                        
                        wb = load_workbook(st.session_state.filled_template_path, data_only=True)
                        sheet_names = wb.sheetnames
                        
                        if len(sheet_names) == 1:
                            sheet_name = sheet_names[0]
                            ws = wb[sheet_name]
                            display_excel_sheet(ws, sheet_name, st.session_state.template_schema)
                        else:
                            selected_sheet = st.selectbox("选择工作表", sheet_names, key="sheet_selector")
                            if selected_sheet:
                                ws = wb[selected_sheet]
                                display_excel_sheet(ws, selected_sheet, st.session_state.template_schema)
                        
                        wb.close()
                    except Exception as e:
                        st.error(f"无法显示表格: {e}")
                        logger.error("Failed to display Excel: %s", e)
                        import traceback
                        with st.expander("错误详情"):
                            st.code(traceback.format_exc())
                
                if st.session_state.template_schema:
                    st.divider()
                    with st.expander("Template Schema", expanded=False):
                        st.json(st.session_state.template_schema.model_dump(mode='json', exclude_none=True))
                
                if st.session_state.fill_plan:
                    st.divider()
                    fill_plan = st.session_state.fill_plan
                    debug_info = fill_plan.get("debug", {}) if isinstance(fill_plan, dict) else {}
                    llm_status = debug_info.get("llm_status")
                    records_before = debug_info.get("records_count_before_filter")
                    records_after = debug_info.get("records_count_after_filter")
                    cells_written = debug_info.get("cells_written")
                    selected_sources = debug_info.get("selected_sources")
                    target_info = fill_plan.get("target", {})
                    row_writes = fill_plan.get("row_writes", [])
                    clear_ranges = fill_plan.get("clear_ranges", [])
                    
                    with st.expander("Debug", expanded=False):
                        if st.session_state.record_routing_debug is not None:
                            st.json(st.session_state.record_routing_debug)
                        st.write(f"**llm_used**: {fill_plan.get('llm_used', False)}")
                        st.write(f"**llm_status**: {llm_status}")
                        st.write(f"**records_count**: {records_before} -> {records_after}")
                        records_source_files = debug_info.get("records_source_files", [])
                        records_count_per_source = debug_info.get("records_count_per_source", {})
                        if records_source_files:
                            st.write(f"**records_source_files**: {records_source_files}")
                        if records_count_per_source:
                            st.write(f"**records_count_per_source**: {records_count_per_source}")
                        st.write(f"**cells_written**: {cells_written}")
                        if selected_sources:
                            st.write(f"**selected_sources**: {selected_sources}")
                        st.write(f"**fill_plan.target**: {target_info}")
                        if row_writes:
                            st.write(f"**row_writes[0].start_cell**: {row_writes[0].get('start_cell')}")
                        if clear_ranges:
                            st.write(f"**clear_ranges[0]**: {clear_ranges[0]}")
                    
                    row_writes_count = len(fill_plan.get("row_writes", []))
                    writes_count = len(fill_plan.get("writes", []))
                    clear_ranges_count = len(fill_plan.get("clear_ranges", []))
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Row Writes", row_writes_count)
                    with col2:
                        st.metric("Cell Writes", writes_count)
                    with col3:
                        st.metric("Clear Ranges", clear_ranges_count)
                    
                    if row_writes_count > 0:
                        for idx, row_write in enumerate(fill_plan.get("row_writes", [])):
                            with st.expander(f"Row Write {idx + 1}: {row_write.get('start_cell', 'N/A')}", expanded=False):
                                st.write(f"**Start Cell**: {row_write.get('start_cell', 'N/A')}")
                                st.write(f"**Rows Count**: {len(row_write.get('rows', []))}")
                                st.write(f"**Column Mapping**: {row_write.get('column_mapping', {})}")
                                if len(row_write.get('rows', [])) > 0:
                                    st.write("**First Row Data**:")
                                    st.json(row_write.get('rows', [])[0] if row_write.get('rows') else {})
                    
                    if writes_count > 0:
                        with st.expander(f"Cell Writes ({writes_count} cells)", expanded=False):
                            st.json(fill_plan.get("writes", [])[:10])
                            if writes_count > 10:
                                st.info(f"... and {writes_count - 10} more cells")
                    
                    with st.expander("Full Fill Plan JSON", expanded=False):
                        st.json(fill_plan)
                    
                    if fill_plan.get("warnings"):
                        st.warning("Warnings: " + "; ".join(fill_plan["warnings"]))
            elif st.session_state.final_json is None:
                st.info("Upload template and click 'Fill Template' to generate output.")
                
                if st.button("运行映射 (Legacy)"):
                    with st.spinner("正在运行映射..."):
                        try:
                            settings = get_settings()
                            llm = get_llm_client()  # Use singleton
                            prompts = get_prompts()
                            
                            eval_dir = Path(__file__).parent.parent / "eval" / "fixtures"
                            target_schema_csv = eval_dir / "target_schema_sample.csv"
                            
                            if target_schema_csv.exists():
                                import csv
                                header_rows = []
                                sample_rows = []
                                
                                with open(target_schema_csv, 'r', encoding='utf-8') as f:
                                    reader = csv.reader(f)
                                    rows = list(reader)
                                    if len(rows) >= 2:
                                        header_rows = [rows[0], rows[1]]
                                    if len(rows) >= 4:
                                        sample_rows = [rows[2], rows[3]]
                                
                                attribute_set = build_attribute_set(header_rows, sample_rows, llm, prompts)
                                st.session_state.final_json = map_to_schema(st.session_state.ir, attribute_set, None, llm, prompts)
                                update_ir_scores(st.session_state.ir, st.session_state.final_json, attribute_set)
                                st.success("Mapping completed!")
                                st.rerun()
                            else:
                                st.error("target_schema_sample.csv not found.")
                        except Exception as e:
                            st.error(f"Mapping failed: {e}")
            else:
                st.subheader("Final JSON Output")
                st.json(st.session_state.final_json)
                
                if st.session_state.ir.scores:
                    st.divider()
                    st.subheader("Validation Results")
                    
                    scores = st.session_state.ir.scores
                    st.metric("Constraint Pass Rate", f"{scores.get('constraint_pass_rate', 0):.2%}")
                    st.metric("Needs Human Review", "Yes" if scores.get('needs_human_review', False) else "No")
                    
                    if scores.get('cells'):
                        cells_data = []
                        for cell in scores['cells']:
                            cells_data.append({
                                "Name": cell.get('name', ''),
                                "Value": json.dumps(cell.get('value', ''), ensure_ascii=False) if cell.get('value') is not None else '',
                                "OK": "✅" if cell.get('ok', False) else "❌",
                                "Reason": cell.get('reason', '') or ''
                            })
                        
                        df = pd.DataFrame(cells_data)
                        st.dataframe(_coerce_df_for_display(df), use_container_width=True, hide_index=True)
