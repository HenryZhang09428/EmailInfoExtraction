"""
ExcelReader: low-level file I/O and sheet management.

Encapsulates all the complexity of:
- pandas vs openpyxl vs xlrd engine selection
- Header-row-loss fallback (row-index-zero bug)
- Sheet listing and best-sheet heuristic
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from core.extractors.excel.config import ExtractorConfig, DEFAULT_CONFIG
from core.extractors.excel.data_cleaner import DataCleaner
from core.extractors.excel.header_detector import HeaderDetector
from core.logger import get_logger

logger = get_logger(__name__)


class ExcelReader:
    """
    Load Excel data into a ``pandas.DataFrame`` with automatic fallback
    strategies when the primary engine produces broken output.
    """

    def __init__(
        self,
        cfg: ExtractorConfig = DEFAULT_CONFIG,
        header_detector: Optional[HeaderDetector] = None,
    ):
        self._cfg = cfg
        self._hd = header_detector or HeaderDetector(cfg)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def choose_best_sheet(
        self,
        file_path: str,
        preferred_sheet: Optional[str] = None,
    ) -> Tuple[str, Dict[str, Any]]:
        """
        Inspect all sheets and return the name of the most promising one.

        Scoring favours sheets that contain a clear header row; ties
        are broken by non-empty cell count.
        """
        suffix = Path(file_path).suffix.lower()
        if suffix == ".xls":
            return self._choose_best_sheet_xls(file_path, preferred_sheet)
        return self._choose_best_sheet_xlsx(file_path, preferred_sheet)

    def read_df(
        self,
        file_path: str,
        sheet_name: Any = 0,
        warnings: Optional[List[str]] = None,
    ) -> Tuple[pd.DataFrame, str, bool, bool]:
        """
        Read a single sheet into a DataFrame, applying fallback when needed.

        Returns ``(df, backend_label, fallback_used, row_index_zero_detected)``.
        """
        row_index_zero_detected = False
        suffix = Path(file_path).suffix.lower()
        df, backend, df_warnings = self._extract_sheet_df(file_path, sheet_name, suffix)
        if warnings is not None and df_warnings:
            warnings.extend(df_warnings)
        fallback_used = False

        if suffix == ".xls":
            return df, backend, fallback_used, row_index_zero_detected

        # --- Fallback 1: re-read with header=None if top row looks empty ---
        if len(df) > 0:
            top_all_empty = all(DataCleaner.cell_to_str(c) == "" for c in df.iloc[0].tolist())
            next_has_data = False
            for idx in range(1, min(6, len(df))):
                feats = self._hd.extract_row_features(
                    [DataCleaner.cell_to_str(c) for c in df.iloc[idx].tolist()]
                )
                if feats.get("non_empty_ratio", 0.0) > 0.2:
                    next_has_data = True
                    break
            if top_all_empty and next_has_data:
                try:
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=None,
                        keep_default_na=False,
                        engine="openpyxl",
                        engine_kwargs={"read_only": False, "data_only": False},
                    )
                except TypeError:
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=None,
                        keep_default_na=False,
                        engine="openpyxl",
                    )
                backend = "pandas_openpyxl_full"
                fallback_used = True

        # --- Fallback 2: full openpyxl values when row-0 is still empty ---
        if df.shape[0] >= 2:
            row0 = df.iloc[0].tolist()
            empty_ratio = sum(1 for c in row0 if DataCleaner.cell_to_str(c) == "") / max(1, len(row0))
            row1_feats = self._hd.extract_row_features(
                [DataCleaner.cell_to_str(c) for c in df.iloc[1].tolist()]
            )
            if empty_ratio > 0.95 and row1_feats.get("non_empty_ratio", 0.0) > 0.3:
                try:
                    resolved = self._resolve_sheet_name(file_path, sheet_name)
                    df = self._read_openpyxl_values(
                        file_path, str(resolved), self._cfg.max_rows_to_process + 20,
                    )
                    backend = "openpyxl_values_full"
                    fallback_used = True
                    if warnings is not None:
                        warnings.append("excel_pandas_openpyxl_header_row_lost_fallback_openpyxl_values")
                    if df.shape[0] >= 2:
                        r0_empty = all(DataCleaner.cell_to_str(c) == "" for c in df.iloc[0].tolist())
                        later_data = any(
                            any(DataCleaner.cell_to_str(c) != "" for c in df.iloc[r].tolist())
                            for r in range(1, min(df.shape[0], 10))
                        )
                        if r0_empty and later_data:
                            row_index_zero_detected = True
                except Exception:
                    pass

        return df, backend, fallback_used, row_index_zero_detected

    # ------------------------------------------------------------------
    # Smart multi-sheet selection (automatic)
    # ------------------------------------------------------------------

    # Very low bar — we'd rather over-extract than miss useful data.
    _DEFAULT_MIN_USEFUL_CELLS = 3

    def select_useful_sheets(
        self,
        file_path: str,
        min_non_empty: int = _DEFAULT_MIN_USEFUL_CELLS,
    ) -> Tuple[List[str], Dict[str, Any]]:
        """
        Scan all sheets and return those that likely contain useful data.

        Uses the same cheap heuristic as ``choose_best_sheet`` (first
        30 rows × 80 cols) but returns **all** sheets that pass a lenient
        content threshold, rather than picking just one winner.

        A sheet is kept if:

        - It has a detected header-like row (``best_header_like_score``
          is not ``None``), **OR**
        - It has at least *min_non_empty* non-empty cells in the scan
          window.

        The threshold is deliberately very low so that we prefer
        over-extraction over missing any sheet that contains real data.

        Returns ``(useful_sheet_names, selection_debug)``.
        """
        best_name, debug = self.choose_best_sheet(file_path, preferred_sheet=None)
        sheets_info = debug.get("sheets", [])

        useful: List[str] = []
        skipped: List[str] = []

        for s in sheets_info:
            name = s.get("sheet_name")
            if not isinstance(name, str):
                continue
            header_score = s.get("best_header_like_score")
            non_empty = s.get("non_empty_cells_count", 0) or 0

            has_header = header_score is not None
            has_content = non_empty >= min_non_empty

            if has_header or has_content:
                useful.append(name)
            else:
                skipped.append(name)

        # If nothing passed (shouldn't happen often), keep the best one.
        if not useful and best_name:
            useful = [best_name]

        # Sort: sheets with a header first (by score desc), then by
        # non_empty desc.  This way the "best" sheet is always first.
        sheet_lookup = {
            s["sheet_name"]: s
            for s in sheets_info
            if isinstance(s.get("sheet_name"), str)
        }

        def _sort_key(name: str) -> tuple:
            info = sheet_lookup.get(name, {})
            hs = info.get("best_header_like_score")
            ne = info.get("non_empty_cells_count", 0) or 0
            return (
                1 if hs is not None else 0,
                float(hs) if hs is not None else -1e18,
                ne,
            )

        useful.sort(key=_sort_key, reverse=True)

        selection_debug = {
            "useful_sheets": useful,
            "skipped_sheets": skipped,
            "best_sheet": best_name,
            "scan_details": sheets_info,
            "min_non_empty_threshold": min_non_empty,
        }

        logger.info(
            "Sheet selection: %d useful / %d total (useful=%s, skipped=%s)",
            len(useful), len(sheets_info), useful, skipped,
        )

        return useful, selection_debug

    # ------------------------------------------------------------------

    def list_sheet_names(self, file_path: str) -> Tuple[List[str], str]:
        suffix = Path(file_path).suffix.lower()
        if suffix == ".xls":
            import xlrd
            wb = xlrd.open_workbook(file_path)
            return wb.sheet_names(), "xlrd"
        wb = load_workbook(file_path, read_only=True, data_only=False)
        return wb.sheetnames or [], "openpyxl"

    @staticmethod
    def xlsx_has_row_index_zero(file_path: str) -> bool:
        """Fast XML-level check for a ``<row r="0">`` element."""
        try:
            with zipfile.ZipFile(file_path, "r") as zf:
                for name in zf.namelist():
                    if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                        continue
                    with zf.open(name) as fh:
                        if b'<row r="0"' in fh.read():
                            return True
        except Exception:
            return False
        return False

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_sheet_df(
        file_path: str,
        sheet_name: Any,
        suffix: str,
    ) -> Tuple[pd.DataFrame, str, List[str]]:
        if suffix == ".xls":
            df = pd.read_excel(
                file_path, sheet_name=sheet_name,
                header=None, engine="xlrd", keep_default_na=False,
            )
            return df, "pandas_xlrd", []
        df = pd.read_excel(
            file_path, sheet_name=sheet_name,
            header=None, engine="openpyxl", keep_default_na=False,
        )
        return df, "pandas_openpyxl", []

    @staticmethod
    def _read_openpyxl_values(
        file_path: str,
        sheet_name: str,
        max_rows: Optional[int] = None,
    ) -> pd.DataFrame:
        wb = load_workbook(file_path, data_only=False, read_only=False)
        try:
            ws = wb[sheet_name]
            mr = min(ws.max_row, max_rows) if max_rows else ws.max_row
            rows = list(ws.iter_rows(min_row=1, max_row=mr, max_col=ws.max_column, values_only=True))
            return pd.DataFrame(rows)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    @staticmethod
    def _resolve_sheet_name(file_path: str, sheet_name: Any) -> Any:
        if not isinstance(sheet_name, int):
            return sheet_name
        wb = load_workbook(file_path, data_only=False, read_only=False)
        try:
            if 0 <= sheet_name < len(wb.sheetnames):
                return wb.sheetnames[sheet_name]
            return wb.sheetnames[0]
        finally:
            try:
                wb.close()
            except Exception:
                pass

    # --- Best-sheet heuristics (xlsx / xls) --------------------------------

    def _choose_best_sheet_xlsx(
        self,
        file_path: str,
        preferred_sheet: Optional[str],
    ) -> Tuple[str, Dict[str, Any]]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        names = wb.sheetnames or []
        debug: Dict[str, Any] = {"sheets": [], "chosen_reason": "", "selected_sheet_name": ""}
        if not names:
            return "Sheet1", debug
        if preferred_sheet and preferred_sheet in names:
            debug["chosen_reason"] = "profile_sheet"
            debug["selected_sheet_name"] = preferred_sheet
            return preferred_sheet, debug
        if preferred_sheet and preferred_sheet not in names:
            debug["profile_sheet"] = preferred_sheet
            debug["profile_sheet_not_found"] = True
        for sn in names:
            ws = wb[sn]
            best_score: Optional[float] = None
            non_empty = 0
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=80, values_only=True):
                row_str = [DataCleaner.cell_to_str(c) for c in (list(row) if row else [])]
                feats = self._hd.extract_row_features(row_str)
                if self._hd.is_header_like(feats):
                    s = self._hd.header_score(feats)
                    if best_score is None or s > best_score:
                        best_score = s
                for c in (list(row) if row else []):
                    if not DataCleaner.is_empty(c):
                        non_empty += 1
            debug["sheets"].append({
                "sheet_name": sn,
                "best_header_like_score": best_score,
                "non_empty_cells_count": non_empty,
            })
        return self._pick_from_debug(debug, names)

    def _choose_best_sheet_xls(
        self,
        file_path: str,
        preferred_sheet: Optional[str],
    ) -> Tuple[str, Dict[str, Any]]:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        names = wb.sheet_names() or []
        debug: Dict[str, Any] = {"sheets": [], "chosen_reason": "", "selected_sheet_name": ""}
        if not names:
            return "Sheet1", debug
        if preferred_sheet and preferred_sheet in names:
            debug["chosen_reason"] = "profile_sheet"
            debug["selected_sheet_name"] = preferred_sheet
            return preferred_sheet, debug
        if preferred_sheet and preferred_sheet not in names:
            debug["profile_sheet"] = preferred_sheet
            debug["profile_sheet_not_found"] = True
        for sn in names:
            ws = wb.sheet_by_name(sn)
            best_score: Optional[float] = None
            non_empty = 0
            for ri in range(min(30, ws.nrows)):
                row_str = [DataCleaner.cell_to_str(c) for c in ws.row_values(ri, 0, min(80, ws.ncols))]
                feats = self._hd.extract_row_features(row_str)
                if self._hd.is_header_like(feats):
                    s = self._hd.header_score(feats)
                    if best_score is None or s > best_score:
                        best_score = s
                for c in ws.row_values(ri, 0, min(80, ws.ncols)):
                    if not DataCleaner.is_empty(c):
                        non_empty += 1
            debug["sheets"].append({
                "sheet_name": sn,
                "best_header_like_score": best_score,
                "non_empty_cells_count": non_empty,
            })
        return self._pick_from_debug(debug, names)

    @staticmethod
    def _pick_from_debug(debug: Dict[str, Any], names: List[str]) -> Tuple[str, Dict[str, Any]]:
        with_header = [s for s in debug["sheets"] if s["best_header_like_score"] is not None]
        if with_header:
            with_header.sort(key=lambda s: (-s["best_header_like_score"], names.index(s["sheet_name"])))
            chosen = with_header[0]
            debug["chosen_reason"] = "best_header_like_score"
        else:
            debug["sheets"].sort(key=lambda s: (-s["non_empty_cells_count"], names.index(s["sheet_name"])))
            chosen = debug["sheets"][0]
            debug["chosen_reason"] = "non_empty_cells_count"
        debug["selected_sheet_name"] = chosen["sheet_name"]
        return chosen["sheet_name"], debug
