"""
Excel 提取器模块 (Excel Extractor Module)
=========================================

ExcelExtractor：Excel 文件抽取的高层编排器。

整体流程（从文件到结构化输出）：
Reader → HeaderDetector → SchemaMapper → DataCleaner → SourceDoc

复杂逻辑被下沉到 `core.extractors.excel` 子包内的专用类，本模块负责：
- 组织调用顺序与异常兜底
- 组装抽取结果的元信息（metadata）、告警（warnings）与调试用逗号分隔样本文本
- 保持对外 `ExcelExtractor.extract()` / `safe_extract()` 接口稳定

职责划分（按模块）：
- `ExcelReader`：文件读取、工作表选择、读取失败的回退策略
- `HeaderDetector`：行特征提取、表头识别与布局选择
- `SchemaMapper`：语义键推断（LLM + 规则）、冲突消解、覆盖率评估与兜底映射
- `DataCleaner`：单元格转字符串、值归一化、表头规范化
- `ExtractorConfig`：可调参数集中管理
"""

from __future__ import annotations

import json
import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from openpyxl import load_workbook  # noqa: F401（保留该导入，便于测试时动态替换/打桩）

from core.extractors.base import BaseExtractor
from core.extractors.excel.config import ExtractorConfig, DEFAULT_CONFIG
from core.extractors.excel.data_cleaner import DataCleaner
from core.extractors.excel.header_detector import HeaderDetector
from core.extractors.excel.reader import ExcelReader
from core.extractors.excel.schema_mapper import SchemaMapper
from core.ir import SourceDoc, SourceBlock, BlockType
from core.llm import LLMClient
from core.logger import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# 向后兼容：测试用例可能直接从本模块导入的符号别名
# ---------------------------------------------------------------------------
_cell_to_str = DataCleaner.cell_to_str
deterministic_header_map = SchemaMapper.deterministic_header_map
_normalize_value = DataCleaner.normalize_value
_parse_month_start = DataCleaner.parse_month_start

# 重新导出：其它模块（例如 `social_security.py`）依赖的独立辅助函数
_normalize_header_compact = DataCleaner.normalize_header_compact
_normalize_header_for_semantic_key = DataCleaner.normalize_header_for_semantic_key
_normalize_reason_text = DataCleaner.normalize_header_compact
_infer_remove_intent = SchemaMapper._is_remove_intent
_is_empty_cell = DataCleaner.is_empty
_extract_row_features = HeaderDetector.extract_row_features
_header_score = HeaderDetector(DEFAULT_CONFIG).header_score
_is_header_like_row = HeaderDetector(DEFAULT_CONFIG).is_header_like
_looks_like_header_row = HeaderDetector.looks_like_header_row
_infer_key_from_header = SchemaMapper.infer_key_from_header

# 配置常量（向后兼容：避免外部引用的常量名失效）
MAX_ROWS_TO_PROCESS = DEFAULT_CONFIG.max_rows_to_process
MAX_RECORDS_PER_WORKBOOK = DEFAULT_CONFIG.max_records_per_workbook
MAX_CSV_CHARS = DEFAULT_CONFIG.max_csv_chars
SAMPLE_HEAD_ROWS = DEFAULT_CONFIG.sample_head_rows
SAMPLE_SPREAD_ROWS = DEFAULT_CONFIG.sample_spread_rows
SAMPLE_MAX_PER_COLUMN = DEFAULT_CONFIG.sample_max_per_column
SCHEMA_INFER_COVERAGE_THRESHOLD = DEFAULT_CONFIG.schema_infer_coverage_threshold
FAST_ROWS_THRESHOLD = DEFAULT_CONFIG.fast_rows_threshold
MIN_ROWS_FOR_SAMPLING = DEFAULT_CONFIG.min_rows_for_sampling
_REMOVE_INTENT_KEYWORDS = ("减员", "离职", "退工", "退保", "终止", "停保")
_TERMINATION_KEYS = ("termination_date", "terminationdate", "leave_date", "leavedate", "end_date", "enddate")

# 向后兼容：测试可能会动态替换本模块的全局函数，因此保留这些薄封装。
def _list_sheet_names(file_path: str) -> Tuple[List[str], str]:
    """
    列出 Excel 文件中的工作表名称。

    返回：
    - sheet_names：工作表名列表
    - backend：实际使用的读取后端标识（例如 `xlrd` / `openpyxl`）
    """
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        return wb.sheet_names(), "xlrd"
    wb = load_workbook(file_path, read_only=True, data_only=False)
    return wb.sheetnames or [], "openpyxl"


def _extract_sheet_df(
    file_path: str,
    sheet_name: Any,
    suffix: str,
) -> Tuple[pd.DataFrame, str, List[str]]:
    """
    读取指定工作表为数据帧（不设表头行），并返回读取后端信息。

    说明：
    - `.xls` 使用 `xlrd`
    - `.xlsx/.xlsm` 使用 `openpyxl`

    返回：
    - df：读取到的数据帧（header=None）
    - backend：读取后端标识（例如 `pandas_xlrd` / `pandas_openpyxl`）
    - warnings：预留告警列表（与旧逻辑保持一致）
    """
    if suffix == ".xls":
        df = pd.read_excel(
            file_path, sheet_name=sheet_name,
            header=None, engine="xlrd", keep_default_na=False,
        )
        return df, "pandas_xlrd", []
    df = pd.read_excel(
        file_path, sheet_name=sheet_name,
        header=None, engine="openpyxl", keep_default_na=False,
    )
    return df, "pandas_openpyxl", []


# ---------------------------------------------------------------------------
# 采样辅助函数（用于结构推断的输入构造与调试样本生成）
# ---------------------------------------------------------------------------

def _evenly_spaced_indices(start: int, end_exclusive: int, count: int) -> List[int]:
    """
    在区间 `[start, end_exclusive)` 内选取近似等间隔的若干索引。

    用途：
    - 生成“分布采样行”索引，避免只看表头附近的少量行导致结构推断偏差。
    """
    if count <= 0 or end_exclusive <= start:
        return []
    if count == 1:
        return [start]
    span = end_exclusive - start - 1
    if span <= 0:
        return [start]
    step = span / (count - 1)
    indices = [start + int(round(i * step)) for i in range(count)]
    return sorted(set(idx for idx in indices if start <= idx < end_exclusive))


def _build_column_summaries(
    df: pd.DataFrame,
    header1: List[Any],
    header2: List[Any],
    data_start_idx: int,
    head_rows: int = SAMPLE_HEAD_ROWS,
    spread_rows: int = SAMPLE_SPREAD_ROWS,
    max_samples: int = SAMPLE_MAX_PER_COLUMN,
) -> Tuple[List[Dict[str, Any]], List[str], int]:
    """
    构造“按列”的采样摘要，用于结构推断（LLM 输入）与覆盖率评估。

    主要输出：
    - summaries：每列的 header_path、非空占比、去重数量、样本值列表等
    - header_paths：每列规范化后的 header_path
    - data_rows：数据行数量（用于截断与诊断）
    """
    total_rows = len(df)
    num_cols = df.shape[1] if total_rows > 0 else max(len(header1), len(header2))
    header1 = list(header1) + [""] * (num_cols - len(header1))
    header2 = list(header2) + [""] * (num_cols - len(header2))

    data_rows = max(0, total_rows - data_start_idx)
    head_indices = list(range(data_start_idx, min(total_rows, data_start_idx + head_rows)))
    spread_indices = _evenly_spaced_indices(data_start_idx, total_rows, spread_rows)
    sample_indices = sorted(set(head_indices + spread_indices))

    header_paths: List[str] = []
    summaries: List[Dict[str, Any]] = []

    for col_idx in range(num_cols):
        hp = HeaderDetector.make_header_path(header1[col_idx], header2[col_idx], f"col_{col_idx + 1}")
        header_paths.append(hp)

        samples: List[str] = []
        for idx in sample_indices:
            v = DataCleaner.cell_to_str(df.iat[idx, col_idx])
            if v and v not in samples:
                samples.append(v)
            if len(samples) >= max_samples:
                break

        unique_count = None
        non_empty_ratio = 0.0
        if data_rows > 0:
            sample_end = min(total_rows, data_start_idx + MAX_ROWS_TO_PROCESS)
            col_values = df.iloc[data_start_idx:sample_end, col_idx].tolist()
            non_empty_values = [DataCleaner.cell_to_str(v) for v in col_values if not DataCleaner.is_empty(v)]
            non_empty_ratio = len(non_empty_values) / max(1, len(col_values))
            if non_empty_values:
                unique_count = len(set(non_empty_values))

        summaries.append({
            "column_index": col_idx + 1,
            "header_path": hp,
            "non_empty_ratio": round(non_empty_ratio, 4),
            "unique_count": unique_count,
            "samples": samples,
        })

    return summaries, header_paths, data_rows


def _apply_row_filter(
    record: Dict[str, str],
    row_values: List[str],
    row_filter: Dict[str, Any],
) -> bool:
    """
    应用 LLM 给出的“行过滤规则”，判断当前行记录是否保留。

    说明：
    - 该过滤规则通常由结构推断阶段返回，用于丢弃空行、说明行、汇总行等无效数据；
    - 过滤逻辑尽量保持确定性：不依赖外部状态，仅基于当前行值与已组装的 record。
    """
    if not row_values or all(v == "" for v in row_values):
        return False
    min_ratio = row_filter.get("min_nonempty_ratio")
    if isinstance(min_ratio, (int, float)):
        ratio = sum(1 for v in row_values if v != "") / max(1, len(row_values))
        if ratio < float(min_ratio):
            return False
    required_any = row_filter.get("required_fields_any") or []
    if isinstance(required_any, list) and required_any:
        if not any(record.get(f, "") != "" for f in required_any):
            return False
    exclude_terms = row_filter.get("exclude_if_contains_any") or []
    if isinstance(exclude_terms, list) and exclude_terms:
        lowered = [v.lower() for v in row_values if v]
        for term in exclude_terms:
            if not term:
                continue
            t = str(term).lower()
            if any(t in v for v in lowered):
                return False
    return True


# ---------------------------------------------------------------------------
# ExcelExtractor（Excel 抽取器）
# ---------------------------------------------------------------------------

class ExcelExtractor(BaseExtractor):
    """
    Excel 抽取的高层编排器。

    对外接口：
    - `extract()`：主抽取入口（支持单表/全表）
    - `safe_extract()`：带异常兜底的抽取入口（失败时返回错误版 `SourceDoc`）

    内部实现会委派给：
    - `ExcelReader`：读取与选表
    - `HeaderDetector`：表头检测
    - `SchemaMapper`：语义键推断与覆盖率评估
    - `DataCleaner`：值归一化与表头规范化
    """

    def __init__(
        self,
        llm: LLMClient,
        prompts: Optional[dict] = None,
        header_force_map: Optional[Dict[str, str]] = None,
    ):
        super().__init__(llm, prompts)
        self._cfg = DEFAULT_CONFIG
        self._cleaner = DataCleaner()
        self._header = HeaderDetector(self._cfg)
        self._reader = ExcelReader(self._cfg, self._header)
        self._mapper = SchemaMapper(self._cfg)
        self._header_force_map = header_force_map if isinstance(header_force_map, dict) else {}

    # ------------------------------------------------------------------
    # 对外接口（保持稳定）
    # ------------------------------------------------------------------

    def safe_extract(
        self,
        file_path: str,
        extract_all_sheets: bool = False,
        preferred_sheet: Optional[str] = None,
        sheet_names: Optional[List[str]] = None,
    ) -> SourceDoc:
        """
        安全抽取入口：捕获内部异常，保证始终返回 `SourceDoc`。

        参数：
        - file_path：Excel 文件路径
        - extract_all_sheets：是否抽取所有工作表并合并到一个结果中
        - preferred_sheet：指定优先抽取的工作表名称（找不到会回退到自动选表）
        - sheet_names：当 `extract_all_sheets=True` 时，可指定要遍历的工作表名称列表（用于过滤无用 sheet 或 Top-K 选择）
        """
        try:
            return self.extract(
                file_path,
                extract_all_sheets=extract_all_sheets,
                preferred_sheet=preferred_sheet,
                sheet_names=sheet_names,
            )
        except Exception as e:
            logger.error("Extraction failed for %s: %s", file_path, e, exc_info=True)
            return self._create_error_source_doc(file_path, e)

    def extract(
        self,
        file_path: str,
        extract_all_sheets: bool = False,
        preferred_sheet: Optional[str] = None,
        sheet_names: Optional[List[str]] = None,
    ) -> SourceDoc:
        """
        抽取 Excel 文件内容并返回 `SourceDoc`。

        行为说明：
        - 默认只抽取“最佳工作表”（由 `ExcelReader.choose_best_sheet()` 决定）；
        - 当 `extract_all_sheets=True` 时，会遍历所有工作表并合并记录（受最大记录数限制）。
          - 若传入 `sheet_names`，则只遍历该列表中的工作表（可用于过滤无用 sheet 或 Top-K 选择）。

        返回结果包含：
        - `TABLE_CSV`：抽样的逗号分隔文本（用于调试/可观测，可能为空）
        - `EXTRACTED_JSON`：结构化数据（`data`）+ 元信息（`metadata`）+ 告警（`warnings`）
        """
        self.clear_derived_files()
        filename = Path(file_path).name
        source_id = str(uuid.uuid4())
        selected_sheet, sheet_debug = self._reader.choose_best_sheet(file_path, preferred_sheet)

        top_warnings: List[str] = []
        if sheet_debug.get("profile_sheet_not_found"):
            top_warnings.append("sheet_not_found_fallback_to_auto")

        # Resolve which sheets to iterate when extract_all_sheets is requested.
        resolved_sheet_names: Optional[List[str]] = None
        if extract_all_sheets:
            # Callers may provide an explicit sheet list (e.g., from
            # select_useful_sheets); otherwise fall back to all sheets.
            provided = [
                str(s).strip()
                for s in (sheet_names or [])
                if isinstance(s, (str, int, float)) and str(s).strip()
            ]
            if provided:
                # Validate against actual workbook sheet names.
                try:
                    actual_names, _backend = self._reader.list_sheet_names(file_path)
                    actual_set = set(actual_names or [])
                    filtered = [sn for sn in provided if sn in actual_set]
                    missing = [sn for sn in provided if sn not in actual_set]
                    if missing:
                        top_warnings.append(f"sheet_names_filtered_missing: {missing[:10]}")
                    resolved_sheet_names = filtered
                except Exception as e:
                    top_warnings.append(f"sheet_names_read_failed: {e}")
                    resolved_sheet_names = provided
            else:
                try:
                    resolved_sheet_names, _ = self._reader.list_sheet_names(file_path)
                except Exception as e:
                    top_warnings.append(f"sheet_names_read_failed: {e}")
            if not resolved_sheet_names:
                extract_all_sheets = False

        if not extract_all_sheets:
            return self._single_sheet(file_path, selected_sheet, filename, source_id, sheet_debug, top_warnings)
        return self._all_sheets(file_path, resolved_sheet_names, selected_sheet, filename, source_id, sheet_debug, top_warnings)

    # ------------------------------------------------------------------
    # 单工作表路径
    # ------------------------------------------------------------------

    def _single_sheet(
        self,
        file_path: str,
        sheet_name: Any,
        filename: str,
        source_id: str,
        sheet_debug: Dict[str, Any],
        top_warnings: List[str],
    ) -> SourceDoc:
        """
        抽取单个工作表并直接返回结果。

        说明：
        - 该路径用于默认模式（不合并所有工作表）；
        - 会将选表阶段产生的 `top_warnings` 合并进当前工作表的 `warnings`。
        """
        extracted, csv_text, _, table_meta = self._extract_sheet(
            file_path, sheet_name, filename, source_id,
            sheet_name, sheet_debug, add_sheet_name=False,
        )
        warnings = extracted.get("warnings") or []
        for w in top_warnings:
            if w not in warnings:
                warnings.append(w)
        extracted["warnings"] = warnings
        blocks = self._make_blocks(extracted, csv_text, table_meta)
        return SourceDoc(
            source_id=source_id, filename=filename,
            file_path=file_path, source_type="excel",
            blocks=blocks, extracted=extracted,
        )

    # ------------------------------------------------------------------
    # 全工作表路径（工作簿合并）
    # ------------------------------------------------------------------

    def _all_sheets(
        self,
        file_path: str,
        sheet_names: List[str],
        selected_sheet: str,
        filename: str,
        source_id: str,
        sheet_debug: Dict[str, Any],
        top_warnings: List[str],
    ) -> SourceDoc:
        """
        遍历并抽取所有工作表，将记录合并到一个 `SourceDoc` 中。

        合并策略：
        - 每条记录会补充 `__source_file__` 与 `__sheet_name__` 方便溯源；
        - 受 `max_records_per_workbook` 限制，超过后会截断并在 metadata 中标记原因；
        - 最终返回的 `TABLE_CSV` 与 `table_meta` 默认取“选中工作表”（或第一个成功抽取的工作表），用于调试展示。
        """
        cfg = self._cfg
        combined: List[Dict[str, Any]] = []
        combined_warnings: List[str] = []
        sheet_summaries: List[Dict[str, Any]] = []
        records_by_sheet: Dict[str, int] = {}
        per_sheet: Dict[str, Dict[str, Any]] = {}
        truncated_wb = False
        truncation_reason: Optional[str] = None
        total_seen = 0

        selected_extracted: Optional[Dict[str, Any]] = None
        selected_csv: Optional[str] = None
        selected_meta: Optional[Dict[str, Any]] = None
        first_extracted: Optional[Dict[str, Any]] = None
        first_csv: Optional[str] = None
        first_meta: Optional[Dict[str, Any]] = None

        for sn in sheet_names:
            s_ext, csv_text, s_sum, t_meta = self._extract_sheet(
                file_path, sn, filename, source_id,
                selected_sheet, sheet_debug, add_sheet_name=True,
            )
            data = s_ext.get("data")
            sheet_count = 0
            skipped = 0
            filtered_demo_records_count = 0
            if isinstance(data, list):
                for rec in data:
                    if not isinstance(rec, dict):
                        continue
                    if self._should_filter_demo_record(str(sn), rec):
                        filtered_demo_records_count += 1
                        continue
                    total_seen += 1
                    if cfg.max_records_per_workbook and len(combined) >= cfg.max_records_per_workbook:
                        truncated_wb = True
                        truncation_reason = "workbook_records_limit"
                        skipped += 1
                        continue
                    rec.setdefault("__source_file__", filename)
                    rec.setdefault("__sheet_name__", sn)
                    combined.append(rec)
                    sheet_count += 1

            records_by_sheet[str(sn)] = sheet_count
            s_meta = s_ext.get("metadata") if isinstance(s_ext, dict) else {}
            s_warn = list(s_ext.get("warnings") or []) if isinstance(s_ext, dict) else []

            no_reason: Optional[str] = None
            if sheet_count == 0:
                dr = (s_meta or {}).get("data_rows")
                hi = (s_meta or {}).get("header_row_idx")
                if skipped > 0:
                    no_reason = "workbook_records_limit_reached"
                elif "schema_infer_failed" in s_warn:
                    no_reason = "schema_infer_failed_and_no_fallback"
                elif "no_header_row_detected" in s_warn and (dr == 0 or hi is None or hi < 0):
                    no_reason = "no_header_detected"
                elif isinstance(dr, int) and dr == 0:
                    no_reason = "no_data_rows"
                elif isinstance(dr, int) and dr > 0:
                    no_reason = "filtered_all_rows"
                else:
                    no_reason = "read_failed"

            per_sheet[str(sn)] = {
                "header_row_idx": (s_meta or {}).get("header_row_idx"),
                "data_start_idx": (s_meta or {}).get("data_start_idx"),
                "data_rows": (s_meta or {}).get("data_rows"),
                "coverage": (s_meta or {}).get("coverage"),
                "semantic_key_by_header": (s_meta or {}).get("semantic_key_by_header"),
                "warnings": s_warn,
                "records_count": sheet_count,
                "read_backend": (s_meta or {}).get("read_backend"),
                "no_records_reason": no_reason,
                "filtered_demo_records_count": filtered_demo_records_count,
            }
            combined_warnings.extend(s_warn)
            sheet_summaries.append(s_sum)

            if first_extracted is None:
                first_extracted, first_csv, first_meta = s_ext, csv_text, t_meta
            if sn == selected_sheet:
                selected_extracted, selected_csv, selected_meta = s_ext, csv_text, t_meta

        if selected_extracted is None:
            selected_extracted = first_extracted or {"data": [], "metadata": {}, "warnings": []}
            selected_csv = first_csv
            selected_meta = first_meta

        wb_meta = {
            "sheets": sheet_summaries,
            "processed_rows": len(combined),
            "records_count_by_sheet": records_by_sheet,
            "per_sheet": per_sheet,
            "truncated_workbook_level": truncated_wb,
            "workbook_records_seen": total_seen,
            "selected_sheet_name": selected_sheet,
            "sheet_selection_debug": sheet_debug,
        }
        if cfg.max_records_per_workbook:
            wb_meta["workbook_records_limit"] = cfg.max_records_per_workbook
        if truncated_wb and truncation_reason:
            wb_meta["truncation_reason"] = truncation_reason

        extracted = {
            "data": combined,
            "metadata": wb_meta,
            "warnings": list(selected_extracted.get("warnings") or []),
        }
        warnings = extracted["warnings"]
        for w in combined_warnings + top_warnings:
            if w not in warnings:
                warnings.append(w)

        blocks = self._make_blocks(extracted, selected_csv, selected_meta)
        return SourceDoc(
            source_id=source_id, filename=filename,
            file_path=file_path, source_type="excel",
            blocks=blocks, extracted=extracted,
        )

    # ------------------------------------------------------------------
    # 单工作表抽取核心流程
    # ------------------------------------------------------------------

    def _extract_sheet(
        self,
        file_path: str,
        sheet_name: Any,
        filename: str,
        source_id: str,
        selected_sheet_name: str,
        sheet_debug: Dict[str, Any],
        add_sheet_name: bool,
    ) -> Tuple[Dict[str, Any], Optional[str], Dict[str, Any], Optional[Dict[str, Any]]]:
        """
        抽取单个工作表。

        核心流水线：
        读取 → 表头检测 → 语义键推断（LLM + 规则）→ 值归一化 → 生成记录列表

        返回值（四元组）：
        - extracted：结构化抽取结果（`data`/`metadata`/`warnings`）
        - csv_text：用于调试展示的抽样逗号分隔文本（可能为 None）
        - sheet_sum：该工作表的摘要信息（用于工作簿级合并时的汇总）
        - table_meta：`TABLE_CSV` block 的 meta（例如是否截断、抽样行数等；可能为 None）
        """
        cfg = self._cfg
        hd = self._header
        mapper = self._mapper
        warnings: List[str] = []

        # ---- 1）读取：将工作表读入数据帧，并记录读取后端/回退情况 ----
        df, backend, read_fb, row0_det = self._reader.read_df(file_path, sheet_name, warnings)
        if read_fb and "excel_read_fallback_full_openpyxl" not in warnings:
            warnings.append("excel_read_fallback_full_openpyxl")
        if row0_det and "xlsx_row_index_zero_detected" not in warnings:
            warnings.append("xlsx_row_index_zero_detected")
        total_rows = len(df)

        # ---- 2）表头检测：定位表头行/表头布局，并推断数据起始行 ----
        h_idx, h_debug = hd.select_header_row_index(df)
        header_mode = "header"
        if h_idx >= 0:
            row_str = [DataCleaner.cell_to_str(c) for c in df.iloc[h_idx].tolist()]
            feats = hd.extract_row_features(row_str)
            if (
                (feats.get("long_digit_ratio", 0.0) > 0.03 or feats.get("date_ratio", 0.0) > 0.03)
                and feats.get("text_ratio", 0.0) >= 0.6
            ):
                h_idx = -1
                header_mode = "no_header"
                if "header_row_looks_like_data_forced_no_header" not in warnings:
                    warnings.append("header_row_looks_like_data_forced_no_header")

        if h_idx < 0:
            if "no_header_row_detected" not in warnings:
                warnings.append("no_header_row_detected")
            header_mode = "no_header"
            num_cols = df.shape[1] if df is not None and len(df.shape) > 1 else 0
            h1 = [f"col_{i}" for i in range(num_cols)]
            h2: List[Any] = []
            ds = hd.first_non_empty_row_idx(df)
        else:
            h1, h2, ds = hd.get_header_rows(df, h_idx, warnings)

        summaries, hpaths, data_rows = _build_column_summaries(df, h1, h2, ds)
        hpaths, summaries = mapper.normalize_header_paths_and_summaries(hpaths, summaries)

        # 若识别出的数据行数为 0：尝试替代布局（例如双表头/错位表头）以提升命中率
        if data_rows == 0 and total_rows >= 2:
            h_idx, h1, h2, ds = hd.pick_header_layout(df, h_idx, _build_column_summaries, warnings)
            summaries, hpaths, data_rows = _build_column_summaries(df, h1, h2, ds)
            hpaths, summaries = mapper.normalize_header_paths_and_summaries(hpaths, summaries)
        if data_rows == 0 and total_rows > 0:
            h_idx = -1
            header_mode = "no_header"
            num_cols = df.shape[1] if df is not None and len(df.shape) > 1 else 0
            h1 = [f"col_{i}" for i in range(num_cols)]
            h2 = []
            ds = hd.first_non_empty_row_idx(df)
            summaries, hpaths, data_rows = _build_column_summaries(df, h1, h2, ds)
            hpaths, summaries = mapper.normalize_header_paths_and_summaries(hpaths, summaries)

        # ---- 3）语义键推断：LLM 推断为主，规则兜底与强制覆盖 ----
        schema_input = {
            "total_rows": total_rows,
            "data_rows": data_rows,
            "columns": len(hpaths),
            "header_row_1": [str(c) for c in h1],
            "header_row_2": [str(c) for c in h2],
            "column_summaries": summaries,
        }
        schema_infer = None
        fallback_used = False
        row_filter = None
        sem_map: Dict[str, str] = {}

        try:
            prompt = self.prompts["EXCEL_SCHEMA_INFER_PROMPT"] + "\n\nINPUT_JSON:\n" + json.dumps(schema_input, ensure_ascii=False)
            schema_infer = self.llm.chat_json(
                prompt, system=None, step="excel_schema_infer",
                filename=filename, source_id=source_id, mode="schema_infer",
            )
        except Exception as e:
            warnings.append(f"schema_infer_failed: {e}")

        if isinstance(schema_infer, dict):
            sem_map = schema_infer.get("semantic_key_by_header") or schema_infer.get("column_semantics") or {}
            if not isinstance(sem_map, dict):
                sem_map = {}
            row_filter = schema_infer.get("row_filter")
            normalization = schema_infer.get("normalization")
            if not isinstance(row_filter, dict):
                row_filter = None
            if not isinstance(normalization, dict):
                normalization = None

        sem_map = mapper.sanitize(sem_map, hpaths)
        sem_map, det_overrides = mapper.apply_deterministic_overrides(sem_map, hpaths)
        coverage = mapper.compute_coverage(sem_map, hpaths)

        if coverage < cfg.schema_infer_coverage_threshold:
            fallback_used = True
            warnings.append("schema_infer_low_coverage")
            sem_map = mapper.fallback_infer(hpaths, summaries)
            sem_map, det_overrides = mapper.apply_deterministic_overrides(sem_map, hpaths)
            coverage = mapper.compute_coverage(sem_map, hpaths)

        sem_map = mapper.resolve_conflicts(sem_map, hpaths, warnings)
        hpaths, sem_map = mapper.apply_forced_mappings(
            hpaths,
            sem_map,
            extra_force_map=self._header_force_map,
        )
        sem_map = mapper.ensure_termination_reason(df, hpaths, sem_map, ds, sheet_name, cfg.max_rows_to_process)
        sem_map = self._sanitize_boolean_id_check_columns(df, hpaths, sem_map, ds, cfg.max_rows_to_process, warnings)

        # ---- 4）覆盖率过低则提前返回：避免生成大量“无键/错键”记录 ----
        if coverage < cfg.schema_infer_coverage_threshold:
            logger.info(
                "Excel mode=fallback_failed | filename=%s | source_id=%s | rows=%d | cols=%d | coverage=%.3f",
                filename, source_id, total_rows, len(hpaths), coverage,
            )
            warnings.append("schema_infer_failed")
            ext = self._build_extracted(
                [], 0, data_rows, total_rows, h_idx, ds, header_mode, False,
                fallback_used, 0, det_overrides, sem_map, h_debug, backend,
                selected_sheet_name, sheet_debug, warnings, coverage,
            )
            logger.info("Excel schema infer failed: coverage=%.3f keys=0 records=0 fallback=%s", coverage, fallback_used)
            sheet_sum = {"sheet_name": sheet_name, "header_row_idx": h_idx, "data_rows": data_rows, "coverage": round(coverage, 4), "warnings": list(warnings)}
            return ext, None, sheet_sum, None

        # ---- 5）生成记录：按语义键映射逐行抽取，并应用行过滤规则 ----
        sem_keys = mapper.build_semantic_keys(hpaths, sem_map, warnings)
        sem_keys = self._stabilize_grouped_semantic_keys(sem_keys, h1, h2, hpaths, sheet_name, warnings)
        logger.info(
            "Excel mode=%s | filename=%s | source_id=%s | rows=%d | cols=%d | coverage=%.3f",
            "fallback" if fallback_used else "schema_infer",
            filename, source_id, total_rows, len(hpaths), coverage,
        )

        records: List[Dict[str, Any]] = []
        truncated = False
        end_idx = min(total_rows, ds + cfg.max_rows_to_process)
        if data_rows > cfg.max_rows_to_process:
            truncated = True
            warnings.append(f"数据行 {data_rows} 行，仅处理前 {cfg.max_rows_to_process} 行")

        for ri in range(ds, end_idx):
            row_vals = [DataCleaner.cell_to_str(df.iat[ri, ci]) for ci in range(len(sem_keys))]
            if not row_vals or all(v == "" for v in row_vals):
                continue
            rec: Dict[str, Any] = {}
            for ci, sk in enumerate(sem_keys):
                if sk:
                    rec[sk] = DataCleaner.normalize_value(df.iat[ri, ci])
            if row_filter and not _apply_row_filter(rec, row_vals, row_filter):
                # Guard: avoid dropping real first rows right below header
                # when LLM row_filter is too aggressive (e.g. "合计" token appears in tail cells).
                if ri < ds + 3 and self._is_strong_early_record_candidate(rec):
                    if "row_filter_bypass_for_early_strong_record" not in warnings:
                        warnings.append("row_filter_bypass_for_early_strong_record")
                else:
                    continue
            if rec:
                rec["__source_file__"] = filename
                if add_sheet_name:
                    rec["__sheet_name__"] = sheet_name
                records.append(rec)

        # ---- 6）生成调试用逗号分隔样本文本：包含表头与若干数据行样本 ----
        sample_indices = _evenly_spaced_indices(ds, end_idx, cfg.sample_head_rows)
        sample_lines: List[str] = []
        if h1:
            sample_lines.append(",".join(DataCleaner.cell_to_str(c) for c in h1))
        if h2:
            sample_lines.append(",".join(DataCleaner.cell_to_str(c) for c in h2))
        for idx in sample_indices:
            sample_lines.append(",".join(DataCleaner.cell_to_str(c) for c in df.iloc[idx].tolist()))
        csv_text = "\n".join(sample_lines)

        ext = self._build_extracted(
            records, len(records), data_rows, total_rows, h_idx, ds,
            header_mode, truncated, fallback_used,
            sum(1 for k in sem_keys if k), det_overrides, sem_map,
            h_debug, backend, selected_sheet_name, sheet_debug, warnings, coverage,
        )
        logger.info(
            "Excel schema infer: coverage=%.3f keys=%d records=%d fallback=%s",
            coverage, sum(1 for k in sem_keys if k), len(records), fallback_used,
        )
        sheet_sum = {"sheet_name": sheet_name, "header_row_idx": h_idx, "data_rows": data_rows, "coverage": round(coverage, 4), "warnings": list(warnings)}
        table_meta = {"total_rows": total_rows, "truncated": truncated, "mode": "schema_infer", "sampled_rows": len(sample_indices)}
        return ext, csv_text, sheet_sum, table_meta

    # ------------------------------------------------------------------
    # 辅助函数
    # ------------------------------------------------------------------

    @staticmethod
    def _is_demo_sheet_name(sheet_name: str) -> bool:
        sn = str(sheet_name or "").strip().lower()
        return any(k in sn for k in ("示例", "样例", "example", "demo"))

    @staticmethod
    def _record_has_demo_placeholder(rec: Dict[str, Any]) -> bool:
        marker_tokens = ("****", "xxx", "测试", "sample", "demo")
        id_number = str(rec.get("id_number", "") or "")
        if "*" in id_number:
            return True
        for v in rec.values():
            text = str(v or "").strip().lower()
            if not text:
                continue
            if any(tok in text for tok in marker_tokens):
                return True
        return False

    @classmethod
    def _should_filter_demo_record(cls, sheet_name: str, rec: Dict[str, Any]) -> bool:
        if not cls._is_demo_sheet_name(sheet_name):
            return False
        return cls._record_has_demo_placeholder(rec)

    @staticmethod
    def _is_strong_early_record_candidate(rec: Dict[str, Any]) -> bool:
        name = DataCleaner.cell_to_str(rec.get("name", ""))
        if len(name) < 2:
            return False
        for key in ("id_number", "employee_id"):
            value = DataCleaner.cell_to_str(rec.get(key, "")).replace(" ", "")
            if len(value) >= 15 and all(ch.isdigit() or ch in ("X", "x") for ch in value):
                return True
        return False

    @staticmethod
    def _is_boolean_check_value(value: str) -> bool:
        v = str(value or "").strip().lower()
        return v in {"true", "false", "是", "否", "通过", "不通过", "y", "n", "yes", "no", "1", "0"}

    @staticmethod
    def _infer_group_tag(*texts: Any) -> str:
        merged = DataCleaner.normalize_header_compact(" ".join(str(t or "") for t in texts))
        if any(k in merged for k in ("公积金", "住房公积金", "公积")):
            return "hf"
        if any(k in merged for k in ("社保", "社会保险")):
            return "ss"
        if any(k in merged for k in ("医保", "医疗")):
            return "med"
        return ""

    @classmethod
    def _stabilize_grouped_semantic_keys(
        cls,
        sem_keys: List[str],
        h1: List[Any],
        h2: List[Any],
        hpaths: List[str],
        sheet_name: Any,
        warnings: List[str],
    ) -> List[str]:
        keys = list(sem_keys or [])
        if not keys:
            return keys

        amount_map = {"ss": "ss_base", "hf": "hf_base", "med": "med_base"}
        end_map = {"ss": "ss_end_month", "hf": "hf_end_month", "med": "med_end_month"}
        changed = False
        used = set(k for k in keys if k)

        for ci, key in enumerate(keys):
            if not key:
                continue
            base = re.sub(r"__\d+$", "", key)
            if base not in {"amount", "end_date"}:
                continue
            group = cls._infer_group_tag(
                hpaths[ci] if ci < len(hpaths) else "",
                h1[ci] if ci < len(h1) else "",
                h2[ci] if ci < len(h2) else "",
            )
            if not group:
                continue
            target = amount_map.get(group) if base == "amount" else end_map.get(group)
            if not target:
                continue
            if target in used and target != key:
                continue
            used.discard(key)
            keys[ci] = target
            used.add(target)
            changed = True

        # Minimal fallback for remove-intent sheets when group tags are weak:
        # split first two duplicate amount/end_date columns into SS/HF by order.
        is_remove_like = any(k in DataCleaner.normalize_header_compact(sheet_name) for k in ("减员", "离职", "停保"))
        if is_remove_like:
            amount_idx = [i for i, k in enumerate(keys) if re.sub(r"__\d+$", "", k) == "amount"]
            end_idx = [i for i, k in enumerate(keys) if re.sub(r"__\d+$", "", k) == "end_date"]
            if len(amount_idx) >= 2:
                keys[amount_idx[0]] = "ss_base"
                keys[amount_idx[1]] = "hf_base"
                changed = True
            if len(end_idx) >= 2:
                keys[end_idx[0]] = "ss_end_month"
                keys[end_idx[1]] = "hf_end_month"
                changed = True

        if changed and "semantic_group_split_applied" not in warnings:
            warnings.append("semantic_group_split_applied")
        return keys

    @classmethod
    def _sanitize_boolean_id_check_columns(
        cls,
        df: pd.DataFrame,
        header_paths: List[str],
        mapping: Dict[str, str],
        data_start_idx: int,
        max_rows: int,
        warnings: List[str],
    ) -> Dict[str, str]:
        updated = dict(mapping or {})
        has_real_id_column = any(
            (k or "").strip() in {"id_number", "employee_id"} for k in updated.values()
        )
        if not has_real_id_column:
            return updated

        end = min(len(df), data_start_idx + max_rows)
        warned = False
        for ci, hp in enumerate(header_paths):
            header_text = str(hp or "")
            mapped = (updated.get(hp) or "").strip()
            if mapped not in {"id_number", "id_card", "employee_id", ""}:
                continue
            if "身份证" not in header_text and "身份证" not in DataCleaner.normalize_header_compact(header_text):
                continue

            values = [
                DataCleaner.cell_to_str(df.iat[r, ci])
                for r in range(data_start_idx, end)
                if DataCleaner.cell_to_str(df.iat[r, ci]) != ""
            ]
            if not values:
                continue

            bool_like_ratio = sum(1 for v in values if cls._is_boolean_check_value(v)) / max(1, len(values))
            if bool_like_ratio >= 0.7:
                updated[hp] = "id_check_passed"
                if not warned:
                    warnings.append("boolean_id_check_column_skipped")
                    warned = True
        return updated

    @staticmethod
    def _build_extracted(
        records: List[Dict[str, Any]],
        processed_rows: int,
        data_rows: int,
        total_rows: int,
        header_row_idx: int,
        data_start_idx: int,
        header_mode: str,
        truncated: bool,
        fallback_used: bool,
        semantic_keys_count: int,
        deterministic_overrides_count: int,
        semantic_key_by_header: Dict[str, str],
        header_selection_debug: Dict[str, Any],
        read_backend: str,
        selected_sheet_name: str,
        sheet_selection_debug: Dict[str, Any],
        warnings: List[str],
        coverage: float,
    ) -> Dict[str, Any]:
        """
        统一构造抽取结果的 JSON 结构。

        该结构会被写入 `SourceDoc.extracted`，并作为 `EXTRACTED_JSON` block 的内容：
        - `data`：记录列表（每条记录为字典）
        - `metadata`：表头位置、覆盖率、读取后端、回退状态、调试信息等
        - `warnings`：告警列表（字符串数组）
        """
        return {
            "data": records,
            "metadata": {
                "mode": "schema_infer",
                "coverage": round(coverage, 4),
                "total_rows": total_rows,
                "processed_rows": processed_rows,
                "data_rows": data_rows,
                "header_row_idx": header_row_idx,
                "data_start_idx": data_start_idx,
                "header_mode": header_mode,
                "truncated": truncated,
                "fallback_used": fallback_used,
                "semantic_keys_count": semantic_keys_count,
                "deterministic_overrides_count": deterministic_overrides_count,
                "semantic_key_by_header": semantic_key_by_header,
                "header_selection_debug": header_selection_debug,
                "read_backend": read_backend,
                "selected_sheet_name": selected_sheet_name,
                "sheet_selection_debug": sheet_selection_debug,
            },
            "warnings": warnings,
        }

    @staticmethod
    def _make_blocks(
        extracted: Dict[str, Any],
        csv_text: Optional[str],
        table_meta: Optional[Dict[str, Any]],
    ) -> List[SourceBlock]:
        """
        将抽取结果组装为 `SourceBlock` 列表。

        约定：
        - 若 `csv_text` 为空，则只输出 `EXTRACTED_JSON`；
        - 否则先输出 `TABLE_CSV`（含 `table_meta`），再输出 `EXTRACTED_JSON`。
        """
        if csv_text is None:
            return [SourceBlock(order=1, type=BlockType.EXTRACTED_JSON, content=extracted, meta={})]
        return [
            SourceBlock(order=1, type=BlockType.TABLE_CSV, content=csv_text, meta=table_meta or {}),
            SourceBlock(order=2, type=BlockType.EXTRACTED_JSON, content=extracted, meta={}),
        ]
