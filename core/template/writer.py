"""
模板写入模块 (Template Writer Module)
====================================

将填充计划应用到 Excel 模板，执行清除、行写入、单元格写入等操作。
"""

from __future__ import annotations

from datetime import datetime
import re
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from core.logger import get_logger

logger = get_logger(__name__)

# ------------------------------
# Constants / precompiled regex
# ------------------------------

DATE_FORMATS: Tuple[str, ...] = (
    "%Y-%m-%d",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y年%m月%d日",
    "%Y年%m月%d号",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y.%m.%d",
    "%d.%m.%Y",
    "%Y%m%d",
)

DATE_FORMAT_INDICATORS: Tuple[str, ...] = ("d", "m", "y", "h", "s", "D", "M", "Y", "H", "S")

RE_CELL_REF = re.compile(r"^([A-Z]+)(\d+)$")
RE_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}")
RE_ISO_DATE_FULL = re.compile(r"^\d{4}-\d{2}-\d{2}(T\d{2}:\d{2}:\d{2}(\.\d+)?(Z|[+-]\d{2}:\d{2})?)?$")
RE_CHINESE_DATE = re.compile(r"^(\d{4})年(\d{1,2})月(\d{1,2})[日号]?$")
RE_SLASH_DATE = re.compile(r"^\d{1,4}/\d{1,2}/\d{1,4}$")
RE_DOT_DATE = re.compile(r"^\d{4}\.\d{2}\.\d{2}$")
RE_COMPACT_DATE = re.compile(r"^\d{8}$")
RE_LONG_DIGITS = re.compile(r"^\d{11,}$")
RE_LONG_DIGITS_WITH_X = re.compile(r"^\d{14,}[Xx]$")
RE_DECIMAL = re.compile(r"^[+-]?\d+\.\d+$")


class ValueConverter:
    """
    值规范化与类型转换工具类。

    处理日期、长数字、小数等，避免 Excel 科学计数法等问题。
    """

    @staticmethod
    def convert(value: Any, number_format: str) -> Any:
        """
        根据单元格数字格式转换值。
        长数字保持字符串，日期解析为 datetime，数字解析为 int/float。
        """
        if value is None:
            return None
        if isinstance(value, (int, float, datetime)):
            return value
        if not isinstance(value, str):
            return value

        text = value.strip()
        if not text:
            return None

        # Preserve long IDs / phone-like strings to prevent Excel scientific notation.
        if RE_LONG_DIGITS.match(text) or RE_LONG_DIGITS_WITH_X.match(text):
            return text

        if _is_date_format(number_format) or _looks_like_date(text):
            parsed = ValueConverter._parse_date(text)
            if parsed is not None:
                return parsed
            logger.warning("Date parsing failed for value '%s', keeping as string", text)
            return text

        if ValueConverter._looks_like_number(text):
            normalized = text.replace(",", "").replace(" ", "")
            try:
                if RE_DECIMAL.match(normalized):
                    return float(normalized)
                return int(normalized)
            except Exception:
                return text

        return text

    @staticmethod
    def _parse_date(value: Optional[str]) -> Optional[datetime]:
        if not value or not isinstance(value, str):
            return None
        text = value.strip()
        if not text:
            return None

        try:
            normalized = text.replace("Z", "+00:00")
            return datetime.fromisoformat(normalized)
        except (ValueError, TypeError):
            pass

        m = RE_CHINESE_DATE.match(text)
        if m:
            try:
                return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except (ValueError, TypeError):
                pass

        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(text, fmt)
            except (ValueError, TypeError):
                continue
        return None

    @staticmethod
    def _looks_like_number(value: str) -> bool:
        text = value.strip()
        if not text:
            return False
        try:
            float(text.replace(",", "").replace(" ", ""))
            return True
        except Exception:
            return False


class ExcelFillExecutor:
    """
    Excel 填充执行器：加载模板、应用填充计划、保存输出。
    """

    def __init__(self, template_path: str, output_path: str):
        self.template_path = template_path
        self.output_path = output_path
        self.wb: Optional[Workbook] = None
        self.ws: Optional[Worksheet] = None

    def load(self, fill_plan: Mapping[str, Any]) -> None:
        """加载模板工作簿并解析目标工作表。"""
        self.wb = load_workbook(self.template_path, data_only=False)
        self.ws = self._resolve_sheet(fill_plan)
        logger.debug("ExcelFillExecutor.load: template=%s, sheet=%s", self.template_path, self.ws.title)

    def save(self) -> None:
        """保存工作簿到输出路径并关闭。"""
        if self.wb is None:
            raise RuntimeError("Workbook is not loaded")
        self.wb.save(self.output_path)
        self.wb.close()
        logger.debug("ExcelFillExecutor.save: output=%s", self.output_path)

    def execute(self, plan: Mapping[str, Any]) -> int:
        """执行填充计划：清除、行写入、直接写入，返回写入单元格数。"""
        if self.ws is None:
            raise RuntimeError("Worksheet is not loaded")

        self._apply_clear_ranges(plan.get("clear_ranges", []))
        cells_written = self._apply_row_writes(plan.get("row_writes", []))
        cells_written += self._apply_direct_writes(plan.get("writes", []))
        logger.info("apply_fill_plan: total cells_written=%d", cells_written)
        return cells_written

    def _resolve_sheet(self, fill_plan: Mapping[str, Any]) -> Worksheet:
        if self.wb is None:
            raise RuntimeError("Workbook is not loaded")
        target = fill_plan.get("target", {})
        sheet_name: Optional[str] = None
        if isinstance(target, Mapping):
            sheet_name = target.get("sheet_name") or target.get("sheet")
        if not sheet_name:
            sheet_name = fill_plan.get("sheet_name") or fill_plan.get("sheet")
        if isinstance(sheet_name, str):
            sheet_name = sheet_name.strip()
        if not sheet_name:
            sheet_name = self.wb.sheetnames[0] if self.wb.sheetnames else None
        if not sheet_name or sheet_name not in self.wb.sheetnames:
            available = ", ".join(self.wb.sheetnames) if self.wb.sheetnames else "none"
            raise ValueError(f"Sheet '{sheet_name}' not found in template. Available sheets: {available}")
        return self.wb[sheet_name]

    def _apply_clear_ranges(self, clear_ranges: Any) -> None:
        if self.ws is None or not isinstance(clear_ranges, Iterable):
            return
        for range_str in clear_ranges:
            if isinstance(range_str, str):
                _clear_range(self.ws, range_str)

    def _apply_row_writes(self, row_writes: Any) -> int:
        if self.ws is None or not isinstance(row_writes, list):
            return 0
        logger.debug("apply_fill_plan: %d row_writes", len(row_writes))
        total = 0
        for rw_idx, row_write in enumerate(row_writes):
            if not isinstance(row_write, Mapping):
                continue
            start_cell = row_write.get("start_cell")
            if not start_cell:
                logger.debug("apply_fill_plan: row_write %d has no start_cell, skip", rw_idx)
                continue
            _, start_row = _parse_cell(str(start_cell))
            rows = row_write.get("rows", [])
            column_mapping = row_write.get("column_mapping", {})
            if not isinstance(rows, list) or not rows:
                logger.debug("row_write %d (%s): 0 rows, skip", rw_idx, start_cell)
                continue

            first = rows[0]
            if isinstance(first, Mapping) and "column_letter" in first:
                total += self._write_row_format_cells_same_row(start_row, rows)
            elif isinstance(first, list):
                total += self._write_row_format_nested_rows(start_row, rows)
            elif isinstance(first, Mapping) and isinstance(column_mapping, Mapping):
                total += self._write_row_format_mapped_rows(start_row, rows, column_mapping)
            else:
                logger.debug("row_write %d: unhandled row format=%s", rw_idx, type(first).__name__)
        return total

    def _apply_direct_writes(self, writes: Any) -> int:
        if self.ws is None or not isinstance(writes, list):
            return 0
        written = 0
        for write in writes:
            if not isinstance(write, Mapping):
                continue
            cell_ref = write.get("cell")
            if not cell_ref:
                continue
            col, row = _parse_cell(str(cell_ref))
            if self._write_cell_safe(row, col, write.get("value")):
                written += 1
        return written

    def _write_row_format_cells_same_row(self, start_row: int, rows: Sequence[Any]) -> int:
        written = 0
        for cell_data in rows:
            if not isinstance(cell_data, Mapping):
                continue
            col_letter = cell_data.get("column_letter")
            if not col_letter:
                continue
            try:
                col_idx = column_index_from_string(str(col_letter))
            except Exception:
                logger.warning("Invalid column letter '%s' at row %d", col_letter, start_row)
                continue
            if self._write_cell_safe(start_row, col_idx, cell_data.get("value")):
                written += 1
        return written

    def _write_row_format_nested_rows(self, start_row: int, rows: Sequence[Any]) -> int:
        written = 0
        current_row = start_row
        for row_data in rows:
            if isinstance(row_data, list):
                for cell_data in row_data:
                    if not isinstance(cell_data, Mapping):
                        continue
                    col_letter = cell_data.get("column_letter")
                    if not col_letter:
                        continue
                    try:
                        col_idx = column_index_from_string(str(col_letter))
                    except Exception:
                        logger.warning("Invalid column letter '%s' at row %d", col_letter, current_row)
                        continue
                    if self._write_cell_safe(current_row, col_idx, cell_data.get("value")):
                        written += 1
            current_row += 1
        return written

    def _write_row_format_mapped_rows(
        self,
        start_row: int,
        rows: Sequence[Any],
        column_mapping: Mapping[str, Any],
    ) -> int:
        written = 0
        current_row = start_row
        logger.debug(
            "apply_fill_plan: mapped row_writes=%d, start_row=%d, mapping_keys=%d",
            len(rows),
            start_row,
            len(column_mapping),
        )

        for row_idx, row_data in enumerate(rows):
            row_written = 0
            if isinstance(row_data, list):
                for cell_data in row_data:
                    if not isinstance(cell_data, Mapping):
                        continue
                    col_letter = cell_data.get("column_letter")
                    if not col_letter:
                        continue
                    try:
                        col_idx = column_index_from_string(str(col_letter))
                    except Exception:
                        logger.warning("Invalid column letter '%s' at row %d", col_letter, current_row)
                        continue
                    if self._write_cell_safe(current_row, col_idx, cell_data.get("value")):
                        written += 1
                        row_written += 1
            elif isinstance(row_data, Mapping):
                for key, col_letter in column_mapping.items():
                    if key not in row_data or not col_letter:
                        continue
                    try:
                        col_idx = column_index_from_string(str(col_letter))
                    except Exception:
                        logger.warning("Invalid column letter '%s' for key '%s'", col_letter, key)
                        continue
                    if self._write_cell_safe(current_row, col_idx, row_data.get(key)):
                        written += 1
                        row_written += 1
            if row_idx < 3 or row_idx == len(rows) - 1:
                logger.debug(
                    "apply_fill_plan: row %d/%d (excel row %d), cells_written=%d",
                    row_idx + 1,
                    len(rows),
                    current_row,
                    row_written,
                )
            current_row += 1
        return written

    def _write_cell_safe(self, row: int, col: int, value: Any) -> bool:
        if self.ws is None:
            return False
        try:
            cell = self.ws.cell(row, col)
            if _is_formula_cell(cell):
                return False
            converted = ValueConverter.convert(value, cell.number_format)
            cell.value = converted
            return True
        except Exception as exc:
            logger.warning("Failed to write cell r=%d c=%d: %s", row, col, exc)
            return False


def apply_fill_plan(template_path: str, fill_plan: Dict[str, Any], output_path: str) -> int:
    """
    公开入口：将填充计划应用到模板并保存到输出路径。
    返回写入的单元格数量。
    """
    executor = ExcelFillExecutor(template_path=template_path, output_path=output_path)
    executor.load(fill_plan)
    cells_written = executor.execute(fill_plan)
    executor.save()
    return cells_written


def _clear_range(ws: Worksheet, range_str: str) -> None:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        for row in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                cell = ws.cell(row, col_idx)
                if not _is_formula_cell(cell):
                    cell.value = None
    except Exception:
        logger.debug("Ignore invalid clear range: %s", range_str)


def _parse_cell(cell_ref: str) -> Tuple[int, int]:
    match = RE_CELL_REF.match(cell_ref.strip().upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    col_idx = column_index_from_string(match.group(1))
    row = int(match.group(2))
    return col_idx, row


def _is_formula_cell(cell: Cell) -> bool:
    if cell.data_type == "f":
        return True
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return True
    return False


# Compatibility wrappers for existing tests/imports.
def _convert_value(value: Any, number_format: str) -> Any:
    return ValueConverter.convert(value, number_format)


def _parse_date(value: Optional[str]) -> Optional[datetime]:
    return ValueConverter._parse_date(value)


def _looks_like_date(value: str) -> bool:
    if not value:
        return False
    text = value.strip()
    return bool(
        RE_ISO_DATE.match(text)
        or RE_CHINESE_DATE.match(text)
        or RE_SLASH_DATE.match(text)
        or RE_DOT_DATE.match(text)
        or RE_COMPACT_DATE.match(text)
    )


def _is_iso_date(s: str) -> bool:
    return bool(RE_ISO_DATE_FULL.match(str(s or "").strip()))


def _is_date_format(fmt: str) -> bool:
    if not fmt:
        return False
    return any(indicator in fmt for indicator in DATE_FORMAT_INDICATORS)


def _looks_like_number(s: str) -> bool:
    return ValueConverter._looks_like_number(s)
