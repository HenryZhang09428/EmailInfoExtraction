"""
保险业务域层 (Insurance Domain — Layer 2)
==========================================

本模块封装所有社保增减员的业务规则、中文关键词和特定策略。

包含:
  - 领域工具函数 (_parse_year_month, _format_year_month 等)
  - InsuranceRecord           — 单条记录的领域封装 (意图推断)
  - InsuranceBusinessLogic    — 意图检测/记录过滤/去重/费用年月计算
  - InsuranceSourceSelector   — LLM + 启发式数据源选择
  - InsuranceFillPlanner      — 编排保险业务流程 → FillPlan
  - 领域管线函数 (_apply_record_type_filter, _auto_infer_common_fields 等)

依赖: config.py (InsuranceConfig, CONFIG) + common.py (通用工具 + 类)
      不依赖 fill_planner.py (单向依赖, 避免循环引用)
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import (
    Any,
    Dict,
    List,
    Optional,
    Set,
    Tuple,
)

from openpyxl.utils import get_column_letter, range_boundaries

from core.ir import FillPlan, FillPlanTarget
from core.llm import LLMClient
from core.logger import get_logger
from core.template.prompts import build_insurance_param_prompt
from core.template.schema import TemplateSchema

from core.template.logic.config import InsuranceConfig, CONFIG
from core.template.logic.common import (
    SourceDataExtractor,
    HeaderMatcher,
    SchemaInspector,
    GenericFillPlanner,
    _normalize_match_text,
    _normalize_header,
    _normalize_text,
    _header_variants,
    _is_non_empty_value,
    _parse_any_date,
    _looks_like_chinese_name,
    _get_record_value,
    _trim_record_fields,
    _merge_debug_info,
    _dict_to_fill_plan,
)

logger = get_logger(__name__)

def _parse_year_month(value: Any) -> Optional[Tuple[int, int]]:
    """从多种格式中解析 (year, month)。"""
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, (int, float)):
        value = str(int(value))
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    patterns = [
        r'^(\d{4})[-/](\d{1,2})[-/]\d{1,2}',  # YYYY-MM-DD
        r'^(\d{4})[-/](\d{1,2})$',              # YYYY-MM
        r'^(\d{4})年(\d{1,2})月$',              # YYYY年MM月
        r'^(\d{4})(\d{2})\d{2}$',               # YYYYMMDD
        r'^(\d{4})(\d{2})$',                     # YYYYMM
    ]
    for pat in patterns:
        m = re.match(pat, text)
        if m:
            year, month = int(m.group(1)), int(m.group(2))
            if 1 <= month <= 12:
                return year, month
    return None


def _format_year_month(year: int, month: int, fmt: str) -> str:
    """按指定格式输出年月字符串。"""
    formats = {
        "YYYYMM": f"{year}{month:02d}",
        "YYYY/MM": f"{year}/{month:02d}",
        "YYYY年MM月": f"{year}年{month:02d}月",
    }
    return formats.get(fmt, f"{year}-{month:02d}")


def _detect_year_month_format_from_value(value: Any) -> Optional[str]:
    """从样本值推断年月格式。"""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "年" in text and "月" in text:
        return "YYYY年MM月"
    if "/" in text:
        return "YYYY/MM"
    if "-" in text:
        return "YYYY-MM"
    if re.match(r'^\d{6}$', text):
        return "YYYYMM"
    return None


def _next_month_yyyymm(value: datetime) -> Optional[str]:
    """
    计算下月 YYYYMM。

    通用算法——给定一个日期，返回下个自然月的 YYYYMM 字符串。
    """
    if not isinstance(value, datetime):
        return None
    year, month = value.year, value.month + 1
    if month > 12:
        year += 1
        month = 1
    return f"{year}{month:02d}"



# ============================================================================
# 2.2  InsuranceRecord — 保险记录领域模型
# ============================================================================


class InsuranceRecord:
    """
    单条保险记录的领域封装。

    封装意图推断逻辑:
      - 从"申报类型"等显式字段推断 (最高优先级)
      - 从英文归一化键名推断 (terminationdate → remove)
      - 从键值关键词扫描推断
      - 从 sheet 名称推断 (最低优先级)
    """
    __slots__ = ("_data", "_config")

    def __init__(self, data: Dict[str, Any], config: InsuranceConfig = CONFIG):
        self._data = data if isinstance(data, dict) else {}
        self._config = config

    @property
    def raw(self) -> Dict[str, Any]:
        return self._data

    def get(self, key: str, default: Any = None) -> Any:
        return self._data.get(key, default)

    @property
    def source_file(self) -> str:
        return str(self._data.get("__source_file__", "") or "")

    @property
    def sheet_name(self) -> str:
        return str(self._data.get("__sheet_name__", "") or "")

    def infer_explicit_intent(self) -> Optional[str]:
        """从"申报类型"等显式字段推断意图。"""
        cfg = self._config
        for key, value in self._data.items():
            kt = _normalize_text(key)
            if not any(_normalize_text(k) in kt for k in cfg.declare_type_header_synonyms):
                continue
            if isinstance(value, str):
                vt = _normalize_text(value)
                a = any(k in vt for k in cfg.add_keywords)
                r = any(k in vt for k in cfg.remove_keywords)
                s = any(k in vt for k in cfg.in_service_keywords)
                if a and not r: return "add"
                if r and not a: return "remove"
                if s and not a and not r: return "in_service"
        return None

    def infer_intent_from_keys_and_values(self) -> Optional[str]:
        """扫描所有键名和值中的关键词推断意图。"""
        cfg = self._config
        ah = rh = sh = 0
        for key, value in self._data.items():
            kt = _normalize_text(key)
            ah += any(k in kt for k in cfg.add_keywords)
            rh += any(k in kt for k in cfg.remove_keywords)
            sh += any(k in kt for k in cfg.in_service_keywords)
            if isinstance(value, str):
                vt = _normalize_text(value)
                ah += any(k in vt for k in cfg.add_keywords)
                rh += any(k in vt for k in cfg.remove_keywords)
                sh += any(k in vt for k in cfg.in_service_keywords)
        if ah > 0 and rh == 0: return "add"
        if rh > 0 and ah == 0: return "remove"
        if sh > 0 and ah == 0 and rh == 0: return "in_service"
        return None

    def infer_record_type(self) -> Optional[str]:
        """综合推断记录类型 (显式字段 → 英文键名 → 键值扫描)。"""
        explicit = self.infer_explicit_intent()
        if explicit in ("add", "remove"):
            return explicit
        cfg = self._config
        for key in self._data.keys():
            if not isinstance(key, str): continue
            n = _normalize_header(key)
            if any(n == rk or n.startswith(rk) for rk in cfg.remove_semantic_keys):
                return "remove"
        for key in self._data.keys():
            if not isinstance(key, str): continue
            n = _normalize_header(key)
            if any(n == ak or n.startswith(ak) for ak in cfg.add_semantic_keys):
                return "add"
        return None

    def infer_type_from_sheet(self) -> Optional[str]:
        """从 sheet 名推断。"""
        s = self.sheet_name
        if "减员" in s or "离职" in s: return "remove"
        if "增员" in s: return "add"
        return None

    def infer_type_combined(self) -> Optional[str]:
        """最终综合推断。"""
        return self.infer_record_type() or self.infer_type_from_sheet()


# ============================================================================
# 2.3  InsuranceBusinessLogic — 业务规则
# ============================================================================


class InsuranceBusinessLogic:
    """
    封装所有社保增减员的业务规则。

    职责:
      - 意图检测 (模板 / 数据源 / 记录级别)
      - 记录过滤 (按意图 / 按 sheet 名)
      - 邮件+Excel 离职记录合并与去重 (减员专用)
      - 费用年月计算 (下月规则)
      - 数据源概要构建 (LLM prompt 用)
    """

    def __init__(self, config: InsuranceConfig = CONFIG):
        self._config = config

    # -- 模板意图 --

    def infer_intent_from_filename(self, filename: str) -> Optional[str]:
        """从模板文件名推断增员/减员意图。"""
        text = _normalize_match_text(filename or "")
        a, r = "增员" in text, "减员" in text
        if a and not r: return "add"
        if r and not a: return "remove"
        return None

    def detect_template_intent(self, filename: Optional[str], schema: TemplateSchema) -> Optional[str]:
        """
        综合推断模板意图: 先看文件名，再看 sheet 名 + 表头内容。

        业务背景: 社保增员表和减员表是两张不同的模板。
        """
        cfg = self._config
        ft = _normalize_text(filename or "")
        if ft:
            a = any(k in ft for k in cfg.add_keywords)
            r = any(k in ft for k in cfg.remove_keywords)
            if a and not r: return "add"
            if r and not a: return "remove"
        sn = [ss.sheet for ss in schema.sheet_schemas]
        hp = []
        for ss in schema.sheet_schemas:
            for reg in ss.regions:
                if reg.table and reg.table.header:
                    hp.extend(h.header_path for h in reg.table.header if h.header_path)
        st = _normalize_text(" ".join(sn + hp))
        a = any(k in st for k in cfg.add_keywords)
        r = any(k in st for k in cfg.remove_keywords)
        if a and not r: return "add"
        if r and not a: return "remove"
        return None

    def is_insurance_template(self, schema: TemplateSchema) -> bool:
        """
        检测模板是否为社保增减员类型。

        判据: 表头中须同时含"姓名"、"申报类型"、"费用年月"后缀的列。
        """
        inspector = SchemaInspector(schema)
        region = inspector.get_main_table_region()
        if not region or not region.table or not region.table.header:
            return False
        found = {s: False for s in self._config.required_header_suffixes}
        for h in region.table.header:
            path = getattr(h, "header_path", None)
            if not isinstance(path, str): continue
            n = path.replace("\u3000", "").strip()
            for s in self._config.required_header_suffixes:
                if not found[s] and n.endswith(s):
                    found[s] = True
            if all(found.values()):
                break
        return all(found.values())

    # -- 数据源意图 --

    def infer_source_intent(self, filename: Optional[str]) -> Optional[str]:
        cfg = self._config
        ft = _normalize_text(filename or "")
        if not ft: return None
        a = any(k in ft for k in cfg.add_keywords)
        r = any(k in ft for k in cfg.remove_keywords)
        s = any(k in ft for k in cfg.in_service_keywords)
        if a and not r: return "add"
        if r and not a: return "remove"
        if s and not a and not r: return "in_service"
        return None

    # -- 记录过滤 --

    def filter_by_sheet_intent(self, records: List[Dict[str, Any]], intent: str) -> Tuple[List[Dict[str, Any]], bool]:
        """
        按 sheet 名关键词过滤记录。

        业务背景: Excel 可能有多个 sheet，增员/减员 sheet 中的记录分别对应不同模板。
        """
        if intent == "add": kw = ("增员",)
        elif intent == "remove": kw = ("减员", "离职")
        else: return records, False
        return [r for r in records if any(k in str(r.get("__sheet_name__", "") or "") for k in kw)], True

    def filter_by_template_intent(self, records: List[Dict[str, Any]], intent: Optional[str]) -> List[Dict[str, Any]]:
        """按模板意图过滤记录，保留匹配或未知意图的记录。"""
        if intent not in ("add", "remove"):
            return records
        cfg = self._config

        def _intents(r: Dict[str, Any]) -> List[str]:
            out: List[str] = []
            si = self.infer_source_intent(r.get("__source_file__"))
            if si: out.append(si)
            rec = InsuranceRecord(r, cfg)
            ei = rec.infer_explicit_intent()
            if ei: out.append(ei)
            ki = rec.infer_intent_from_keys_and_values()
            if ki: out.append(ki)
            return out

        has_signal = any(intent in _intents(r) for r in records if isinstance(r, dict))

        def ok(r: Dict[str, Any]) -> bool:
            rec = InsuranceRecord(r, cfg)
            rt = rec.infer_type_combined()
            if rt is not None: return rt == intent
            si = self.infer_source_intent(r.get("__source_file__"))
            if si and si != intent: return False
            ei = rec.infer_explicit_intent()
            if ei and ei != intent: return False
            ki = rec.infer_intent_from_keys_and_values()
            if ki is None:
                return (si == intent or ei == intent) if has_signal else True
            return ki == intent

        strict = [r for r in records if isinstance(r, dict) and ok(r)]
        return strict if strict else records

    # -- 邮件离职记录 (减员专用) --

    def collect_email_leave_records(self, extractor: SourceDataExtractor) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        从邮件数据源收集离职记录。

        业务背景: HR 常在邮件正文中列出离职人员清单（姓名+离职日期），
        需要与 Excel 附件中的结构化数据合并。
        """
        records: List[Dict[str, Any]] = []
        warnings: List[str] = []
        count = 0
        for source in extractor.get_sources("email"):
            count += 1
            ex = source.get("extracted")
            if not isinstance(ex, dict): continue
            data = ex.get("data")
            if not isinstance(data, list): continue
            fn = source.get("filename", "unknown")
            for item in data:
                if not isinstance(item, dict): continue
                if item.get("intent") == "remove" or item.get("leave_date"):
                    records.append({
                        "name": str(item.get("name", "") or "").strip(),
                        "employee_id": str(item.get("employee_id", "") or "").strip(),
                        "leave_date": str(item.get("leave_date", "") or "").strip(),
                        "leave_date_text": str(item.get("leave_date_text", "") or "").strip(),
                        "intent": "remove", "__source_file__": fn, "__source_type__": "email",
                    })
        if count:
            warnings.append(f"email_sources_scanned:{count}")
        return records, warnings

    # -- 去重 (关键业务逻辑) --

    @staticmethod
    def deduplicate_leave_records(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
        """
        对离职/减员记录去重。

        去重策略 (按优先级):
          1. employee_id 相同 → 同一人 → 保留字段最完整的记录
          2. (name, leave_date) 相同 → 同一次离职
          3. 无可用 key 的记录无法去重，直接保留

        业务背景:
          邮件正文 + Excel 附件可能都含同一批离职人员信息，但格式不同。
          邮件可能只有姓名+离职日期，Excel 有完整字段。
          去重时优先保留字段更完整的那条。
        """
        if not records:
            return [], 0
        by_eid: Dict[str, List[Dict[str, Any]]] = {}
        by_nd: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
        no_key: List[Dict[str, Any]] = []
        for r in records:
            eid = r.get("employee_id", "").strip()
            name = r.get("name", "").strip()
            ld = r.get("leave_date", "").strip()
            if eid: by_eid.setdefault(eid, []).append(r)
            elif name: by_nd.setdefault((name, ld), []).append(r)
            else: no_key.append(r)

        def _best(g: List[Dict[str, Any]]) -> Dict[str, Any]:
            return max(g, key=lambda r: sum(1 for v in r.values() if v and str(v).strip())) if len(g) > 1 else g[0]

        result: List[Dict[str, Any]] = []
        seen_eid: Set[str] = set()
        seen_nd: Set[Tuple[str, str]] = set()
        for eid, g in by_eid.items():
            best = _best(g)
            result.append(best)
            seen_eid.add(eid)
            n = best.get("name", "").strip()
            if n: seen_nd.add((n, best.get("leave_date", "").strip()))
        for (n, ld), g in by_nd.items():
            if (n, ld) in seen_nd: continue
            best = _best(g)
            if best.get("employee_id", "").strip() in seen_eid: continue
            result.append(best)
            seen_nd.add((n, ld))
        result.extend(no_key)
        return result, len(records) - len(result)

    # -- 费用年月 --

    def extract_fee_month(self, record: Dict[str, Any], intent: Optional[str]) -> Optional[Tuple[int, int]]:
        """
        从记录提取费用年月。

        策略: 直接费用字段 → 按意图日期字段 → 任意日期字段。
        """
        cfg = self._config
        for key in record:
            if not isinstance(key, str): continue
            if key in cfg.fee_month_direct_keys or _normalize_header(key) in [_normalize_header(k) for k in cfg.fee_month_direct_keys]:
                ym = _parse_year_month(record.get(key))
                if ym: return ym
        for key in (cfg.remove_date_keys if intent == "remove" else cfg.add_date_keys):
            if key in record:
                ym = _parse_year_month(record.get(key))
                if ym: return ym
        for key, val in record.items():
            if isinstance(key, str) and ("日期" in key or "时间" in key):
                ym = _parse_year_month(val)
                if ym: return ym
        return None

    def find_fee_month_headers(self, schema: TemplateSchema) -> List[str]:
        """在模板中查找所有费用年月类表头。"""
        cfg = self._config
        headers: List[str] = []
        for ss in schema.sheet_schemas:
            for reg in ss.regions:
                if not reg.table or not reg.table.header: continue
                for h in reg.table.header:
                    hp = getattr(h, "header_path", None)
                    if not hp: continue
                    n = _normalize_header(hp)
                    if ("年月" in n or "月份" in n) and any(a in n for a in cfg.fee_month_header_anchors):
                        headers.append(hp)
                    elif n in [_normalize_header(k) for k in cfg.fee_month_header_exact]:
                        headers.append(hp)
        return list(dict.fromkeys(headers))

    def infer_fee_month_for_records(
        self, schema: TemplateSchema, records: List[Dict[str, Any]], intent: Optional[str],
        column_mapping: Dict[str, str], constant_values: Dict[str, str], warnings: List[str],
    ) -> Tuple[Dict[str, str], Dict[str, str]]:
        """为每条记录推算费用年月并注入 __fee_month__。"""
        if not records: return column_mapping, constant_values
        fee_headers = self.find_fee_month_headers(schema)
        if not fee_headers: return column_mapping, constant_values
        fmt = None
        inspector = SchemaInspector(schema)
        for row in inspector.get_sample_rows():
            if not isinstance(row, dict): continue
            for h in fee_headers:
                if h in row and row[h]:
                    fmt = _detect_year_month_format_from_value(row[h])
                    if fmt: break
            if fmt: break
        if not fmt: fmt = "YYYYMM"
        for rec in records:
            if not isinstance(rec, dict): continue
            ym = self.extract_fee_month(rec, intent)
            rec["__fee_month__"] = _format_year_month(ym[0], ym[1], fmt) if ym else ""
        for h in fee_headers:
            if h not in column_mapping:
                column_mapping[h] = "__fee_month__"
        return column_mapping, constant_values

    # -- 同义词匹配 --

    def synonym_resolver(self, header: str, keys: List[str]) -> Optional[str]:
        """同义词 + 精确/包含匹配 (用于 fallback)。"""
        cfg = self._config
        h0 = header.split("/")[-1].strip() if "/" in header else header
        hc = h0.replace(" ", "").replace("_", "").replace("-", "")
        for k in keys:
            kc = k.replace(" ", "").replace("_", "").replace("-", "")
            if k == header or k == h0 or header in k or k in header or h0 in k or k in h0 or hc == kc:
                return k
        for k in keys:
            kc = k.replace(" ", "").replace("_", "").replace("-", "")
            if h0 in cfg.synonym_map.get(k, ()): return k
            if k in cfg.synonym_map.get(h0, ()): return k
            for sl in cfg.synonym_map.values():
                sc = [s.replace(" ", "").replace("_", "").replace("-", "") for s in sl]
                if hc in sc and kc in sc: return k
        return None

    # -- 数据源概要 (LLM prompt 用) --

    def build_sources_profile(self, extractor: SourceDataExtractor) -> List[Dict[str, Any]]:
        """为 LLM prompt 构建 Excel 数据源概要。"""
        cfg = self._config
        profile: List[Dict[str, Any]] = []
        for source in extractor.get_sources("excel"):
            ex = source.get("extracted") or {}
            data = ex.get("data") if isinstance(ex, dict) else None
            all_recs = data if isinstance(data, list) else []
            rk: List[str] = []
            seen: Set[str] = set()
            for rec in all_recs:
                if not isinstance(rec, dict): continue
                for k in rec:
                    if isinstance(k, str) and k not in seen:
                        seen.add(k); rk.append(k)
                        if len(rk) >= cfg.max_source_keys_in_profile: break
                if len(rk) >= cfg.max_source_keys_in_profile: break
            skbh = {}
            if isinstance(ex, dict):
                skbh = (ex.get("metadata") or {}).get("semantic_key_by_header") or {}
            hc = list(skbh.keys())[:cfg.max_source_keys_in_profile] if isinstance(skbh, dict) else []
            sr = [_trim_record_fields(r, cfg.record_trim_limit) for r in (all_recs or [])[:cfg.sample_records_count]]
            profile.append({"source_id": source.get("source_id"), "filename": source.get("filename"),
                            "record_keys": rk, "header_candidates": hc, "sample_records": sr})
        return profile

    # -- 公积金/社保分区 --

    @staticmethod
    def insurance_col_tag(header_path: str) -> str:
        """判断表头属于公积金(gj)还是社保(ss)分区。"""
        n = header_path.replace("\u3000", "").strip()
        return "gj" if ("公积金" in n or "公积" in n) else "ss"


# ============================================================================
# 2.4  InsuranceSourceSelector — 数据源选择
# ============================================================================


class InsuranceSourceSelector:
    """
    负责选择最合适的 Excel 数据源和推断关键字段 (name_key, effective_date_key)。

    先尝试 LLM，失败后降级到启发式 (文件名关键词 / 日期字段评分)。
    """

    def __init__(self, llm: LLMClient, config: InsuranceConfig = CONFIG):
        self._llm = llm
        self._config = config

    def select(
        self,
        sources: List[dict],
        extractor: SourceDataExtractor,
        template_intent: str,
        template_headers: List[str],
        sources_profile: List[Dict[str, Any]],
        warnings: List[str],
    ) -> Tuple[Optional[dict], Optional[str], Optional[str], bool, Optional[Any], Optional[str]]:
        """
        选择数据源 + 推断 name_key / effective_date_key。

        返回 (source, name_key, date_key, llm_used, confidence, notes)
        """
        response = self._call_llm(template_intent, template_headers, sources_profile)
        src, nk, dk, conf, notes = None, None, None, None, None
        llm_used = False

        if isinstance(response, dict):
            sid = response.get("selected_source_id")
            nk, dk = response.get("name_key"), response.get("effective_date_key")
            conf, notes = response.get("confidence"), response.get("notes")
            if isinstance(sid, str) and sid.strip():
                s = next((s for s in sources if s.get("source_id") == sid), None)
                if s:
                    recs = extractor.get_source_records(s)
                    if (recs and isinstance(nk, str) and nk.strip() and isinstance(dk, str) and dk.strip()
                            and any(nk in r for r in recs) and any(dk in r for r in recs)):
                        src, llm_used = s, True

        if not llm_used:
            if response is not None:
                warnings.append("insurance_llm_invalid_params")
            src = self._heuristic_source(sources, template_intent, extractor)
            if src:
                recs = extractor.get_source_records(src)
                keys = HeaderMatcher.collect_keys(recs)
                if keys:
                    nk = self._heuristic_name_key(recs, keys)
                    dk = self._heuristic_date_key(keys, template_intent)
        return src, nk, dk, llm_used, conf, notes

    def _call_llm(self, intent: str, headers: List[str], profile: List[Dict[str, Any]]) -> Optional[dict]:
        prompt = build_insurance_param_prompt(intent, headers, profile)
        try:
            resp = self._llm.chat_json_once(prompt, system=None, temperature=0, step="insurance_template_param", timeout=30)
        except Exception:
            return None
        if isinstance(resp, dict) and resp.get("error"): return None
        return resp if isinstance(resp, dict) else None

    def _heuristic_source(self, sources: List[dict], intent: str, ext: SourceDataExtractor) -> Optional[dict]:
        cfg = self._config
        # by filename
        kw = cfg.add_keywords[:2] if intent == "add" else cfg.remove_keywords[:2]
        for s in sources:
            fn = _normalize_match_text(s.get("filename") or "")
            if any(_normalize_match_text(k) in fn for k in kw):
                return s
        # by date keys
        dk = cfg.add_date_semantic_keys if intent == "add" else cfg.remove_date_semantic_keys
        best, best_sc = None, -1.0
        for s in sources:
            keys = HeaderMatcher.collect_keys(ext.get_source_records(s))
            if not keys: continue
            m = sum(1 for k in keys if any(_normalize_match_text(d) in _normalize_match_text(k) for d in dk))
            sc = m / max(len(keys), 1)
            if sc > best_sc: best_sc, best = sc, s
        return best

    def _heuristic_name_key(self, records: List[Dict[str, Any]], keys: List[str]) -> Optional[str]:
        for k in keys:
            if k.lower() == "name": return k
        for k in keys:
            if "姓名" in k or "姓名" in _normalize_match_text(k): return k
        bk, bh = None, 0
        for k in keys:
            h = sum(1 for r in records if _looks_like_chinese_name(r.get(k)))
            if h > bh: bh, bk = h, k
        return bk if bk and bh > 0 else None

    def _heuristic_date_key(self, keys: List[str], intent: str) -> Optional[str]:
        if intent == "add":
            for k in keys:
                if k.lower() == "start_date": return k
            for k in keys:
                if any(d in _normalize_match_text(k) for d in ["入职日期", "到岗日期"]): return k
        else:
            for k in keys:
                if k.lower() in ("leave_date", "end_date"): return k
            # 减员场景补充月度字段关键词，避免遗漏“减员月份/停保月份/费用年月”类列
            for k in keys:
                if any(d in _normalize_match_text(k) for d in ["减员月份", "减员年月", "停保月份", "停保年月", "费用年月"]): return k
            for k in keys:
                if any(d in _normalize_match_text(k) for d in ["离职日期", "终止日期"]): return k
        return None


# ============================================================================
# 2.5  InsuranceFillPlanner — 保险业务编排器
# ============================================================================


class InsuranceFillPlanner:
    """
    社保增减员填充规划的主编排器。

    协调 InsuranceBusinessLogic / InsuranceSourceSelector / GenericFillPlanner。
    """

    def __init__(
        self, schema: TemplateSchema, extracted_json: dict, llm: LLMClient,
        filename: Optional[str], intent: str, config: InsuranceConfig = CONFIG,
    ):
        self._schema = schema
        self._json = extracted_json
        self._filename = filename
        self._intent = intent
        self._config = config
        self._logic = InsuranceBusinessLogic(config)
        self._selector = InsuranceSourceSelector(llm, config)
        self._extractor = SourceDataExtractor(extracted_json)
        self._inspector = SchemaInspector(schema)

    def plan(self) -> FillPlan:
        """执行完整的保险模板填充规划。"""
        debug: Dict[str, Any] = {
            "template_intent": self._intent, "template_filename": self._filename,
            "planner_mode": "insurance_constrained_llm",
            "selected_source_filename": None, "name_key": None, "effective_date_key": None,
            "kept_records": 0, "skipped_records": 0, "llm_confidence": None, "notes": None,
            "remove_sheet_counts_before_filter": {},
            "remove_sheet_counts_after_filter": {},
            "date_key_fallback_hits": 0,
        }
        warnings: List[str] = []

        # 1. 验证模板
        region = self._inspector.get_main_table_region()
        if not region or not region.table or not region.table.header:
            return FillPlan(target=FillPlanTarget(), warnings=["insurance_template_missing_table"], llm_used=False, debug=debug)
        headers = region.table.header or []
        name_cols = SchemaInspector.find_columns_by_suffix(headers, "姓名")
        fee_cols = SchemaInspector.find_columns_with_paths_by_suffix(headers, "费用年月")
        type_cols = SchemaInspector.find_columns_with_paths_by_suffix(headers, "申报类型")
        if not name_cols or not fee_cols or not type_cols:
            return FillPlan(target=FillPlanTarget(), warnings=["insurance_template_headers_missing"], llm_used=False, debug=debug)

        # 2. 选择数据源
        sources = self._extractor.get_sources("excel")
        if not sources:
            return FillPlan(target=FillPlanTarget(), warnings=["insurance_no_sources"], llm_used=False, debug=debug)
        t_headers = [getattr(h, "header_path", None) for h in headers if isinstance(getattr(h, "header_path", None), str)]
        profile = self._logic.build_sources_profile(self._extractor)
        src, nk, dk, llm_used, conf, notes = self._selector.select(
            sources, self._extractor, self._intent, t_headers, profile, warnings
        )
        if self._intent != "remove" and (not src or not nk or not dk):
            return FillPlan(target=FillPlanTarget(), warnings=["insurance_fallback_failed"], llm_used=False, debug=debug)

        # 3. 收集/去重记录
        new_records, extra_debug, extra_warns = self._collect_records(src, nk, dk)
        warnings.extend(extra_warns)
        debug.update(extra_debug)
        if not new_records and self._intent == "remove":
            return FillPlan(target=FillPlanTarget(), warnings=["insurance_no_leave_records_found"], llm_used=llm_used, debug=debug)

        # 4. 组装
        try:
            min_col, min_row, max_col, max_row = range_boundaries(region.table.range)
        except Exception:
            return FillPlan(target=FillPlanTarget(), warnings=["insurance_template_range_invalid"], llm_used=llm_used, debug=debug)
        mhr = max(region.header_rows) if region.header_rows else min_row
        dsr = mhr + 1
        sc = f"{get_column_letter(min_col)}{dsr}"
        cer = min(max_row, dsr + len(new_records) + 30)
        cr = f"{get_column_letter(min_col)}{dsr}:{get_column_letter(max_col)}{cer}"

        cm: Dict[str, str] = {}
        declare_keys: List[str] = []
        fee_keys: List[str] = []
        if name_cols: cm["__name__"] = name_cols[0]
        if type_cols:
            cl, hp = type_cols[0]
            tag = InsuranceBusinessLogic.insurance_col_tag(hp)
            k = f"__declare_type__{tag}"; cm[k] = cl; declare_keys.append(k)
        if fee_cols:
            cl, hp = fee_cols[0]
            tag = InsuranceBusinessLogic.insurance_col_tag(hp)
            k = f"__fee_month__{tag}"; cm[k] = cl; fee_keys.append(k)

        rows = []
        for r in new_records:
            rd: Dict[str, Any] = {"__name__": r["__name__"]}
            for k in declare_keys: rd[k] = r["__declare_type__"]
            for k in fee_keys: rd[k] = r["__fee_month__"]
            rows.append(rd)

        debug.update({
            "selected_source_filename": src.get("filename") if isinstance(src, dict) else None,
            "name_key": nk, "effective_date_key": dk, "kept_records": len(new_records),
            "skipped_records": debug.get("skipped_records", 0), "llm_confidence": conf, "notes": notes,
        })
        sn = self._inspector.resolve_sheet_name(self._schema.sheet_schemas[0] if self._schema.sheet_schemas else None)
        fp = {
            "target": {"sheet_name": sn, "sheet": sn, "region_id": region.region_id,
                        "layout_type": "table", "clear_policy": "clear_values_keep_format"},
            "clear_ranges": [cr],
            "row_writes": [{"start_cell": sc, "rows": rows, "column_mapping": cm}],
            "writes": [], "warnings": warnings, "llm_used": llm_used, "constant_values_count": 0,
        }
        _merge_debug_info(fp, debug)
        return _dict_to_fill_plan(fp)

    def _collect_records(self, src: Optional[dict], nk: Optional[str], dk: Optional[str]) -> Tuple[List[Dict[str, Any]], Dict[str, Any], List[str]]:
        """根据意图收集/合并/去重记录。"""
        new: List[Dict[str, Any]] = []
        warns: List[str] = []
        skipped = dpf = fmf = 0
        dt = "增" if self._intent == "add" else "减"
        erc = xrc = dup = 0
        dbg: Dict[str, Any] = {}

        if self._intent == "remove":
            # 减员: 合并邮件 + Excel 离职记录
            email_recs, ew = self._logic.collect_email_leave_records(self._extractor)
            warns.extend(ew); erc = len(email_recs)
            excel_recs: List[Dict[str, Any]] = []
            remove_sheet_counts_before_filter: Dict[str, int] = {}
            remove_sheet_counts_after_filter: Dict[str, int] = {}
            date_key_fallback_hits = 0
            if src and nk:
                raw_all = self._extractor.get_source_records(src)
                for rec in raw_all:
                    if not isinstance(rec, dict):
                        continue
                    sheet_name = str(rec.get("__sheet_name__", "") or "")
                    remove_sheet_counts_before_filter[sheet_name] = remove_sheet_counts_before_filter.get(sheet_name, 0) + 1

                raw, _ = self._logic.filter_by_sheet_intent(raw_all, self._intent)
                for rec in raw:
                    if not isinstance(rec, dict):
                        continue
                    sheet_name = str(rec.get("__sheet_name__", "") or "")
                    remove_sheet_counts_after_filter[sheet_name] = remove_sheet_counts_after_filter.get(sheet_name, 0) + 1

                fallback_date_keys = ["leave_date", "end_date", "ss_end_month", "hf_end_month", "termination_date", "减员月份", "停保月份"]
                candidate_date_keys: List[str] = []
                if isinstance(dk, str) and dk.strip():
                    candidate_date_keys.append(dk.strip())
                for k in fallback_date_keys:
                    if k not in candidate_date_keys:
                        candidate_date_keys.append(k)

                for r in raw:
                    if not isinstance(r, dict): continue
                    name = str(r.get(nk, "")).strip()
                    if not name: continue
                    ldr = None
                    d = None
                    used_date_key: Optional[str] = None
                    for date_key in candidate_date_keys:
                        if not _is_non_empty_value(r.get(date_key)):
                            continue
                        candidate_value = r.get(date_key)
                        parsed = _parse_any_date(candidate_value)
                        if parsed:
                            ldr = candidate_value
                            d = parsed
                            used_date_key = date_key
                            break
                        if ldr is None:
                            ldr = candidate_value
                    if (
                        used_date_key
                        and isinstance(dk, str)
                        and dk.strip()
                        and used_date_key != dk.strip()
                    ):
                        date_key_fallback_hits += 1
                    excel_recs.append({
                        "name": name, "employee_id": str(r.get("employee_id", "") or r.get("工号", "") or "").strip(),
                        "leave_date": d.strftime("%Y-%m-%d") if d else "",
                        "leave_date_text": str(ldr or ""), "intent": "remove",
                        "__source_file__": src.get("filename", "excel"), "__source_type__": "excel",
                    })
                xrc = len(excel_recs)
            deduped, dup = InsuranceBusinessLogic.deduplicate_leave_records(email_recs + excel_recs)
            if dup: warns.append(f"insurance_duplicates_removed:{dup}")
            for r in deduped:
                name = r.get("name", "").strip()
                if not name: skipped += 1; continue
                ld = r.get("leave_date", "").strip()
                if not ld: dpf += 1; skipped += 1; continue
                d = _parse_any_date(ld)
                if not d: dpf += 1; skipped += 1; continue
                fm = _next_month_yyyymm(d)
                if not fm: fmf += 1; skipped += 1; continue
                new.append({"__name__": name, "__fee_month__": fm, "__declare_type__": "减"})
            dbg.update({
                "email_records_count": erc,
                "excel_records_count": xrc,
                "duplicate_count": dup,
                "remove_sheet_counts_before_filter": remove_sheet_counts_before_filter,
                "remove_sheet_counts_after_filter": remove_sheet_counts_after_filter,
                "date_key_fallback_hits": date_key_fallback_hits,
            })
        else:
            if not src or not nk or not dk:
                return [], {"skipped_records": 0}, ["insurance_fallback_failed"]
            recs = self._extractor.get_source_records(src)
            recs, _ = self._logic.filter_by_sheet_intent(recs, self._intent)
            for r in recs:
                if not isinstance(r, dict): continue
                name = str(r.get(nk, "")).strip()
                if not name: skipped += 1; continue
                d = _parse_any_date(r.get(dk))
                if not d: dpf += 1; skipped += 1; continue
                fm = _next_month_yyyymm(d)
                if not fm: fmf += 1; skipped += 1; continue
                new.append({"__name__": name, "__fee_month__": fm, "__declare_type__": dt})
        if skipped: warns.append(f"insurance_records_skipped:{skipped}")
        if dpf: warns.append(f"insurance_date_parse_failed:{dpf}")
        if fmf: warns.append(f"insurance_fee_month_failed:{fmf}")
        dbg["skipped_records"] = skipped
        return new, dbg, warns




# ============================================================================
# 领域管线函数 — 向后兼容 (从 fill_planner.py 迁移)
# ============================================================================
# 这些函数以前定义在 fill_planner.py 的兼容层中，因为它们包含大量领域逻辑，
# 按"瘦编排器"原则迁移到本模块。fill_planner.py 通过 import 重新导出它们。

def _apply_derived_fields_to_records(
    schema: TemplateSchema, records: List[Dict[str, Any]], derived_fields: List[Dict[str, Any]],
    template_intent: Optional[str], warnings: List[str],
) -> List[Dict[str, Any]]:
    if not records or not derived_fields: return records
    for ri, rec in enumerate(records):
        if not isinstance(rec, dict): continue
        for di, item in enumerate(derived_fields):
            if not isinstance(item, dict): continue
            nk, op, args = item.get("new_key"), item.get("op"), item.get("args", {})
            if not isinstance(nk, str) or not nk or op != "MONTH_FROM_DATE": continue
            if not isinstance(args, dict):
                warnings.append(f"derived_fields[{di}] invalid args; skipped"); rec[nk] = ""; continue
            sv = None
            for k in (args.get("source_keys") or []):
                if isinstance(k, str):
                    v = rec.get(k)
                    if _is_non_empty_value(v): sv = v; break
            if not _is_non_empty_value(sv): rec[nk] = ""; continue
            ym = _parse_year_month(sv)
            if not ym:
                warnings.append(f"derived_fields[{di}] parse failed for record {ri}"); rec[nk] = ""; continue
            fmt = args.get("output_format", "YYYY-MM")
            if fmt == "from_template_sample":
                fmt = _resolve_year_month_format_from_template(schema, item.get("_template_headers", []))
            rec[nk] = _format_year_month(ym[0], ym[1], fmt)
    return records



def _resolve_year_month_format_from_template(schema: TemplateSchema, headers: List[str]) -> str:
    for row in SchemaInspector(schema).get_sample_rows():
        if not isinstance(row, dict): continue
        for h in (headers or []):
            if h in row and row[h]:
                f = _detect_year_month_format_from_value(row[h])
                if f: return f
    return "YYYY-MM"



def _apply_record_gating_to_extracted_json(extracted_json: dict, schema: TemplateSchema, warnings: List[str]) -> dict:
    if not isinstance(extracted_json, dict): return extracted_json
    matcher = HeaderMatcher(CONFIG.fuzzy_match_threshold)
    t_headers = SchemaInspector(schema).get_all_header_paths()
    def gate(recs: list) -> list:
        if not recs: return recs
        sk: Set[str] = set()
        for r in recs:
            if isinstance(r, dict): sk.update(k for k in r if isinstance(k, str))
        matched = matcher.match_headers_to_keys(t_headers, sorted(sk))
        mk = [k for k, _ in matched.values()]
        if not mk: warnings.append("record_gating: no matched keys; skipping"); return recs
        scored = [(sum(1 for k in mk if _is_non_empty_value(r.get(k))) / len(mk) if isinstance(r, dict) else 0.0, r) for r in recs]
        kept = [r for ratio, r in scored if ratio >= CONFIG.record_gating_min_ratio]
        if not kept and recs:
            scored.sort(key=lambda x: x[0], reverse=True)
            kept = [r for _, r in scored[:CONFIG.record_gating_fallback_top_n]]
            warnings.append("record_gating: all below threshold; keeping top")
        return kept
    filt: Dict[str, Any] = {}
    if "sources" in extracted_json:
        filt["sources"] = []
        for s in extracted_json.get("sources", []):
            if not isinstance(s, dict): filt["sources"].append(s); continue
            fs = dict(s); ex = s.get("extracted")
            if isinstance(ex, dict):
                ue = dict(ex)
                for k, v in ex.items():
                    if isinstance(v, list) and v and isinstance(v[0], dict): ue[k] = gate(v)
                fs["extracted"] = ue
            elif isinstance(ex, list) and ex and isinstance(ex[0], dict):
                fs["extracted"] = gate(ex)
            filt["sources"].append(fs)
    if "merged" in extracted_json: filt["merged"] = extracted_json.get("merged", {})
    for k in extracted_json:
        if k not in filt: filt[k] = extracted_json[k]
    return filt



def _record_filter_conflicts_with_intent(record_filter: Dict[str, Any], template_intent: str) -> bool:
    if not template_intent or not record_filter: return False
    vt = _normalize_text(" ".join(str(v) for v in (record_filter.get("values") or [])))
    excl = bool(record_filter.get("exclude", False))
    cfg = CONFIG
    ha = any(k in vt for k in cfg.add_keywords)
    hr = any(k in vt for k in cfg.remove_keywords)
    if template_intent == "add":
        return (not excl and hr and not ha) or (excl and ha and not hr)
    if template_intent == "remove":
        return (not excl and ha and not hr) or (excl and hr and not ha)
    return False



def _infer_template_intent_from_mapping(
    schema: TemplateSchema, filename: Optional[str],
    column_mapping: Optional[Dict[str, str]], constant_values: Optional[Dict[str, str]],
) -> Optional[str]:
    for k, v in (constant_values or {}).items():
        kt = _normalize_text(k)
        if "申报类型" not in kt and "declaretype" not in kt: continue
        vt = _normalize_text(v)
        a = any(x in vt for x in ("增", "增员", "入职", "新增"))
        r = any(x in vt for x in ("减", "减员", "离职", "退工", "退保"))
        if a and not r: return "add"
        if r and not a: return "remove"
    for k in (column_mapping or {}):
        n = _normalize_header(k) if isinstance(k, str) else ""
        if n in ("terminationdate", "terminationreason"): return "remove"
    for k in (column_mapping or {}):
        if isinstance(k, str) and _normalize_header(k) == "startdate": return "add"
    return _detect_template_intent(filename, schema)



def _apply_record_type_filter(extracted_json: dict, template_intent: Optional[str]) -> dict:
    if template_intent not in ("add", "remove") or not isinstance(extracted_json, dict): return extracted_json
    all_recs = _extract_records(extracted_json)
    has = any((_infer_record_type(r) or _infer_record_type_from_sheet(r)) == template_intent for r in all_recs if isinstance(r, dict))
    if not has: return extracted_json
    def ok(r: Dict[str, Any]) -> bool: return (_infer_record_type(r) or _infer_record_type_from_sheet(r)) == template_intent
    def fl(recs: list) -> list: return [r for r in recs if isinstance(r, dict) and ok(r)]
    def fe(ex: Any) -> Any:
        if ex is None: return None
        if isinstance(ex, list): return fl(ex) if ex and isinstance(ex[0], dict) else ex
        if isinstance(ex, dict):
            return {k: (fl(v) if isinstance(v, list) and v and isinstance(v[0], dict) else v) for k, v in ex.items()}
        return ex
    filt: Dict[str, Any] = {}
    if "sources" in extracted_json:
        filt["sources"] = [dict(s, extracted=fe(s.get("extracted"))) if isinstance(s, dict) else s for s in extracted_json["sources"]]
    if "merged" in extracted_json and isinstance(extracted_json["merged"], dict):
        filt["merged"] = {k: (fl(v) if isinstance(v, list) and v and isinstance(v[0], dict) else v) for k, v in extracted_json["merged"].items()}
    for k in extracted_json:
        if k not in filt: filt[k] = extracted_json[k]
    return filt



def _auto_filter_records_by_template_intent(extracted_json: dict, template_intent: Optional[str]) -> dict:
    if not template_intent or not isinstance(extracted_json, dict): return extracted_json
    logic = InsuranceBusinessLogic(CONFIG)
    return _apply_auto_intent_filter(extracted_json, template_intent, logic)


def _apply_auto_intent_filter(extracted_json: dict, intent: str, logic: InsuranceBusinessLogic) -> dict:
    """Internal implementation of auto intent filtering."""
    all_recs = _extract_records(extracted_json)
    def _intents(r: Dict[str, Any]) -> List[str]:
        out: List[str] = []
        si = logic.infer_source_intent(r.get("__source_file__"))
        if si: out.append(si)
        rec = InsuranceRecord(r, CONFIG)
        ei = rec.infer_explicit_intent()
        if ei: out.append(ei)
        ki = rec.infer_intent_from_keys_and_values()
        if ki: out.append(ki)
        return out
    has_sig = any(intent in _intents(r) for r in all_recs if isinstance(r, dict))
    def ok(r: Dict[str, Any], strict: bool = True) -> bool:
        rec = InsuranceRecord(r, CONFIG)
        rt = rec.infer_record_type() or rec.infer_type_from_sheet()
        if rt is not None: return rt == intent
        si = logic.infer_source_intent(r.get("__source_file__"))
        if si and si != intent: return False
        ei = rec.infer_explicit_intent()
        if ei and ei != intent: return False
        if strict:
            ki = rec.infer_intent_from_keys_and_values()
            if ki is None: return (si == intent or ei == intent) if has_sig else True
            return ki == intent
        return True
    def fl(recs: list, strict: bool = True) -> list: return [r for r in recs if isinstance(r, dict) and ok(r, strict)]
    def fe(ex: Any) -> Any:
        if ex is None: return None
        if isinstance(ex, list) and ex and isinstance(ex[0], dict):
            f = fl(ex, True); return f if f else fl(ex, False)
        if isinstance(ex, dict):
            out = {}
            for k, v in ex.items():
                if isinstance(v, list) and v and isinstance(v[0], dict):
                    f = fl(v, True); out[k] = f if f else fl(v, False)
                else: out[k] = v
            return out
        return ex
    filt: Dict[str, Any] = {}
    if "sources" in extracted_json:
        filt["sources"] = [dict(s, extracted=fe(s.get("extracted"))) if isinstance(s, dict) else s for s in extracted_json["sources"]]
    if "merged" in extracted_json and isinstance(extracted_json["merged"], dict):
        filt["merged"] = {k: (fl(v) if isinstance(v, list) and v and isinstance(v[0], dict) else v) for k, v in extracted_json["merged"].items()}
    for k in extracted_json:
        if k not in filt: filt[k] = extracted_json[k]
    return filt



def _auto_infer_common_fields(
    schema: TemplateSchema, extracted_json: dict,
    column_mapping: Optional[Dict[str, str]], constant_values: Optional[Dict[str, str]],
    template_intent: Optional[str], template_sample_rows: List[Dict[str, Any]],
) -> Tuple[Dict[str, str], Dict[str, str]]:
    cfg = CONFIG
    cm = dict(column_mapping or {})
    cv = dict(constant_values or {})
    recs = SourceDataExtractor(extracted_json).get_all_records()
    ek = HeaderMatcher.collect_keys(recs)
    headers = SchemaInspector(schema).get_all_header_paths()

    def find_key(cands: List[str]) -> Optional[str]:
        for c in cands:
            cn = _normalize_header(c)
            for k in ek:
                kn = _normalize_header(k)
                if cn == kn or cn in kn or kn in cn: return k
        return None

    def sample_val(h: str) -> Optional[str]:
        for row in template_sample_rows:
            if h in row and row[h]: return str(row[h])
        return None

    def ym_fmt(sv: Optional[str]) -> str:
        if not sv: return "YYYY-MM"
        if "年" in sv and "月" in sv: return "YYYY年MM月"
        if "-" in sv: return "YYYY-MM"
        if "/" in sv: return "YYYY/MM"
        if len(sv) == 6 and sv.isdigit(): return "YYYYMM"
        return "YYYY-MM"

    def ym_text(t: Any) -> Optional[Tuple[int, int]]:
        if t is None: return None
        m = re.search(r"(20\d{2})[年\-\./]?(\d{1,2})", str(t))
        if not m: return None
        y, mo = int(m.group(1)), int(m.group(2))
        return (y, mo) if 1 <= mo <= 12 else None

    def ym_recs() -> Optional[Tuple[int, int]]:
        for r in recs:
            for k in list(cfg.fee_month_direct_keys) + ["离职日期", "入职日期", "变动日期", "申报日期", "生效日期", "开始日期", "结束日期", "办理日期"]:
                if k in r and r[k]:
                    ym = ym_text(r[k])
                    if ym: return ym
            for k, v in r.items():
                if isinstance(k, str) and ("日期" in k or "时间" in k):
                    ym = ym_text(v)
                    if ym: return ym
        return None

    def ym_fns() -> Optional[Tuple[int, int]]:
        for s in extracted_json.get("sources", []):
            if isinstance(s, dict):
                ym = ym_text(s.get("filename"))
                if ym: return ym
        return None

    for h in headers:
        if h in cm or h in cv: continue
        hn = _normalize_header(h)
        if any(k in hn for k in [_normalize_header(x) for x in cfg.fee_month_header_exact]):
            k = find_key(list(cfg.fee_month_header_exact))
            if k: cm[h] = k; continue
            ym = ym_recs() or ym_fns()
            if ym: cv[h] = _format_year_month(ym[0], ym[1], ym_fmt(sample_val(h))); continue
        if any(k in hn for k in [_normalize_header(x) for x in cfg.declare_type_header_synonyms]):
            k = find_key(list(cfg.declare_type_header_synonyms))
            if k: cm[h] = k; continue
            if template_intent == "add": cv[h] = "增员"
            elif template_intent == "remove": cv[h] = "减员"
    return cm, cv

