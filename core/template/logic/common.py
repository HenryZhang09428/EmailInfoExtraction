"""
通用核心层 (Generic Core — Layer 1)
====================================

本模块包含所有与业务域无关的通用逻辑，可直接复用于任意 Excel 模板填充场景。

严格约束: 本文件中 **不得出现** 任何保险/社保业务关键词
（如"增员"、"减员"、"费用年月"、"申报类型"等中文业务术语）。

包含:
  - 通用工具函数 (归一化、日期解析、值判定等)
  - FillPlan 构造辅助
  - SourceDataExtractor  — 从嵌套 JSON 中扁平化提取记录 + 来源溯源
  - HeaderMatcher        — 字符串归一化 + rapidfuzz 模糊匹配
  - SchemaInspector      — 解析 TemplateSchema，定位区域/表头/样本行
  - LLMResponseParser    — 解析 LLM 返回的映射方案
  - GenericFillPlanner   — 基于 header→key 映射 + 记录列表 → FillPlan

依赖: 仅依赖标准库、openpyxl、rapidfuzz 和项目内的 core.ir / core.template.schema。
      不依赖 config.py 或 insurance.py (单向依赖)。
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from typing import (
    Any,
    Callable,
    Dict,
    List,
    Optional,
    Set,
    Tuple,
)

from openpyxl.utils import get_column_letter, range_boundaries
from rapidfuzz import fuzz, process

from core.ir import CellWrite, FillPlan, FillPlanTarget, RowWrite
from core.logger import get_logger
from core.template.schema import TemplateSchema

logger = get_logger(__name__)


# ============================================================================
# 通用工具函数
# ============================================================================


def _normalize_match_text(text: Any) -> str:
    """
    NFKC Unicode 归一化 + 小写 + 去除空白与连接符。

    用于关键词匹配时忽略全角/半角、大小写、空格/下划线差异。
    例如: 'Ａ Ｂ Ｃ' → 'abc', 'Start_Date' → 'startdate'
    """
    if text is None:
        return ""
    normalized = unicodedata.normalize("NFKC", str(text)).lower()
    normalized = re.sub(r'[\s_\-/]+', '', normalized)
    return normalized


def _normalize_header(text: str) -> str:
    """简化版表头归一化: 去空格/下划线/连字符 + 小写。"""
    return text.replace(" ", "").replace("_", "").replace("-", "").strip().lower()


def _normalize_text(text: Any) -> str:
    """通用文本归一化（同 _normalize_header，适用更广泛场景）。"""
    return str(text).lower().replace(" ", "").replace("_", "").replace("-", "").strip()


def _header_variants(header_path: str) -> List[str]:
    """
    为一个表头路径生成匹配变体列表。

    对于多级路径（以 '/' 分隔），额外生成每级的归一化文本，
    提高模糊匹配时的命中率。
    """
    variants: Set[str] = set()
    if not header_path:
        return []
    variants.add(_normalize_match_text(header_path))
    if "/" in header_path:
        parts = [p.strip() for p in header_path.split("/") if p.strip()]
        for part in parts:
            variants.add(_normalize_match_text(part))
        if parts:
            variants.add(_normalize_match_text(parts[-1]))
    return [v for v in variants if v]


def _is_non_empty_value(value: Any) -> bool:
    """判断值是否"非空"——None 或空白字符串视为空。"""
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _parse_any_date(value: Any) -> Optional[datetime]:
    """
    尽最大努力把各种格式的日期值解析为 datetime。

    支持: datetime / date 对象、ISO 字符串(含Z)、YYYY-MM-DD、YYYYMMDD、纯数字等。
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        value = str(int(value))
    if not isinstance(value, str):
        value = str(value)
    text = value.strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        pass
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', text)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except (ValueError, TypeError):
            return None
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', text)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except (ValueError, TypeError):
            return None
    return None


def _looks_like_chinese_name(value: Any) -> bool:
    """判断字符串是否像中文姓名 (2-4 个汉字)。"""
    if not isinstance(value, str):
        return False
    text = value.strip()
    return bool(text and re.match(r'^[\u4e00-\u9fff]{2,4}$', text))


def _get_record_value(record: Dict[str, Any], key: str) -> Any:
    """从记录中获取值，支持 '.' 或 '/' 分隔的嵌套路径。"""
    if key in record:
        return record.get(key)
    if "." in key or "/" in key:
        parts = [p for p in key.replace("/", ".").split(".") if p]
        current: Any = record
        for part in parts:
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                return None
        return current
    return None


def _trim_record_fields(record: Dict[str, Any], limit: int) -> Dict[str, Any]:
    """截断记录字段数到 limit。"""
    if not isinstance(record, dict):
        return {}
    return dict(list(record.items())[:limit])


# ============================================================================
# FillPlan 构造辅助
# ============================================================================


def _dict_to_fill_plan(data: dict) -> FillPlan:
    """将字典转换为经验证的 FillPlan 对象。"""
    try:
        return FillPlan.from_dict(data)
    except Exception as e:
        logger.warning("Failed to convert dict to FillPlan: %s", e)
        return FillPlan(
            target=FillPlanTarget(),
            warnings=[f"Failed to validate fill plan: {str(e)}"],
        )


def _empty_fill_plan(warnings: Optional[List[str]] = None, debug: Optional[dict] = None) -> FillPlan:
    """创建空 FillPlan。"""
    return FillPlan(target=FillPlanTarget(), warnings=warnings or [], debug=debug)


def _merge_debug_info(fill_plan_dict: dict, debug_info: dict) -> None:
    """合并调试信息。"""
    if not isinstance(fill_plan_dict, dict) or not isinstance(debug_info, dict):
        return
    existing = fill_plan_dict.get("debug")
    if isinstance(existing, dict):
        merged = dict(existing)
        merged.update({k: v for k, v in debug_info.items() if v is not None})
        fill_plan_dict["debug"] = merged
    else:
        fill_plan_dict["debug"] = {k: v for k, v in debug_info.items() if v is not None}


def _append_warning(warnings: List[str], message: str) -> None:
    """追加去重警告。"""
    if message and warnings is not None and message not in warnings:
        warnings.append(message)


# ============================================================================
# SourceDataExtractor — 通用数据提取器
# ============================================================================


class SourceDataExtractor:
    """
    从嵌套的 extracted_json 结构中扁平化提取记录。

    该类完全与业务域无关——它只理解 extracted_json 的嵌套结构
    (sources 列表 / extracted 字典 / data 数组)，
    不做任何基于关键词的过滤或推断。

    核心规则:
      - 记录只从 sources[*].extracted 中提取 (merged 不提供记录)
      - 每条记录保证有 __source_file__ 字段标识来源
    """

    CONTAINER_PRIORITY_KEYS = ("data", "records", "rows", "items", "extracted_data", "table_data")

    def __init__(self, extracted_json: Any):
        self._json = extracted_json if isinstance(extracted_json, dict) else {}

    @property
    def raw(self) -> dict:
        return self._json

    def get_sources(self, source_type: Optional[str] = None) -> List[dict]:
        """获取所有（或指定类型的）数据源字典。"""
        sources = self._json.get("sources", [])
        if not isinstance(sources, list):
            return []
        result = [s for s in sources if isinstance(s, dict)]
        if source_type:
            result = [s for s in result if s.get("source_type") == source_type]
        return result

    @staticmethod
    def extract_from_container(container: Any) -> List[Dict[str, Any]]:
        """
        从一个 extracted 容器中提取记录列表。

        容器可能是:
          - 直接的 list[dict]
          - dict 中 "data"/"records"/"rows" 等键对应的 list[dict]
        """
        if isinstance(container, list) and len(container) > 0:
            if isinstance(container[0], dict):
                return list(container)
            if isinstance(container[0], list):
                return [item for item in container if isinstance(item, dict)]
        if isinstance(container, dict):
            for key in SourceDataExtractor.CONTAINER_PRIORITY_KEYS:
                value = container.get(key)
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    return list(value)
            for value in container.values():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    return list(value)
        return []

    def get_source_records(self, source: dict) -> List[Dict[str, Any]]:
        """获取单个数据源的记录列表。"""
        extracted = source.get("extracted") or {}
        if not isinstance(extracted, dict):
            return []
        data_records = extracted.get("data")
        if isinstance(data_records, list):
            return [r for r in data_records if isinstance(r, dict)]
        return []

    def get_all_records(self) -> List[Dict[str, Any]]:
        """从所有数据源提取全部记录。"""
        all_records: List[Dict[str, Any]] = []
        sources = self._json.get("sources", [])
        if not isinstance(sources, list):
            return all_records
        for s in sources:
            if not isinstance(s, dict):
                continue
            records = self.extract_from_container(s.get("extracted"))
            if records:
                all_records.extend(records)
        return all_records

    def get_all_records_with_provenance(
        self,
    ) -> Tuple[List[Dict[str, Any]], Dict[str, int], List[str]]:
        """
        提取全部记录并注入 __source_file__ 来源标记。

        返回 (records, counts_per_source, contributing_filenames)
        """
        all_records: List[Dict[str, Any]] = []
        counts: Dict[str, int] = {}
        contributing_files: List[str] = []
        sources = self._json.get("sources", [])
        if not isinstance(sources, list):
            return all_records, counts, contributing_files
        for s_idx, s in enumerate(sources):
            if not isinstance(s, dict):
                continue
            filename = s.get("filename", f"source_{s_idx}")
            records = self.extract_from_container(s.get("extracted")) or []
            normalized: List[Dict[str, Any]] = []
            for r in records:
                if not isinstance(r, dict):
                    continue
                rec = dict(r)
                if "__source_file__" not in rec or not rec.get("__source_file__"):
                    rec["__source_file__"] = filename
                normalized.append(rec)
            if normalized:
                all_records.extend(normalized)
                counts[filename] = counts.get(filename, 0) + len(normalized)
                if filename not in contributing_files:
                    contributing_files.append(filename)
        return all_records, counts, contributing_files

    def count_records_in_sources(self) -> int:
        """统计所有数据源的记录总数。"""
        total = 0
        for s in self._json.get("sources", []):
            if isinstance(s, dict):
                total += len(self.extract_from_container(s.get("extracted")) or [])
        return total


# ============================================================================
# HeaderMatcher — 通用模糊匹配器
# ============================================================================


class HeaderMatcher:
    """
    纯粹的字符串匹配引擎。

    使用 rapidfuzz (WRatio 算法) 在模板表头与数据源字段名之间建立映射。
    不包含任何业务关键词——匹配阈值通过参数注入。
    """

    def __init__(self, fuzzy_threshold: int = 80):
        self._threshold = fuzzy_threshold

    def match_headers_to_keys(
        self, template_headers: List[str], source_keys: List[str]
    ) -> Dict[str, Tuple[str, int]]:
        """
        将模板表头与数据源字段做模糊匹配。

        返回 {template_header: (best_source_key, score)}。
        """
        if not template_headers or not source_keys:
            return {}
        norm_key_map: Dict[str, str] = {}
        for key in sorted(source_keys):
            norm = _normalize_match_text(key)
            if norm and norm not in norm_key_map:
                norm_key_map[norm] = key
        norm_keys = list(norm_key_map.keys())
        matched: Dict[str, Tuple[str, int]] = {}
        for header in template_headers:
            best_score, best_norm = -1, None
            for variant in _header_variants(header):
                result = process.extractOne(variant, norm_keys, scorer=fuzz.WRatio)
                if result and result[1] > best_score:
                    best_score = result[1]
                    best_norm = result[0]
            if best_norm and best_score >= self._threshold:
                matched[header] = (norm_key_map[best_norm], int(best_score))
        return matched

    def enrich_mapping(
        self,
        headers: List[Any],
        extracted_keys: List[str],
        existing_mapping: Dict[str, str],
    ) -> Tuple[Dict[str, str], List[str]]:
        """对 LLM 未覆盖的表头做补充模糊匹配。"""
        if not headers or not extracted_keys:
            return existing_mapping, []
        mapped = set(existing_mapping.keys())
        enriched = dict(existing_mapping)
        warnings: List[str] = []
        for header in headers:
            path = getattr(header, "header_path", None)
            if not path or path in mapped:
                continue
            result = process.extractOne(path, extracted_keys, scorer=fuzz.WRatio)
            if not result:
                continue
            best_key, score, _ = result
            if score > self._threshold and best_key not in enriched.values():
                enriched[path] = best_key
                warnings.append(f"Auto-mapped '{path}' to '{best_key}' (score: {int(score)})")
        return enriched, warnings

    @staticmethod
    def collect_keys(records: List[Dict[str, Any]]) -> List[str]:
        """从记录列表收集所有唯一字段名。"""
        keys: List[str] = []
        seen: Set[str] = set()
        for record in records:
            if not isinstance(record, dict):
                continue
            for key in record.keys():
                if isinstance(key, str) and key not in seen:
                    seen.add(key)
                    keys.append(key)
        return keys


# ============================================================================
# SchemaInspector — 通用模板结构分析
# ============================================================================


class SchemaInspector:
    """
    分析 TemplateSchema 的通用工具。

    只处理结构信息（区域、表头、列字母、样本行），
    不做任何基于关键词的业务判断。
    """

    def __init__(self, template_schema: TemplateSchema):
        self._schema = template_schema

    def get_main_table_region(self) -> Optional[Any]:
        """获取第一个含表格的区域。"""
        if not self._schema.sheet_schemas:
            return None
        sheet = self._schema.sheet_schemas[0]
        if not sheet.regions:
            return None
        for region in sheet.regions:
            if region.table and region.table.header:
                return region
        return None

    def get_all_header_paths(self) -> List[str]:
        """获取所有表头路径。"""
        result: List[str] = []
        for ss in self._schema.sheet_schemas:
            for region in ss.regions:
                if not region.table or not region.table.header:
                    continue
                for h in region.table.header:
                    if getattr(h, "header_path", None):
                        result.append(h.header_path)
        return result

    def get_sample_rows(self) -> List[Dict[str, Any]]:
        """提取模板样本行 (最多 3 条)。"""
        rows: List[Dict[str, Any]] = []
        for ss in self._schema.sheet_schemas:
            for region in ss.regions:
                if region.table and region.table.sample_rows:
                    rows.extend(region.table.sample_rows)
                    if len(rows) >= 3:
                        return rows[:3]
        return rows

    @staticmethod
    def find_columns_by_suffix(headers: List[Any], suffix: str) -> List[str]:
        """收集所有以指定后缀结尾的列字母。"""
        if not headers or not suffix:
            return []
        suffix_norm = _normalize_match_text(suffix)
        columns: List[str] = []
        for header in headers:
            for attr in ("header_path", "header_text", "header", "text", "name"):
                val = getattr(header, attr, None)
                if not isinstance(val, str) or not val.strip():
                    continue
                for variant in _header_variants(val):
                    if variant.endswith(suffix_norm):
                        col = getattr(header, "col_letter", None)
                        if col:
                            columns.append(col)
                        break
        return columns

    @staticmethod
    def find_columns_with_paths_by_suffix(
        headers: List[Any], suffix: str
    ) -> List[Tuple[str, str]]:
        """收集匹配后缀的 (col_letter, header_path) 对。"""
        if not headers or not suffix:
            return []
        suffix_norm = _normalize_match_text(suffix)
        columns: List[Tuple[str, str]] = []
        for header in headers:
            path = getattr(header, "header_path", None) or ""
            for attr in ("header_path", "header_text", "header", "text", "name"):
                val = getattr(header, attr, None)
                if not isinstance(val, str) or not val.strip():
                    continue
                for variant in _header_variants(val):
                    if variant.endswith(suffix_norm):
                        col = getattr(header, "col_letter", None)
                        if col:
                            columns.append((col, path))
                        break
        return columns

    @staticmethod
    def find_header_path_by_suffix(headers: List[Any], suffix: str) -> Optional[str]:
        """查找以指定后缀结尾的表头路径。"""
        if not headers or not suffix:
            return None
        suffix_norm = _normalize_match_text(suffix)
        for header in headers:
            for attr in ("header_path", "header_text", "header", "text", "name"):
                val = getattr(header, attr, None)
                if not isinstance(val, str) or not val.strip():
                    continue
                for variant in _header_variants(val):
                    if variant.endswith(suffix_norm):
                        return getattr(header, "header_path", None) or val
        return None

    @staticmethod
    def extract_sheet_title(sheet_schema: Any) -> Optional[str]:
        if not sheet_schema:
            return None
        for attr in ("sheet_name", "sheet", "title", "name"):
            val = getattr(sheet_schema, attr, None)
            if isinstance(val, str) and val.strip():
                return val.strip()
        return None

    def resolve_sheet_name(self, sheet_schema: Any = None) -> str:
        name = self.extract_sheet_title(sheet_schema)
        if name:
            return name
        for ss in self._schema.sheet_schemas or []:
            name = self.extract_sheet_title(ss)
            if name:
                return name
        return "Sheet1"

    @staticmethod
    def build_header_lookup(headers: List[Any]) -> Dict[str, Any]:
        """构建表头查找字典: 精确 / 归一化 / 短名。"""
        lookup: Dict[str, Any] = {}
        for header in headers:
            path = getattr(header, "header_path", None)
            if not path:
                continue
            variants: Set[str] = {path, _normalize_header(path)}
            if "/" in path:
                short = path.split("/")[-1].strip()
                if short:
                    variants.update({short, _normalize_header(short)})
            for v in variants:
                if v and v not in lookup:
                    lookup[v] = header
        return lookup

    @staticmethod
    def lookup_header(lookup: Dict[str, Any], template_header: str) -> Optional[Any]:
        if template_header in lookup:
            return lookup[template_header]
        return lookup.get(_normalize_header(template_header))


# ============================================================================
# LLMResponseParser — 通用 LLM 响应解析
# ============================================================================


class LLMResponseParser:
    """
    解析 LLM 返回的 JSON 映射方案。

    纯解析逻辑——不包含任何业务关键词，不执行 LLM 调用本身。
    """

    @staticmethod
    def parse_column_mapping(resp: Any) -> Optional[Dict[str, str]]:
        """从 LLM 响应提取 column_mapping。"""
        if not isinstance(resp, dict):
            return None
        mapping = resp.get("column_mapping", resp) if "column_mapping" in resp else resp
        if not isinstance(mapping, dict):
            return None
        clean: Dict[str, str] = {}
        for k, v in mapping.items():
            if k in ("column_mapping", "constant_values"):
                continue
            if isinstance(k, str) and isinstance(v, str) and k.strip() and v.strip():
                clean[k.strip()] = v.strip()
        return clean or None

    @staticmethod
    def parse_constant_values(resp: Any) -> Dict[str, str]:
        """从 LLM 响应提取 constant_values。"""
        if not isinstance(resp, dict):
            return {}
        cv = resp.get("constant_values", {})
        if not isinstance(cv, dict):
            return {}
        return {
            k.strip(): (str(v) if not isinstance(v, str) else v)
            for k, v in cv.items()
            if isinstance(k, str) and k.strip() and v is not None
        }

    @staticmethod
    def parse_record_filter(resp: Any) -> Optional[Dict[str, Any]]:
        """从 LLM 响应提取 record_filter。"""
        if not isinstance(resp, dict):
            return None
        rf = resp.get("record_filter")
        if not isinstance(rf, dict):
            return None
        fld = rf.get("field")
        vals = rf.get("values")
        if not isinstance(fld, str) or not fld or not isinstance(vals, list) or not vals:
            return None
        return {
            "field": fld.strip(),
            "values": [str(v).strip() for v in vals if v is not None],
            "exclude": bool(rf.get("exclude", False)),
        }

    @staticmethod
    def parse_derived_fields(resp: Any) -> Tuple[List[Dict[str, Any]], List[str]]:
        """解析并校验 derived_fields。"""
        warnings: List[str] = []
        if not isinstance(resp, dict):
            return [], warnings
        df = resp.get("derived_fields")
        if df is None:
            return [], warnings
        if not isinstance(df, list):
            warnings.append("derived_fields must be a list; ignoring")
            return [], warnings
        allowed_ops = {"MONTH_FROM_DATE"}
        allowed_fmts = {"from_template_sample", "YYYYMM", "YYYY-MM", "YYYY/MM", "YYYY年MM月"}
        parsed: List[Dict[str, Any]] = []
        for idx, item in enumerate(df):
            if not isinstance(item, dict):
                warnings.append(f"derived_fields[{idx}] not an object; dropped")
                continue
            new_key, op, args = item.get("new_key"), item.get("op"), item.get("args")
            if not isinstance(new_key, str) or not new_key.strip():
                warnings.append(f"derived_fields[{idx}] missing new_key; dropped")
                continue
            if not isinstance(op, str) or op not in allowed_ops:
                warnings.append(f"derived_fields[{idx}] unsupported op '{op}'; dropped")
                continue
            if not isinstance(args, dict):
                warnings.append(f"derived_fields[{idx}] args must be object; dropped")
                continue
            if op == "MONTH_FROM_DATE":
                skeys = args.get("source_keys")
                if not isinstance(skeys, list) or not all(isinstance(k, str) and k.strip() for k in skeys):
                    warnings.append(f"derived_fields[{idx}] invalid source_keys; dropped")
                    continue
                if args.get("strategy") != "first_non_empty":
                    warnings.append(f"derived_fields[{idx}] invalid strategy; dropped")
                    continue
                if args.get("output_format") not in allowed_fmts:
                    warnings.append(f"derived_fields[{idx}] invalid output_format; dropped")
                    continue
            parsed.append({"new_key": new_key.strip(), "op": op, "args": args})
        return parsed, warnings


# ============================================================================
# GenericFillPlanner — 通用填充规划器
# ============================================================================


class GenericFillPlanner:
    """
    通用 Excel 模板填充规划器。

    给定:
      - TemplateSchema (模板结构)
      - 记录列表 (records)
      - header→key 映射 (column_mapping)
      - 常量值 (constant_values)
    生成 FillPlan (包含 row_writes / clear_ranges 等)。

    完全不含业务关键词——业务层通过预处理记录 + 映射后调用此类。
    """

    def __init__(self, template_schema: TemplateSchema):
        self._schema = template_schema
        self._inspector = SchemaInspector(template_schema)

    def build_fill_plan_from_mapping(
        self,
        records: List[Dict[str, Any]],
        header_to_key: Dict[str, str],
        constant_values: Optional[Dict[str, str]] = None,
        mapping_warnings: Optional[List[str]] = None,
    ) -> Optional[dict]:
        """
        从 header→key 映射 + 记录列表构建 fill_plan 字典。

        这是通用层的核心方法——接收已预处理好的记录和映射关系，
        计算行坐标、清除范围等，输出填充方案。
        """
        if not self._schema.sheet_schemas:
            return None
        sheet = self._schema.sheet_schemas[0]
        if not sheet.regions or not sheet.regions[0].table:
            return None
        region = sheet.regions[0]
        table = region.table
        headers = table.header or []
        if not headers or not records:
            return None
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.range)
        except Exception:
            return None

        max_header_row = max(region.header_rows) if region.header_rows else min_row
        data_start_row = max_header_row + 1
        start_cell = f"{get_column_letter(min_col)}{data_start_row}"
        warnings: List[str] = list(mapping_warnings or [])

        header_lookup = SchemaInspector.build_header_lookup(headers)
        column_mapping: Dict[str, str] = {}
        constant_column_mapping: Dict[str, Tuple[str, str]] = {}

        # 常规映射
        for template_header, extracted_key in header_to_key.items():
            if not extracted_key:
                continue
            info = SchemaInspector.lookup_header(header_lookup, template_header)
            if not info:
                warnings.append(f"Template header '{template_header}' not found in schema")
                continue
            col = info.col_letter
            if extracted_key in column_mapping and column_mapping[extracted_key] != col:
                warnings.append(f"Duplicate key '{extracted_key}' for columns {column_mapping[extracted_key]} and {col}")
                continue
            column_mapping[extracted_key] = col

        # 常量映射
        if constant_values:
            for th, cv in constant_values.items():
                if cv is None or (isinstance(cv, str) and not cv.strip()):
                    continue
                info = SchemaInspector.lookup_header(header_lookup, th)
                if not info:
                    warnings.append(f"Constant header '{th}' not found in schema")
                    continue
                const_key = f"__const__{th}"
                constant_column_mapping[const_key] = (info.col_letter, cv)

        if not column_mapping and not constant_column_mapping:
            return None

        rows = []
        for rec in records:
            row: Dict[str, Any] = {}
            for ek in column_mapping:
                val = _get_record_value(rec, ek)
                row[ek] = val if val is not None else ""
            for ck, (_, cv) in constant_column_mapping.items():
                row[ck] = cv
            rows.append(row)
        for ck, (cl, _) in constant_column_mapping.items():
            column_mapping[ck] = cl

        clear_end_row = min(max_row, data_start_row + len(rows) + 10)
        clear_range = f"{get_column_letter(min_col)}{data_start_row}:{get_column_letter(max_col)}{clear_end_row}"
        sheet_name = self._inspector.resolve_sheet_name(sheet)
        return {
            "target": {
                "sheet_name": sheet_name, "sheet": sheet_name,
                "region_id": region.region_id, "layout_type": "table",
                "clear_policy": "clear_values_keep_format",
            },
            "clear_ranges": [clear_range],
            "row_writes": [{"start_cell": start_cell, "rows": rows, "column_mapping": column_mapping}],
            "writes": [], "warnings": warnings,
        }

    def build_fallback_plan(
        self,
        extracted_json: dict,
        synonym_resolver: Optional[Callable[[str, List[str]], Optional[str]]] = None,
    ) -> Optional[dict]:
        """
        兜底填充方案: 同义词匹配 + 位置匹配。

        synonym_resolver 是可选的外部同义词匹配函数 (header, keys) → Optional[key]。
        如果不提供，退而使用纯位置匹配。
        """
        if not self._schema.sheet_schemas:
            return None
        sheet = self._schema.sheet_schemas[0]
        if not sheet.regions or not sheet.regions[0].table:
            return None
        region = sheet.regions[0]
        table = region.table
        headers = table.header or []
        if not headers:
            return None
        try:
            min_col, min_row, max_col, max_row = range_boundaries(table.range)
        except Exception:
            return None
        max_header_row = max(region.header_rows) if region.header_rows else min_row
        data_start_row = max_header_row + 1
        start_cell = f"{get_column_letter(min_col)}{data_start_row}"

        extractor = SourceDataExtractor(extracted_json)
        records: List[Dict[str, Any]] = []
        for s in extracted_json.get("sources", []):
            if not isinstance(s, dict):
                continue
            found = extractor.extract_from_container(s.get("extracted"))
            if found:
                records.extend(found)
                break
        if not records:
            merged = extracted_json.get("merged", {})
            if isinstance(merged, dict):
                for key in SourceDataExtractor.CONTAINER_PRIORITY_KEYS:
                    if key in merged and isinstance(merged[key], list) and merged[key] and isinstance(merged[key][0], dict):
                        records = list(merged[key])
                        break
                if not records:
                    for v in merged.values():
                        if isinstance(v, list) and v and isinstance(v[0], dict):
                            records = list(v)
                            break
        if not records:
            return None

        rk = list(records[0].keys())
        column_mapping: Dict[str, str] = {}
        if synonym_resolver:
            for th in headers:
                if not th.header_path:
                    continue
                match = synonym_resolver(th.header_path, rk)
                if match and match not in column_mapping:
                    column_mapping[match] = th.col_letter
        if not column_mapping:
            for i, th in enumerate(headers):
                if th.header_path and i < len(rk):
                    column_mapping[rk[i]] = th.col_letter
        rows = []
        for rec in records:
            rd = {ek: rec[ek] for ek, col in column_mapping.items() if ek in rec and rec[ek] is not None}
            if rd:
                rows.append(rd)
        if not rows:
            return None
        clear_end_row = min(max_row, data_start_row + len(rows) + 10)
        clear_range = f"{get_column_letter(min_col)}{data_start_row}:{get_column_letter(max_col)}{clear_end_row}"
        sheet_name = self._inspector.resolve_sheet_name(sheet)
        return {
            "target": {
                "sheet_name": sheet_name, "sheet": sheet_name,
                "region_id": region.region_id, "layout_type": "table",
                "clear_policy": "clear_values_keep_format",
            },
            "clear_ranges": [clear_range],
            "row_writes": [{"start_cell": start_cell, "rows": rows, "column_mapping": column_mapping}],
            "writes": [],
            "warnings": [f"Fallback fill plan used; LLM returned insufficient rows. Filling {len(rows)} rows."],
        }
