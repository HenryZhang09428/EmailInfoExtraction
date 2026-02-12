import json
import re
import unicodedata
from datetime import datetime, date
from typing import Any, Optional, List, Dict, Tuple, Union
from openpyxl.utils import get_column_letter, range_boundaries
from rapidfuzz import process, fuzz
from core.llm import LLMClient, get_llm_client
from core.config import get_settings
from core.template.schema import TemplateSchema
from core.template.prompts import TEMPLATE_FILL_PROMPT, build_insurance_param_prompt
from core.ir import FillPlan, FillPlanTarget, RowWrite, CellWrite
from core.logger import get_logger

logger = get_logger(__name__)

def _dict_to_fill_plan(data: dict) -> FillPlan:
    """Convert a dictionary to a validated FillPlan object."""
    try:
        return FillPlan.from_dict(data)
    except Exception as e:
        logger.warning("Failed to convert dict to FillPlan: %s", e)
        # Return a minimal valid FillPlan with error
        return FillPlan(
            target=FillPlanTarget(),
            warnings=[f"Failed to validate fill plan: {str(e)}"]
        )


def _empty_fill_plan(warnings: List[str] = None, debug: Optional[dict] = None) -> FillPlan:
    """Create an empty FillPlan with optional warnings."""
    return FillPlan(
        target=FillPlanTarget(),
        warnings=warnings or [],
        debug=debug
    )


def _merge_debug_info(fill_plan_dict: dict, debug_info: dict) -> None:
    if not isinstance(fill_plan_dict, dict) or not isinstance(debug_info, dict):
        return
    existing = fill_plan_dict.get("debug")
    if isinstance(existing, dict):
        merged = dict(existing)
        merged.update({k: v for k, v in debug_info.items() if v is not None})
        fill_plan_dict["debug"] = merged
    else:
        fill_plan_dict["debug"] = {k: v for k, v in debug_info.items() if v is not None}


def _append_warning(warnings: List[str], message: str) -> None:
    if not message:
        return
    if warnings is None:
        return
    if message not in warnings:
        warnings.append(message)


def plan_fill(
    template_schema: TemplateSchema,
    extracted: Any,
    llm: LLMClient,
    template_filename: Optional[str] = None,
    require_llm: bool = False
) -> FillPlan:
    """
    Generate a fill plan for populating a template with extracted data.
    
    Returns a validated FillPlan object instead of a raw dictionary.
    """
    extracted_json = extracted if isinstance(extracted, (dict, list)) else {"data": extracted}

    template_intent = _infer_insurance_template_intent_from_filename(template_filename or "")
    if not template_intent:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["insurance_intent_unknown"],
            llm_used=False
        )
    
    try:
        from core.template.profiles.social_security import (
            detect_social_security_template,
            build_social_security_fill_plan,
        )
        
        profile = detect_social_security_template(template_schema)
        if profile.is_detected:
            logger.info(
                "Social security template detected: %s (intent=%s, ss=%s, hf=%s)",
                template_filename,
                template_intent,
                profile.has_social_security_section,
                profile.has_housing_fund_section
            )
            return build_social_security_fill_plan(
                template_schema,
                extracted_json,
                llm,
                template_filename or "",
                profile,
                template_intent
            )
    except ImportError:
        logger.debug("Social security profile module not available, using default planner")
    except Exception as e:
        logger.warning("Social security profile detection failed: %s", e)
    
    return _plan_insurance_template_with_llm(
        template_schema,
        extracted_json,
        llm,
        template_filename,
        template_intent
    )

    schema_json = template_schema.model_dump(mode='json', exclude_none=True)
    
    template_headers = _get_template_headers(template_schema)
    template_intent = _infer_template_intent_simple(template_filename or "", template_headers)
    
    selection_warnings: List[str] = []
    selected_sources, selection_details = _select_sources_for_template(
        extracted_json,
        template_schema,
        selection_warnings,
        template_intent=template_intent,
        strict_intent=True,
        excel_only=True,
        max_sources=1,
    )
    if selection_details:
        selection_warnings.append(f"source_selection_details: {json.dumps(selection_details, ensure_ascii=False)}")
    if not selected_sources:
        available_sources = []
        for detail in selection_details or []:
            if not isinstance(detail, dict):
                continue
            if not detail.get("filename"):
                continue
            available_sources.append({
                "filename": detail.get("filename"),
                "intent": detail.get("intent"),
                "intent_confidence": detail.get("intent_confidence"),
                "intent_evidence": detail.get("intent_evidence"),
            })
        debug_info = {
            "template_intent": template_intent,
            "available_sources": available_sources,
        }
        warnings = ["no_source_matches_template_intent"]
        warnings.extend(selection_warnings)
        return FillPlan(
            target=FillPlanTarget(),
            warnings=warnings,
            llm_used=False,
            constant_values_count=0,
            debug=debug_info
        )
    selected_extracted_json = _build_selected_extracted_json(extracted_json, selected_sources)
    selected_source_names = [
        s.get("filename") for s in selected_sources
        if isinstance(s, dict) and s.get("filename")
    ]
    
    target_sheet_name = _resolve_sheet_name(
        template_schema,
        template_schema.sheet_schemas[0] if template_schema.sheet_schemas else None
    )
    records_before_filter = _count_records_in_sources(selected_extracted_json)
    _, pre_counts, pre_source_files = _extract_records_with_provenance(selected_extracted_json)
    debug_info = {
        "llm_status": None,
        "records_count_before_filter": records_before_filter,
        "records_count_after_filter": records_before_filter,
        "mapping_size": 0,
        "constants_size": 0,
        "target_sheet_name": target_sheet_name,
        "selected_sources": selected_source_names or None,
        "selected_source_files": selected_source_names or [],
        "template_intent": template_intent,
        # Observability: which sources currently contribute records (pre-filter snapshot)
        "records_source_files": sorted(pre_source_files) if pre_source_files else [],
        "records_count_per_source": pre_counts or {}
    }
    
    selected_extracted_json = _auto_filter_records_by_template_intent(selected_extracted_json, template_intent)
    
    source_context = _extract_filename_context(selected_extracted_json, llm, is_template=False)
    template_context = _extract_template_context(template_schema, template_filename, llm)
    
    # Get sample records from EACH source file (not just the first 3 overall)
    source_samples = _extract_samples_from_each_source(
        selected_extracted_json,
        samples_per_source=2,
        sources_override=selected_sources
    )
    
    # Extract template sample rows as few-shot examples
    template_sample_rows = _extract_template_sample_rows(template_schema)
    
    inputs = f"""
=== TemplateSchema ===
{json.dumps(schema_json, ensure_ascii=False, indent=2)}

=== Extracted Data (samples from each source file) ===
{json.dumps(source_samples, ensure_ascii=False, indent=2)}
"""
    
    # Add template sample rows as few-shot examples
    if template_sample_rows:
        inputs += f"""

=== Template Sample Data (few-shot examples) ===
These are existing data rows from the template. Use them to understand the expected format and field mapping:
{json.dumps(template_sample_rows, ensure_ascii=False, indent=2)}
"""
    
    if template_context:
        inputs += f"""

=== Template Context Analysis ===
{template_context}
"""
    
    if source_context:
        inputs += f"""

=== Source Files Context Analysis ===
{source_context}
"""
    
    prompt = TEMPLATE_FILL_PROMPT + "\n\n" + inputs
    
    settings = get_settings()
    if not settings.OPENAI_API_KEY:
        debug_info["llm_status"] = "missing_api_key"
        if require_llm:
            return _empty_fill_plan(["Failed to generate fill plan: missing_api_key"], debug_info)
        fallback = build_fallback_fill_plan(template_schema, extracted_json)
        if fallback:
            fallback_rows = sum(len(r.get("rows", [])) for r in fallback.get("row_writes", []))
            logger.debug("plan_fill: Fallback fill plan generated with %d rows", fallback_rows)
            fallback["llm_used"] = False
            _append_warning(fallback.setdefault("warnings", []), "llm_status: missing_api_key")
            _merge_debug_info(fallback, debug_info)
            return _dict_to_fill_plan(fallback)
        logger.error("plan_fill: LLM disabled and fallback failed, returning empty plan")
        return _empty_fill_plan(["Failed to generate fill plan: missing_api_key"], debug_info)
    
    try:
        logger.debug("plan_fill: Calling LLM with prompt length: %d chars", len(prompt))
        mapping_response = llm.chat_json(prompt, system=None, temperature=0, step="template_fill_plan")
        logger.debug("plan_fill: LLM call succeeded")
    except Exception as e:
        # 如果LLM调用失败（超时或其他错误），使用fallback
        error_msg = str(e)
        error_type = type(e).__name__
        
        if "timeout" in error_msg.lower() or "timed out" in error_msg.lower() or "Timeout" in error_type:
            logger.warning("plan_fill: LLM request timed out (%s), using fallback fill plan", error_type)
        else:
            logger.warning("plan_fill: LLM request failed (%s: %s), using fallback fill plan", error_type, error_msg[:200])
        
        debug_info["llm_status"] = "llm_exception"
        if require_llm:
            return _empty_fill_plan([f"Failed to generate fill plan: {error_type}: {error_msg[:200]}"], debug_info)
        fallback = build_fallback_fill_plan(template_schema, extracted_json)
        if fallback:
            fallback_rows = sum(len(r.get("rows", [])) for r in fallback.get("row_writes", []))
            logger.debug("plan_fill: Fallback fill plan generated with %d rows", fallback_rows)
            fallback["llm_used"] = False
            _append_warning(fallback.setdefault("warnings", []), "llm_status: llm_exception")
            _merge_debug_info(fallback, debug_info)
            return _dict_to_fill_plan(fallback)
        else:
            # 如果fallback也失败，返回一个空的fill_plan
            logger.error("plan_fill: Both LLM and fallback failed, returning empty plan")
            return _empty_fill_plan([f"Failed to generate fill plan: {error_type}: {error_msg[:200]}"], debug_info)
    
    # Log LLM response for debugging
    logger.info("plan_fill: LLM response type: %s", type(mapping_response).__name__)
    if isinstance(mapping_response, dict):
        logger.info("plan_fill: LLM response keys: %s", list(mapping_response.keys()))
    
    if isinstance(mapping_response, dict) and mapping_response.get("error") == "json_parse_error":
        debug_info["llm_status"] = "json_parse_error"
        if require_llm:
            return _empty_fill_plan(["Failed to generate fill plan: json_parse_error"], debug_info)
        fallback = build_fallback_fill_plan(template_schema, extracted_json)
        if fallback:
            fallback["llm_used"] = False
            _append_warning(fallback.setdefault("warnings", []), "llm_status: json_parse_error")
            _merge_debug_info(fallback, debug_info)
            return _dict_to_fill_plan(fallback)
        return _empty_fill_plan(["Failed to generate fill plan: json_parse_error"], debug_info)
    
    column_mapping = _parse_column_mapping(mapping_response)
    constant_values = _parse_constant_values(mapping_response)
    record_filter = _parse_record_filter(mapping_response)
    derived_fields, derived_warnings = _parse_derived_fields(mapping_response)
    record_filter_warnings: List[str] = []
    if template_intent and record_filter and _record_filter_conflicts_with_intent(record_filter, template_intent):
        record_filter_warnings.append(
            f"record_filter conflicts with template_intent '{template_intent}'; ignoring record_filter"
        )
        record_filter = None
    
    logger.info("plan_fill: Parsed column_mapping=%d items, constant_values=%d items, record_filter=%s", 
                len(column_mapping) if column_mapping else 0,
                len(constant_values) if constant_values else 0,
                record_filter is not None)
    
    if constant_values:
        logger.debug("plan_fill: Constant values: %s", list(constant_values.keys()))
    
    if record_filter:
        logger.info("plan_fill: Record filter - field='%s', values=%s, exclude=%s",
                   record_filter.get("field"), record_filter.get("values"), record_filter.get("exclude", False))
    if derived_fields:
        logger.info("plan_fill: Parsed derived_fields=%d items", len(derived_fields))
    
    if not column_mapping and not constant_values:
        logger.warning("plan_fill: LLM returned invalid mapping (no column_mapping or constant_values), using fallback")
        debug_info["llm_status"] = "empty_mapping"
        if require_llm:
            return _empty_fill_plan(["Failed to generate fill plan: empty_mapping"], debug_info)
        fallback = build_fallback_fill_plan(template_schema, extracted_json)
        if fallback:
            fallback["llm_used"] = False
            _append_warning(fallback.setdefault("warnings", []), "llm_status: empty_mapping")
            _merge_debug_info(fallback, debug_info)
            return _dict_to_fill_plan(fallback)
        raise ValueError(f"LLM returned invalid column mapping: {type(mapping_response)}")
    
    # Initialize column_mapping if only constants were provided
    if not column_mapping:
        column_mapping = {}
    
    pre_counts = _dbg_count_records(selected_extracted_json)
    filtered_extracted_json = _apply_record_filter(selected_extracted_json, record_filter)
    post_record_filter_counts = _dbg_count_records(filtered_extracted_json)
    inferred_intent = template_intent
    if inferred_intent is None:
        inferred_intent = _infer_template_intent_from_mapping(
            template_schema,
            template_filename,
            column_mapping,
            constant_values
        )
    filtered_extracted_json = _apply_record_type_filter(filtered_extracted_json, inferred_intent)
    post_record_type_filter_counts = _dbg_count_records(filtered_extracted_json)
    dbg_payload = {
        "dbg_stage": "fill_planner_record_routing",
        "template_intent": template_intent,
        "inferred_intent": inferred_intent,
        "record_filter_present": record_filter is not None,
        "counts": {
            "pre": pre_counts,
            "post_record_filter": post_record_filter_counts,
            "post_record_type_filter": post_record_type_filter_counts,
        }
    }
    print(json.dumps(dbg_payload, ensure_ascii=False))
    debug_info["record_routing_debug"] = dbg_payload
    records, per_source_counts, source_files = _extract_records_with_provenance(filtered_extracted_json)
    
    # Apply derived fields after record filtering
    derived_exec_warnings: List[str] = []
    if derived_fields:
        derived_fields = _attach_template_headers_to_derived_fields(derived_fields, column_mapping)
        _apply_derived_fields_to_records(
            template_schema,
            records,
            derived_fields,
            template_intent,
            derived_exec_warnings
        )
        column_mapping, constant_values = _promote_derived_constants(
            derived_fields,
            records,
            column_mapping,
            constant_values,
            derived_exec_warnings
        )
    
    fee_month_warnings: List[str] = []
    column_mapping, constant_values = _infer_fee_month(
        template_schema,
        records,
        template_intent,
        column_mapping,
        constant_values,
        fee_month_warnings
    )
    
    record_gating_warnings: List[str] = []
    filtered_extracted_json = _apply_record_gating_to_extracted_json(
        filtered_extracted_json,
        template_schema,
        record_gating_warnings
    )
    records, per_source_counts, source_files = _extract_records_with_provenance(filtered_extracted_json)
    debug_info["records_count_after_filter"] = sum(per_source_counts.values()) if per_source_counts else 0
    debug_info["records_source_files"] = sorted(source_files) if source_files else []
    debug_info["records_count_per_source"] = per_source_counts or {}
    debug_info["selected_source_files"] = selected_source_names or []
    
    # Auto-infer common fields like 费用年月 / 社保申报类型
    column_mapping, constant_values = _auto_infer_common_fields(
        template_schema,
        filtered_extracted_json,
        column_mapping,
        constant_values,
        template_intent,
        template_sample_rows
    )
    
    enriched_mapping, mapping_warnings = _enrich_mapping_with_fuzzy_match(
        template_schema,
        filtered_extracted_json,
        column_mapping
    )
    if derived_warnings or derived_exec_warnings or record_filter_warnings or fee_month_warnings or selection_warnings or record_gating_warnings:
        mapping_warnings = list(mapping_warnings or [])
        mapping_warnings.extend(derived_warnings)
        mapping_warnings.extend(derived_exec_warnings)
        mapping_warnings.extend(record_filter_warnings)
        mapping_warnings.extend(fee_month_warnings)
        mapping_warnings.extend(selection_warnings)
        mapping_warnings.extend(record_gating_warnings)
    debug_info["mapping_size"] = len(enriched_mapping) if enriched_mapping else 0
    debug_info["constants_size"] = len(constant_values) if constant_values else 0
    debug_info["llm_status"] = debug_info["llm_status"] or "ok"
    
    fill_plan = _build_fill_plan_from_mapping(
        template_schema,
        filtered_extracted_json,
        enriched_mapping,
        mapping_warnings,
        constant_values,
        inferred_intent
    )
    if fill_plan:
        rw = fill_plan.get("row_writes", [])
        total_rows = sum(len(r.get("rows", [])) for r in rw)
        logger.debug("plan_fill: column_mapping=%d constants=%d rows=%d", len(column_mapping), len(constant_values), total_rows)
        # Mark as LLM-generated plan with metadata
        fill_plan["llm_used"] = True
        fill_plan["constant_values_count"] = len(constant_values) if constant_values else 0
        _merge_debug_info(fill_plan, debug_info)
        return _dict_to_fill_plan(fill_plan)
    
    fallback = build_fallback_fill_plan(template_schema, extracted_json)
    if fallback:
        logger.debug("plan_fill: using fallback fill plan (mapping build failed)")
        fallback["llm_used"] = False
        _merge_debug_info(fallback, debug_info)
        return _dict_to_fill_plan(fallback)
    
    return _empty_fill_plan(["Failed to generate fill plan from column mapping"], debug_info)

def _parse_column_mapping(mapping_response: Any) -> Optional[Dict[str, str]]:
    if isinstance(mapping_response, dict):
        if "column_mapping" in mapping_response and isinstance(mapping_response["column_mapping"], dict):
            mapping = mapping_response["column_mapping"]
        else:
            mapping = mapping_response
        clean_mapping = {}
        for k, v in mapping.items():
            if k in ("column_mapping", "constant_values"):
                continue
            if not isinstance(k, str) or not isinstance(v, str):
                continue
            key = k.strip()
            val = v.strip()
            if key and val:
                clean_mapping[key] = val
        return clean_mapping if clean_mapping else None
    return None


def _parse_constant_values(mapping_response: Any) -> Dict[str, str]:
    """Extract constant_values from LLM response."""
    if not isinstance(mapping_response, dict):
        return {}
    
    constant_values = mapping_response.get("constant_values", {})
    if not isinstance(constant_values, dict):
        return {}
    
    clean_constants = {}
    for k, v in constant_values.items():
        if not isinstance(k, str):
            continue
        key = k.strip()
        # Value can be string, number, etc.
        if key and v is not None:
            clean_constants[key] = str(v) if not isinstance(v, str) else v
    
    return clean_constants


def _parse_record_filter(mapping_response: Any) -> Optional[Dict[str, Any]]:
    """Extract record_filter from LLM response."""
    if not isinstance(mapping_response, dict):
        return None
    
    record_filter = mapping_response.get("record_filter")
    if record_filter is None or not isinstance(record_filter, dict):
        return None
    
    field = record_filter.get("field")
    values = record_filter.get("values")
    
    if not field or not isinstance(field, str):
        return None
    if not values or not isinstance(values, list):
        return None
    
    return {
        "field": field.strip(),
        "values": [str(v).strip() for v in values if v is not None],
        "exclude": bool(record_filter.get("exclude", False))
    }


def _dbg_count_records(extracted_json: dict) -> dict:
    records = _extract_records(extracted_json) if isinstance(extracted_json, dict) else []
    counts = {"total": 0, "add": 0, "remove": 0, "unknown": 0}
    for record in records:
        if not isinstance(record, dict):
            continue
        counts["total"] += 1
        record_type = _infer_record_type(record) or _infer_record_type_from_sheet(record)
        if record_type == "add":
            counts["add"] += 1
        elif record_type == "remove":
            counts["remove"] += 1
        else:
            counts["unknown"] += 1
    return counts


def _parse_derived_fields(mapping_response: Any) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse and validate derived_fields from LLM response.
    
    Returns:
        (derived_fields, warnings)
    """
    warnings: List[str] = []
    if not isinstance(mapping_response, dict):
        return [], warnings
    
    derived_fields = mapping_response.get("derived_fields")
    if derived_fields is None:
        return [], warnings
    if not isinstance(derived_fields, list):
        warnings.append("derived_fields must be a list; ignoring invalid derived_fields")
        return [], warnings
    
    allowed_ops = {"MONTH_FROM_DATE"}
    allowed_formats = {"from_template_sample", "YYYYMM", "YYYY-MM", "YYYY/MM", "YYYY年MM月"}
    
    parsed: List[Dict[str, Any]] = []
    for idx, item in enumerate(derived_fields):
        if not isinstance(item, dict):
            warnings.append(f"derived_fields[{idx}] is not an object; dropped")
            continue
        
        new_key = item.get("new_key")
        op = item.get("op")
        args = item.get("args")
        
        if not isinstance(new_key, str) or not new_key.strip():
            warnings.append(f"derived_fields[{idx}] missing valid new_key; dropped")
            continue
        if not isinstance(op, str) or op not in allowed_ops:
            warnings.append(f"derived_fields[{idx}] has unsupported op '{op}'; dropped")
            continue
        if not isinstance(args, dict):
            warnings.append(f"derived_fields[{idx}] args must be an object; dropped")
            continue
        
        if op == "MONTH_FROM_DATE":
            source_keys = args.get("source_keys")
            strategy = args.get("strategy")
            output_format = args.get("output_format")
            
            if not isinstance(source_keys, list) or not all(isinstance(k, str) and k.strip() for k in source_keys):
                warnings.append(f"derived_fields[{idx}] invalid source_keys; dropped")
                continue
            if strategy != "first_non_empty":
                warnings.append(f"derived_fields[{idx}] invalid strategy '{strategy}'; dropped")
                continue
            if output_format not in allowed_formats:
                warnings.append(f"derived_fields[{idx}] invalid output_format '{output_format}'; dropped")
                continue
        
        parsed.append({
            "new_key": new_key.strip(),
            "op": op,
            "args": args
        })
    
    return parsed, warnings


def _attach_template_headers_to_derived_fields(
    derived_fields: List[Dict[str, Any]],
    column_mapping: Dict[str, str]
) -> List[Dict[str, Any]]:
    """
    Attach template headers that map to derived new_key.
    This helps resolve output_format from template samples.
    """
    if not derived_fields or not column_mapping:
        return derived_fields
    
    headers_by_new_key: Dict[str, List[str]] = {}
    for header, key in column_mapping.items():
        if not isinstance(key, str):
            continue
        headers_by_new_key.setdefault(key, []).append(header)
    
    enriched = []
    for item in derived_fields:
        if not isinstance(item, dict):
            enriched.append(item)
            continue
        new_key = item.get("new_key")
        headers = headers_by_new_key.get(new_key, []) if isinstance(new_key, str) else []
        enriched_item = dict(item)
        enriched_item["_template_headers"] = headers
        enriched.append(enriched_item)
    
    return enriched


def _apply_derived_fields_to_records(
    template_schema: TemplateSchema,
    records: List[Dict[str, Any]],
    derived_fields: List[Dict[str, Any]],
    template_intent: Optional[str],
    warnings: List[str]
) -> List[Dict[str, Any]]:
    """
    Apply derived_fields to each record deterministically.
    
    - Does not raise exceptions.
    - Any failure logs a warning and leaves the field empty.
    """
    if not records or not derived_fields:
        return records
    
    for record_idx, record in enumerate(records):
        if not isinstance(record, dict):
            continue
        for df_idx, item in enumerate(derived_fields):
            if not isinstance(item, dict):
                continue
            new_key = item.get("new_key")
            op = item.get("op")
            args = item.get("args", {})
            if not isinstance(new_key, str) or not new_key:
                continue
            if op != "MONTH_FROM_DATE":
                continue
            if not isinstance(args, dict):
                warnings.append(f"derived_fields[{df_idx}] invalid args; skipped")
                record[new_key] = ""
                continue
            
            source_keys = args.get("source_keys", [])
            output_format = args.get("output_format", "YYYY-MM")
            
            source_value = None
            if isinstance(source_keys, list):
                for key in source_keys:
                    if not isinstance(key, str):
                        continue
                    val = record.get(key)
                    if _is_non_empty_value(val):
                        source_value = val
                        break
            
            if not _is_non_empty_value(source_value):
                record[new_key] = ""
                continue
            
            year_month = _parse_year_month(source_value)
            if not year_month:
                warnings.append(
                    f"derived_fields[{df_idx}] failed to parse date for record {record_idx}"
                )
                record[new_key] = ""
                continue
            
            fmt = output_format
            if output_format == "from_template_sample":
                template_headers = item.get("_template_headers", [])
                fmt = _resolve_year_month_format_from_template(template_schema, template_headers)
            
            formatted = _format_year_month(year_month[0], year_month[1], fmt)
            record[new_key] = formatted
    
    return records


def _promote_derived_constants(
    derived_fields: List[Dict[str, Any]],
    records: List[Dict[str, Any]],
    column_mapping: Dict[str, str],
    constant_values: Dict[str, str],
    warnings: List[str]
) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    If all derived values are identical and non-empty, promote to constant_values.
    """
    if not derived_fields or not records:
        return column_mapping, constant_values
    
    updated_mapping = dict(column_mapping)
    updated_constants = dict(constant_values)
    
    for item in derived_fields:
        if not isinstance(item, dict):
            continue
        new_key = item.get("new_key")
        if not isinstance(new_key, str) or not new_key:
            continue
        
        values = []
        for record in records:
            if not isinstance(record, dict):
                continue
            val = record.get(new_key)
            if _is_non_empty_value(val):
                values.append(str(val).strip())
            else:
                values.append("")
        
        if not values:
            continue
        
        non_empty = [v for v in values if v]
        if not non_empty:
            continue
        
        first_value = non_empty[0]
        if all(v == first_value for v in non_empty) and len(non_empty) == len(values):
            headers = [h for h, k in updated_mapping.items() if k == new_key]
            for header in headers:
                if header in updated_constants:
                    if updated_constants[header] != first_value:
                        warnings.append(
                            f"derived constant for '{header}' differs from existing constant; keeping existing"
                        )
                    continue
                updated_constants[header] = first_value
                updated_mapping.pop(header, None)
    
    return updated_mapping, updated_constants


def _is_non_empty_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _parse_year_month(value: Any) -> Optional[Tuple[int, int]]:
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, (int, float)):
        value = str(int(value))
    if not isinstance(value, str):
        return None
    
    text = value.strip()
    if not text:
        return None
    
    match = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    
    match = re.match(r'^(\d{4})[-/](\d{1,2})$', text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    
    match = re.match(r'^(\d{4})年(\d{1,2})月$', text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    
    match = re.match(r'^(\d{4})(\d{2})(\d{2})$', text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    
    match = re.match(r'^(\d{4})(\d{2})$', text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month
    
    return None


def _format_year_month(year: int, month: int, fmt: str) -> str:
    if fmt == "YYYYMM":
        return f"{year}{month:02d}"
    if fmt == "YYYY/MM":
        return f"{year}/{month:02d}"
    if fmt == "YYYY年MM月":
        return f"{year}年{month:02d}月"
    return f"{year}-{month:02d}"


def _resolve_year_month_format_from_template(
    template_schema: TemplateSchema,
    template_headers: List[str]
) -> str:
    sample_rows = _extract_template_sample_rows(template_schema)
    for row in sample_rows:
        if not isinstance(row, dict):
            continue
        for header in template_headers or []:
            if header in row and row[header]:
                fmt = _detect_year_month_format_from_value(row[header])
                if fmt:
                    return fmt
    return "YYYY-MM"


def _detect_year_month_format_from_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "年" in text and "月" in text:
        return "YYYY年MM月"
    if "/" in text:
        return "YYYY/MM"
    if "-" in text:
        return "YYYY-MM"
    if re.match(r'^\d{6}$', text):
        return "YYYYMM"
    return None


def _normalize_match_text(text: Any) -> str:
    if text is None:
        return ""
    normalized = unicodedata.normalize("NFKC", str(text)).lower()
    normalized = re.sub(r'[\s_\-/]+', '', normalized)
    return normalized


def _header_variants(header_path: str) -> List[str]:
    variants = set()
    if not header_path:
        return []
    variants.add(_normalize_match_text(header_path))
    if "/" in header_path:
        parts = [p.strip() for p in header_path.split("/") if p.strip()]
        for part in parts:
            variants.add(_normalize_match_text(part))
        if parts:
            variants.add(_normalize_match_text(parts[-1]))
    return [v for v in variants if v]


def _get_template_headers(template_schema: TemplateSchema) -> List[str]:
    headers: List[str] = []
    for sheet_schema in template_schema.sheet_schemas:
        for region in sheet_schema.regions:
            if not region.table or not region.table.header:
                continue
            for header in region.table.header:
                header_path = getattr(header, "header_path", None)
                if header_path:
                    headers.append(header_path)
    return headers


def _is_insurance_add_remove_template(template_schema: TemplateSchema) -> bool:
    if not template_schema.sheet_schemas:
        return False
    sheet = template_schema.sheet_schemas[0]
    if not sheet.regions:
        return False
    region = None
    for candidate in sheet.regions:
        if candidate.table and candidate.table.header:
            region = candidate
            break
    if not region or not region.table or not region.table.header:
        return False

    required_suffixes = ("姓名", "申报类型", "费用年月")
    found = {suffix: False for suffix in required_suffixes}
    for header in region.table.header:
        header_path = getattr(header, "header_path", None)
        if not isinstance(header_path, str):
            continue
        normalized = header_path.replace("\u3000", "").strip()
        if not normalized:
            continue
        for suffix in required_suffixes:
            if not found[suffix] and normalized.endswith(suffix):
                found[suffix] = True
        if all(found.values()):
            break

    return all(found.values())


def _infer_insurance_template_intent_from_filename(template_filename: str) -> Optional[str]:
    filename_text = _normalize_match_text(template_filename or "")
    has_add = "增员" in filename_text
    has_remove = "减员" in filename_text
    if has_add and not has_remove:
        return "add"
    if has_remove and not has_add:
        return "remove"
    return None


def _get_main_table_region(template_schema: TemplateSchema) -> Optional[Any]:
    if not template_schema.sheet_schemas:
        return None
    sheet = template_schema.sheet_schemas[0]
    if not sheet.regions:
        return None
    for region in sheet.regions:
        if region.table and region.table.header:
            return region
    return None


def _find_header_path_by_suffix(headers: List[Any], suffix: str) -> Optional[str]:
    if not headers or not suffix:
        return None
    suffix_norm = _normalize_match_text(suffix)
    for header in headers:
        header_texts: List[str] = []
        for attr in ("header_path", "header_text", "header", "text", "name"):
            value = getattr(header, attr, None)
            if isinstance(value, str) and value.strip():
                header_texts.append(value)
        for text in header_texts:
            for variant in _header_variants(text):
                if variant.endswith(suffix_norm):
                    header_path = getattr(header, "header_path", None)
                    return header_path or text
    return None


def _trim_record_fields(record: Dict[str, Any], limit: int) -> Dict[str, Any]:
    if not isinstance(record, dict):
        return {}
    items = list(record.items())[:limit]
    return {k: v for k, v in items}


def _build_insurance_sources_profile(extracted_json: dict) -> List[Dict[str, Any]]:
    sources = extracted_json.get("sources", []) if isinstance(extracted_json, dict) else []
    profile: List[Dict[str, Any]] = []
    for source in sources:
        if not isinstance(source, dict):
            continue
        if source.get("source_type") != "excel":
            continue
        extracted = source.get("extracted") or {}
        data_records = extracted.get("data") if isinstance(extracted, dict) else None
        all_records = data_records if isinstance(data_records, list) else []
        record_keys: List[str] = []
        seen_keys = set()
        for record in all_records or []:
            if not isinstance(record, dict):
                continue
            for key in record.keys():
                if not isinstance(key, str) or key in seen_keys:
                    continue
                seen_keys.add(key)
                record_keys.append(key)
                if len(record_keys) >= 60:
                    break
            if len(record_keys) >= 60:
                break
        semantic_key_by_header = {}
        if isinstance(extracted, dict):
            semantic_key_by_header = (extracted.get("metadata") or {}).get("semantic_key_by_header") or {}
        header_candidates = list(semantic_key_by_header.keys())[:60] if isinstance(semantic_key_by_header, dict) else []
        sample_records = [_trim_record_fields(r, 30) for r in (all_records or [])[:2]]
        profile.append({
            "source_id": source.get("source_id"),
            "filename": source.get("filename"),
            "record_keys": record_keys,
            "header_candidates": header_candidates,
            "sample_records": sample_records,
        })
    return profile


def _looks_like_chinese_name(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text:
        return False
    return bool(re.match(r'^[\u4e00-\u9fff]{2,4}$', text))


def _collect_header_columns_by_suffix(headers: List[Any], suffix: str) -> List[str]:
    if not headers or not suffix:
        return []
    suffix_norm = _normalize_match_text(suffix)
    columns: List[str] = []
    for header in headers:
        header_texts: List[str] = []
        for attr in ("header_path", "header_text", "header", "text", "name"):
            value = getattr(header, attr, None)
            if isinstance(value, str) and value.strip():
                header_texts.append(value)
        for text in header_texts:
            for variant in _header_variants(text):
                if variant.endswith(suffix_norm):
                    col_letter = getattr(header, "col_letter", None)
                    if col_letter:
                        columns.append(col_letter)
                    break
    return columns


def _collect_header_columns_with_paths_by_suffix(headers: List[Any], suffix: str) -> List[Tuple[str, str]]:
    if not headers or not suffix:
        return []
    suffix_norm = _normalize_match_text(suffix)
    columns: List[Tuple[str, str]] = []
    for header in headers:
        header_path = getattr(header, "header_path", None)
        if not isinstance(header_path, str):
            header_path = ""
        header_texts: List[str] = []
        for attr in ("header_path", "header_text", "header", "text", "name"):
            value = getattr(header, attr, None)
            if isinstance(value, str) and value.strip():
                header_texts.append(value)
        for text in header_texts:
            for variant in _header_variants(text):
                if variant.endswith(suffix_norm):
                    col_letter = getattr(header, "col_letter", None)
                    if col_letter:
                        columns.append((col_letter, header_path))
                    break
    return columns


def _replace_records_in_extracted(extracted: Any, records: List[Dict[str, Any]]) -> Any:
    if isinstance(extracted, list):
        return list(records)
    if isinstance(extracted, dict):
        updated = dict(extracted)
        target_key = None
        priority_keys = ["data", "records", "rows", "items", "extracted_data", "table_data"]
        for key in priority_keys:
            value = updated.get(key)
            if isinstance(value, list) and value and isinstance(value[0], dict):
                target_key = key
                break
        if target_key is None:
            for key, value in updated.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    target_key = key
                    break
        if target_key is not None:
            updated[target_key] = list(records)
        else:
            updated["records"] = list(records)
        return updated
    return extracted


def _find_source_by_id(extracted_json: dict, source_id: str) -> Optional[dict]:
    if not isinstance(extracted_json, dict) or not source_id:
        return None
    sources = extracted_json.get("sources", [])
    if not isinstance(sources, list):
        return None
    for source in sources:
        if isinstance(source, dict) and source.get("source_id") == source_id:
            return source
    return None


def _key_present_in_records(records: List[Dict[str, Any]], key: str) -> bool:
    if not key:
        return False
    for record in records:
        if not isinstance(record, dict):
            continue
        if _get_record_value(record, key) is not None:
            return True
    return False


def _parse_any_date(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        value = str(int(value))
    if not isinstance(value, str):
        value = str(value)
    text = value.strip()
    if not text:
        return None
    try:
        normalized = text.replace("Z", "+00:00")
        return datetime.fromisoformat(normalized)
    except (ValueError, TypeError):
        pass
    common_match = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', text)
    if common_match:
        try:
            return datetime(
                int(common_match.group(1)),
                int(common_match.group(2)),
                int(common_match.group(3))
            )
        except (ValueError, TypeError):
            return None
    compact_match = re.match(r'^(\d{4})(\d{2})(\d{2})$', text)
    if compact_match:
        try:
            return datetime(
                int(compact_match.group(1)),
                int(compact_match.group(2)),
                int(compact_match.group(3))
            )
        except (ValueError, TypeError):
            return None
    return None


def _next_month_yyyymm(value: datetime) -> Optional[str]:
    if not isinstance(value, datetime):
        return None
    year = value.year
    month = value.month + 1
    if month > 12:
        year += 1
        month = 1
    return f"{year}{month:02d}"


def _collect_email_leave_records(extracted_json: dict) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Collect leave/removal records from all email sources.
    
    Returns records where:
    - source_type == "email"
    - source["extracted"]["data"] is a list of dicts
    - record has intent == "remove" OR contains leave_date
    
    Returns:
        (records, warnings)
    """
    records: List[Dict[str, Any]] = []
    warnings: List[str] = []
    
    sources = extracted_json.get("sources", []) if isinstance(extracted_json, dict) else []
    email_source_count = 0
    
    for source in sources:
        if not isinstance(source, dict):
            continue
        if source.get("source_type") != "email":
            continue
        
        email_source_count += 1
        extracted = source.get("extracted")
        if not isinstance(extracted, dict):
            continue
        
        data = extracted.get("data")
        if not isinstance(data, list):
            continue
        
        source_filename = source.get("filename", "unknown")
        
        for item in data:
            if not isinstance(item, dict):
                continue
            
            # Check if record has intent == "remove" OR contains leave_date
            intent = item.get("intent", "")
            has_leave_date = bool(item.get("leave_date"))
            
            if intent == "remove" or has_leave_date:
                # Normalize the record to have consistent keys
                record = {
                    "name": str(item.get("name", "") or "").strip(),
                    "employee_id": str(item.get("employee_id", "") or "").strip(),
                    "leave_date": str(item.get("leave_date", "") or "").strip(),
                    "leave_date_text": str(item.get("leave_date_text", "") or "").strip(),
                    "intent": "remove",
                    "__source_file__": source_filename,
                    "__source_type__": "email",
                }
                records.append(record)
    
    if email_source_count > 0:
        warnings.append(f"email_sources_scanned:{email_source_count}")
    
    return records, warnings


def _deduplicate_leave_records(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
    """
    De-duplicate leave records.
    
    Priority:
    1. Use employee_id if present and non-empty
    2. Else use (name, leave_date) tuple
    
    When duplicates exist, prefer records with more complete data.
    
    Returns:
        (deduplicated_records, duplicate_count)
    """
    if not records:
        return [], 0
    
    # Group by dedup key
    by_employee_id: Dict[str, List[Dict[str, Any]]] = {}
    by_name_date: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    no_key_records: List[Dict[str, Any]] = []
    
    for record in records:
        emp_id = record.get("employee_id", "").strip()
        name = record.get("name", "").strip()
        leave_date = record.get("leave_date", "").strip()
        
        if emp_id:
            by_employee_id.setdefault(emp_id, []).append(record)
        elif name and leave_date:
            by_name_date.setdefault((name, leave_date), []).append(record)
        elif name:
            # Has name but no leave_date - use name as key
            by_name_date.setdefault((name, ""), []).append(record)
        else:
            no_key_records.append(record)
    
    def _pick_best(group: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Pick record with most non-empty fields."""
        if len(group) == 1:
            return group[0]
        
        def score(r: Dict[str, Any]) -> int:
            return sum(1 for v in r.values() if v and str(v).strip())
        
        return max(group, key=score)
    
    result: List[Dict[str, Any]] = []
    seen_emp_ids: set = set()
    seen_name_dates: set = set()
    
    # Process employee_id groups
    for emp_id, group in by_employee_id.items():
        best = _pick_best(group)
        result.append(best)
        seen_emp_ids.add(emp_id)
        # Also mark name+date as seen to avoid duplicates
        name = best.get("name", "").strip()
        leave_date = best.get("leave_date", "").strip()
        if name:
            seen_name_dates.add((name, leave_date))
    
    # Process name+date groups (skip if already seen via employee_id)
    for (name, leave_date), group in by_name_date.items():
        if (name, leave_date) in seen_name_dates:
            continue
        best = _pick_best(group)
        # Check if this record's employee_id was already processed
        emp_id = best.get("employee_id", "").strip()
        if emp_id and emp_id in seen_emp_ids:
            continue
        result.append(best)
        seen_name_dates.add((name, leave_date))
    
    # Add no-key records (can't deduplicate)
    result.extend(no_key_records)
    
    duplicate_count = len(records) - len(result)
    return result, duplicate_count


def _plan_insurance_template_with_llm(
    template_schema: TemplateSchema,
    extracted_json: dict,
    llm: LLMClient,
    template_filename: Optional[str],
    template_intent: str
) -> FillPlan:
    debug_info: Dict[str, Any] = {
        "template_intent": template_intent,
        "template_filename": template_filename,
        "planner_mode": "insurance_constrained_llm",
        "selected_source_filename": None,
        "name_key": None,
        "effective_date_key": None,
        "kept_records": 0,
        "skipped_records": 0,
        "llm_confidence": None,
        "notes": None,
    }
    warnings: List[str] = []

    region = _get_main_table_region(template_schema)
    if not region or not region.table or not region.table.header:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["insurance_template_missing_table"],
            llm_used=False,
            debug=debug_info
        )

    headers = region.table.header or []
    name_columns = _collect_header_columns_by_suffix(headers, "姓名")
    fee_columns = _collect_header_columns_with_paths_by_suffix(headers, "费用年月")
    type_columns = _collect_header_columns_with_paths_by_suffix(headers, "申报类型")
    if not name_columns or not fee_columns or not type_columns:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["insurance_template_headers_missing"],
            llm_used=False,
            debug=debug_info
        )

    sources = [
        s for s in (extracted_json.get("sources", []) if isinstance(extracted_json, dict) else [])
        if isinstance(s, dict) and s.get("source_type") == "excel"
    ]
    if not sources:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["insurance_no_sources"],
            llm_used=False,
            debug=debug_info
        )

    def _source_records(source: dict) -> List[Dict[str, Any]]:
        extracted = source.get("extracted") or {}
        if not isinstance(extracted, dict):
            return []
        data_records = extracted.get("data")
        if isinstance(data_records, list):
            return [r for r in data_records if isinstance(r, dict)]
        return []
    
    def _filter_records_by_sheet_intent(
        records: List[Dict[str, Any]],
        intent: str
    ) -> Tuple[List[Dict[str, Any]], bool]:
        if intent == "add":
            keywords = ("增员",)
        elif intent == "remove":
            keywords = ("减员", "离职")
        else:
            return records, False
        filtered = [
            record for record in records
            if any(keyword in str(record.get("__sheet_name__", "") or "") for keyword in keywords)
        ]
        return filtered, True

    sources_profile = _build_insurance_sources_profile({"sources": sources})
    template_headers = []
    for header in headers:
        header_path = getattr(header, "header_path", None)
        if isinstance(header_path, str) and header_path.strip():
            template_headers.append(header_path)

    prompt = build_insurance_param_prompt(template_intent, template_headers, sources_profile)
    response = None
    llm_used = False
    try:
        response = llm.chat_json_once(
            prompt,
            system=None,
            temperature=0,
            step="insurance_template_param",
            timeout=30
        )
    except Exception:
        response = None

    if isinstance(response, dict) and response.get("error"):
        response = None

    selected_source = None
    name_key = None
    effective_date_key = None
    llm_confidence = None
    notes = None

    def _validate_llm_params(resp: dict) -> bool:
        nonlocal selected_source, name_key, effective_date_key, llm_confidence, notes
        selected_source_id = resp.get("selected_source_id")
        name_key = resp.get("name_key")
        effective_date_key = resp.get("effective_date_key")
        llm_confidence = resp.get("confidence")
        notes = resp.get("notes")
        if not isinstance(selected_source_id, str) or not selected_source_id.strip():
            return False
        source_by_id = next((s for s in sources if s.get("source_id") == selected_source_id), None)
        if not source_by_id:
            return False
        records = _source_records(source_by_id)
        if not records:
            return False
        if not isinstance(name_key, str) or not name_key.strip():
            return False
        if not isinstance(effective_date_key, str) or not effective_date_key.strip():
            return False
        if not any(name_key in r for r in records):
            return False
        if not any(effective_date_key in r for r in records):
            return False
        selected_source = source_by_id
        return True

    if isinstance(response, dict) and _validate_llm_params(response):
        llm_used = True
    else:
        if response is not None:
            warnings.append("insurance_llm_invalid_params")

        def _choose_source_by_filename() -> Optional[dict]:
            keywords = ("入职", "增员") if template_intent == "add" else ("离职", "减员")
            for source in sources:
                filename = source.get("filename") or ""
                filename_norm = _normalize_match_text(filename)
                if any(_normalize_match_text(k) in filename_norm for k in keywords):
                    return source
            return None

        def _choose_source_by_date_keys() -> Optional[dict]:
            keywords = (
                ("start_date", "入职日期", "到岗日期")
                if template_intent == "add"
                else ("leave_date", "end_date", "离职日期", "终止日期")
            )
            best_source = None
            best_score = -1.0
            for source in sources:
                records = _source_records(source)
                keys: List[str] = []
                seen = set()
                for record in records:
                    for key in record.keys():
                        if isinstance(key, str) and key not in seen:
                            seen.add(key)
                            keys.append(key)
                if not keys:
                    continue
                matches = 0
                for key in keys:
                    key_norm = _normalize_match_text(key)
                    if any(_normalize_match_text(k) in key_norm for k in keywords):
                        matches += 1
                score = matches / max(len(keys), 1)
                if score > best_score:
                    best_score = score
                    best_source = source
            return best_source

        selected_source = _choose_source_by_filename() or _choose_source_by_date_keys()
        if selected_source:
            records = _source_records(selected_source)
            keys = []
            seen = set()
            for record in records:
                for key in record.keys():
                    if isinstance(key, str) and key not in seen:
                        seen.add(key)
                        keys.append(key)
            if keys:
                for key in keys:
                    if key == "name" or key.lower() == "name":
                        name_key = key
                        break
                if not name_key:
                    for key in keys:
                        if "姓名" in key or "姓名" in _normalize_match_text(key):
                            name_key = key
                            break
                if not name_key:
                    best_key = None
                    best_hits = 0
                    for key in keys:
                        hits = sum(1 for r in records if _looks_like_chinese_name(r.get(key)))
                        if hits > best_hits:
                            best_hits = hits
                            best_key = key
                    if best_key and best_hits > 0:
                        name_key = best_key
                if template_intent == "add":
                    for key in keys:
                        if key == "start_date" or key.lower() == "start_date":
                            effective_date_key = key
                            break
                    if not effective_date_key:
                        for key in keys:
                            key_norm = _normalize_match_text(key)
                            if any(k in key_norm for k in ["入职日期", "到岗日期"]):
                                effective_date_key = key
                                break
                else:
                    for key in keys:
                        key_lower = key.lower()
                        if key_lower in ("leave_date", "end_date"):
                            effective_date_key = key
                            break
                    if not effective_date_key:
                        for key in keys:
                            key_norm = _normalize_match_text(key)
                            if any(k in key_norm for k in ["离职日期", "终止日期"]):
                                effective_date_key = key
                                break

        # For "remove" intent, we can proceed with just email records if no Excel source
        if template_intent == "remove":
            # Will try to merge email records below, so don't fail yet
            pass
        elif not selected_source or not name_key or not effective_date_key:
            return FillPlan(
                target=FillPlanTarget(),
                warnings=["insurance_fallback_failed"],
                llm_used=False,
                debug=debug_info
            )

    # ===== Collect and merge records for "remove" intent =====
    new_records: List[Dict[str, Any]] = []
    skipped = 0
    date_parse_failed = 0
    fee_month_failed = 0
    declare_type = "增" if template_intent == "add" else "减"
    email_records_count = 0
    excel_records_count = 0
    duplicate_count = 0

    if template_intent == "remove":
        # Step 1: Collect email records with intent="remove" or leave_date
        email_leave_records, email_warnings = _collect_email_leave_records(extracted_json)
        warnings.extend(email_warnings)
        email_records_count = len(email_leave_records)
        
        # Step 2: Collect Excel records if source was selected
        excel_leave_records: List[Dict[str, Any]] = []
        if selected_source and name_key and effective_date_key:
            excel_raw_records = _source_records(selected_source)
            excel_raw_records, sheet_filter_applied = _filter_records_by_sheet_intent(
                excel_raw_records,
                template_intent
            )
            if sheet_filter_applied:
                warnings.append("remove_intent_ignore_non_leave_sheets")
            for record in excel_raw_records:
                if not isinstance(record, dict):
                    continue
                name = str(record.get(name_key, "")).strip()
                if not name:
                    continue
                # Get leave_date from effective_date_key
                leave_date_raw = record.get(effective_date_key)
                dt = _parse_any_date(leave_date_raw)
                leave_date = dt.strftime("%Y-%m-%d") if dt else ""
                
                excel_leave_records.append({
                    "name": name,
                    "employee_id": str(record.get("employee_id", "") or record.get("工号", "") or "").strip(),
                    "leave_date": leave_date,
                    "leave_date_text": str(leave_date_raw or ""),
                    "intent": "remove",
                    "__source_file__": selected_source.get("filename", "excel"),
                    "__source_type__": "excel",
                })
            excel_records_count = len(excel_leave_records)
        
        # Step 3: Merge email + excel records
        all_leave_records = email_leave_records + excel_leave_records
        
        # Step 4: De-duplicate
        deduplicated_records, duplicate_count = _deduplicate_leave_records(all_leave_records)
        if duplicate_count > 0:
            warnings.append(f"insurance_duplicates_removed:{duplicate_count}")
        
        # Step 5: Generate final records with __name__, __declare_type__, __fee_month__
        for record in deduplicated_records:
            name = record.get("name", "").strip()
            if not name:
                skipped += 1
                continue
            
            leave_date = record.get("leave_date", "").strip()
            if not leave_date:
                date_parse_failed += 1
                skipped += 1
                continue
            
            dt = _parse_any_date(leave_date)
            if not dt:
                date_parse_failed += 1
                skipped += 1
                continue
            
            fee_month = _next_month_yyyymm(dt)
            if not fee_month:
                fee_month_failed += 1
                skipped += 1
                continue
            
            new_records.append({
                "__name__": name,
                "__fee_month__": fee_month,
                "__declare_type__": "减"
            })
        
        debug_info["email_records_count"] = email_records_count
        debug_info["excel_records_count"] = excel_records_count
        debug_info["duplicate_count"] = duplicate_count
        
        # If no records at all, return error
        if not new_records and not all_leave_records:
            return FillPlan(
                target=FillPlanTarget(),
                warnings=["insurance_no_leave_records_found"],
                llm_used=llm_used,
                debug=debug_info
            )
    else:
        # Original logic for "add" intent (Excel only)
        if not selected_source or not name_key or not effective_date_key:
            return FillPlan(
                target=FillPlanTarget(),
                warnings=["insurance_fallback_failed"],
                llm_used=False,
                debug=debug_info
            )
        
        records = _source_records(selected_source)
        records, sheet_filter_applied = _filter_records_by_sheet_intent(records, template_intent)
        if sheet_filter_applied:
            warnings.append("add_intent_ignore_non_add_sheets")
        for record in records:
            if not isinstance(record, dict):
                continue
            name = str(record.get(name_key, "")).strip()
            if not name:
                skipped += 1
                continue
            dt = _parse_any_date(record.get(effective_date_key))
            if not dt:
                date_parse_failed += 1
                skipped += 1
                continue
            fee_month = _next_month_yyyymm(dt)
            if not fee_month:
                fee_month_failed += 1
                skipped += 1
                continue
            new_records.append({
                "__name__": name,
                "__fee_month__": fee_month,
                "__declare_type__": declare_type
            })

    if skipped:
        warnings.append(f"insurance_records_skipped:{skipped}")
    if date_parse_failed:
        warnings.append(f"insurance_date_parse_failed:{date_parse_failed}")
    if fee_month_failed:
        warnings.append(f"insurance_fee_month_failed:{fee_month_failed}")

    try:
        min_col, min_row, max_col, max_row = range_boundaries(region.table.range)
    except Exception:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["insurance_template_range_invalid"],
            llm_used=llm_used,
            debug=debug_info
        )

    max_header_row = max(region.header_rows) if region.header_rows else min_row
    data_start_row = max_header_row + 1
    start_cell = f"{get_column_letter(min_col)}{data_start_row}"
    clear_end_row = min(max_row, data_start_row + len(new_records) + 30)
    clear_range = f"{get_column_letter(min_col)}{data_start_row}:{get_column_letter(max_col)}{clear_end_row}"

    def _insurance_col_tag(header_path: str) -> str:
        normalized = header_path.replace("\u3000", "").strip()
        if "公积金" in normalized or "公积" in normalized:
            return "gj"
        if "社保" in normalized:
            return "ss"
        return "ss"

    column_mapping: Dict[str, str] = {}
    declare_keys: List[str] = []
    fee_keys: List[str] = []

    if name_columns:
        column_mapping["__name__"] = name_columns[0]

    if type_columns:
        col_letter, header_path = type_columns[0]
        tag = _insurance_col_tag(header_path)
        key = f"__declare_type__{tag}"
        column_mapping[key] = col_letter
        declare_keys.append(key)

    if fee_columns:
        col_letter, header_path = fee_columns[0]
        tag = _insurance_col_tag(header_path)
        key = f"__fee_month__{tag}"
        column_mapping[key] = col_letter
        fee_keys.append(key)

    rows = []
    for record in new_records:
        row_data: Dict[str, Any] = {"__name__": record["__name__"]}
        for key in declare_keys:
            row_data[key] = record["__declare_type__"]
        for key in fee_keys:
            row_data[key] = record["__fee_month__"]
        rows.append(row_data)

    debug_info.update({
        "selected_source_filename": selected_source.get("filename") if isinstance(selected_source, dict) else None,
        "name_key": name_key,
        "effective_date_key": effective_date_key,
        "kept_records": len(new_records),
        "skipped_records": skipped,
        "llm_confidence": llm_confidence,
        "notes": notes,
    })

    sheet_name = _resolve_sheet_name(template_schema, template_schema.sheet_schemas[0] if template_schema.sheet_schemas else None)
    fill_plan = {
        "target": {
            "sheet_name": sheet_name,
            "sheet": sheet_name,
            "region_id": region.region_id,
            "layout_type": "table",
            "clear_policy": "clear_values_keep_format"
        },
        "clear_ranges": [clear_range],
        "row_writes": [{
            "start_cell": start_cell,
            "rows": rows,
            "column_mapping": column_mapping
        }],
        "writes": [],
        "warnings": warnings
    }
    fill_plan["llm_used"] = llm_used
    fill_plan["constant_values_count"] = 0
    _merge_debug_info(fill_plan, debug_info)
    return _dict_to_fill_plan(fill_plan)


def _extract_records_from_source_extracted(extracted_obj: Any, limit: int = 20) -> List[Dict[str, Any]]:
    records = _extract_records_from_container(extracted_obj)
    if not records:
        return []
    return records[:limit]


def _infer_value_type(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if re.match(r'^\d{11}$', text):
        return "phone"
    if re.match(r'^\d{15,18}$', text) or re.match(r'^\d{17}[0-9Xx]$', text):
        return "id"
    if _parse_year_month(text):
        return "date"
    if re.search(r'\d+\.\d+', text) or re.search(r'\d{1,3}(,\d{3})+', text):
        return "amount"
    return "text"


def _value_matches_type(value: Any, expected: str) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if expected == "date":
        return _parse_year_month(text) is not None
    if expected == "id":
        return bool(re.match(r'^\d{15,18}$', text) or re.match(r'^\d{17}[0-9Xx]$', text))
    if expected == "phone":
        return bool(re.match(r'^\d{11}$', text))
    if expected == "amount":
        return bool(re.search(r'\d+\.\d+', text) or re.search(r'\d{1,3}(,\d{3})+', text) or text.isdigit())
    return True


def _infer_template_field_types(template_schema: TemplateSchema) -> Dict[str, str]:
    field_types: Dict[str, str] = {}
    sample_rows = _extract_template_sample_rows(template_schema)
    if not sample_rows:
        return field_types
    for row in sample_rows:
        if not isinstance(row, dict):
            continue
        for key, value in row.items():
            if key in field_types:
                continue
            inferred = _infer_value_type(value)
            if inferred:
                field_types[key] = inferred
    return field_types


def _match_template_headers_to_source_keys(
    template_headers: List[str],
    source_keys: List[str]
) -> Dict[str, Tuple[str, int]]:
    if not template_headers or not source_keys:
        return {}
    norm_key_map: Dict[str, str] = {}
    for key in sorted(source_keys):
        norm_key = _normalize_match_text(key)
        if norm_key and norm_key not in norm_key_map:
            norm_key_map[norm_key] = key
    norm_keys = list(norm_key_map.keys())
    matched: Dict[str, Tuple[str, int]] = {}
    for header in template_headers:
        best_score = -1
        best_norm = None
        for variant in _header_variants(header):
            match = process.extractOne(variant, norm_keys, scorer=fuzz.WRatio)
            if not match:
                continue
            norm_key, score, _ = match
            if score > best_score:
                best_score = score
                best_norm = norm_key
        if best_norm and best_score >= 80:
            matched[header] = (norm_key_map[best_norm], int(best_score))
    return matched


def _score_source_against_template(
    template_schema: TemplateSchema,
    template_headers: List[str],
    records: List[Dict[str, Any]]
) -> Dict[str, Any]:
    if not template_headers or not records:
        return {
            "coverage": 0.0,
            "completeness": 0.0,
            "type_fit": 0.5,
            "score": 0.0,
            "matched_pairs": []
        }
    
    source_keys = set()
    for record in records:
        if isinstance(record, dict):
            source_keys.update([k for k in record.keys() if isinstance(k, str)])
    source_keys_list = sorted(source_keys)
    
    matched = _match_template_headers_to_source_keys(template_headers, source_keys_list)
    coverage = len(matched) / len(template_headers) if template_headers else 0.0
    
    completeness_scores = []
    for header, (key, _) in matched.items():
        non_empty = 0
        for record in records:
            if isinstance(record, dict) and _is_non_empty_value(record.get(key)):
                non_empty += 1
        completeness_scores.append(non_empty / len(records))
    completeness = sum(completeness_scores) / len(completeness_scores) if completeness_scores else 0.0
    
    template_types = _infer_template_field_types(template_schema)
    type_scores = []
    for header, (key, _) in matched.items():
        expected_type = template_types.get(header)
        if not expected_type:
            continue
        matched_count = 0
        for record in records:
            if isinstance(record, dict) and _value_matches_type(record.get(key), expected_type):
                matched_count += 1
        type_scores.append(matched_count / len(records))
    type_fit = sum(type_scores) / len(type_scores) if type_scores else 0.5
    
    score = 0.55 * coverage + 0.35 * completeness + 0.10 * type_fit
    matched_pairs = [{"header": h, "key": k, "score": s} for h, (k, s) in matched.items()]
    
    return {
        "coverage": coverage,
        "completeness": completeness,
        "type_fit": type_fit,
        "score": score,
        "matched_pairs": matched_pairs
    }


def _select_sources_for_template(
    extracted_json: dict,
    template_schema: TemplateSchema,
    warnings: List[str],
    score_threshold: float = 0.65,
    max_sources: int = 1,
    template_intent: Optional[str] = None,
    strict_intent: bool = True,
    excel_only: bool = True,
) -> Tuple[List[dict], List[dict]]:
    sources = extracted_json.get("sources", []) if isinstance(extracted_json, dict) else []
    template_headers = _get_template_headers(template_schema)
    scored_sources = []
    
    for source in sources:
        if not isinstance(source, dict):
            continue
        if excel_only and source.get("source_type") != "excel":
            continue
        records = _extract_records_from_source_extracted(source.get("extracted"))
        detail = _score_source_against_template(template_schema, template_headers, records)
        extracted = source.get("extracted") if isinstance(source, dict) else None
        semantic_key_by_header = {}
        if isinstance(extracted, dict):
            semantic_key_by_header = (extracted.get("metadata") or {}).get("semantic_key_by_header") or {}
        source_headers = list(semantic_key_by_header.keys()) if isinstance(semantic_key_by_header, dict) else []
        source_intent, intent_conf, intent_evidence = _infer_source_intent_simple(
            source.get("filename") or "",
            source_headers,
            records or []
        )
        base_score = detail["score"]
        intent_bonus = 0.0
        if template_intent in ("add", "remove"):
            if source_intent == template_intent:
                intent_bonus = 2.0 * intent_conf
            elif source_intent in ("add", "remove"):
                intent_bonus = -3.0
        final_score = base_score + intent_bonus
        scored_sources.append({
            "source": source,
            "score": final_score,
            "base_score": base_score,
            "coverage": detail["coverage"],
            "completeness": detail["completeness"],
            "type_fit": detail["type_fit"],
            "matched_pairs": detail["matched_pairs"],
            "intent": source_intent,
            "intent_confidence": intent_conf,
            "intent_evidence": intent_evidence,
        })
    
    if strict_intent and template_intent in ("add", "remove"):
        for s in scored_sources:
            if s["intent"] in ("add", "remove") and s["intent"] != template_intent:
                s["filtered_by_intent"] = True
        filtered_candidates = [
            s for s in scored_sources
            if not (s.get("filtered_by_intent") and s["intent"] in ("add", "remove"))
        ]
        if not any(s["intent"] == template_intent for s in filtered_candidates):
            details = [
                {
                    "filename": s["source"].get("filename"),
                    "source_type": s["source"].get("source_type"),
                    "parent_source_id": s["source"].get("parent_source_id"),
                    "score": s["score"],
                    "base_score": s["base_score"],
                    "coverage": s["coverage"],
                    "completeness": s["completeness"],
                    "type_fit": s["type_fit"],
                    "intent": s["intent"],
                    "intent_confidence": s["intent_confidence"],
                    "intent_evidence": s["intent_evidence"],
                    "filtered_by_intent": s.get("filtered_by_intent", False),
                }
                for s in scored_sources
            ]
            details.append({"no_intent_matched": True, "template_intent": template_intent})
            return [], details
        scored_sources = filtered_candidates

    scored_sources.sort(
        key=lambda s: (
            -s["score"],
            -float(s.get("intent_confidence") or 0.0),
            -float(s.get("base_score") or 0.0),
            (s["source"].get("filename") or "")
        )
    )
    
    selected = [s for s in scored_sources if s["score"] >= score_threshold][:max_sources]
    if not selected and scored_sources:
        selected = [scored_sources[0]]
        warnings.append("source_selection: no source met score threshold; using top-1 source")
    
    selected_sources = [s["source"] for s in selected]
    details = [
        {
            "filename": s["source"].get("filename"),
            "source_type": s["source"].get("source_type"),
            "parent_source_id": s["source"].get("parent_source_id"),
            "score": s["score"],
            "base_score": s["base_score"],
            "coverage": s["coverage"],
            "completeness": s["completeness"],
            "type_fit": s["type_fit"],
            "intent": s["intent"],
            "intent_confidence": s["intent_confidence"],
            "intent_evidence": s["intent_evidence"],
        }
        for s in scored_sources
    ]
    return selected_sources, details


def _build_selected_extracted_json(extracted_json: dict, selected_sources: List[dict]) -> dict:
    def _is_scalar_value(value: Any) -> bool:
        return value is None or isinstance(value, (str, int, float, bool))
    
    merged = {}
    for source in selected_sources:
        extracted = source.get("extracted") if isinstance(source, dict) else None
        if isinstance(extracted, dict):
            for key, value in extracted.items():
                if not isinstance(key, str):
                    continue
                if not _is_scalar_value(value):
                    continue
                if key not in merged or not merged[key]:
                    merged[key] = value
    
    return {
        "sources": selected_sources,
        "merged": merged
    }


def _gate_records_by_filled_ratio(
    template_schema: TemplateSchema,
    records: List[Dict[str, Any]],
    warnings: List[str],
    min_ratio: float = 0.4,
    fallback_top_n: int = 5
) -> List[Dict[str, Any]]:
    if not records:
        return records
    template_headers = _get_template_headers(template_schema)
    source_keys = set()
    for record in records:
        if isinstance(record, dict):
            source_keys.update([k for k in record.keys() if isinstance(k, str)])
    matched = _match_template_headers_to_source_keys(template_headers, sorted(source_keys))
    matched_keys = [key for key, _ in matched.values()]
    
    if not matched_keys:
        warnings.append("record_gating: no matched keys found; skipping gating")
        return records
    
    scored = []
    for record in records:
        if not isinstance(record, dict):
            scored.append((0.0, record))
            continue
        filled = 0
        for key in matched_keys:
            if _is_non_empty_value(record.get(key)):
                filled += 1
        ratio = filled / len(matched_keys) if matched_keys else 0.0
        scored.append((ratio, record))
    
    kept = [record for ratio, record in scored if ratio >= min_ratio]
    if not kept and records:
        scored.sort(key=lambda x: x[0], reverse=True)
        kept = [record for _, record in scored[:fallback_top_n]]
        warnings.append("record_gating: all records below threshold; keeping top ratios")
    
    return kept


def _apply_record_gating_to_extracted_json(
    extracted_json: dict,
    template_schema: TemplateSchema,
    warnings: List[str]
) -> dict:
    if not isinstance(extracted_json, dict):
        return extracted_json
    
    def filter_records_list(records: list) -> list:
        gated = _gate_records_by_filled_ratio(template_schema, records, warnings)
        return gated
    
    filtered = {}
    if "sources" in extracted_json:
        filtered["sources"] = []
        for source in extracted_json.get("sources", []):
            if not isinstance(source, dict):
                filtered["sources"].append(source)
                continue
            filtered_source = dict(source)
            extracted = source.get("extracted")
            if isinstance(extracted, dict):
                updated_extracted = dict(extracted)
                for key, value in extracted.items():
                    if isinstance(value, list) and value and isinstance(value[0], dict):
                        updated_extracted[key] = filter_records_list(value)
                filtered_source["extracted"] = updated_extracted
            elif isinstance(extracted, list) and extracted and isinstance(extracted[0], dict):
                filtered_source["extracted"] = filter_records_list(extracted)
            filtered["sources"].append(filtered_source)
    
    if "merged" in extracted_json:
        filtered["merged"] = extracted_json.get("merged", {})
    
    for key in extracted_json:
        if key not in filtered:
            filtered[key] = extracted_json[key]
    
    return filtered


def _infer_fee_month(
    template_schema: TemplateSchema,
    records: List[Dict[str, Any]],
    template_intent: Optional[str],
    column_mapping: Dict[str, str],
    constant_values: Dict[str, str],
    warnings: List[str]
) -> Tuple[Dict[str, str], Dict[str, str]]:
    if not records:
        return column_mapping, constant_values
    
    fee_headers = _find_fee_month_headers(template_schema)
    if not fee_headers:
        return column_mapping, constant_values
    
    output_format = None
    sample_rows = _extract_template_sample_rows(template_schema)
    for row in sample_rows:
        if not isinstance(row, dict):
            continue
        for header in fee_headers:
            if header in row and row[header]:
                output_format = _detect_year_month_format_from_value(row[header])
                if output_format:
                    break
        if output_format:
            break
    if not output_format:
        output_format = "YYYYMM"
    
    derived_values = []
    for record in records:
        if not isinstance(record, dict):
            derived_values.append("")
            continue
        ym = _extract_fee_month_from_record(record, template_intent)
        if not ym:
            derived_values.append("")
            continue
        derived_values.append(_format_year_month(ym[0], ym[1], output_format))

    # IMPORTANT (deterministic): always write per-record derived fee month.
    # Do NOT promote to constant_values here, otherwise downstream mapping may
    # skip record-level provenance and cause instability across multi-source inputs.
    
    for record, value in zip(records, derived_values):
        if isinstance(record, dict):
            record["__fee_month__"] = value
    
    for header in fee_headers:
        if header not in column_mapping:
            column_mapping[header] = "__fee_month__"
    
    return column_mapping, constant_values


def _find_fee_month_headers(template_schema: TemplateSchema) -> List[str]:
    headers: List[str] = []
    anchors = ("费用", "缴费", "参保", "申报", "社保", "所属")
    for sheet_schema in template_schema.sheet_schemas:
        for region in sheet_schema.regions:
            if not region.table or not region.table.header:
                continue
            for header in region.table.header:
                header_path = getattr(header, "header_path", None)
                if not header_path:
                    continue
                normalized = _normalize_header(header_path)
                if ("年月" in normalized or "月份" in normalized) and any(a in normalized for a in anchors):
                    headers.append(header_path)
                elif normalized in ("费用年月", "缴费年月", "参保月份", "申报月份", "社保月份", "所属期", "费用月份", "缴费月份"):
                    headers.append(header_path)
    return list(dict.fromkeys(headers))


def _extract_fee_month_from_record(record: Dict[str, Any], template_intent: Optional[str]) -> Optional[Tuple[int, int]]:
    direct_keys = (
        "费用年月", "缴费年月", "参保月份", "申报月份", "社保月份", "所属期", "费用月份", "缴费月份", "参保年月", "申报年月"
    )
    for key in record.keys():
        if not isinstance(key, str):
            continue
        if key in direct_keys or _normalize_header(key) in [_normalize_header(k) for k in direct_keys]:
            ym = _parse_year_month(record.get(key))
            if ym:
                return ym
    
    if template_intent == "remove":
        date_keys = ["离职日期", "退保日期", "退场日期", "退工日期", "变动日期", "办理日期"]
    else:
        date_keys = ["入职日期", "签订日期", "生效日期", "变动日期", "办理日期"]
    
    for key in date_keys:
        if key in record:
            ym = _parse_year_month(record.get(key))
            if ym:
                return ym
    
    for key, value in record.items():
        if isinstance(key, str) and ("日期" in key or "时间" in key):
            ym = _parse_year_month(value)
            if ym:
                return ym
    
    return None


def _apply_record_filter(extracted_json: dict, record_filter: Optional[Dict[str, Any]]) -> dict:
    """
    Apply record filter to extracted_json, returning filtered version.
    Filters ALL sources and merged data based on the filter criteria.
    """
    if record_filter is None:
        return extracted_json
    
    field = record_filter.get("field")
    values = record_filter.get("values", [])
    exclude = record_filter.get("exclude", False)
    
    if not field or not values:
        return extracted_json
    
    # Normalize values for comparison
    normalized_values = [str(v).lower().strip() for v in values]
    
    def should_include(record: dict) -> bool:
        record_value = record.get(field)
        if record_value is None:
            # If field doesn't exist, include only if we're excluding (not matching = include)
            return exclude
        
        normalized_record_value = str(record_value).lower().strip()
        matches = normalized_record_value in normalized_values
        
        # If exclude=True, include records that DON'T match
        # If exclude=False, include records that DO match
        return not matches if exclude else matches
    
    def filter_records_list(records: list) -> list:
        """Filter a list of records."""
        return [r for r in records if isinstance(r, dict) and should_include(r)]
    
    def filter_extracted_data(extracted: Any) -> Any:
        """Filter records within an extracted data structure."""
        if extracted is None:
            return None
        
        # If extracted is directly a list of records
        if isinstance(extracted, list):
            if len(extracted) > 0 and isinstance(extracted[0], dict):
                original_count = len(extracted)
                filtered = filter_records_list(extracted)
                if original_count != len(filtered):
                    logger.debug("_apply_record_filter: Filtered list %d -> %d records", original_count, len(filtered))
                return filtered
            return extracted
        
        # If extracted is a dict, look for list fields to filter
        if isinstance(extracted, dict):
            filtered_extracted = {}
            for key, value in extracted.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    original_count = len(value)
                    filtered_value = filter_records_list(value)
                    filtered_extracted[key] = filtered_value
                    if original_count != len(filtered_value):
                        logger.debug("_apply_record_filter: Filtered '%s' %d -> %d records", key, original_count, len(filtered_value))
                else:
                    filtered_extracted[key] = value
            return filtered_extracted
        
        return extracted
    
    # Create filtered version of extracted_json
    filtered = {}
    total_original = 0
    total_filtered = 0
    
    # Filter records in ALL sources
    if "sources" in extracted_json:
        filtered["sources"] = []
        for source_idx, source in enumerate(extracted_json["sources"]):
            if not isinstance(source, dict):
                filtered["sources"].append(source)
                continue
            
            filtered_source = dict(source)
            source_filename = source.get("filename", f"source_{source_idx}")
            extracted = source.get("extracted")
            
            if extracted is not None:
                # Count original records
                original_records = _count_records_in_extracted(extracted)
                total_original += original_records
                
                # Apply filter
                filtered_source["extracted"] = filter_extracted_data(extracted)
                
                # Count filtered records
                filtered_records = _count_records_in_extracted(filtered_source["extracted"])
                total_filtered += filtered_records
                
                if original_records != filtered_records:
                    logger.info("_apply_record_filter: Source '%s': %d -> %d records (field='%s')", 
                               source_filename, original_records, filtered_records, field)
            
            filtered["sources"].append(filtered_source)
    
    # Filter records in merged data
    if "merged" in extracted_json:
        merged = extracted_json["merged"]
        if isinstance(merged, dict):
            filtered["merged"] = {}
            for key, value in merged.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    filtered["merged"][key] = filter_records_list(value)
                else:
                    filtered["merged"][key] = value
    
    # Copy other keys
    for key in extracted_json:
        if key not in filtered:
            filtered[key] = extracted_json[key]
    
    logger.info("_apply_record_filter: Total filtered %d -> %d records across all sources", total_original, total_filtered)
    return filtered


def _count_records_in_extracted(extracted: Any) -> int:
    """Count the number of records in an extracted data structure."""
    if extracted is None:
        return 0
    
    if isinstance(extracted, list):
        if len(extracted) > 0 and isinstance(extracted[0], dict):
            return len(extracted)
        return 0
    
    if isinstance(extracted, dict):
        # Look for common list fields
        for key in ["data", "records", "rows", "items"]:
            if key in extracted and isinstance(extracted[key], list):
                return len(extracted[key])
        # Check any list field
        for value in extracted.values():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                return len(value)
    
    return 0


def _extract_sheet_title(sheet_schema: Any) -> Optional[str]:
    if not sheet_schema:
        return None
    for attr in ("sheet_name", "sheet", "title", "name"):
        value = getattr(sheet_schema, attr, None)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def _resolve_sheet_name(template_schema: TemplateSchema, sheet_schema: Any = None) -> str:
    sheet_name = _extract_sheet_title(sheet_schema)
    if sheet_name:
        return sheet_name
    for ss in template_schema.sheet_schemas or []:
        sheet_name = _extract_sheet_title(ss)
        if sheet_name:
            return sheet_name
    return "Sheet1"


def _build_fill_plan_from_mapping(
    template_schema: TemplateSchema,
    extracted_json: dict,
    header_to_key: Dict[str, str],
    mapping_warnings: Optional[List[str]] = None,
    constant_values: Optional[Dict[str, str]] = None,
    template_intent: Optional[str] = None
) -> Optional[dict]:
    if not template_schema.sheet_schemas:
        return None
    sheet = template_schema.sheet_schemas[0]
    if not sheet.regions or not sheet.regions[0].table:
        return None
    region = sheet.regions[0]
    table = region.table
    headers = table.header or []
    if not headers:
        return None
    
    try:
        min_col, min_row, max_col, max_row = range_boundaries(table.range)
    except Exception:
        return None
    
    max_header_row = max(region.header_rows) if region.header_rows else min_row
    data_start_row = max_header_row + 1
    start_cell = f"{get_column_letter(min_col)}{data_start_row}"
    warnings: List[str] = list(mapping_warnings or [])
    
    # IMPORTANT: records must only come from sources; also ensure __source_file__ is present.
    records, _, _ = _extract_records_with_provenance(extracted_json)
    if not records:
        return None
    if template_intent in ("add", "remove"):
        filtered_records = [
            r for r in records
            if (_infer_record_type(r) or _infer_record_type_from_sheet(r)) == template_intent
        ]
        if not filtered_records:
            warnings.append("record_type_filtered_empty")
            records = []
        else:
            records = filtered_records
    
    header_lookup = _build_header_lookup(headers)
    column_mapping: Dict[str, str] = {}
    constant_column_mapping: Dict[str, Tuple[str, str]] = {}  # header -> (col_letter, value)
    
    logger.debug("_build_fill_plan_from_mapping: header_to_key has %d items", len(header_to_key))
    logger.debug("_build_fill_plan_from_mapping: template headers: %s", list(header_lookup.keys())[:10])
    
    # Process regular column mappings (header -> extracted_key)
    matched_count = 0
    for template_header, extracted_key in header_to_key.items():
        if not extracted_key:
            continue
        header_info = _lookup_header(header_lookup, template_header)
        if not header_info:
            warnings.append(f"Template header '{template_header}' not found in schema")
            logger.debug("_build_fill_plan_from_mapping: Header '%s' not found in template", template_header)
            continue
        col_letter = header_info.col_letter
        if extracted_key in column_mapping and column_mapping[extracted_key] != col_letter:
            warnings.append(
                f"Duplicate extracted key '{extracted_key}' for columns {column_mapping[extracted_key]} and {col_letter}"
            )
            continue
        column_mapping[extracted_key] = col_letter
        matched_count += 1
        logger.debug("_build_fill_plan_from_mapping: Matched '%s' -> '%s' -> column %s", template_header, extracted_key, col_letter)
    
    logger.info("_build_fill_plan_from_mapping: Matched %d of %d header mappings", matched_count, len(header_to_key))
    
    # Process constant values (header -> fixed_value for all rows)
    if constant_values:
        for template_header, const_value in constant_values.items():
            if const_value is None:
                continue
            if isinstance(const_value, str) and not const_value.strip():
                continue
            header_info = _lookup_header(header_lookup, template_header)
            if not header_info:
                warnings.append(f"Constant header '{template_header}' not found in schema")
                continue
            col_letter = header_info.col_letter
            # Use a special key prefix to avoid collision with extracted keys
            const_key = f"__const__{template_header}"
            constant_column_mapping[const_key] = (col_letter, const_value)
            logger.debug("_build_fill_plan_from_mapping: Constant '%s' = '%s' -> column %s", 
                        template_header, const_value, col_letter)
    
    logger.info("_build_fill_plan_from_mapping: Final column_mapping=%d, constant_column_mapping=%d", 
                len(column_mapping), len(constant_column_mapping))
    
    if not column_mapping and not constant_column_mapping:
        logger.warning("_build_fill_plan_from_mapping: No mappings found, returning None")
        return None
    
    rows = []
    for rec in records:
        record = rec if isinstance(rec, dict) else {}
        row = {}
        # Add values from extracted data
        for extracted_key in column_mapping.keys():
            value = _get_record_value(record, extracted_key)
            if value is None:
                value = ""
            row[extracted_key] = value
        # Add constant values (same for all rows)
        for const_key, (col_letter, const_value) in constant_column_mapping.items():
            row[const_key] = const_value
        rows.append(row)
    
    # Merge constant_column_mapping into column_mapping for the fill plan
    for const_key, (col_letter, _) in constant_column_mapping.items():
        column_mapping[const_key] = col_letter
    
    clear_end_row = min(max_row, data_start_row + len(rows) + 10)
    clear_range = f"{get_column_letter(min_col)}{data_start_row}:{get_column_letter(max_col)}{clear_end_row}"
    
    sheet_name = _resolve_sheet_name(template_schema, sheet)
    fill_plan = {
        "target": {
            "sheet_name": sheet_name,
            "sheet": sheet_name,
            "region_id": region.region_id,
            "layout_type": "table",
            "clear_policy": "clear_values_keep_format"
        },
        "clear_ranges": [clear_range],
        "row_writes": [{
            "start_cell": start_cell,
            "rows": rows,
            "column_mapping": column_mapping
        }],
        "writes": [],
        "warnings": warnings
    }
    
    return fill_plan

def _enrich_mapping_with_fuzzy_match(
    template_schema: TemplateSchema,
    extracted_json: dict,
    header_to_key: Dict[str, str]
) -> Tuple[Dict[str, str], List[str]]:
    if not template_schema.sheet_schemas:
        return header_to_key, []
    sheet = template_schema.sheet_schemas[0]
    if not sheet.regions or not sheet.regions[0].table:
        return header_to_key, []
    headers = sheet.regions[0].table.header or []
    if not headers:
        return header_to_key, []
    
    records = _extract_records(extracted_json)
    extracted_keys = _collect_extracted_keys(records)
    if not extracted_keys:
        return header_to_key, []
    
    mapped_headers = set(header_to_key.keys())
    warnings: List[str] = []
    enriched_mapping = dict(header_to_key)
    
    for header in headers:
        header_path = getattr(header, "header_path", None)
        if not header_path or header_path in mapped_headers:
            continue
        match = process.extractOne(
            header_path,
            extracted_keys,
            scorer=fuzz.WRatio
        )
        if not match:
            continue
        best_key, score, _ = match
        if score > 80 and best_key not in enriched_mapping.values():
            enriched_mapping[header_path] = best_key
            warnings.append(f"Auto-mapped '{header_path}' to '{best_key}' (score: {int(score)})")
    
    return enriched_mapping, warnings

def _collect_extracted_keys(records: List[Dict[str, Any]]) -> List[str]:
    keys: List[str] = []
    seen = set()
    for record in records:
        if not isinstance(record, dict):
            continue
        for key in record.keys():
            if isinstance(key, str) and key not in seen:
                seen.add(key)
                keys.append(key)
    return keys

def _normalize_header(text: str) -> str:
    return text.replace(" ", "").replace("_", "").replace("-", "").strip().lower()

def _build_header_lookup(headers: List[Any]) -> Dict[str, Any]:
    lookup: Dict[str, Any] = {}
    for header in headers:
        header_path = getattr(header, "header_path", None)
        if not header_path:
            continue
        variants = set()
        variants.add(header_path)
        variants.add(_normalize_header(header_path))
        if "/" in header_path:
            short = header_path.split("/")[-1].strip()
            if short:
                variants.add(short)
                variants.add(_normalize_header(short))
        for v in variants:
            if v and v not in lookup:
                lookup[v] = header
    return lookup

def _lookup_header(lookup: Dict[str, Any], template_header: str) -> Optional[Any]:
    if template_header in lookup:
        return lookup[template_header]
    normalized = _normalize_header(template_header)
    return lookup.get(normalized)

def _get_record_value(record: Dict[str, Any], key: str) -> Any:
    if key in record:
        return record.get(key)
    if "." in key or "/" in key:
        normalized = key.replace("/", ".")
        parts = [p for p in normalized.split(".") if p]
        current: Any = record
        for part in parts:
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                return None
        return current
    return None

def _extract_samples_from_each_source(
    extracted_json: dict,
    samples_per_source: int = 2,
    sources_override: Optional[List[Dict[str, Any]]] = None
) -> List[Dict[str, Any]]:
    """
    Extract sample records from EACH source file.
    This ensures LLM sees data from all sources to make correct filtering decisions.
    """
    if not isinstance(extracted_json, dict):
        return []
    
    all_samples = []
    sources = sources_override if sources_override is not None else extracted_json.get("sources", [])
    
    for s_idx, s in enumerate(sources):
        if not isinstance(s, dict):
            continue
        
        source_filename = s.get("filename", f"source_{s_idx}")
        ex = s.get("extracted")
        records = _extract_records_from_container(ex)
        
        if records:
            # Take samples from this source
            samples = records[:samples_per_source]
            # Add source info to each sample for context
            for sample in samples:
                sample_with_source = dict(sample)
                sample_with_source["__source_file__"] = source_filename
                all_samples.append(sample_with_source)
            
            logger.debug("_extract_samples_from_each_source: Added %d samples from '%s'", len(samples), source_filename)
    
    logger.info("_extract_samples_from_each_source: Total %d samples from %d sources", len(all_samples), len(sources))
    return all_samples


def _extract_template_sample_rows(template_schema: TemplateSchema) -> List[Dict[str, Any]]:
    """
    Extract sample rows from template as few-shot examples.
    
    These examples help the LLM understand:
    1. The expected data format for each column
    2. The mapping between header names and actual data
    3. Common patterns in the data (date formats, ID formats, etc.)
    """
    sample_rows = []
    
    for sheet_schema in template_schema.sheet_schemas:
        for region in sheet_schema.regions:
            if region.table and region.table.sample_rows:
                sample_rows.extend(region.table.sample_rows)
                # Return first 3 sample rows to keep context manageable
                if len(sample_rows) >= 3:
                    return sample_rows[:3]
    
    return sample_rows


def _extract_records(extracted_json: dict) -> List[Dict[str, Any]]:
    """
    Extract records from ALL sources in extracted_json.
    Combines records from multiple source files.
    """
    if not isinstance(extracted_json, dict):
        return []
    
    all_records: List[Dict[str, Any]] = []
    
    # Extract records from ALL sources
    sources = extracted_json.get("sources", [])
    for s_idx, s in enumerate(sources):
        if not isinstance(s, dict):
            continue
        source_filename = s.get("filename", f"source_{s_idx}")
        ex = s.get("extracted")
        records = _extract_records_from_container(ex)
        if records:
            logger.debug("_extract_records: Found %d records from source '%s'", len(records), source_filename)
            all_records.extend(records)
    
    if all_records:
        logger.info("_extract_records: Total %d records from %d sources", len(all_records), len(sources))
    # IMPORTANT: records must ONLY come from sources[*].extracted.
    # merged is reserved for constant inference and must never provide records.
    return all_records

def _extract_records_from_container(container: Any) -> List[Dict[str, Any]]:
    if isinstance(container, list) and len(container) > 0:
        if isinstance(container[0], dict):
            return list(container)
        if isinstance(container[0], list):
            records: List[Dict[str, Any]] = []
            for item in container:
                if isinstance(item, dict):
                    records.append(item)
            return records
    if isinstance(container, dict):
        priority_keys = ["data", "records", "rows", "items", "extracted_data", "table_data"]
        for key in priority_keys:
            value = container.get(key)
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                return list(value)
        for value in container.values():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                return list(value)
    return []


def _count_records_in_sources(extracted_json: Any) -> int:
    """Count records by scanning extracted_json['sources'][*]['extracted'] only."""
    if not isinstance(extracted_json, dict):
        return 0
    sources = extracted_json.get("sources", [])
    if not isinstance(sources, list):
        return 0
    total = 0
    for s in sources:
        if not isinstance(s, dict):
            continue
        ex = s.get("extracted")
        total += len(_extract_records_from_container(ex) or [])
    return total


def _extract_records_with_provenance(extracted_json: Any) -> Tuple[List[Dict[str, Any]], Dict[str, int], List[str]]:
    """
    Aggregate records ONLY from extracted_json['sources'][*]['extracted'].
    Inject __source_file__ if missing, using the corresponding source filename.
    
    Returns:
      - all_records: list of records
      - counts_per_source: filename -> count
      - source_files: list of filenames that contributed >=1 record
    """
    if not isinstance(extracted_json, dict):
        return [], {}, []
    sources = extracted_json.get("sources", [])
    if not isinstance(sources, list):
        return [], {}, []
    
    all_records: List[Dict[str, Any]] = []
    counts: Dict[str, int] = {}
    contributing_files: List[str] = []
    
    for s_idx, s in enumerate(sources):
        if not isinstance(s, dict):
            continue
        filename = s.get("filename", f"source_{s_idx}")
        ex = s.get("extracted")
        records = _extract_records_from_container(ex) or []
        if not records:
            continue
        injected = 0
        normalized_records = []
        for r in records:
            if not isinstance(r, dict):
                continue
            rec = dict(r)
            if "__source_file__" not in rec or not rec.get("__source_file__"):
                rec["__source_file__"] = filename
                injected += 1
            normalized_records.append(rec)
        if normalized_records:
            all_records.extend(normalized_records)
            counts[filename] = counts.get(filename, 0) + len(normalized_records)
            if filename not in contributing_files:
                contributing_files.append(filename)
            if injected:
                logger.debug("_extract_records_with_provenance: injected __source_file__ for %d records from %s", injected, filename)
    
    return all_records, counts, contributing_files


def _normalize_text(text: Any) -> str:
    return str(text).lower().replace(" ", "").replace("_", "").replace("-", "").strip()


def _infer_template_intent_simple(template_filename: str, template_headers: List[str]) -> Optional[str]:
    add_keywords = ("增员", "入职", "新增", "录用", "入司")
    remove_keywords = ("减员", "离职", "退场", "终止", "解除", "注销")
    filename_text = _normalize_text(template_filename or "")
    header_texts = [_normalize_text(h) for h in (template_headers or []) if h]

    filename_add_hits = [k for k in add_keywords if k in filename_text]
    filename_remove_hits = [k for k in remove_keywords if k in filename_text]
    header_add_hits = [k for k in add_keywords if any(k in ht for ht in header_texts)]
    header_remove_hits = [k for k in remove_keywords if any(k in ht for ht in header_texts)]

    has_add = bool(filename_add_hits or header_add_hits)
    has_remove = bool(filename_remove_hits or header_remove_hits)

    if has_add and not has_remove:
        return "add"
    if has_remove and not has_add:
        return "remove"
    if has_add and has_remove:
        if len(header_add_hits) > len(header_remove_hits):
            return "add"
        if len(header_remove_hits) > len(header_add_hits):
            return "remove"
    return None


def _infer_source_intent_simple(
    source_filename: str,
    source_headers: List[str],
    sample_records: List[Dict[str, Any]]
) -> Tuple[Optional[str], float, Dict[str, List[str]]]:
    add_keywords = ("入职", "增员", "新增")
    remove_keywords = ("离职", "减员", "退场")
    in_service_keywords = ("在职", "花名册", "人员信息", "名册")
    header_add_keywords = ("入职日期", "到岗", "录用")
    header_remove_keywords = ("离职日期", "离岗", "终止日期", "解除日期")

    filename_text = _normalize_text(source_filename or "")
    header_texts = [_normalize_text(h) for h in (source_headers or []) if h]

    filename_add_hits = [k for k in add_keywords if k in filename_text]
    filename_remove_hits = [k for k in remove_keywords if k in filename_text]
    filename_in_service_hits = [k for k in in_service_keywords if k in filename_text]

    header_add_hits = [k for k in header_add_keywords if any(k in ht for ht in header_texts)]
    header_remove_hits = [k for k in header_remove_keywords if any(k in ht for ht in header_texts)]

    value_remove_hits = []
    for record in sample_records or []:
        if not isinstance(record, dict):
            continue
        for value in record.values():
            if not isinstance(value, str):
                continue
            val_text = _normalize_text(value)
            for k in remove_keywords:
                if k in val_text and k not in value_remove_hits:
                    value_remove_hits.append(k)

    evidence = {
        "filename_hits": sorted(set(filename_add_hits + filename_remove_hits + filename_in_service_hits)),
        "header_hits": sorted(set(header_add_hits + header_remove_hits)),
        "value_hits": sorted(set(value_remove_hits)),
    }

    intent: Optional[str] = None
    confidence = 0.0

    if filename_add_hits and not filename_remove_hits:
        intent = "add"
        confidence = 0.9
    elif filename_remove_hits and not filename_add_hits:
        intent = "remove"
        confidence = 0.9
    elif filename_add_hits and filename_remove_hits:
        if len(header_add_hits) > len(header_remove_hits):
            intent = "add"
            confidence = 0.7
        elif len(header_remove_hits) > len(header_add_hits):
            intent = "remove"
            confidence = 0.7
        else:
            intent = None
            confidence = 0.0
    else:
        if header_add_hits and not header_remove_hits:
            intent = "add"
            confidence = 0.6
        elif header_remove_hits and not header_add_hits:
            intent = "remove"
            confidence = 0.6
        elif header_add_hits and header_remove_hits:
            if len(header_add_hits) > len(header_remove_hits):
                intent = "add"
                confidence = 0.5
            elif len(header_remove_hits) > len(header_add_hits):
                intent = "remove"
                confidence = 0.5
        elif value_remove_hits:
            intent = "remove"
            confidence = 0.3

    if filename_in_service_hits and not (filename_add_hits or filename_remove_hits):
        if intent is None:
            confidence = 0.2
        else:
            confidence = max(0.0, confidence - 0.2)

    return intent, float(min(1.0, max(0.0, confidence))), evidence


def _detect_template_intent(template_filename: Optional[str], template_schema: TemplateSchema) -> Optional[str]:
    """
    Detect whether template is for add (增员/入职) or remove (减员/离职) records.
    Returns "add", "remove", or None if ambiguous.
    """
    add_keywords = ("增员", "入职", "新增", "新入职", "新增人员", "加员", "扩员")
    remove_keywords = ("减员", "离职", "退场", "辞退", "解除", "退工", "停保", "退保")
    
    filename_text = _normalize_text(template_filename or "")
    if filename_text:
        has_add = any(k in filename_text for k in add_keywords)
        has_remove = any(k in filename_text for k in remove_keywords)
        if has_add and not has_remove:
            return "add"
        if has_remove and not has_add:
            return "remove"
    
    sheet_names = [ss.sheet for ss in template_schema.sheet_schemas]
    header_paths = []
    for sheet_schema in template_schema.sheet_schemas:
        for region in sheet_schema.regions:
            if region.table and region.table.header:
                for header in region.table.header:
                    if header.header_path:
                        header_paths.append(header.header_path)
    
    schema_text = _normalize_text(" ".join(sheet_names + header_paths))
    has_add = any(k in schema_text for k in add_keywords)
    has_remove = any(k in schema_text for k in remove_keywords)
    if has_add and not has_remove:
        return "add"
    if has_remove and not has_add:
        return "remove"
    
    return None


def _infer_source_intent(filename: Optional[str]) -> Optional[str]:
    add_keywords = ("增员", "入职", "新增", "新入职", "新增人员", "加员", "扩员")
    remove_keywords = ("减员", "离职", "退场", "辞退", "解除", "退工", "停保", "退保")
    in_service_keywords = ("在职", "在岗", "在册")
    filename_text = _normalize_text(filename or "")
    if not filename_text:
        return None
    has_add = any(k in filename_text for k in add_keywords)
    has_remove = any(k in filename_text for k in remove_keywords)
    has_in_service = any(k in filename_text for k in in_service_keywords)
    if has_add and not has_remove:
        return "add"
    if has_remove and not has_add:
        return "remove"
    if has_in_service and not has_add and not has_remove:
        return "in_service"
    return None


def _infer_explicit_record_intent(record: Dict[str, Any]) -> Optional[str]:
    intent_keys = ("社保/申报类型", "申报类型", "社保类型", "变动类型", "办理类型", "业务类型", "操作类型")
    add_keywords = ("增员", "入职", "新增", "新入职", "新增人员", "加员", "扩员")
    remove_keywords = ("减员", "离职", "退场", "辞退", "解除", "退工", "停保", "退保")
    in_service_keywords = ("在职", "在岗", "在册")
    
    for key, value in record.items():
        key_text = _normalize_text(key)
        if not any(_normalize_text(k) in key_text for k in intent_keys):
            continue
        if isinstance(value, str):
            val_text = _normalize_text(value)
            has_add = any(k in val_text for k in add_keywords)
            has_remove = any(k in val_text for k in remove_keywords)
            has_in_service = any(k in val_text for k in in_service_keywords)
            if has_add and not has_remove:
                return "add"
            if has_remove and not has_add:
                return "remove"
            if has_in_service and not has_add and not has_remove:
                return "in_service"
    return None


def _record_filter_conflicts_with_intent(record_filter: Dict[str, Any], template_intent: str) -> bool:
    if not template_intent or not record_filter:
        return False
    values = record_filter.get("values") or []
    values_text = _normalize_text(" ".join([str(v) for v in values]))
    exclude = bool(record_filter.get("exclude", False))
    add_keywords = ("增员", "入职", "新增", "新入职", "新增人员", "加员", "扩员")
    remove_keywords = ("减员", "离职", "退场", "辞退", "解除", "退工", "停保", "退保")
    has_add = any(k in values_text for k in add_keywords)
    has_remove = any(k in values_text for k in remove_keywords)
    
    if template_intent == "add":
        if not exclude and has_remove and not has_add:
            return True
        if exclude and has_add and not has_remove:
            return True
    if template_intent == "remove":
        if not exclude and has_add and not has_remove:
            return True
        if exclude and has_remove and not has_add:
            return True
    return False


def _infer_record_intent(record: Dict[str, Any]) -> Optional[str]:
    add_keywords = ("增员", "入职", "新增", "新入职", "新增人员", "加员", "扩员")
    remove_keywords = ("减员", "离职", "退场", "辞退", "解除", "退工", "停保", "退保")
    in_service_keywords = ("在职", "在岗", "在册")
    
    add_hits = 0
    remove_hits = 0
    in_service_hits = 0
    
    for key, value in record.items():
        key_text = _normalize_text(key)
        if any(k in key_text for k in add_keywords):
            add_hits += 1
        if any(k in key_text for k in remove_keywords):
            remove_hits += 1
        if any(k in key_text for k in in_service_keywords):
            in_service_hits += 1
        
        if isinstance(value, str):
            val_text = _normalize_text(value)
            if any(k in val_text for k in add_keywords):
                add_hits += 1
            if any(k in val_text for k in remove_keywords):
                remove_hits += 1
            if any(k in val_text for k in in_service_keywords):
                in_service_hits += 1
    
    if add_hits > 0 and remove_hits == 0:
        return "add"
    if remove_hits > 0 and add_hits == 0:
        return "remove"
    if in_service_hits > 0 and add_hits == 0 and remove_hits == 0:
        return "in_service"
    return None


def _infer_record_type(record: Dict[str, Any]) -> Optional[str]:
    explicit = _infer_explicit_record_intent(record)
    if explicit in ("add", "remove"):
        return explicit
    remove_keys = ("terminationdate", "terminationreason")
    add_keys = (
        "startdate",
        "employeestatus",
        "employmentstatus",
        "monthlycontribution",
        "contribution",
        "feemonth"
    )
    for key in record.keys():
        if not isinstance(key, str):
            continue
        normalized = _normalize_header(key)
        if any(normalized == rk or normalized.startswith(rk) for rk in remove_keys):
            return "remove"
    for key in record.keys():
        if not isinstance(key, str):
            continue
        normalized = _normalize_header(key)
        if any(normalized == ak or normalized.startswith(ak) for ak in add_keys):
            return "add"
    return None


def _infer_record_type_from_sheet(record: Dict[str, Any]) -> Optional[str]:
    sheet_name_text = str(record.get("__sheet_name__", "") or "")
    if "减员" in sheet_name_text or "离职" in sheet_name_text:
        return "remove"
    if "增员" in sheet_name_text:
        return "add"
    return None


def _infer_template_intent_from_mapping(
    template_schema: TemplateSchema,
    template_filename: Optional[str],
    column_mapping: Optional[Dict[str, str]],
    constant_values: Optional[Dict[str, str]]
) -> Optional[str]:
    constants = dict(constant_values or {})
    for key, value in constants.items():
        key_text = _normalize_text(key)
        if "申报类型" not in key_text and "declaretype" not in key_text:
            continue
        value_text = _normalize_text(value)
        has_add = any(k in value_text for k in ("增", "增员", "入职", "新增"))
        has_remove = any(k in value_text for k in ("减", "减员", "离职", "退工", "退保"))
        if has_add and not has_remove:
            return "add"
        if has_remove and not has_add:
            return "remove"
    mapping_keys = [k for k in (column_mapping or {}).keys() if isinstance(k, str)]
    for key in mapping_keys:
        normalized = _normalize_header(key)
        if normalized in ("terminationdate", "terminationreason"):
            return "remove"
    for key in mapping_keys:
        normalized = _normalize_header(key)
        if normalized == "startdate":
            return "add"
    return _detect_template_intent(template_filename, template_schema)


def _apply_record_type_filter(extracted_json: dict, template_intent: Optional[str]) -> dict:
    if template_intent not in ("add", "remove") or not isinstance(extracted_json, dict):
        return extracted_json
    all_records = _extract_records(extracted_json)
    has_target = any(
        (lambda t: t == template_intent)(
            _infer_record_type(r) or _infer_record_type_from_sheet(r)
        )
        for r in all_records
        if isinstance(r, dict)
    )
    if not has_target:
        return extracted_json
    def record_matches(record: Dict[str, Any]) -> bool:
        record_type = _infer_record_type(record)
        if record_type is None:
            record_type = _infer_record_type_from_sheet(record)
        return record_type == template_intent
    def filter_records_list(records: list) -> list:
        return [r for r in records if isinstance(r, dict) and record_matches(r)]
    def filter_extracted_data(extracted: Any) -> Any:
        if extracted is None:
            return None
        if isinstance(extracted, list):
            if len(extracted) > 0 and isinstance(extracted[0], dict):
                return filter_records_list(extracted)
            return extracted
        if isinstance(extracted, dict):
            filtered_extracted = {}
            for key, value in extracted.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    filtered_extracted[key] = filter_records_list(value)
                else:
                    filtered_extracted[key] = value
            return filtered_extracted
        return extracted
    filtered = {}
    if "sources" in extracted_json:
        filtered["sources"] = []
        for source in extracted_json["sources"]:
            if not isinstance(source, dict):
                filtered["sources"].append(source)
                continue
            filtered_source = dict(source)
            extracted = source.get("extracted")
            if extracted is not None:
                filtered_source["extracted"] = filter_extracted_data(extracted)
            filtered["sources"].append(filtered_source)
    if "merged" in extracted_json and isinstance(extracted_json["merged"], dict):
        filtered["merged"] = {}
        for key, value in extracted_json["merged"].items():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                filtered["merged"][key] = filter_records_list(value)
            else:
                filtered["merged"][key] = value
    for key in extracted_json:
        if key not in filtered:
            filtered[key] = extracted_json[key]
    return filtered


def _auto_filter_records_by_template_intent(extracted_json: dict, template_intent: Optional[str]) -> dict:
    """
    Filter records by inspecting source intent and record content.
    Keeps records with matching intent or unknown intent.
    """
    if not template_intent or not isinstance(extracted_json, dict):
        return extracted_json
    
    def _collect_record_intents(record: Dict[str, Any]) -> List[str]:
        intents = []
        source_intent = _infer_source_intent(record.get("__source_file__"))
        if source_intent:
            intents.append(source_intent)
        explicit_intent = _infer_explicit_record_intent(record)
        if explicit_intent:
            intents.append(explicit_intent)
        record_intent = _infer_record_intent(record)
        if record_intent:
            intents.append(record_intent)
        return intents
    
    all_records = _extract_records(extracted_json)
    has_target_signal = any(
        template_intent in _collect_record_intents(r)
        for r in all_records
        if isinstance(r, dict)
    )
    
    def should_include_record(record: Dict[str, Any], strict: bool = True) -> bool:
        record_type = _infer_record_type(record)
        if record_type is None:
            record_type = _infer_record_type_from_sheet(record)
        if record_type is not None:
            return record_type == template_intent
        source_intent = _infer_source_intent(record.get("__source_file__"))
        if source_intent and source_intent != template_intent:
            return False
        
        explicit_intent = _infer_explicit_record_intent(record)
        if explicit_intent and explicit_intent != template_intent:
            return False
        
        if strict:
            record_intent = _infer_record_intent(record)
            if record_intent is None:
                # If any target-intent signal exists, exclude unknowns to avoid mixing
                if has_target_signal:
                    return (source_intent == template_intent) or (explicit_intent == template_intent)
                return True
            return record_intent == template_intent
        
        return True
    
    def filter_records_list(records: list, strict: bool = True) -> list:
        return [r for r in records if isinstance(r, dict) and should_include_record(r, strict)]
    
    def filter_extracted_data(extracted: Any) -> Any:
        if extracted is None:
            return None
        if isinstance(extracted, list):
            if len(extracted) > 0 and isinstance(extracted[0], dict):
                original_count = len(extracted)
                filtered = filter_records_list(extracted, strict=True)
                if original_count > 0 and len(filtered) == 0:
                    filtered = filter_records_list(extracted, strict=False)
                if original_count != len(filtered):
                    logger.debug("_auto_filter_records_by_template_intent: Filtered list %d -> %d records", original_count, len(filtered))
                return filtered
            return extracted
        if isinstance(extracted, dict):
            filtered_extracted = {}
            for key, value in extracted.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    original_count = len(value)
                    filtered_value = filter_records_list(value, strict=True)
                    if original_count > 0 and len(filtered_value) == 0:
                        filtered_value = filter_records_list(value, strict=False)
                    filtered_extracted[key] = filtered_value
                    if original_count != len(filtered_value):
                        logger.debug("_auto_filter_records_by_template_intent: Filtered '%s' %d -> %d records", key, original_count, len(filtered_value))
                else:
                    filtered_extracted[key] = value
            return filtered_extracted
        return extracted
    
    filtered = {}
    total_original = 0
    total_filtered = 0
    
    if "sources" in extracted_json:
        filtered["sources"] = []
        for source_idx, source in enumerate(extracted_json["sources"]):
            if not isinstance(source, dict):
                filtered["sources"].append(source)
                continue
            filtered_source = dict(source)
            extracted = source.get("extracted")
            if extracted is not None:
                original_records = _count_records_in_extracted(extracted)
                total_original += original_records
                filtered_source["extracted"] = filter_extracted_data(extracted)
                filtered_records = _count_records_in_extracted(filtered_source["extracted"])
                total_filtered += filtered_records
            filtered["sources"].append(filtered_source)
    
    if "merged" in extracted_json and isinstance(extracted_json["merged"], dict):
        filtered["merged"] = {}
        for key, value in extracted_json["merged"].items():
            if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                total_original += len(value)
                filtered_value = filter_records_list(value)
                total_filtered += len(filtered_value)
                filtered["merged"][key] = filtered_value
            else:
                filtered["merged"][key] = value
    
    for key in extracted_json:
        if key not in filtered:
            filtered[key] = extracted_json[key]
    
    logger.info("_auto_filter_records_by_template_intent: intent='%s' total %d -> %d", template_intent, total_original, total_filtered)
    return filtered


def _auto_infer_common_fields(
    template_schema: TemplateSchema,
    extracted_json: dict,
    column_mapping: Optional[Dict[str, str]],
    constant_values: Optional[Dict[str, str]],
    template_intent: Optional[str],
    template_sample_rows: List[Dict[str, Any]]
) -> Tuple[Dict[str, str], Dict[str, str]]:
    column_mapping = dict(column_mapping or {})
    constant_values = dict(constant_values or {})
    
    records = _extract_records(extracted_json)
    extracted_keys = _collect_extracted_keys(records)
    
    headers: List[str] = []
    for sheet_schema in template_schema.sheet_schemas:
        for region in sheet_schema.regions:
            if region.table and region.table.header:
                for header in region.table.header:
                    if header.header_path:
                        headers.append(header.header_path)
    
    def find_key_by_candidates(candidates: List[str]) -> Optional[str]:
        if not extracted_keys:
            return None
        for cand in candidates:
            cand_norm = _normalize_header(cand)
            if not cand_norm:
                continue
            for key in extracted_keys:
                key_norm = _normalize_header(key)
                if cand_norm == key_norm or cand_norm in key_norm or key_norm in cand_norm:
                    return key
        return None
    
    def get_template_sample_value(header: str) -> Optional[str]:
        for row in template_sample_rows:
            if header in row and row[header]:
                return str(row[header])
        return None
    
    def detect_year_month_format(sample_value: Optional[str]) -> str:
        if not sample_value:
            return "YYYY-MM"
        v = str(sample_value)
        if "年" in v and "月" in v:
            return "YYYY年MM月"
        if "-" in v:
            return "YYYY-MM"
        if "/" in v:
            return "YYYY/MM"
        if len(v) == 6 and v.isdigit():
            return "YYYYMM"
        return "YYYY-MM"
    
    def extract_year_month_from_text(text: Any) -> Optional[Tuple[int, int]]:
        if text is None:
            return None
        v = str(text)
        match = re.search(r"(20\\d{2})[年\\-\\./]?(\\d{1,2})", v)
        if not match:
            return None
        year = int(match.group(1))
        month = int(match.group(2))
        if month < 1 or month > 12:
            return None
        return year, month
    
    def format_year_month(year: int, month: int, fmt: str) -> str:
        if fmt == "YYYY年MM月":
            return f"{year}年{month:02d}月"
        if fmt == "YYYY/MM":
            return f"{year}/{month:02d}"
        if fmt == "YYYYMM":
            return f"{year}{month:02d}"
        return f"{year}-{month:02d}"
    
    def infer_year_month_from_records() -> Optional[Tuple[int, int]]:
        if not records:
            return None
        primary_keys = ["费用年月", "缴费年月", "参保月份", "申报年月", "费用月份", "缴费月份", "参保年月", "申报月份"]
        date_keys = ["离职日期", "入职日期", "变动日期", "申报日期", "生效日期", "开始日期", "结束日期", "办理日期"]
        for record in records:
            for key in primary_keys + date_keys:
                if key in record and record[key]:
                    ym = extract_year_month_from_text(record[key])
                    if ym:
                        return ym
            for key, value in record.items():
                if isinstance(key, str) and ("日期" in key or "时间" in key):
                    ym = extract_year_month_from_text(value)
                    if ym:
                        return ym
        return None
    
    def infer_year_month_from_filenames() -> Optional[Tuple[int, int]]:
        sources = extracted_json.get("sources", []) if isinstance(extracted_json, dict) else []
        for source in sources:
            if not isinstance(source, dict):
                continue
            filename = source.get("filename")
            ym = extract_year_month_from_text(filename)
            if ym:
                return ym
        return None
    
    for header in headers:
        if header in column_mapping or header in constant_values:
            continue
        
        header_norm = _normalize_header(header)
        
        # 费用年月 / 缴费年月等
        if any(k in header_norm for k in ["费用年月", "缴费年月", "参保月份", "申报年月", "费用月份", "缴费月份"]):
            key = find_key_by_candidates(["费用年月", "缴费年月", "参保月份", "申报年月", "费用月份", "缴费月份"])
            if key:
                column_mapping[header] = key
                continue
            
            ym = infer_year_month_from_records() or infer_year_month_from_filenames()
            if ym:
                sample_val = get_template_sample_value(header)
                fmt = detect_year_month_format(sample_val)
                constant_values[header] = format_year_month(ym[0], ym[1], fmt)
                continue
        
        # 社保/申报类型
        if any(k in header_norm for k in ["社保类型", "申报类型", "社保/申报类型"]):
            key = find_key_by_candidates(["社保类型", "申报类型", "社保/申报类型", "变动类型", "办理类型", "业务类型", "操作类型"])
            if key:
                column_mapping[header] = key
                continue
            if template_intent == "add":
                constant_values[header] = "增员"
            elif template_intent == "remove":
                constant_values[header] = "减员"
    
    if constant_values:
        logger.debug("_auto_infer_common_fields: constants=%s", list(constant_values.keys()))
    if column_mapping:
        logger.debug("_auto_infer_common_fields: mappings=%s", list(column_mapping.keys()))
    
    return column_mapping, constant_values


def _normalize_text(text: Any) -> str:
    return str(text).lower().replace(" ", "").replace("_", "").replace("-", "").strip()


def _detect_template_intent(template_filename: Optional[str], template_schema: TemplateSchema) -> Optional[str]:
    """
    Detect whether template is for add (增员/入职) or remove (减员/离职) records.
    Returns "add", "remove", or None if ambiguous.
    """
    add_keywords = ("增员", "入职", "新增", "新入职", "新增人员", "加员", "扩员")
    remove_keywords = ("减员", "离职", "退场", "辞退", "解除", "退工", "停保", "退保")
    
    filename_text = _normalize_text(template_filename or "")
    if filename_text:
        has_add = any(k in filename_text for k in add_keywords)
        has_remove = any(k in filename_text for k in remove_keywords)
        if has_add and not has_remove:
            return "add"
        if has_remove and not has_add:
            return "remove"
    
    sheet_names = [ss.sheet for ss in template_schema.sheet_schemas]
    header_paths = []
    for sheet_schema in template_schema.sheet_schemas:
        for region in sheet_schema.regions:
            if region.table and region.table.header:
                for header in region.table.header:
                    if header.header_path:
                        header_paths.append(header.header_path)
    
    schema_text = _normalize_text(" ".join(sheet_names + header_paths))
    has_add = any(k in schema_text for k in add_keywords)
    has_remove = any(k in schema_text for k in remove_keywords)
    if has_add and not has_remove:
        return "add"
    if has_remove and not has_add:
        return "remove"
    
    return None


def _infer_record_intent(record: Dict[str, Any]) -> Optional[str]:
    add_keywords = ("增员", "入职", "新增", "新入职", "新增人员", "加员", "扩员")
    remove_keywords = ("减员", "离职", "退场", "辞退", "解除", "退工", "停保", "退保")
    
    add_hits = 0
    remove_hits = 0
    
    for key, value in record.items():
        key_text = _normalize_text(key)
        if any(k in key_text for k in add_keywords):
            add_hits += 1
        if any(k in key_text for k in remove_keywords):
            remove_hits += 1
        
        if isinstance(value, str):
            val_text = _normalize_text(value)
            if any(k in val_text for k in add_keywords):
                add_hits += 1
            if any(k in val_text for k in remove_keywords):
                remove_hits += 1
    
    if add_hits > 0 and remove_hits == 0:
        return "add"
    if remove_hits > 0 and add_hits == 0:
        return "remove"
    return None


def build_fallback_fill_plan(template_schema: TemplateSchema, extracted_json: dict) -> Optional[dict]:
    if not template_schema.sheet_schemas:
        return None
    sheet = template_schema.sheet_schemas[0]
    if not sheet.regions or not sheet.regions[0].table:
        return None
    region = sheet.regions[0]
    table = region.table
    headers = table.header or []
    if not headers:
        return None
    
    try:
        min_col, min_row, max_col, max_row = range_boundaries(table.range)
    except Exception:
        return None
    
    max_header_row = max(region.header_rows) if region.header_rows else min_row
    data_start_row = max_header_row + 1
    start_cell = f"{get_column_letter(min_col)}{data_start_row}"
    
    records: List[Dict[str, Any]] = []
    
    logger.debug("build_fallback_fill_plan: Searching for records in %d sources", len(extracted_json.get('sources', [])))
    
    # 首先检查sources中的extracted数据
    for s_idx, s in enumerate(extracted_json.get("sources", [])):
        if not isinstance(s, dict):
            continue
        ex = s.get("extracted")
        filename = s.get("filename", f"source_{s_idx}")
        logger.debug("build_fallback_fill_plan: Checking source %d: %s, extracted type: %s", s_idx, filename, type(ex).__name__)
        
        # 如果extracted本身就是列表（数组格式）
        if isinstance(ex, list) and len(ex) > 0:
            if isinstance(ex[0], dict):
                records.extend(ex)
                logger.debug("build_fallback_fill_plan: Found %d records in sources[%d].extracted (list)", len(ex), s_idx)
                break
            elif isinstance(ex[0], list):
                # 嵌套列表的情况
                for item in ex:
                    if isinstance(item, dict):
                        records.append(item)
                if records:
                    logger.debug("build_fallback_fill_plan: Found %d records in sources[%d].extracted (nested list)", len(records), s_idx)
                    break
        
        # 如果extracted是字典，查找其中的列表字段
        if isinstance(ex, dict):
            logger.debug("build_fallback_fill_plan: sources[%d].extracted is dict, keys: %s", s_idx, list(ex.keys())[:10])
            # 优先查找常见的字段名
            priority_keys = ["data", "records", "rows", "items", "extracted_data", "table_data"]
            found = False
            for key in priority_keys:
                if key in ex:
                    value = ex[key]
                    logger.debug("build_fallback_fill_plan: Found key '%s', type: %s, length: %s", key, type(value).__name__, len(value) if isinstance(value, (list, dict)) else 'N/A')
                    if isinstance(value, list) and len(value) > 0:
                        if isinstance(value[0], dict):
                            records.extend(value)
                            logger.debug("build_fallback_fill_plan: Found %d records in sources[%d].extracted['%s']", len(value), s_idx, key)
                            found = True
                            break
            
            # 如果没找到，遍历所有字段
            if not found:
                for k, v in ex.items():
                    if isinstance(v, list) and len(v) > 0:
                        if isinstance(v[0], dict):
                            records.extend(v)
                            logger.debug("build_fallback_fill_plan: Found %d records in sources[%d].extracted['%s']", len(v), s_idx, k)
                            found = True
                            break
        
        if records:
            break
    
    # 如果sources中没找到，检查merged数据
    if not records:
        merged = extracted_json.get("merged", {})
        if isinstance(merged, dict):
            # 优先查找常见的字段名
            priority_keys = ["data", "records", "rows", "items", "extracted_data", "table_data"]
            found = False
            for key in priority_keys:
                if key in merged and isinstance(merged[key], list) and len(merged[key]) > 0:
                    if isinstance(merged[key][0], dict):
                        records = list(merged[key])
                        logger.debug("build_fallback_fill_plan: Found %d records in merged['%s']", len(records), key)
                        found = True
                        break
            
            # 如果没找到，遍历所有字段
            if not found:
                for k, v in merged.items():
                    if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                        records = list(v)
                        logger.debug("build_fallback_fill_plan: Found %d records in merged['%s']", len(records), k)
                        break
    
    if not records:
        return None
    
    def _best_key(h: str, keys: List[str]) -> Optional[str]:
        h0 = h.split("/")[-1].strip() if "/" in h else h
        h_clean = h0.replace(" ", "").replace("_", "").replace("-", "")
        
        # 同义词映射
        synonyms = {
            "身份证号码": ["证件号码", "身份证号", "身份证", "证件号"],
            "证件号码": ["身份证号码", "身份证号", "身份证", "证件号"],
            "姓名": ["名字", "名称"],
            "名字": ["姓名", "名称"],
            "联系方式": ["电话", "手机", "联系电话", "手机号"],
            "电话": ["联系方式", "手机", "联系电话", "手机号"],
            "入职日期": ["入职时间", "入职"],
            "离职退场时间": ["离职时间", "退场时间", "离职日期", "退场日期"],
        }
        
        # 首先尝试精确匹配和包含匹配
        for k in keys:
            k_clean = k.replace(" ", "").replace("_", "").replace("-", "")
            if k == h or k == h0 or h in k or k in h or h0 in k or k in h0 or h_clean == k_clean:
                return k
        
        # 然后尝试同义词匹配
        for k in keys:
            k_clean = k.replace(" ", "").replace("_", "").replace("-", "")
            # 检查h0是否在k的同义词列表中
            if h0 in synonyms.get(k, []):
                return k
            # 检查k是否在h0的同义词列表中
            if k in synonyms.get(h0, []):
                return k
            # 检查同义词的清理版本
            for syn_list in synonyms.values():
                if h_clean in [s.replace(" ", "").replace("_", "").replace("-", "") for s in syn_list]:
                    if k_clean in [s.replace(" ", "").replace("_", "").replace("-", "") for s in syn_list]:
                        return k
        
        return None
    
    rk = list(records[0].keys()) if records else []
    logger.debug("build_fallback_fill_plan: Record keys: %s", rk[:10])
    logger.debug("build_fallback_fill_plan: Template headers: %s", [th.header_path for th in headers[:10]])
    
    column_mapping: Dict[str, str] = {}
    for th in headers:
        if not th.header_path:
            continue
        match = _best_key(th.header_path, rk)
        if match and match not in column_mapping:
            column_mapping[match] = th.col_letter
            logger.debug("build_fallback_fill_plan: Mapped '%s' -> '%s' -> column %s", th.header_path, match, th.col_letter)
    
    if not column_mapping:
        logger.debug("build_fallback_fill_plan: No matches found, using positional mapping")
        for i, th in enumerate(headers):
            if th.header_path and i < len(rk):
                column_mapping[rk[i]] = th.col_letter
                logger.debug("build_fallback_fill_plan: Positional mapping: '%s' -> column %s", rk[i], th.col_letter)
    
    logger.debug("build_fallback_fill_plan: Final column_mapping: %d mappings", len(column_mapping))
    
    rows = []
    for rec_idx, rec in enumerate(records):
        row_dict = {}
        for ek, col in column_mapping.items():
            if ek in rec and rec[ek] is not None:
                row_dict[ek] = rec[ek]
        if row_dict:
            rows.append(row_dict)
        if rec_idx < 3:  # 打印前3条记录的信息
            logger.debug("build_fallback_fill_plan: record %d: %d fields mapped", rec_idx, len(row_dict))
    
    logger.debug("build_fallback_fill_plan: Total %d records, %d rows prepared", len(records), len(rows))
    
    if not rows:
        logger.debug("build_fallback_fill_plan: No rows prepared, returning None")
        return None
    
    clear_end_row = min(max_row, data_start_row + len(rows) + 10)
    clear_range = f"{get_column_letter(min_col)}{data_start_row}:{get_column_letter(max_col)}{clear_end_row}"
    
    sheet_name = _resolve_sheet_name(template_schema, sheet)
    fallback = {
        "target": {
            "sheet_name": sheet_name,
            "sheet": sheet_name,
            "region_id": region.region_id,
            "layout_type": "table",
            "clear_policy": "clear_values_keep_format"
        },
        "clear_ranges": [clear_range],
        "row_writes": [{
            "start_cell": start_cell,
            "rows": rows,
            "column_mapping": column_mapping
        }],
        "writes": [],
        "warnings": [f"Fallback fill plan used; LLM returned insufficient rows. Filling {len(rows)} rows."]
    }
    logger.info("build_fallback_fill_plan: Returning fallback with %d rows, start_cell=%s", len(rows), start_cell)
    return fallback

def _extract_template_context(template_schema: TemplateSchema, template_filename: Optional[str], llm: LLMClient) -> Optional[str]:
    sheet_names = [ss.sheet for ss in template_schema.sheet_schemas]
    header_paths = []
    for sheet_schema in template_schema.sheet_schemas:
        for region in sheet_schema.regions:
            if region.table and region.table.header:
                for header in region.table.header:
                    if header.header_path:
                        header_paths.append(header.header_path)
    
    context_info = {
        "template_filename": template_filename or "unknown",
        "sheet_names": sheet_names,
        "sample_headers": list(set(header_paths))[:15]
    }
    
    context_prompt = f"""Analyze the following template information and determine its business purpose:

Template filename: {context_info['template_filename']}
Sheet names: {', '.join(context_info['sheet_names'])}
Sample column headers: {', '.join(context_info['sample_headers'])}

Provide a brief analysis (2-3 sentences) about:
1. What business operation this template is designed for (e.g., 社保增员, 社保减员, 员工入职, etc.)
2. What type of data records should be filled (use broad semantic equivalence)
3. Synonyms: 减员≈离职≈退场≈辞退≈解除; 增员≈入职≈新入职≈新增. Only exclude clearly opposite data."""
    
    try:
        context_analysis = llm.chat_text(context_prompt, system=None, temperature=0, step="template_context_analysis")
        if context_analysis and len(context_analysis.strip()) > 0:
            return f"""Template Business Context:
{context_analysis.strip()}

SEMANTIC MATCHING: 减员/离职/退场/辞退/解除 etc. are equivalent — include all. Only exclude data that clearly opposes the template (e.g. 增员 when template is 减员). Prefer including records over filtering them out."""
    except Exception as e:
        logger.warning("Failed to generate template context: %s", e)
    
    context_note = f"Template: {context_info['template_filename']}\n"
    context_note += f"Sheets: {', '.join(context_info['sheet_names'])}\n"
    context_note += f"Headers: {', '.join(context_info['sample_headers'][:10])}\n"
    context_note += "\n减员≈离职≈退场≈辞退≈解除 (include all). Only exclude 增员 when template is 减员. Prefer inclusion."
    
    return context_note

def _extract_filename_context(extracted_json: dict, llm: LLMClient, is_template: bool = False) -> Optional[str]:
    if not isinstance(extracted_json, dict):
        return None
    
    sources = extracted_json.get("sources", [])
    if not sources:
        return None
    
    filenames_info = []
    for source in sources:
        if isinstance(source, dict):
            filename = source.get("filename", "")
            source_type = source.get("source_type", "")
            if filename:
                extracted_keys = list(source.get("extracted", {}).keys()) if isinstance(source.get("extracted"), dict) else []
                filenames_info.append({
                    "filename": filename,
                    "type": source_type,
                    "extracted_keys": extracted_keys[:20]
                })
    
    if not filenames_info:
        return None
    
    context_prompt = f"""Analyze the following source filenames and provide brief context hints for data mapping:

Filenames and their extracted data keys:
{chr(10).join([f"- {info['filename']} ({info['type']}): {', '.join(info['extracted_keys']) if info['extracted_keys'] else 'no keys'}" for info in filenames_info])}

Provide a brief analysis (2-3 sentences) about:
1. What business context these filenames suggest
2. Which extracted fields are likely most relevant for mapping to template columns
3. Note: 离职/退场/辞退/解除 ≈ 减员 (same semantic); 入职/新增 ≈ 增员. Use these when matching to template purpose."""
    
    try:
        context_analysis = llm.chat_text(context_prompt, system=None, temperature=0, step="template_context_analysis")
        if context_analysis and len(context_analysis.strip()) > 0:
            return f"Source Context:\n{context_analysis.strip()}\n\n离职/退场/辞退/解除≈减员; 入职/新增≈增员. Map by semantic match."
    except Exception as e:
        logger.warning("Failed to generate filename context: %s", e)
    
    context_note = "Source filenames and their extracted data keys:\n"
    for info in filenames_info:
        context_note += f"- {info['filename']} ({info['type']}): {', '.join(info['extracted_keys'][:10])}"
        if len(info['extracted_keys']) > 10:
            context_note += f" ... ({len(info['extracted_keys'])} total keys)"
        context_note += "\n"
    
    context_note += "\n离职/退场/辞退/解除≈减员; 入职/新增≈增员. Map by semantic match."
    
    return context_note

def _optimize_extracted_for_llm(extracted_json: dict) -> dict:
    """
    优化extracted_json格式，保留所有数据但优化结构
    不减少数据量，只是优化格式以便LLM更好地处理
    """
    # 目前不减少数据，保留所有记录给LLM处理
    # 如果未来需要优化，可以在这里添加逻辑
    return extracted_json
