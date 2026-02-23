"""
模板解析模块 (Template Parser Module)
====================================

解析 Excel 模板文件，提取表结构、表头、约束等，生成 TemplateSchema。

Refactored into an OO design following SOLID principles:
  - ``ParserConfig``          – externalises every tunable constant.
  - ``ExcelTemplateParser``   – single-responsibility parser class.
  - ``parse_template_xlsx()`` – backward-compatible module-level wrapper.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from core.logger import get_logger
from core.template.schema import (
    Constraints,
    KeyValuePair,
    RegionSchema,
    SheetSchema,
    TableHeader,
    TableInfo,
    TemplateSchema,
)

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ParserConfig:
    """
    Immutable configuration that externalises every "magic number" used
    during template parsing.

    Attributes:
        max_header_scan_rows: Maximum number of rows to scan when
            searching for a header row in a used-range region.
        max_data_preview_rows: Maximum number of data rows (after the
            header) to consider when analysing a used-range region.
        num_sample_rows: Number of example / few-shot rows to extract
            beneath each detected header.
        min_header_cells: Minimum number of non-empty cells a row must
            have to be considered a valid header row.
        min_header_cells_early_stop: When a candidate row already has
            this many non-empty cells *and* data exists below it, stop
            the header search early.
    """

    max_header_scan_rows: int = 100
    max_data_preview_rows: int = 500
    num_sample_rows: int = 3
    min_header_cells: int = 2
    min_header_cells_early_stop: int = 3


# ---------------------------------------------------------------------------
# Parser class
# ---------------------------------------------------------------------------

class ExcelTemplateParser:
    """Object-oriented Excel template parser.

    Usage::

        parser = ExcelTemplateParser()          # default config
        schema = parser.parse("template.xlsx")

        # …or with custom tuning:
        cfg = ParserConfig(num_sample_rows=5)
        schema = ExcelTemplateParser(cfg).parse("template.xlsx")
    """

    def __init__(self, config: Optional[ParserConfig] = None) -> None:
        self._config: ParserConfig = config or ParserConfig()

    # -- public entry point -------------------------------------------------

    def parse(self, path: str) -> TemplateSchema:
        """Parse an Excel template file and return a :class:`TemplateSchema`.

        The parser first attempts to use explicitly defined tables
        (``ws.tables``).  If a sheet has none, it falls back to
        detecting a region from the used range.
        """
        wb = load_workbook(path, data_only=False)
        sheet_schemas: List[SheetSchema] = []

        try:
            for sheet_name in wb.sheetnames:
                ws: Worksheet = wb[sheet_name]
                regions = self._parse_sheet(ws, sheet_name)
                if regions:
                    sheet_schemas.append(
                        SheetSchema(sheet=sheet_name, regions=regions)
                    )
        finally:
            wb.close()

        return TemplateSchema(sheet_schemas=sheet_schemas)

    # -- sheet-level --------------------------------------------------------

    def _parse_sheet(
        self, ws: Worksheet, sheet_name: str
    ) -> List[RegionSchema]:
        """Return all detected regions for a single worksheet."""
        regions: List[RegionSchema] = []

        # 1. Try explicitly-defined tables first.
        if ws.tables:
            for table_name, table in ws.tables.items():
                region = self._parse_table_region(
                    ws, table.ref, f"table_{table_name}", sheet_name
                )
                if region is not None:
                    regions.append(region)

        # 2. Fall back to used-range detection if no table regions found.
        if not regions and ws.max_row > 0 and ws.max_column > 0:
            region = self._detect_region_from_used_range(ws, sheet_name)
            if region is not None:
                regions.append(region)

        return regions

    # -- region parsers -----------------------------------------------------

    def _parse_table_region(
        self,
        ws: Worksheet,
        ref: str,
        region_id: str,
        sheet_name: str,
    ) -> Optional[RegionSchema]:
        """Parse a named-table region given its cell reference string."""
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)

            header_rows = self._detect_header_rows_in_range(
                ws, min_row, max_row, min_col, max_col
            )

            table_info, constraints = self._build_table_and_constraints(
                ws, min_row, max_row, min_col, max_col, header_rows
            )

            return RegionSchema(
                region_id=region_id,
                layout_type="table",
                header_rows=header_rows,
                table=table_info,
                constraints=constraints,
            )
        except Exception:
            logger.warning(
                "Sheet '%s': failed to parse table region '%s' (ref=%s). "
                "Skipping this region.",
                sheet_name,
                region_id,
                ref,
                exc_info=True,
            )
            return None

    def _detect_region_from_used_range(
        self, ws: Worksheet, sheet_name: str
    ) -> Optional[RegionSchema]:
        """Detect a table region from the worksheet's used range."""
        try:
            min_row = 1
            min_col = 1
            max_row: int = ws.max_row
            max_col: int = ws.max_column

            header_row = self._find_header_row(
                ws, min_row, max_row, min_col, max_col
            )
            if header_row is None:
                logger.info(
                    "Sheet '%s': no header row detected in used range. "
                    "Skipping.",
                    sheet_name,
                )
                return None

            header_rows = [header_row]
            if header_row + 1 <= max_row:
                second_row_has_content = any(
                    ws.cell(header_row + 1, c).value
                    for c in range(min_col, max_col + 1)
                )
                if second_row_has_content and self._is_header_row(
                    ws, header_row + 1, min_col, max_col
                ):
                    header_rows.append(header_row + 1)

            max_data_row = min(
                max_row, header_row + self._config.max_data_preview_rows
            )

            table_info, constraints = self._build_table_and_constraints(
                ws, header_row, max_data_row, min_col, max_col, header_rows
            )

            return RegionSchema(
                region_id=f"region_{header_row}",
                layout_type="table",
                header_rows=header_rows,
                table=table_info,
                constraints=constraints,
            )
        except Exception:
            logger.warning(
                "Sheet '%s': failed to detect region from used range. "
                "Skipping.",
                sheet_name,
                exc_info=True,
            )
            return None

    # -- header detection ---------------------------------------------------

    def _detect_header_rows_in_range(
        self,
        ws: Worksheet,
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> List[int]:
        """Return a list of header row numbers for a known table range.

        At minimum the first row is assumed to be a header.  A second
        row is added when it also contains data.
        """
        header_rows: List[int] = [min_row]
        if min_row + 1 <= max_row:
            second_row_has_content = any(
                ws.cell(min_row + 1, c).value
                for c in range(min_col, max_col + 1)
            )
            if second_row_has_content:
                header_rows.append(min_row + 1)
        return header_rows

    def _find_header_row(
        self,
        ws: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> Optional[int]:
        """Scan the first *max_header_scan_rows* rows and return the most
        likely header row number, or ``None`` if nothing qualifies.
        """
        best_row: Optional[int] = None
        best_count: int = 0
        scan_limit = min(
            end_row + 1, start_row + self._config.max_header_scan_rows
        )

        for row in range(start_row, scan_limit):
            non_empty_count = sum(
                1
                for col in range(start_col, end_col + 1)
                if ws.cell(row, col).value
            )
            if non_empty_count > best_count:
                best_count = non_empty_count
                best_row = row

            # Early stop: enough header cells and data exists below.
            if (
                best_count >= self._config.min_header_cells_early_stop
                and row + 1 <= end_row
            ):
                has_data_below = any(
                    ws.cell(row + 1, col).value
                    for col in range(start_col, end_col + 1)
                )
                if has_data_below:
                    return best_row

        return best_row if best_count >= self._config.min_header_cells else None

    @staticmethod
    def _is_header_row(
        ws: Worksheet, row: int, start_col: int, end_col: int
    ) -> bool:
        """Return ``True`` if *row* has at least 2 non-empty cells."""
        non_empty = sum(
            1
            for col in range(start_col, end_col + 1)
            if ws.cell(row, col).value
        )
        return non_empty >= 2

    # -- composite builder (replaces old "god function") --------------------

    def _build_table_and_constraints(
        self,
        ws: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        header_rows: List[int],
    ) -> Tuple[TableInfo, Constraints]:
        """Orchestrate header, sample-row, and constraint extraction.

        This replaces the former monolithic ``_extract_table_info``.
        """
        data_start_row = max(header_rows) + 1

        headers, header_paths = self._extract_headers(
            ws, start_col, end_col, data_start_row, end_row, header_rows
        )

        sample_rows = self._extract_sample_rows(
            ws, data_start_row, end_row, start_col, end_col, header_paths
        )

        constraints = self._detect_constraints(
            ws, start_row, end_row, start_col, end_col, data_start_row
        )

        range_str = (
            f"{get_column_letter(start_col)}{start_row}"
            f":{get_column_letter(end_col)}{end_row}"
        )

        table_info = TableInfo(
            range=range_str,
            header=headers,
            sample_rows=sample_rows if sample_rows else None,
        )

        return table_info, constraints

    # -- decomposed helpers -------------------------------------------------

    def _extract_headers(
        self,
        ws: Worksheet,
        start_col: int,
        end_col: int,
        data_start_row: int,
        end_row: int,
        header_rows: List[int],
    ) -> Tuple[List[TableHeader], List[str]]:
        """Build :class:`TableHeader` objects and a parallel list of header
        path strings (used later to key sample-row dicts).
        """
        headers: List[TableHeader] = []
        header_paths: List[str] = []

        for col_idx in range(start_col, end_col + 1):
            col_letter = get_column_letter(col_idx)
            header_path = self._build_header_path(ws, header_rows, col_idx)
            header_paths.append(header_path)

            # Collect per-column sample values
            sample_values: List[str] = []
            sample_end = min(
                data_start_row + self._config.num_sample_rows, end_row + 1
            )
            for sample_row in range(data_start_row, sample_end):
                cell: Cell = ws.cell(sample_row, col_idx)
                if cell.value is not None:
                    sample_values.append(str(cell.value))

            headers.append(
                TableHeader(
                    col_letter=col_letter,
                    col_index=col_idx,
                    header_path=header_path,
                    sample_values=sample_values,
                )
            )

        return headers, header_paths

    def _extract_sample_rows(
        self,
        ws: Worksheet,
        data_start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        header_paths: List[str],
    ) -> List[Dict[str, str]]:
        """Extract up to *num_sample_rows* complete data rows as
        ``{header_path: value}`` dicts (few-shot examples).

        Formula cells are intentionally skipped.
        """
        sample_rows: List[Dict[str, str]] = []
        sample_end = min(
            data_start_row + self._config.num_sample_rows, end_row + 1
        )

        for row_idx in range(data_start_row, sample_end):
            row_data: Dict[str, str] = {}
            has_data = False

            for col_offset, col_idx in enumerate(
                range(start_col, end_col + 1)
            ):
                cell: Cell = ws.cell(row_idx, col_idx)
                header_path = (
                    header_paths[col_offset]
                    if col_offset < len(header_paths)
                    else f"Col_{col_idx}"
                )

                if not header_path or cell.value is None:
                    continue

                # Skip formula cells
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    continue

                row_data[header_path] = str(cell.value)
                has_data = True

            if has_data:
                sample_rows.append(row_data)

        return sample_rows

    def _detect_constraints(
        self,
        ws: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        data_start_row: int,
    ) -> Constraints:
        """Scan the region for formula cells and non-default number
        formats, returning a :class:`Constraints` instance.
        """
        formula_cells: List[str] = []
        number_formats: Dict[str, str] = {}

        for col_idx in range(start_col, end_col + 1):
            col_letter = get_column_letter(col_idx)

            # Number format – check the first data row
            if data_start_row <= end_row:
                sample_cell: Cell = ws.cell(data_start_row, col_idx)
                if (
                    sample_cell.number_format
                    and sample_cell.number_format != "General"
                ):
                    number_formats[col_letter] = sample_cell.number_format

        # Formula cells – scan entire region
        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell: Cell = ws.cell(row, col_idx)
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    formula_cells.append(
                        f"{get_column_letter(col_idx)}{row}"
                    )

        return Constraints(
            has_formulas=len(formula_cells) > 0,
            formula_cells=formula_cells,
            validations=[],
            number_formats=number_formats,
        )

    # -- low-level utility --------------------------------------------------

    @staticmethod
    def _build_header_path(
        ws: Worksheet, header_rows: List[int], col_idx: int
    ) -> str:
        """Build a ``/``-separated header path for multi-row headers."""
        if len(header_rows) == 1:
            cell: Cell = ws.cell(header_rows[0], col_idx)
            return str(cell.value) if cell.value else ""

        parts: List[str] = []
        prev_value = ""
        for row in header_rows:
            cell = ws.cell(row, col_idx)
            value = str(cell.value) if cell.value else prev_value
            if value:
                parts.append(value)
                prev_value = value

        return "/".join(parts) if parts else ""


# ---------------------------------------------------------------------------
# Backward-compatible module-level wrapper
# ---------------------------------------------------------------------------

def parse_template_xlsx(
    path: str,
    config: Optional[ParserConfig] = None,
) -> TemplateSchema:
    """Parse an Excel template file and return a :class:`TemplateSchema`.

    This is a thin wrapper around :class:`ExcelTemplateParser` kept for
    backward compatibility.  Existing call-sites that do
    ``from core.template.parser import parse_template_xlsx`` will
    continue to work without modification.
    """
    return ExcelTemplateParser(config).parse(path)
