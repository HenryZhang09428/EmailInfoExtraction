"""
社保模板专用处理模块 (Social Security Template Profile)
=====================================================

针对社保增员/减员模板的专用处理逻辑。

检测包含 姓名、证件号码、申报类型、费用年月 等表头的模板，
构建仅填充特定列的约束性填充计划。
"""
import json
import re
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple
from dataclasses import dataclass, field
from openpyxl.utils import get_column_letter, range_boundaries

from core.template.schema import TemplateSchema
from core.ir import FillPlan, FillPlanTarget
from core.llm import LLMClient
from core.logger import get_logger
from core.prompts_loader import get_prompts

logger = get_logger(__name__)


@dataclass
class SocialSecurityProfile:
    is_detected: bool = False
    template_intent: Optional[str] = None
    name_columns: List[Tuple[str, str]] = field(default_factory=list)
    id_columns: List[Tuple[str, str]] = field(default_factory=list)
    declare_type_columns: List[Tuple[str, str]] = field(default_factory=list)
    fee_month_columns: List[Tuple[str, str]] = field(default_factory=list)
    reason_columns: List[Tuple[str, str]] = field(default_factory=list)
    has_social_security_section: bool = False
    has_housing_fund_section: bool = False
    fee_month_format: Optional[str] = None


@dataclass
class SourceScore:
    source_id: str
    filename: str
    source_type: str
    intent_score: float = 0.0
    name_coverage: float = 0.0
    id_coverage: float = 0.0
    date_coverage: float = 0.0
    record_count: int = 0
    has_exclusion_signal: bool = False
    total_score: float = 0.0
    score_breakdown: Dict[str, Any] = field(default_factory=dict)


def _normalize_header_text(text: str) -> str:
    if not text:
        return ""
    normalized = text.lower()
    normalized = normalized.replace("\u3000", "")
    normalized = normalized.replace(" ", "")
    normalized = normalized.replace("/", "")
    normalized = normalized.replace("\\", "")
    return normalized


def _get_header_texts(header: Any) -> List[str]:
    texts: List[str] = []
    for attr in ("header_path", "header_text", "header", "text", "name"):
        value = getattr(header, attr, None)
        if isinstance(value, str) and value.strip():
            texts.append(value)
    return texts


def _header_contains_any(header: Any, keywords: List[str]) -> bool:
    texts = _get_header_texts(header)
    for text in texts:
        normalized = _normalize_header_text(text)
        for kw in keywords:
            if kw in normalized:
                return True
    return False


def _header_ends_with(header: Any, suffix: str) -> bool:
    texts = _get_header_texts(header)
    suffix_norm = _normalize_header_text(suffix)
    for text in texts:
        normalized = _normalize_header_text(text)
        if normalized.endswith(suffix_norm):
            return True
    return False


def _get_column_info(header: Any) -> Tuple[str, str]:
    col_letter = getattr(header, "col_letter", "")
    header_path = getattr(header, "header_path", "")
    if not isinstance(header_path, str):
        header_path = ""
    return (col_letter, header_path)


def _detect_section_type(header_path: str) -> str:
    normalized = _normalize_header_text(header_path)
    if "公积金" in normalized or "公积" in normalized:
        return "housing_fund"
    if "社保" in normalized or "社会保险" in normalized:
        return "social_security"
    return "unknown"


def detect_social_security_template(template_schema: TemplateSchema) -> SocialSecurityProfile:
    profile = SocialSecurityProfile()
    
    if not template_schema.sheet_schemas:
        return profile
    
    sheet = template_schema.sheet_schemas[0]
    if not sheet.regions:
        return profile
    
    region = None
    for r in sheet.regions:
        if r.table and r.table.header:
            region = r
            break
    
    if not region or not region.table or not region.table.header:
        return profile
    
    headers = region.table.header
    
    name_keywords = ["姓名", "参保人姓名", "员工姓名", "被保险人"]
    id_keywords = ["证件号码", "身份证号", "身份证", "证件号"]
    declare_type_keywords = ["申报类型", "申报类别", "办理类型"]
    fee_month_keywords = ["费用年月", "费款年月", "缴费年月", "参保年月"]
    reason_keywords = ["变动原因", "减员原因", "离职原因", "退工原因", "退保原因", "原因", "reason"]
    
    has_name = False
    has_id = False
    has_declare_type = False
    has_fee_month = False
    
    for header in headers:
        col_letter, header_path = _get_column_info(header)
        if not col_letter:
            continue
        
        if _header_ends_with(header, "姓名") or _header_contains_any(header, name_keywords):
            profile.name_columns.append((col_letter, header_path))
            has_name = True
        
        if _header_ends_with(header, "证件号码") or _header_contains_any(header, id_keywords):
            profile.id_columns.append((col_letter, header_path))
            has_id = True
        
        if _header_ends_with(header, "申报类型") or _header_contains_any(header, declare_type_keywords):
            profile.declare_type_columns.append((col_letter, header_path))
            has_declare_type = True
            section = _detect_section_type(header_path)
            if section == "social_security":
                profile.has_social_security_section = True
            elif section == "housing_fund":
                profile.has_housing_fund_section = True
        
        if _header_ends_with(header, "费用年月") or _header_contains_any(header, fee_month_keywords):
            profile.fee_month_columns.append((col_letter, header_path))
            has_fee_month = True
        if _header_contains_any(header, reason_keywords):
            profile.reason_columns.append((col_letter, header_path))
            section = _detect_section_type(header_path)
            if section == "social_security":
                profile.has_social_security_section = True
            elif section == "housing_fund":
                profile.has_housing_fund_section = True
    
    required_signals = sum([has_name, has_id, has_declare_type, has_fee_month])
    if required_signals >= 3 and has_name and (has_declare_type or has_fee_month):
        profile.is_detected = True
    
    if region.table.sample_rows:
        for row in region.table.sample_rows:
            if not isinstance(row, dict):
                continue
            for key, value in row.items():
                if not value:
                    continue
                value_str = str(value)
                if re.match(r'^\d{6}$', value_str):
                    profile.fee_month_format = "YYYYMM"
                    break
                if re.match(r'^\d{4}-\d{2}$', value_str):
                    profile.fee_month_format = "YYYY-MM"
                    break
            if profile.fee_month_format:
                break
    
    if not profile.fee_month_format:
        profile.fee_month_format = "YYYYMM"
    
    return profile


def _infer_template_intent_from_filename(filename: str) -> Optional[str]:
    if not filename:
        return None
    normalized = _normalize_header_text(filename)
    add_keywords = ["增员", "增加", "入职", "新增", "add"]
    remove_keywords = ["减员", "减少", "离职", "删除", "remove", "停保"]
    has_add = any(k in normalized for k in add_keywords)
    has_remove = any(k in normalized for k in remove_keywords)
    if has_add and not has_remove:
        return "add"
    if has_remove and not has_add:
        return "remove"
    return None


def _parse_any_date(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        value = str(int(value))
    if not isinstance(value, str):
        value = str(value)
    text = value.strip()
    if not text:
        return None
    try:
        normalized = text.replace("Z", "+00:00")
        return datetime.fromisoformat(normalized)
    except (ValueError, TypeError):
        pass
    common_match = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', text)
    if common_match:
        try:
            return datetime(
                int(common_match.group(1)),
                int(common_match.group(2)),
                int(common_match.group(3))
            )
        except (ValueError, TypeError):
            return None
    compact_match = re.match(r'^(\d{4})(\d{2})(\d{2})$', text)
    if compact_match:
        try:
            return datetime(
                int(compact_match.group(1)),
                int(compact_match.group(2)),
                int(compact_match.group(3))
            )
        except (ValueError, TypeError):
            return None
    return None


def _next_month(dt: datetime) -> Tuple[int, int]:
    year = dt.year
    month = dt.month + 1
    if month > 12:
        year += 1
        month = 1
    return year, month


def _format_fee_month(year: int, month: int, fmt: str) -> str:
    if fmt == "YYYY-MM":
        return f"{year:04d}-{month:02d}"
    return f"{year:04d}{month:02d}"


def _get_record_value(record: Dict, key: str) -> Optional[str]:
    if not isinstance(record, dict) or not key:
        return None
    value = record.get(key)
    if value is not None:
        return str(value).strip() if value else None
    key_lower = key.lower()
    for k, v in record.items():
        if isinstance(k, str) and k.lower() == key_lower:
            return str(v).strip() if v else None
    return None


def _normalize_record_key(key: str) -> str:
    return _normalize_header_text(key).replace("_", "")


def _get_record_value_by_key(record: Dict, key: str) -> Optional[str]:
    val = _get_record_value(record, key)
    if val:
        return val
    if not isinstance(record, dict) or not isinstance(key, str):
        return None
    target = _normalize_record_key(key)
    for k, v in record.items():
        if isinstance(k, str) and _normalize_record_key(k) == target:
            return str(v).strip() if v else None
    return None


def _date_candidates(intent: str) -> List[str]:
    if intent == "add":
        return [
            "start_date",
            "startdate",
            "join_date",
            "entry_date",
            "hire_date",
            "参加保险年月",
            "参保年月",
            "增员年月",
            "feemonth",
            "month",
        ]
    if intent == "remove":
        return [
            "end_date",
            "enddate",
            "termination_date",
            "terminationdate",
            "leave_date",
            "leavedate",
            "ss_end_month",
            "hf_end_month",
            "离职日期",
            "减员月份",
            "减员年月",
            "停保月份",
            "停保年月",
            "退工日期",
            "退保日期",
            "终止日期",
            "停保日期",
            "费用年月",
        ]
    return []


def _reason_candidates(intent: str) -> List[str]:
    if intent == "remove":
        return [
            "termination_reason",
            "terminationreason",
            "reason",
            "变动原因",
            "减员原因",
            "离职原因",
            "退工原因",
            "退保原因",
            "原因",
        ]
    return []


def _find_existing_key_by_normalized(records: List[Dict[str, Any]], normalized_prefix: str) -> Optional[str]:
    for record in records:
        if not isinstance(record, dict):
            continue
        for key in record.keys():
            if isinstance(key, str) and _normalize_record_key(key).startswith(normalized_prefix):
                return key
    return None


def _select_date_key(records: List[Dict[str, Any]], intent: str) -> Tuple[Optional[str], Optional[str]]:
    candidates = _date_candidates(intent)
    if not candidates:
        return None, "intent_unknown"
    if intent == "remove":
        end_date_key = _find_existing_key_by_normalized(records, "enddate")
        if end_date_key:
            return end_date_key, "forced_end_date"
        termination_key = _find_existing_key_by_normalized(records, "terminationdate")
        if termination_key:
            return termination_key, "forced_termination"
    for key in candidates:
        for record in records:
            if not isinstance(record, dict):
                continue
            if _get_record_value_by_key(record, key):
                return key, "found"
    return None, f"required_field_missing:{candidates[0]}"


def _validate_chinese_name(name: str) -> bool:
    if not name:
        return False
    return bool(re.match(r'^[\u4e00-\u9fff]{2,4}$', name.strip()))


def _validate_id_number(id_num: str) -> bool:
    if not id_num:
        return True
    id_num = id_num.strip()
    if re.match(r'^\d{15}$', id_num):
        return True
    if re.match(r'^\d{17}[\dXx]$', id_num):
        return True
    if re.match(r'^\d{4,}$', id_num):
        return True
    return False


def _validate_fee_month(fee_month: str) -> bool:
    if not fee_month:
        return False
    fee_month = fee_month.strip()
    if re.match(r'^\d{6}$', fee_month):
        return True
    if re.match(r'^\d{4}-\d{2}$', fee_month):
        return True
    return False


def _validate_declare_type(declare_type: str) -> bool:
    if not declare_type:
        return False
    return declare_type.strip() in ("增", "减")


def _is_valid_cn_id(value: str) -> bool:
    if not value:
        return False
    value = value.strip()
    if re.match(r'^\d{15}$', value):
        return True
    if re.match(r'^\d{17}[\dXx]$', value):
        return True
    return False


def _is_pure_digits(value: str) -> bool:
    if not value:
        return False
    return bool(re.match(r'^\d+$', value.strip()))


def _is_date_like(value: str) -> bool:
    if not value:
        return False
    value = value.strip()
    if re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$', value):
        return True
    if re.match(r'^\d{8}$', value):
        return True
    if re.match(r'^\d{4}年\d{1,2}月\d{1,2}日$', value):
        return True
    if re.match(r'^\d{1,2}月\d{1,2}日$', value):
        return True
    return False


def _is_gender_like(value: str) -> bool:
    if not value:
        return False
    value = value.strip().lower()
    return value in ("男", "女", "male", "female", "m", "f", "1", "2")


def _score_value_as_name(value: str) -> float:
    if not value:
        return 0.0
    value = value.strip()
    if _is_pure_digits(value):
        return 0.0
    if _is_date_like(value):
        return 0.0
    if _is_gender_like(value):
        return 0.0
    if _validate_chinese_name(value):
        return 1.0
    if re.match(r'^[\u4e00-\u9fff]{1,6}$', value):
        return 0.7
    if re.match(r'^[A-Za-z\s]+$', value) and len(value) >= 2:
        return 0.5
    return 0.1


def _score_value_as_id(value: str) -> float:
    if not value:
        return 0.0
    value = value.strip()
    if _is_valid_cn_id(value):
        return 1.0
    if re.match(r'^\d{15}$', value):
        return 1.0
    if re.match(r'^\d{17}[\dXx]$', value):
        return 1.0
    if re.match(r'^\d{6,14}$', value):
        return 0.7
    if re.match(r'^\d{4,}$', value):
        return 0.5
    if _validate_chinese_name(value):
        return 0.0
    return 0.1


def _score_value_as_date(value: str) -> float:
    if not value:
        return 0.0
    value = value.strip()
    if _is_date_like(value):
        return 1.0
    dt = _parse_any_date(value)
    if dt:
        return 1.0
    return 0.0


def _score_key_values_for_field_type(
    records: List[Dict[str, Any]],
    key: str,
    field_type: str
) -> float:
    if not records or not key:
        return 0.0
    
    total = 0
    matched = 0
    
    for rec in records[:20]:
        if not isinstance(rec, dict):
            continue
        value = _get_record_value(rec, key)
        if not value:
            continue
        
        total += 1
        
        if field_type == "name":
            score = _score_value_as_name(value)
        elif field_type == "id":
            score = _score_value_as_id(value)
        elif field_type == "date":
            score = _score_value_as_date(value)
        else:
            score = 0.0
        
        if score >= 0.5:
            matched += 1
    
    if total == 0:
        return 0.0
    
    return matched / total


def _find_best_key_for_field_type(
    records: List[Dict[str, Any]],
    candidate_keys: List[str],
    field_type: str,
    exclude_keys: Optional[List[str]] = None
) -> Tuple[Optional[str], float]:
    if not records or not candidate_keys:
        return None, 0.0
    
    exclude_keys = exclude_keys or []
    best_key = None
    best_score = 0.0
    
    for key in candidate_keys:
        if key in exclude_keys:
            continue
        
        score = _score_key_values_for_field_type(records, key, field_type)
        if score > best_score:
            best_score = score
            best_key = key
    
    return best_key, best_score


@dataclass
class FieldMappingValidation:
    name_key: Optional[str] = None
    name_score: float = 0.0
    name_valid: bool = False
    id_key: Optional[str] = None
    id_score: float = 0.0
    id_valid: bool = False
    id_available: bool = False
    date_key: Optional[str] = None
    date_score: float = 0.0
    date_valid: bool = False
    warnings: List[str] = field(default_factory=list)
    used_fallback: bool = False


def _validate_and_fix_field_mappings(
    records: List[Dict[str, Any]],
    proposed_name_key: str,
    proposed_id_key: str,
    proposed_date_key: str,
    intent: str
) -> FieldMappingValidation:
    result = FieldMappingValidation()
    
    if not records:
        result.warnings.append("no_records_for_validation")
        return result
    
    all_keys: List[str] = []
    seen = set()
    for rec in records[:10]:
        if isinstance(rec, dict):
            for k in rec.keys():
                if k not in seen:
                    seen.add(k)
                    all_keys.append(k)
    
    name_score = _score_key_values_for_field_type(records, proposed_name_key, "name")
    if name_score >= 0.5:
        result.name_key = proposed_name_key
        result.name_score = name_score
        result.name_valid = True
    else:
        best_name_key, best_name_score = _find_best_key_for_field_type(
            records, all_keys, "name", exclude_keys=[]
        )
        if best_name_key and best_name_score >= 0.3:
            result.name_key = best_name_key
            result.name_score = best_name_score
            result.name_valid = True
            result.used_fallback = True
            result.warnings.append(f"name_key_fallback:{proposed_name_key}->{best_name_key}")
        else:
            result.name_key = proposed_name_key
            result.name_score = name_score
            result.name_valid = False
            result.warnings.append(f"name_key_invalid:{proposed_name_key}(score={name_score:.2f})")
    
    exclude_for_id = [result.name_key] if result.name_key else []
    
    id_score = _score_key_values_for_field_type(records, proposed_id_key, "id")
    
    any_id_value = False
    for rec in records[:20]:
        if isinstance(rec, dict):
            val = _get_record_value(rec, proposed_id_key)
            if val:
                any_id_value = True
                break
    
    if not any_id_value:
        for key in all_keys:
            for rec in records[:20]:
                if isinstance(rec, dict):
                    val = _get_record_value(rec, key)
                    if val and _score_value_as_id(val) >= 0.5:
                        any_id_value = True
                        break
            if any_id_value:
                break
    
    result.id_available = any_id_value
    
    if id_score >= 0.5:
        result.id_key = proposed_id_key
        result.id_score = id_score
        result.id_valid = True
    else:
        best_id_key, best_id_score = _find_best_key_for_field_type(
            records, all_keys, "id", exclude_keys=exclude_for_id
        )
        if best_id_key and best_id_score >= 0.3:
            result.id_key = best_id_key
            result.id_score = best_id_score
            result.id_valid = True
            result.used_fallback = True
            result.warnings.append(f"id_key_fallback:{proposed_id_key}->{best_id_key}")
        else:
            if any_id_value:
                result.id_key = proposed_id_key
                result.id_score = id_score
                result.id_valid = False
                result.warnings.append(f"id_key_invalid:{proposed_id_key}(score={id_score:.2f})")
            else:
                result.id_key = None
                result.id_score = 0.0
                result.id_valid = False
                result.warnings.append("required_field_missing:证件号码")
    
    exclude_for_date = [result.name_key, result.id_key]
    exclude_for_date = [k for k in exclude_for_date if k]
    
    date_score = _score_key_values_for_field_type(records, proposed_date_key, "date")
    if date_score >= 0.5:
        result.date_key = proposed_date_key
        result.date_score = date_score
        result.date_valid = True
    else:
        date_key_candidates = _date_candidates(intent)
        best_date_key, best_date_score = _find_best_key_for_field_type(
            records, date_key_candidates, "date", exclude_keys=exclude_for_date
        )
        if best_date_key and best_date_score >= 0.3:
            result.date_key = best_date_key
            result.date_score = best_date_score
            result.date_valid = True
            result.used_fallback = True
            result.warnings.append(f"date_key_fallback:{proposed_date_key}->{best_date_key}")
        else:
            result.date_key = proposed_date_key
            result.date_score = date_score
            result.date_valid = False
            result.warnings.append(f"date_key_invalid:{proposed_date_key}(score={date_score:.2f})")
    
    return result


def _validate_row_values(
    row: Dict[str, Any],
    declare_type_keys: List[str],
    fee_month_keys: List[str],
    expected_declare_type: str
) -> Tuple[bool, List[str]]:
    errors: List[str] = []
    
    name = row.get("__name__", "")
    if not name:
        errors.append("name_empty")
    elif _is_pure_digits(name):
        errors.append(f"name_is_digits:{name}")
    elif not _validate_chinese_name(name):
        if not re.match(r'^[\u4e00-\u9fff]+', name):
            errors.append(f"name_not_chinese_like:{name}")
    
    id_number = row.get("__id_number__", "")
    if id_number:
        if _validate_chinese_name(id_number):
            errors.append(f"id_looks_like_name:{id_number}")
        elif _is_date_like(id_number):
            errors.append(f"id_looks_like_date:{id_number}")
    
    for key in declare_type_keys:
        val = row.get(key, "")
        if val != expected_declare_type:
            errors.append(f"declare_type_not_constant:{key}={val}")
        if val and val not in ("增", "减"):
            errors.append(f"declare_type_invalid:{key}={val}")
    
    for key in fee_month_keys:
        val = row.get(key, "")
        if not val:
            errors.append(f"fee_month_empty:{key}")
        elif _is_date_like(val):
            errors.append(f"fee_month_is_raw_date:{key}={val}")
        elif not _validate_fee_month(val):
            errors.append(f"fee_month_invalid_format:{key}={val}")
    
    is_valid = len(errors) == 0
    return is_valid, errors


def _validate_all_rows(
    rows: List[Dict[str, Any]],
    declare_type_keys: List[str],
    fee_month_keys: List[str],
    expected_declare_type: str
) -> Tuple[List[Dict[str, Any]], List[str]]:
    valid_rows: List[Dict[str, Any]] = []
    all_errors: List[str] = []
    
    for idx, row in enumerate(rows):
        is_valid, errors = _validate_row_values(
            row, declare_type_keys, fee_month_keys, expected_declare_type
        )
        if is_valid:
            valid_rows.append(row)
        else:
            for err in errors:
                all_errors.append(f"row_{idx}:{err}")
    
    return valid_rows, all_errors


def _score_filename_for_intent(filename: str, intent: str) -> Tuple[float, Dict[str, Any]]:
    if not filename:
        return 0.0, {"reason": "no_filename"}
    
    normalized = _normalize_header_text(filename)
    score = 0.0
    matched_signals: List[str] = []
    exclusion_signals: List[str] = []
    
    add_strong = ["入职", "新增", "增员", "add", "onboard", "新入职"]
    add_weak = ["花名册", "员工名单", "人员名单"]
    remove_strong = ["离职", "减员", "停保", "remove", "offboard", "退出"]
    remove_weak = ["减少", "删除"]
    exclusion_keywords = ["在职", "在岗", "active", "current"]
    
    for kw in exclusion_keywords:
        if kw in normalized:
            exclusion_signals.append(kw)
    
    if intent == "add":
        for kw in add_strong:
            if kw in normalized:
                score += 30.0
                matched_signals.append(f"strong:{kw}")
        for kw in add_weak:
            if kw in normalized:
                score += 10.0
                matched_signals.append(f"weak:{kw}")
        for kw in remove_strong:
            if kw in normalized:
                score -= 20.0
                matched_signals.append(f"conflict:{kw}")
    elif intent == "remove":
        for kw in remove_strong:
            if kw in normalized:
                score += 30.0
                matched_signals.append(f"strong:{kw}")
        for kw in remove_weak:
            if kw in normalized:
                score += 10.0
                matched_signals.append(f"weak:{kw}")
        for kw in add_strong:
            if kw in normalized:
                score -= 20.0
                matched_signals.append(f"conflict:{kw}")
    
    if exclusion_signals:
        score -= 15.0
    
    breakdown = {
        "matched_signals": matched_signals,
        "exclusion_signals": exclusion_signals,
    }
    return score, breakdown


def _score_keys_for_intent(keys: List[str], intent: str) -> Tuple[float, Dict[str, Any]]:
    if not keys:
        return 0.0, {"reason": "no_keys"}
    
    score = 0.0
    matched_keys: List[str] = []
    
    name_keys = ["姓名", "name", "员工姓名", "参保人", "被保险人"]
    id_keys = ["身份证", "证件号", "id", "employee_id", "工号"]
    start_date_keys = ["入职日期", "start_date", "入职时间", "开始日期", "到岗日期"]
    end_date_keys = ["离职日期", "end_date", "leave_date", "离职时间", "退出日期"]
    
    keys_lower = [_normalize_header_text(k) for k in keys]
    
    for nk in name_keys:
        if any(nk in k for k in keys_lower):
            score += 10.0
            matched_keys.append(f"name:{nk}")
            break
    
    for ik in id_keys:
        if any(ik in k for k in keys_lower):
            score += 10.0
            matched_keys.append(f"id:{ik}")
            break
    
    if intent == "add":
        for dk in start_date_keys:
            if any(dk in k for k in keys_lower):
                score += 20.0
                matched_keys.append(f"date:{dk}")
                break
        for dk in end_date_keys:
            if any(dk in k for k in keys_lower):
                score -= 5.0
                matched_keys.append(f"conflict_date:{dk}")
    elif intent == "remove":
        for dk in end_date_keys:
            if any(dk in k for k in keys_lower):
                score += 20.0
                matched_keys.append(f"date:{dk}")
                break
        for dk in start_date_keys:
            if any(dk in k for k in keys_lower):
                pass
    
    breakdown = {"matched_keys": matched_keys}
    return score, breakdown


def _compute_field_coverage(
    records: List[Dict[str, Any]],
    intent: str
) -> Tuple[float, float, float, Dict[str, Any]]:
    if not records:
        return 0.0, 0.0, 0.0, {"reason": "no_records"}
    
    name_count = 0
    id_count = 0
    valid_cn_id_count = 0
    date_count = 0
    total = len(records)
    
    name_keys = ["name", "姓名", "员工姓名", "参保人"]
    id_keys = ["id_number", "身份证号", "证件号码", "employee_id", "工号"]
    start_date_keys = ["start_date", "入职日期", "入职时间"]
    end_date_keys = ["leave_date", "end_date", "离职日期", "离职时间"]
    date_keys = end_date_keys if intent == "remove" else start_date_keys
    
    for rec in records:
        if not isinstance(rec, dict):
            continue
        
        has_name = False
        for nk in name_keys:
            val = _get_record_value(rec, nk)
            if val and _validate_chinese_name(val):
                has_name = True
                break
        if has_name:
            name_count += 1
        
        has_id = False
        for ik in id_keys:
            val = _get_record_value(rec, ik)
            if val:
                has_id = True
                if _is_valid_cn_id(val):
                    valid_cn_id_count += 1
                break
        if has_id:
            id_count += 1
        
        has_date = False
        for dk in date_keys:
            val = _get_record_value(rec, dk)
            if val and _parse_any_date(val):
                has_date = True
                break
        if has_date:
            date_count += 1
    
    name_cov = name_count / total if total > 0 else 0.0
    id_cov = id_count / total if total > 0 else 0.0
    date_cov = date_count / total if total > 0 else 0.0
    
    breakdown = {
        "total_records": total,
        "name_count": name_count,
        "id_count": id_count,
        "valid_cn_id_count": valid_cn_id_count,
        "date_count": date_count,
    }
    return name_cov, id_cov, date_cov, breakdown


def _score_source(
    source: Dict[str, Any],
    intent: str
) -> SourceScore:
    source_id = source.get("source_id", "")
    filename = source.get("filename", "")
    source_type = source.get("source_type", "unknown")
    
    score_obj = SourceScore(
        source_id=source_id,
        filename=filename,
        source_type=source_type,
    )
    
    filename_score, filename_breakdown = _score_filename_for_intent(filename, intent)
    score_obj.intent_score = filename_score
    score_obj.score_breakdown["filename"] = filename_breakdown
    
    if filename_breakdown.get("exclusion_signals"):
        score_obj.has_exclusion_signal = True
    
    extracted = source.get("extracted")
    if not isinstance(extracted, dict):
        score_obj.total_score = score_obj.intent_score
        return score_obj
    
    data = extracted.get("data")
    if not isinstance(data, list):
        score_obj.total_score = score_obj.intent_score
        return score_obj
    
    score_obj.record_count = len(data)
    
    record_keys: List[str] = []
    seen_keys = set()
    for rec in data[:10]:
        if isinstance(rec, dict):
            for k in rec.keys():
                if k not in seen_keys:
                    seen_keys.add(k)
                    record_keys.append(k)
    
    keys_score, keys_breakdown = _score_keys_for_intent(record_keys, intent)
    score_obj.score_breakdown["keys"] = keys_breakdown
    
    name_cov, id_cov, date_cov, cov_breakdown = _compute_field_coverage(data, intent)
    score_obj.name_coverage = name_cov
    score_obj.id_coverage = id_cov
    score_obj.date_coverage = date_cov
    score_obj.score_breakdown["coverage"] = cov_breakdown
    
    coverage_score = (name_cov * 25.0) + (id_cov * 15.0) + (date_cov * 30.0)
    
    if source_type == "email" and intent == "remove":
        record_intent_bonus = 0
        for rec in data:
            if isinstance(rec, dict) and rec.get("intent") == "remove":
                record_intent_bonus += 5
        score_obj.score_breakdown["email_intent_bonus"] = record_intent_bonus
        coverage_score += min(record_intent_bonus, 20)
    
    score_obj.total_score = score_obj.intent_score + keys_score + coverage_score
    
    if score_obj.has_exclusion_signal:
        score_obj.total_score *= 0.5
    
    score_obj.score_breakdown["total_components"] = {
        "filename_score": filename_score,
        "keys_score": keys_score,
        "coverage_score": coverage_score,
        "exclusion_penalty_applied": score_obj.has_exclusion_signal,
    }
    
    return score_obj


def _select_sources_for_template(
    extracted_json: dict,
    intent: str,
    max_sources: int = 2
) -> Tuple[List[Dict[str, Any]], List[SourceScore], List[str]]:
    warnings: List[str] = []
    sources = extracted_json.get("sources", []) if isinstance(extracted_json, dict) else []
    
    if not sources:
        return [], [], ["no_sources_available"]
    
    scored_sources: List[Tuple[SourceScore, Dict[str, Any]]] = []
    
    for source in sources:
        if not isinstance(source, dict):
            continue
        
        extracted = source.get("extracted")
        if not isinstance(extracted, dict):
            continue
        
        data = extracted.get("data")
        if not isinstance(data, list) or not data:
            continue
        
        score = _score_source(source, intent)
        scored_sources.append((score, source))
    
    if not scored_sources:
        return [], [], ["no_valid_sources_with_data"]
    
    scored_sources.sort(key=lambda x: x[0].total_score, reverse=True)
    
    all_excluded = all(s[0].has_exclusion_signal for s in scored_sources)
    if all_excluded:
        warnings.append("all_sources_have_exclusion_signals_using_anyway")
    
    selected: List[Tuple[SourceScore, Dict[str, Any]]] = []
    selected_scores: List[SourceScore] = []
    
    for score, source in scored_sources:
        if len(selected) >= max_sources:
            break
        
        if not all_excluded and score.has_exclusion_signal:
            continue
        
        selected.append((score, source))
        selected_scores.append(score)
    
    if not selected and all_excluded:
        for score, source in scored_sources[:max_sources]:
            selected.append((score, source))
            selected_scores.append(score)
    
    selected_sources = [s[1] for s in selected]
    
    for idx, score in enumerate(selected_scores):
        logger.debug(
            "Selected source %d: %s (score=%.1f, type=%s)",
            idx + 1,
            score.filename,
            score.total_score,
            score.source_type
        )
    
    return selected_sources, selected_scores, warnings


def _collect_candidate_source_keys(sources: List[Dict[str, Any]]) -> List[str]:
    seen = set()
    keys: List[str] = []
    for source in sources:
        if not isinstance(source, dict):
            continue
        extracted = source.get("extracted")
        if not isinstance(extracted, dict):
            continue
        data = extracted.get("data")
        if not isinstance(data, list):
            continue
        for item in data[:30]:
            if not isinstance(item, dict):
                continue
            for key in item.keys():
                if isinstance(key, str) and key not in seen:
                    seen.add(key)
                    keys.append(key)
    return keys


def _collect_sample_records_for_mapping(sources: List[Dict[str, Any]], limit: int = 20) -> List[Dict[str, Any]]:
    samples: List[Dict[str, Any]] = []
    for source in sources:
        if not isinstance(source, dict):
            continue
        extracted = source.get("extracted")
        if not isinstance(extracted, dict):
            continue
        data = extracted.get("data")
        if not isinstance(data, list):
            continue
        for item in data:
            if isinstance(item, dict):
                samples.append(dict(item))
            if len(samples) >= limit:
                return samples
    return samples


def _llm_infer_source_key_mapping(
    llm: LLMClient,
    selected_sources: List[Dict[str, Any]],
    profile: SocialSecurityProfile,
    template_intent: str,
) -> Tuple[Dict[str, str], List[str], Dict[str, Any]]:
    warnings: List[str] = []
    debug: Dict[str, Any] = {"enabled": True}
    prompts = get_prompts()
    prompt_tpl = str((prompts or {}).get("TEMPLATE_COLUMN_MAPPING_PROMPT", "") or "").strip()
    if not prompt_tpl:
        warnings.append("llm_mapping_prompt_missing:TEMPLATE_COLUMN_MAPPING_PROMPT")
        debug["llm_called"] = False
        return {}, warnings, debug

    candidate_keys = _collect_candidate_source_keys(selected_sources)
    sample_records = _collect_sample_records_for_mapping(selected_sources)
    if not candidate_keys:
        warnings.append("llm_mapping_no_candidate_keys")
        debug["llm_called"] = False
        return {}, warnings, debug

    payload = {
        "intent": template_intent,
        "target_fields": ["name", "id_number", "event_date", "termination_reason"],
        "candidate_source_keys": candidate_keys,
        "template_columns": {
            "name_columns": profile.name_columns,
            "id_columns": profile.id_columns,
            "declare_type_columns": profile.declare_type_columns,
            "fee_month_columns": profile.fee_month_columns,
            "reason_columns": profile.reason_columns,
        },
        "sample_records": sample_records[:12],
    }
    prompt = f"{prompt_tpl}\n\nINPUT_JSON:\n{json.dumps(payload, ensure_ascii=False)}"
    debug["candidate_key_count"] = len(candidate_keys)
    debug["sample_record_count"] = len(sample_records[:12])
    debug["llm_called"] = True

    try:
        if hasattr(llm, "chat_json_once"):
            raw = llm.chat_json_once(
                prompt,
                system=None,
                temperature=0,
                timeout=30.0,
                step="template_column_mapping_social_security",
            )
        else:
            raw = llm.chat_json(
                prompt,
                system=None,
                temperature=0,
                step="template_column_mapping_social_security",
            )
    except Exception as exc:
        warnings.append(f"llm_mapping_exception:{type(exc).__name__}")
        debug["error"] = str(exc)[:200]
        return {}, warnings, debug

    mapping_obj = {}
    confidence = None
    if isinstance(raw, dict):
        if isinstance(raw.get("target_field_to_source_key"), dict):
            mapping_obj = raw.get("target_field_to_source_key") or {}
        elif isinstance(raw.get("mapping"), dict):
            mapping_obj = raw.get("mapping") or {}
        confidence = raw.get("confidence")
        llm_warnings = raw.get("warnings")
        if isinstance(llm_warnings, list):
            for w in llm_warnings:
                if isinstance(w, str) and w.strip():
                    warnings.append(f"llm_mapping:{w.strip()}")

    valid_targets = {"name", "id_number", "event_date", "termination_reason"}
    mapped: Dict[str, str] = {}
    for target, source_key in mapping_obj.items():
        if target not in valid_targets:
            continue
        if not isinstance(source_key, str) or not source_key.strip():
            continue
        if source_key not in candidate_keys:
            warnings.append(f"llm_mapping_invalid_key:{target}->{source_key}")
            continue
        mapped[target] = source_key.strip()

    debug["confidence"] = confidence if isinstance(confidence, (int, float)) else None
    debug["mapped_targets"] = sorted(mapped.keys())
    return mapped, warnings, debug


def _extract_records_from_source(
    source: Dict[str, Any],
    intent: str,
    mapping_hints: Optional[Dict[str, str]] = None,
) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    warnings: List[str] = []
    debug_info: Dict[str, Any] = {
        "add_sheet_counts_before_filter": {},
        "add_sheet_counts_after_filter": {},
        "add_sheet_counts_blocked": {},
        "remove_sheet_counts_before_filter": {},
        "remove_sheet_counts_after_filter": {},
        "remove_sheet_counts_blocked": {},
        "date_key_fallback_hits": 0,
    }
    
    extracted = source.get("extracted")
    if not isinstance(extracted, dict):
        return records, warnings, debug_info
    
    data = extracted.get("data")
    if not isinstance(data, list):
        return records, warnings, debug_info
    
    source_filename = source.get("filename", "unknown")
    source_type = source.get("source_type", "unknown")
    source_id = source.get("source_id", "")

    if intent in ("add", "remove"):
        for item in data:
            if not isinstance(item, dict):
                continue
            sheet_name = str(item.get("__sheet_name__", "") or "")
            if intent == "add":
                debug_info["add_sheet_counts_before_filter"][sheet_name] = (
                    debug_info["add_sheet_counts_before_filter"].get(sheet_name, 0) + 1
                )
            else:
                debug_info["remove_sheet_counts_before_filter"][sheet_name] = (
                    debug_info["remove_sheet_counts_before_filter"].get(sheet_name, 0) + 1
                )
    
    default_name_keys = ["name", "姓名", "员工姓名", "参保人"]
    default_id_keys = ["id_number", "身份证号", "证件号码", "employee_id", "工号"]
    date_candidates = _date_candidates(intent)
    reason_candidates = _reason_candidates(intent)
    mapping_hints = mapping_hints or {}
    llm_reason_key = mapping_hints.get("reason_key") if isinstance(mapping_hints.get("reason_key"), str) else None
    if llm_reason_key and llm_reason_key not in reason_candidates:
        reason_candidates = [llm_reason_key] + reason_candidates

    proposed_name_key = (
        mapping_hints.get("name_key")
        if isinstance(mapping_hints.get("name_key"), str) and mapping_hints.get("name_key")
        else default_name_keys[0]
    )
    proposed_id_key = (
        mapping_hints.get("id_key")
        if isinstance(mapping_hints.get("id_key"), str) and mapping_hints.get("id_key")
        else default_id_keys[0]
    )
    proposed_date_key = date_candidates[0] if date_candidates else ""
    
    for nk in default_name_keys:
        for item in data[:5]:
            if isinstance(item, dict) and _get_record_value(item, nk):
                proposed_name_key = nk
                break
        else:
            continue
        break
    
    for ik in default_id_keys:
        for item in data[:5]:
            if isinstance(item, dict) and _get_record_value(item, ik):
                proposed_id_key = ik
                break
        else:
            continue
        break
    
    llm_date_key = mapping_hints.get("date_key") if isinstance(mapping_hints.get("date_key"), str) else None
    if llm_date_key and any(_get_record_value_by_key(item, llm_date_key) for item in data if isinstance(item, dict)):
        proposed_date_key = llm_date_key
    else:
        selected_date_key, date_key_reason = _select_date_key(data, intent)
        if selected_date_key:
            proposed_date_key = selected_date_key
        else:
            warnings.append(date_key_reason or "required_field_missing:离职日期")
            return records, warnings, debug_info
    
    validation = _validate_and_fix_field_mappings(
        data, proposed_name_key, proposed_id_key, proposed_date_key, intent
    )
    warnings.extend(validation.warnings)
    
    if validation.used_fallback:
        warnings.append(f"field_mapping_used_fallback:{source_filename}")
    
    name_key = validation.name_key or proposed_name_key
    id_key = validation.id_key
    date_key = validation.date_key or proposed_date_key
    
    id_available = validation.id_available
    
    for item in data:
        if not isinstance(item, dict):
            continue
        
        record_intent = item.get("intent", "")
        if intent == "remove" and record_intent and record_intent != "remove":
            continue
        if intent == "add" and record_intent and record_intent != "add":
            continue

        sheet_name = str(item.get("__sheet_name__", "") or "")
        sheet_name_norm = _normalize_header_text(sheet_name)
        if intent == "remove":
            remove_allow_keywords = ("减员", "离职")
            remove_block_keywords = ("增员", "新增", "入职")
            if any(k in sheet_name_norm for k in remove_block_keywords):
                debug_info["remove_sheet_counts_blocked"][sheet_name] = (
                    debug_info["remove_sheet_counts_blocked"].get(sheet_name, 0) + 1
                )
                continue
            if not any(k in sheet_name_norm for k in remove_allow_keywords):
                debug_info["remove_sheet_counts_blocked"][sheet_name] = (
                    debug_info["remove_sheet_counts_blocked"].get(sheet_name, 0) + 1
                )
                continue
            debug_info["remove_sheet_counts_after_filter"][sheet_name] = (
                debug_info["remove_sheet_counts_after_filter"].get(sheet_name, 0) + 1
            )
        elif intent == "add":
            add_allow_keywords = ("增员", "新增", "入职")
            add_block_keywords = ("减员", "离职", "停保")
            if any(k in sheet_name_norm for k in add_block_keywords):
                debug_info["add_sheet_counts_blocked"][sheet_name] = (
                    debug_info["add_sheet_counts_blocked"].get(sheet_name, 0) + 1
                )
                continue
            # Unknown sheet names are still allowed to avoid over-filtering;
            # we only hard-block known remove-like sheets for add templates.
            if sheet_name_norm and not any(k in sheet_name_norm for k in add_allow_keywords):
                warnings.append(f"add_sheet_name_neutral_kept:{sheet_name}")
            debug_info["add_sheet_counts_after_filter"][sheet_name] = (
                debug_info["add_sheet_counts_after_filter"].get(sheet_name, 0) + 1
            )
        
        name = _get_record_value(item, name_key)
        if not name:
            for nk in default_name_keys:
                val = _get_record_value(item, nk)
                if val and _validate_chinese_name(val):
                    name = val
                    break
        
        if not name or not _validate_chinese_name(name):
            if name and _is_pure_digits(name):
                warnings.append(f"record_skipped_name_is_digits:{name}")
            continue
        
        event_date = None
        used_fallback_date_key = False
        if date_key:
            val = _get_record_value_by_key(item, date_key)
            if val:
                dt = _parse_any_date(val)
                if dt:
                    event_date = dt
        if not event_date:
            for dk in date_candidates:
                if dk == date_key:
                    continue
                val = _get_record_value_by_key(item, dk)
                if val:
                    dt = _parse_any_date(val)
                    if dt:
                        event_date = dt
                        used_fallback_date_key = True
                        break
        if used_fallback_date_key:
            debug_info["date_key_fallback_hits"] = int(debug_info.get("date_key_fallback_hits", 0) or 0) + 1
        
        if not event_date:
            warnings.append(f"record_skipped_no_date:{name}")
            continue
        
        id_number = None
        if id_key and id_available:
            val = _get_record_value(item, id_key)
            if val:
                val = val.strip()
                if _validate_chinese_name(val):
                    warnings.append(f"id_value_looks_like_name:{val}")
                elif _is_date_like(val):
                    warnings.append(f"id_value_looks_like_date:{val}")
                elif _is_valid_cn_id(val) or _validate_id_number(val):
                    id_number = val
        
        if not id_number and id_available:
            for ik in default_id_keys:
                val = _get_record_value(item, ik)
                if val:
                    val = val.strip()
                    if _is_valid_cn_id(val) or _validate_id_number(val):
                        if not _validate_chinese_name(val) and not _is_date_like(val):
                            id_number = val
                            break
        
        reason_value = ""
        if reason_candidates:
            for rk in reason_candidates:
                rv = _get_record_value_by_key(item, rk)
                if rv:
                    reason_value = rv
                    break
        records.append({
            "name": name,
            "id_number": id_number or "",
            "event_date": event_date,
            "termination_reason": reason_value,
            "__source_file__": source_filename,
            "__source_type__": source_type,
            "__source_id__": source_id,
        })
    
    return records, warnings, debug_info


def _deduplicate_records_advanced(
    records: List[Dict[str, Any]]
) -> Tuple[List[Dict[str, Any]], int]:
    if not records:
        return [], 0
    
    by_id: Dict[str, List[Dict[str, Any]]] = {}
    by_name_date: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    no_id_records: List[Dict[str, Any]] = []
    
    for rec in records:
        id_num = rec.get("id_number", "").strip()
        name = rec.get("name", "")
        event_date = rec.get("event_date")
        date_str = event_date.strftime("%Y-%m-%d") if event_date else ""
        
        if id_num and (_is_valid_cn_id(id_num) or len(id_num) >= 4):
            if id_num not in by_id:
                by_id[id_num] = []
            by_id[id_num].append(rec)
        else:
            key = (name, date_str)
            if key not in by_name_date:
                by_name_date[key] = []
            by_name_date[key].append(rec)
            no_id_records.append(rec)
    
    result: List[Dict[str, Any]] = []
    seen_names: set = set()
    duplicates = 0
    
    for id_num, recs in by_id.items():
        best = max(recs, key=lambda r: (1 if r.get("id_number") else 0, len(str(r))))
        result.append(best)
        seen_names.add(best.get("name", ""))
        if len(recs) > 1:
            duplicates += len(recs) - 1
    
    for (name, date_str), recs in by_name_date.items():
        if name in seen_names:
            duplicates += len(recs)
            continue
        
        best = max(recs, key=lambda r: (1 if r.get("id_number") else 0, len(str(r))))
        result.append(best)
        seen_names.add(name)
        if len(recs) > 1:
            duplicates += len(recs) - 1
    
    return result, duplicates


def _collect_records_for_intent(
    extracted_json: dict,
    intent: str,
    name_key: str,
    date_key: str,
    id_key: Optional[str] = None
) -> Tuple[List[Dict[str, Any]], List[str]]:
    records: List[Dict[str, Any]] = []
    warnings: List[str] = []
    
    sources = extracted_json.get("sources", []) if isinstance(extracted_json, dict) else []
    
    date_candidates = _date_candidates(intent)
    for source in sources:
        if not isinstance(source, dict):
            continue
        
        extracted = source.get("extracted")
        if not isinstance(extracted, dict):
            continue
        
        data = extracted.get("data")
        if not isinstance(data, list):
            continue
        
        source_filename = source.get("filename", "unknown")
        source_type = source.get("source_type", "unknown")
        
        for item in data:
            if not isinstance(item, dict):
                continue
            
            record_intent = item.get("intent", "")
            if intent == "remove" and record_intent == "remove":
                pass
            elif intent == "add" and record_intent == "add":
                pass
            elif not record_intent:
                pass
            else:
                continue
            
            name = _get_record_value(item, name_key)
            if not name:
                for fallback in ["name", "姓名", "员工姓名"]:
                    name = _get_record_value(item, fallback)
                    if name:
                        break
            
            if not name or not _validate_chinese_name(name):
                continue
            
            date_value = _get_record_value_by_key(item, date_key)
            if not date_value:
                for fallback in date_candidates:
                    date_value = _get_record_value_by_key(item, fallback)
                    if date_value:
                        break
            
            dt = _parse_any_date(date_value)
            if not dt:
                warnings.append(f"record_skipped: no valid date for {name}")
                continue
            
            id_value = None
            if id_key:
                id_value = _get_record_value(item, id_key)
            if not id_value:
                for fallback in ["id_number", "身份证号", "证件号码", "employee_id"]:
                    id_value = _get_record_value(item, fallback)
                    if id_value:
                        break
            
            if id_value and not _validate_id_number(id_value):
                id_value = None
            
            records.append({
                "name": name,
                "id_number": id_value or "",
                "event_date": dt,
                "__source_file__": source_filename,
                "__source_type__": source_type,
            })
    
    return records, warnings


def _deduplicate_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not records:
        return []
    
    seen: Dict[str, Dict[str, Any]] = {}
    
    for record in records:
        name = record.get("name", "")
        id_num = record.get("id_number", "")
        
        key = id_num if id_num else name
        if not key:
            continue
        
        if key not in seen:
            seen[key] = record
        else:
            existing = seen[key]
            if record.get("id_number") and not existing.get("id_number"):
                seen[key] = record
    
    return list(seen.values())


def build_social_security_fill_plan(
    template_schema: TemplateSchema,
    extracted_json: dict,
    llm: LLMClient,
    template_filename: str,
    profile: SocialSecurityProfile,
    template_intent: str,
    planner_options: Optional[Dict[str, Any]] = None,
    use_llm_mapping: bool = False,
) -> FillPlan:
    debug_info: Dict[str, Any] = {
        "template_intent": template_intent,
        "template_filename": template_filename,
        "planner_mode": "social_security_constrained",
        "profile_detected": profile.is_detected,
        "has_social_security": profile.has_social_security_section,
        "has_housing_fund": profile.has_housing_fund_section,
        "fee_month_format": profile.fee_month_format,
    }
    warnings: List[str] = []
    mapping_hints: Dict[str, str] = {}
    
    if not profile.is_detected:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["social_security_profile_not_detected"],
            llm_used=False,
            debug=debug_info
        )
    
    region = None
    for r in template_schema.sheet_schemas[0].regions:
        if r.table and r.table.header:
            region = r
            break
    
    if not region:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["social_security_no_table_region"],
            llm_used=False,
            debug=debug_info
        )
    
    declare_type = "增" if template_intent == "add" else "减"
    
    max_sources = 2
    if isinstance(planner_options, dict):
        insurance_opts = planner_options.get("insurance")
        if isinstance(insurance_opts, dict):
            configured_max = insurance_opts.get("max_sources")
            if isinstance(configured_max, int) and configured_max > 0:
                max_sources = configured_max

    selected_sources, source_scores, selection_warnings = _select_sources_for_template(
        extracted_json,
        template_intent,
        max_sources=max_sources,
    )
    warnings.extend(selection_warnings)

    template_options = planner_options.get("template") if isinstance(planner_options, dict) and isinstance(planner_options.get("template"), dict) else {}
    if isinstance(template_options.get("use_llm_mapping"), bool):
        use_llm_mapping = bool(template_options.get("use_llm_mapping"))
    if use_llm_mapping and selected_sources:
        mapped, llm_map_warnings, llm_map_debug = _llm_infer_source_key_mapping(
            llm, selected_sources, profile, template_intent
        )
        warnings.extend(llm_map_warnings)
        debug_info["llm_mapping"] = llm_map_debug
        mapping_hints = {
            "name_key": mapped.get("name", ""),
            "id_key": mapped.get("id_number", ""),
            "date_key": mapped.get("event_date", ""),
            "reason_key": mapped.get("termination_reason", ""),
        }
        if not any(v for v in mapping_hints.values()):
            warnings.append("llm_mapping_fallback_to_heuristics")
    else:
        debug_info["llm_mapping"] = {"enabled": False}
    
    debug_info["source_selection"] = {
        "selected_count": len(selected_sources),
        "scores": [
            {
                "source_id": s.source_id,
                "filename": s.filename,
                "source_type": s.source_type,
                "total_score": s.total_score,
                "intent_score": s.intent_score,
                "name_coverage": s.name_coverage,
                "id_coverage": s.id_coverage,
                "date_coverage": s.date_coverage,
                "has_exclusion": s.has_exclusion_signal,
            }
            for s in source_scores
        ],
    }
    debug_info["add_sheet_counts_before_filter"] = {}
    debug_info["add_sheet_counts_after_filter"] = {}
    debug_info["add_sheet_counts_blocked"] = {}
    debug_info["remove_sheet_counts_before_filter"] = {}
    debug_info["remove_sheet_counts_after_filter"] = {}
    debug_info["remove_sheet_counts_blocked"] = {}
    debug_info["date_key_fallback_hits"] = 0
    
    all_records: List[Dict[str, Any]] = []
    for source in selected_sources:
        source_records, extract_warnings, extract_debug = _extract_records_from_source(
            source, template_intent, mapping_hints=mapping_hints
        )
        all_records.extend(source_records)
        warnings.extend(extract_warnings)
        for sn, cnt in (extract_debug.get("add_sheet_counts_before_filter") or {}).items():
            debug_info["add_sheet_counts_before_filter"][sn] = (
                debug_info["add_sheet_counts_before_filter"].get(sn, 0) + int(cnt or 0)
            )
        for sn, cnt in (extract_debug.get("add_sheet_counts_after_filter") or {}).items():
            debug_info["add_sheet_counts_after_filter"][sn] = (
                debug_info["add_sheet_counts_after_filter"].get(sn, 0) + int(cnt or 0)
            )
        for sn, cnt in (extract_debug.get("add_sheet_counts_blocked") or {}).items():
            debug_info["add_sheet_counts_blocked"][sn] = (
                debug_info["add_sheet_counts_blocked"].get(sn, 0) + int(cnt or 0)
            )
        for sn, cnt in (extract_debug.get("remove_sheet_counts_before_filter") or {}).items():
            debug_info["remove_sheet_counts_before_filter"][sn] = (
                debug_info["remove_sheet_counts_before_filter"].get(sn, 0) + int(cnt or 0)
            )
        for sn, cnt in (extract_debug.get("remove_sheet_counts_after_filter") or {}).items():
            debug_info["remove_sheet_counts_after_filter"][sn] = (
                debug_info["remove_sheet_counts_after_filter"].get(sn, 0) + int(cnt or 0)
            )
        for sn, cnt in (extract_debug.get("remove_sheet_counts_blocked") or {}).items():
            debug_info["remove_sheet_counts_blocked"][sn] = (
                debug_info["remove_sheet_counts_blocked"].get(sn, 0) + int(cnt or 0)
            )
        debug_info["date_key_fallback_hits"] = int(debug_info.get("date_key_fallback_hits", 0) or 0) + int(
            extract_debug.get("date_key_fallback_hits", 0) or 0
        )
    
    debug_info["records_before_dedup"] = len(all_records)
    
    records, duplicate_count = _deduplicate_records_advanced(all_records)
    
    if duplicate_count > 0:
        debug_info["duplicates_removed"] = duplicate_count
    
    debug_info["records_collected"] = len(records)
    
    if not records:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["social_security_no_records"],
            llm_used=False,
            debug=debug_info
        )
    
    rows: List[Dict[str, Any]] = []
    skipped = 0
    
    for record in records:
        name = record.get("name", "")
        id_number = record.get("id_number", "")
        event_date = record.get("event_date")
        termination_reason = record.get("termination_reason", "")
        
        if not name or not _validate_chinese_name(name):
            skipped += 1
            continue
        
        if not event_date:
            skipped += 1
            continue
        
        year, month = _next_month(event_date)
        fee_month = _format_fee_month(year, month, profile.fee_month_format or "YYYYMM")
        
        row_data: Dict[str, Any] = {
            "__name__": name,
            "__id_number__": id_number,
            "__declare_type__": declare_type,
            "__fee_month__": fee_month,
        }
        if template_intent == "remove" and termination_reason:
            row_data["__termination_reason__"] = termination_reason
        rows.append(row_data)
    
    if skipped:
        warnings.append(f"social_security_records_skipped:{skipped}")
    
    debug_info["rows_generated"] = len(rows)
    
    if not rows:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["social_security_no_valid_rows"],
            llm_used=False,
            debug=debug_info
        )
    
    try:
        min_col, min_row, max_col, max_row = range_boundaries(region.table.range)
    except Exception:
        return FillPlan(
            target=FillPlanTarget(),
            warnings=["social_security_invalid_range"],
            llm_used=False,
            debug=debug_info
        )
    
    max_header_row = max(region.header_rows) if region.header_rows else min_row
    data_start_row = max_header_row + 1
    start_cell = f"{get_column_letter(min_col)}{data_start_row}"
    clear_end_row = min(max_row, data_start_row + len(rows) + 30)
    clear_range = f"{get_column_letter(min_col)}{data_start_row}:{get_column_letter(max_col)}{clear_end_row}"
    
    column_mapping: Dict[str, str] = {}
    
    if profile.name_columns:
        column_mapping["__name__"] = profile.name_columns[0][0]
    
    if profile.id_columns:
        column_mapping["__id_number__"] = profile.id_columns[0][0]
    
    declare_type_keys: List[str] = []
    for idx, (col_letter, header_path) in enumerate(profile.declare_type_columns):
        section = _detect_section_type(header_path)
        key = f"__declare_type__{section}_{idx}"
        column_mapping[key] = col_letter
        declare_type_keys.append(key)
    
    fee_month_keys: List[str] = []
    for idx, (col_letter, header_path) in enumerate(profile.fee_month_columns):
        section = _detect_section_type(header_path)
        key = f"__fee_month__{section}_{idx}"
        column_mapping[key] = col_letter
        fee_month_keys.append(key)
    
    reason_keys: List[str] = []
    if template_intent == "remove":
        for idx, (col_letter, header_path) in enumerate(profile.reason_columns):
            key = f"__termination_reason__{idx}"
            column_mapping[key] = col_letter
            reason_keys.append(key)
    
    any_id_present = any(row.get("__id_number__") for row in rows)
    if not any_id_present:
        warnings.append("required_field_missing:证件号码")
        if profile.id_columns:
            del column_mapping["__id_number__"]
    
    final_rows: List[Dict[str, Any]] = []
    for row in rows:
        name_val = row["__name__"]
        
        if _is_pure_digits(name_val):
            warnings.append(f"validation_rejected_name_is_digits:{name_val}")
            continue
        
        if _is_date_like(name_val):
            warnings.append(f"validation_rejected_name_is_date:{name_val}")
            continue
        
        if _is_gender_like(name_val):
            warnings.append(f"validation_rejected_name_is_gender:{name_val}")
            continue
        
        id_val = row["__id_number__"]
        if id_val:
            if _validate_chinese_name(id_val):
                warnings.append(f"validation_rejected_id_is_name:{id_val}")
                id_val = ""
            elif _is_date_like(id_val):
                warnings.append(f"validation_rejected_id_is_date:{id_val}")
                id_val = ""
            elif _is_gender_like(id_val):
                warnings.append(f"validation_rejected_id_is_gender:{id_val}")
                id_val = ""
        
        final_row: Dict[str, Any] = {
            "__name__": name_val,
        }
        
        if any_id_present:
            final_row["__id_number__"] = id_val
        
        for key in declare_type_keys:
            final_row[key] = declare_type
        
        for key in fee_month_keys:
            fee_month_val = row["__fee_month__"]
            if _is_date_like(fee_month_val):
                warnings.append(f"validation_rejected_fee_month_is_raw_date:{fee_month_val}")
                continue
            final_row[key] = fee_month_val
        
        if reason_keys:
            reason_val = row.get("__termination_reason__", "")
            for key in reason_keys:
                final_row[key] = reason_val
        final_rows.append(final_row)
    
    validated_rows, validation_errors = _validate_all_rows(
        final_rows, declare_type_keys, fee_month_keys, declare_type
    )
    warnings.extend(validation_errors)
    
    debug_info["rows_before_validation"] = len(final_rows)
    debug_info["rows_after_validation"] = len(validated_rows)
    
    if len(validated_rows) < len(final_rows):
        rejected_count = len(final_rows) - len(validated_rows)
        warnings.append(f"validation_rejected_rows:{rejected_count}")
    
    final_rows = validated_rows
    
    sheet_name = template_schema.sheet_schemas[0].sheet if template_schema.sheet_schemas else "Sheet1"
    
    fill_plan_dict = {
        "target": {
            "sheet_name": sheet_name,
            "sheet": sheet_name,
            "region_id": region.region_id,
            "layout_type": "table",
            "clear_policy": "clear_values_keep_format"
        },
        "clear_ranges": [clear_range],
        "row_writes": [{
            "start_cell": start_cell,
            "rows": final_rows,
            "column_mapping": column_mapping
        }],
        "writes": [],
        "warnings": warnings,
        "llm_used": bool(use_llm_mapping and any(v for v in mapping_hints.values())),
        "constant_values_count": 0,
    }
    
    debug_info["column_mapping"] = column_mapping
    debug_info["declare_type_keys"] = declare_type_keys
    debug_info["fee_month_keys"] = fee_month_keys
    debug_info["termination_reason_keys"] = reason_keys
    fill_plan_dict["debug"] = debug_info
    
    try:
        return FillPlan.from_dict(fill_plan_dict)
    except Exception as e:
        logger.warning("Failed to create FillPlan from dict: %s", e)
        return FillPlan(
            target=FillPlanTarget(sheet_name=sheet_name),
            warnings=warnings + [f"fill_plan_creation_error: {str(e)}"],
            llm_used=False,
            debug=debug_info
        )
