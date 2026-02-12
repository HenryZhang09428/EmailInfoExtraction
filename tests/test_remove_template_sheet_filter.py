from pathlib import Path
from openpyxl import Workbook

from core.ir import SourceDoc
from core.llm import LLMClient
from core.pipeline import _extract_single_doc
from core.template import fill_planner as fp
from core.template.schema import (
    TemplateSchema,
    SheetSchema,
    RegionSchema,
    TableInfo,
    TableHeader,
    Constraints,
)


class MockExcelLLMClient(LLMClient):
    def __init__(self):
        pass
    
    def chat_json_once(self, prompt, system=None, temperature=0, step=None, timeout=None):
        if step == "excel_schema_infer":
            return {
                "semantic_key_by_header": {
                    "姓名": "name",
                    "员工工号": "employee_id",
                    "离职日期": "leave_date",
                }
            }
        return {}
    
    def chat_json(self, prompt, system=None, temperature=0, step=None):
        return self.chat_json_once(prompt, system, temperature, step)


class MockPlannerLLMClient(LLMClient):
    def __init__(self):
        pass
    
    def chat_json_once(self, prompt, system=None, temperature=0, step=None, timeout=None):
        if step == "insurance_template_param":
            return {
                "selected_source_id": "excel_leave_source",
                "name_key": "name",
                "effective_date_key": "leave_date",
                "confidence": "high",
                "notes": "mock selection"
            }
        return {}
    
    def chat_json(self, prompt, system=None, temperature=0, step=None):
        return self.chat_json_once(prompt, system, temperature, step)


def _create_multisheet_workbook(path: Path) -> None:
    wb = Workbook()
    ws_active = wb.active
    ws_active.title = "在职"
    ws_active.append(["姓名", "员工工号", "离职日期"])
    ws_active.append(["张三", "EMP001", "2025-01-10"])
    ws_active.append(["李四", "EMP002", "2025-01-11"])
    
    ws_leave_1 = wb.create_sheet("离职_1")
    ws_leave_1.append(["姓名", "员工工号", "离职日期"])
    ws_leave_1.append(["王五", "EMP003", "2025-02-01"])
    ws_leave_1.append(["赵六", "EMP004", "2025-02-02"])
    
    ws_leave_2 = wb.create_sheet("离职_2")
    ws_leave_2.append(["姓名", "员工工号", "离职日期"])
    ws_leave_2.append(["钱七", "EMP005", "2025-03-01"])
    
    wb.save(path)


def _build_remove_template_schema() -> TemplateSchema:
    headers = [
        TableHeader(col_letter="A", col_index=1, header_path="参保人/姓名", sample_values=[]),
        TableHeader(col_letter="B", col_index=2, header_path="证件号码", sample_values=[]),
        TableHeader(col_letter="C", col_index=3, header_path="社保/申报类型", sample_values=[]),
        TableHeader(col_letter="D", col_index=4, header_path="社保/费用年月", sample_values=[]),
        TableHeader(col_letter="E", col_index=5, header_path="公积金/申报类型", sample_values=[]),
        TableHeader(col_letter="F", col_index=6, header_path="公积金/费用年月", sample_values=[]),
    ]
    table = TableInfo(
        range="A2:F20",
        header=headers,
        sample_rows=[],
    )
    constraints = Constraints(
        has_formulas=False,
        formula_cells=[],
        validations=[],
        number_formats={},
    )
    region = RegionSchema(
        region_id="region_1",
        layout_type="table",
        header_rows=[2],
        table=table,
        constraints=constraints,
    )
    sheet = SheetSchema(sheet="Sheet1", regions=[region])
    return TemplateSchema(sheet_schemas=[sheet])


def test_remove_template_filters_to_leave_sheets(tmp_path):
    file_path = tmp_path / "multisheet_leave.xlsx"
    _create_multisheet_workbook(file_path)
    
    source_doc = SourceDoc(
        source_id="excel_leave_source",
        filename=file_path.name,
        file_path=str(file_path),
        source_type="excel",
        blocks=[],
        extracted=None,
        parent_source_id="email_source_1"
    )
    excel_llm = MockExcelLLMClient()
    prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
    _extract_single_doc(source_doc, excel_llm, prompts)
    
    extracted_json = {
        "sources": [
            {
                "source_id": "excel_leave_source",
                "filename": file_path.name,
                "source_type": "excel",
                "extracted": source_doc.extracted,
            }
        ],
        "merged": {"data": []},
    }
    
    template_schema = _build_remove_template_schema()
    planner_llm = MockPlannerLLMClient()
    
    fill_plan = fp.plan_fill(
        template_schema,
        extracted_json,
        planner_llm,
        template_filename="社保减员模板.xlsx"
    )
    
    names = []
    for row_write in fill_plan.row_writes:
        for row in row_write.rows:
            name = row.get("__name__")
            if name:
                names.append(name)
    
    assert set(names) == {"王五", "赵六", "钱七"}
    assert "张三" not in names
    assert "李四" not in names
    assert "remove_intent_ignore_non_leave_sheets" in fill_plan.warnings
