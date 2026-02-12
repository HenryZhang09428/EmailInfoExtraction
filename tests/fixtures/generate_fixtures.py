"""
Generate test fixtures for regression tests.
Run this script to create the fixture files.
"""
import os
from pathlib import Path
from openpyxl import Workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime

FIXTURES_DIR = Path(__file__).parent


def create_excel_header_detection_fixture():
    """
    Create an Excel file that tests header detection with:
    - Multi-row headers
    - Chinese column names
    - Date columns that need proper parsing
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "入职合同"

    ws["A1"] = "国际客服入职合同花名册"
    ws.merge_cells("A1:H1")

    ws["A2"] = "序号"
    ws["B2"] = "姓名"
    ws["C2"] = "员工工号"
    ws["D2"] = "身份证号"
    ws["E2"] = "入职日期"
    ws["F2"] = "部门"
    ws["G2"] = "岗位"
    ws["H2"] = "合同期限"

    test_data = [
        (1, "张三", "EMP001", "110101199001011234", "2025-11-01", "客服部", "客服专员", "3年"),
        (2, "李四", "EMP002", "110101199002022345", "2025-11-05", "客服部", "客服专员", "3年"),
        (3, "王五", "EMP003", "110101199003033456", "2025-11-10", "技术部", "技术支持", "2年"),
        (4, "赵六", "EMP004", "110101199004044567", "2025-12-01", "客服部", "客服主管", "3年"),
        (5, "钱七", "EMP005", "110101199005055678", "2025-12-15", "运营部", "运营专员", "2年"),
    ]

    for row_idx, row_data in enumerate(test_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output_path = FIXTURES_DIR / "excel_header_detection_sample.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_email_leave_lines_fixture():
    """
    Create an .eml file that contains leave/resignation lines with:
    - Employee names with IDs in parentheses
    - Dates in Chinese format
    - Keywords like 离职, 减员
    """
    msg = MIMEMultipart()
    msg["From"] = "hr@company.com"
    msg["To"] = "insurance@company.com"
    msg["Subject"] = "国际客服中心业务外包花名册202509 - 减员通知"
    msg["Date"] = "Mon, 04 Nov 2025 10:00:00 +0800"

    body_text = """各位同事：

以下人员已办理离职手续，请协助办理社保减员：

1、张三（42648001）申请离职，离职生效日期：11月15日
2、李四（42648002）申请离职，离职生效日期：11月20日
3、王五（42648003）申请离职，离职生效日期：12月1日

另外，以下人员也需要办理减员：
- 赵六（42648004）离职日期：2025-11-25
- 钱七（42648005）离职日期：2025/12/10

请及时处理，谢谢！

人力资源部
2025年11月4日
"""

    msg.attach(MIMEText(body_text, "plain", "utf-8"))

    output_path = FIXTURES_DIR / "email_leave_lines_sample.eml"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(msg.as_string())
    print(f"Created: {output_path}")
    return output_path


def create_all_fixtures():
    create_excel_header_detection_fixture()
    create_email_leave_lines_fixture()
    print("All fixtures created successfully.")


if __name__ == "__main__":
    create_all_fixtures()
