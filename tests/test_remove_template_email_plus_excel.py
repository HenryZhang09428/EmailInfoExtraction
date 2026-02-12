"""
Test remove template fill with email + excel sources merged and deduplicated.

Tests the constrained planner's ability to:
1. Collect leave records from email sources (intent="remove" or has leave_date)
2. Collect leave records from excel source
3. Merge and deduplicate records
4. Generate correct fill plan with __name__, __declare_type__="减", __fee_month__
"""
import os
import tempfile
import pytest
from datetime import datetime
from openpyxl import Workbook, load_workbook

from core.template import fill_planner as fp
from core.template.schema import (
    TemplateSchema,
    SheetSchema,
    RegionSchema,
    TableInfo,
    TableHeader,
    Constraints,
)
from core.template.writer import apply_fill_plan
from core.llm import LLMClient


class MockLLMClient(LLMClient):
    """Mock LLM client that returns predefined responses for testing."""
    
    def __init__(self):
        # Don't call super().__init__() to avoid needing real config
        pass
    
    def chat_json_once(self, prompt, system=None, temperature=0, step=None, timeout=None):
        """Return a mock response that selects the excel source."""
        # For insurance_template_param step, return params that point to excel source
        if step == "insurance_template_param":
            return {
                "selected_source_id": "excel_leave_source",
                "name_key": "姓名",
                "effective_date_key": "离职日期",
                "confidence": "high",
                "notes": "mock selection"
            }
        return {}
    
    def chat_json(self, prompt, system=None, temperature=0, step=None):
        return self.chat_json_once(prompt, system, temperature, step)


def _create_remove_template_workbook(path: str) -> None:
    """
    Create a minimal insurance remove template workbook with required headers.
    
    Headers:
    - A: 参保人/姓名
    - B: 证件号码
    - C: 社保/申报类型
    - D: 社保/费用年月
    - E: 公积金/申报类型
    - F: 公积金/费用年月
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Row 1: Header row 1 (merged headers)
    ws["A1"] = "参保人"
    ws["B1"] = ""
    ws["C1"] = "社保"
    ws["D1"] = "社保"
    ws["E1"] = "公积金"
    ws["F1"] = "公积金"
    
    # Row 2: Header row 2 (sub-headers)
    ws["A2"] = "姓名"
    ws["B2"] = "证件号码"
    ws["C2"] = "申报类型"
    ws["D2"] = "费用年月"
    ws["E2"] = "申报类型"
    ws["F2"] = "费用年月"
    
    # Leave data rows empty (starting from row 3)
    
    wb.save(path)


def _build_template_schema() -> TemplateSchema:
    """Build a template schema matching the remove template workbook."""
    headers = [
        TableHeader(col_letter="A", col_index=1, header_path="参保人/姓名", sample_values=[]),
        TableHeader(col_letter="B", col_index=2, header_path="证件号码", sample_values=[]),
        TableHeader(col_letter="C", col_index=3, header_path="社保/申报类型", sample_values=[]),
        TableHeader(col_letter="D", col_index=4, header_path="社保/费用年月", sample_values=[]),
        TableHeader(col_letter="E", col_index=5, header_path="公积金/申报类型", sample_values=[]),
        TableHeader(col_letter="F", col_index=6, header_path="公积金/费用年月", sample_values=[]),
    ]
    
    table = TableInfo(
        range="A2:F20",
        header=headers,
        sample_rows=[],
    )
    
    constraints = Constraints(
        has_formulas=False,
        formula_cells=[],
        validations=[],
        number_formats={},
    )
    
    region = RegionSchema(
        region_id="region_1",
        layout_type="table",
        header_rows=[2],
        table=table,
        constraints=constraints,
    )
    
    sheet = SheetSchema(sheet="Sheet1", regions=[region])
    return TemplateSchema(sheet_schemas=[sheet])


def _build_extracted_json_with_email_and_excel() -> dict:
    """
    Build extracted_json with:
    - One email source with 2 leave records
    - One excel source with 1 additional leave record
    """
    return {
        "sources": [
            # Email source with 2 leave records
            {
                "source_id": "email_leave_source",
                "filename": "离职通知.eml",
                "source_type": "email",
                "extracted": {
                    "data": [
                        {
                            "name": "张三",
                            "employee_id": "EMP001",
                            "leave_date": "2025-11-15",
                            "leave_date_text": "11月15日",
                            "intent": "remove",
                            "__source_file__": "离职通知.eml",
                            "__extraction_type__": "leave_lines",
                        },
                        {
                            "name": "李四",
                            "employee_id": "EMP002",
                            "leave_date": "2025-11-20",
                            "leave_date_text": "11月20日",
                            "intent": "remove",
                            "__source_file__": "离职通知.eml",
                            "__extraction_type__": "leave_lines",
                        },
                    ],
                    "metadata": {"source": "email_body_leave_lines"},
                    "warnings": [],
                },
            },
            # Excel source with 1 additional leave record
            {
                "source_id": "excel_leave_source",
                "filename": "减员名单.xlsx",
                "source_type": "excel",
                "extracted": {
                    "data": [
                        {
                            "姓名": "王五",
                            "employee_id": "EMP003",
                            "离职日期": "2025-12-01",
                            "__source_file__": "减员名单.xlsx",
                        },
                    ],
                    "metadata": {},
                    "warnings": [],
                },
            },
        ],
        "merged": {
            "data": [],
        },
    }


def _build_extracted_json_with_duplicate() -> dict:
    """
    Build extracted_json with duplicates:
    - Email source has 张三 with employee_id EMP001
    - Excel source also has 张三 with employee_id EMP001 (should be deduplicated)
    """
    return {
        "sources": [
            # Email source
            {
                "source_id": "email_leave_source",
                "filename": "离职通知.eml",
                "source_type": "email",
                "extracted": {
                    "data": [
                        {
                            "name": "张三",
                            "employee_id": "EMP001",
                            "leave_date": "2025-11-15",
                            "leave_date_text": "11月15日",
                            "intent": "remove",
                            "__source_file__": "离职通知.eml",
                        },
                        {
                            "name": "李四",
                            "employee_id": "EMP002",
                            "leave_date": "2025-11-20",
                            "leave_date_text": "11月20日",
                            "intent": "remove",
                            "__source_file__": "离职通知.eml",
                        },
                    ],
                    "metadata": {},
                    "warnings": [],
                },
            },
            # Excel source with duplicate 张三
            {
                "source_id": "excel_leave_source",
                "filename": "减员名单.xlsx",
                "source_type": "excel",
                "extracted": {
                    "data": [
                        {
                            "姓名": "张三",
                            "employee_id": "EMP001",  # Same employee_id as email
                            "离职日期": "2025-11-15",
                            "__source_file__": "减员名单.xlsx",
                        },
                    ],
                    "metadata": {},
                    "warnings": [],
                },
            },
        ],
        "merged": {"data": []},
    }


class TestRemoveTemplateEmailPlusExcel:
    """Tests for remove template filling with email + excel sources."""
    
    def test_collect_email_leave_records(self):
        """Test that email leave records are correctly collected."""
        extracted_json = _build_extracted_json_with_email_and_excel()
        
        records, warnings = fp._collect_email_leave_records(extracted_json)
        
        assert len(records) == 2
        names = [r["name"] for r in records]
        assert "张三" in names
        assert "李四" in names
        
        # Verify records have required fields
        for rec in records:
            assert rec["intent"] == "remove"
            assert rec.get("leave_date")
            assert rec.get("name")
    
    def test_deduplicate_leave_records_by_employee_id(self):
        """Test deduplication by employee_id."""
        records = [
            {"name": "张三", "employee_id": "EMP001", "leave_date": "2025-11-15"},
            {"name": "张三", "employee_id": "EMP001", "leave_date": "2025-11-15"},  # Duplicate
            {"name": "李四", "employee_id": "EMP002", "leave_date": "2025-11-20"},
        ]
        
        deduped, dup_count = fp._deduplicate_leave_records(records)
        
        assert len(deduped) == 2
        assert dup_count == 1
        names = [r["name"] for r in deduped]
        assert "张三" in names
        assert "李四" in names
    
    def test_deduplicate_leave_records_by_name_date(self):
        """Test deduplication by (name, leave_date) when no employee_id."""
        records = [
            {"name": "张三", "employee_id": "", "leave_date": "2025-11-15"},
            {"name": "张三", "employee_id": "", "leave_date": "2025-11-15"},  # Duplicate
            {"name": "李四", "employee_id": "", "leave_date": "2025-11-20"},
        ]
        
        deduped, dup_count = fp._deduplicate_leave_records(records)
        
        assert len(deduped) == 2
        assert dup_count == 1
    
    def test_plan_insurance_template_remove_merges_sources(self):
        """Test that remove template planner merges email + excel sources."""
        template_schema = _build_template_schema()
        extracted_json = _build_extracted_json_with_email_and_excel()
        llm = MockLLMClient()
        
        fill_plan = fp._plan_insurance_template_with_llm(
            template_schema,
            extracted_json,
            llm,
            template_filename="减员模板.xlsx",
            template_intent="remove"
        )
        
        # Check that fill plan was generated
        assert fill_plan is not None
        assert fill_plan.warnings is not None
        
        # Check debug info
        debug = fill_plan.debug or {}
        assert debug.get("template_intent") == "remove"
        
        # Check that we have row_writes
        row_writes = []
        if fill_plan.target and hasattr(fill_plan, 'model_dump'):
            plan_dict = fill_plan.model_dump()
            row_writes = plan_dict.get("row_writes", [])
        
        # Get rows from row_writes
        all_rows = []
        for rw in row_writes:
            if isinstance(rw, dict):
                all_rows.extend(rw.get("rows", []))
        
        # Should have 3 unique records (2 from email + 1 from excel)
        assert len(all_rows) == 3, f"Expected 3 rows, got {len(all_rows)}"
        
        # Check that all declare types are "减"
        for row in all_rows:
            # Check for any declare type key
            declare_type_keys = [k for k in row.keys() if "declare_type" in k]
            for key in declare_type_keys:
                assert row[key] == "减", f"Expected '减' for {key}, got {row[key]}"
    
    def test_plan_insurance_template_remove_deduplicates(self):
        """Test that remove template planner deduplicates records."""
        template_schema = _build_template_schema()
        extracted_json = _build_extracted_json_with_duplicate()
        llm = MockLLMClient()
        
        fill_plan = fp._plan_insurance_template_with_llm(
            template_schema,
            extracted_json,
            llm,
            template_filename="减员模板.xlsx",
            template_intent="remove"
        )
        
        # Get rows from row_writes
        row_writes = []
        if hasattr(fill_plan, 'model_dump'):
            plan_dict = fill_plan.model_dump()
            row_writes = plan_dict.get("row_writes", [])
        
        all_rows = []
        for rw in row_writes:
            if isinstance(rw, dict):
                all_rows.extend(rw.get("rows", []))
        
        # Should have 2 unique records (张三 deduplicated, 李四 unique)
        assert len(all_rows) == 2, f"Expected 2 rows after dedup, got {len(all_rows)}"
        
        # Verify names
        names = [row.get("__name__") for row in all_rows]
        assert "张三" in names
        assert "李四" in names
    
    def test_fee_month_calculation(self):
        """Test that fee_month is calculated as next month of leave_date."""
        # Test case: leave_date = 2025-11-15, expected fee_month = 202512
        dt = datetime(2025, 11, 15)
        fee_month = fp._next_month_yyyymm(dt)
        assert fee_month == "202512"
        
        # Test case: leave_date = 2025-12-01, expected fee_month = 202601 (year rollover)
        dt = datetime(2025, 12, 1)
        fee_month = fp._next_month_yyyymm(dt)
        assert fee_month == "202601"
    
    def test_full_pipeline_with_writer(self):
        """Test the full pipeline: plan + write to template."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create template
            template_path = os.path.join(tmpdir, "减员模板.xlsx")
            _create_remove_template_workbook(template_path)
            
            # Build inputs
            template_schema = _build_template_schema()
            extracted_json = _build_extracted_json_with_email_and_excel()
            llm = MockLLMClient()
            
            # Generate fill plan
            fill_plan = fp._plan_insurance_template_with_llm(
                template_schema,
                extracted_json,
                llm,
                template_filename="减员模板.xlsx",
                template_intent="remove"
            )
            
            # Convert to dict for writer
            fill_plan_dict = fill_plan.model_dump() if hasattr(fill_plan, 'model_dump') else {}
            
            # Apply fill plan
            output_path = os.path.join(tmpdir, "output.xlsx")
            cells_written = apply_fill_plan(template_path, fill_plan_dict, output_path)
            
            # Verify output file was created
            assert os.path.exists(output_path)
            
            # Load and verify output
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Data starts at row 3 (after 2 header rows)
            data_start_row = 3
            
            # Collect written data
            written_names = []
            written_declare_types_ss = []
            written_declare_types_gj = []
            written_fee_months_ss = []
            written_fee_months_gj = []
            
            for row_idx in range(data_start_row, data_start_row + 5):
                name = ws.cell(row_idx, 1).value  # Column A: 姓名
                if name:
                    written_names.append(name)
                    written_declare_types_ss.append(ws.cell(row_idx, 3).value)  # Column C: 社保/申报类型
                    written_fee_months_ss.append(ws.cell(row_idx, 4).value)  # Column D: 社保/费用年月
                    written_declare_types_gj.append(ws.cell(row_idx, 5).value)  # Column E: 公积金/申报类型
                    written_fee_months_gj.append(ws.cell(row_idx, 6).value)  # Column F: 公积金/费用年月
            
            # Assertions
            # 1. Output rows contain all 3 unique people
            assert len(written_names) == 3, f"Expected 3 names, got {len(written_names)}: {written_names}"
            assert "张三" in written_names
            assert "李四" in written_names
            assert "王五" in written_names
            
            # 2. All declare type columns are "减"
            for dt in written_declare_types_ss:
                if dt is not None:
                    assert dt == "减", f"Expected 社保/申报类型='减', got '{dt}'"
            for dt in written_declare_types_gj:
                if dt is not None:
                    assert dt == "减", f"Expected 公积金/申报类型='减', got '{dt}'"
            
            # 3. Fee month columns equal next month of each record's leave_date
            # 张三: leave_date=2025-11-15 -> fee_month=202512
            # 李四: leave_date=2025-11-20 -> fee_month=202512
            # 王五: leave_date=2025-12-01 -> fee_month=202601
            expected_fee_months = {"202512", "202601"}
            for fm in written_fee_months_ss:
                if fm is not None:
                    assert str(fm) in expected_fee_months, f"Unexpected fee_month: {fm}"
            for fm in written_fee_months_gj:
                if fm is not None:
                    assert str(fm) in expected_fee_months, f"Unexpected fee_month: {fm}"
            
            # 4. Other columns (证件号码, column B) remain empty
            for row_idx in range(data_start_row, data_start_row + 3):
                id_value = ws.cell(row_idx, 2).value  # Column B: 证件号码
                assert id_value is None or str(id_value).strip() == "", \
                    f"证件号码 column should be empty, got '{id_value}'"
            
            wb.close()


class TestEdgeCases:
    """Test edge cases for remove template processing."""
    
    def test_email_only_no_excel(self):
        """Test remove template with only email source (no excel)."""
        extracted_json = {
            "sources": [
                {
                    "source_id": "email_only",
                    "filename": "离职通知.eml",
                    "source_type": "email",
                    "extracted": {
                        "data": [
                            {
                                "name": "张三",
                                "employee_id": "EMP001",
                                "leave_date": "2025-11-15",
                                "intent": "remove",
                            },
                        ],
                        "metadata": {},
                        "warnings": [],
                    },
                },
            ],
            "merged": {"data": []},
        }
        
        template_schema = _build_template_schema()
        llm = MockLLMClient()
        
        fill_plan = fp._plan_insurance_template_with_llm(
            template_schema,
            extracted_json,
            llm,
            template_filename="减员模板.xlsx",
            template_intent="remove"
        )
        
        # Should still work with just email records
        row_writes = []
        if hasattr(fill_plan, 'model_dump'):
            plan_dict = fill_plan.model_dump()
            row_writes = plan_dict.get("row_writes", [])
        
        all_rows = []
        for rw in row_writes:
            if isinstance(rw, dict):
                all_rows.extend(rw.get("rows", []))
        
        assert len(all_rows) == 1
        assert all_rows[0].get("__name__") == "张三"
    
    def test_records_without_leave_date_skipped(self):
        """Test that records without leave_date are skipped with warning."""
        extracted_json = {
            "sources": [
                {
                    "source_id": "email_source",
                    "filename": "test.eml",
                    "source_type": "email",
                    "extracted": {
                        "data": [
                            {
                                "name": "张三",
                                "employee_id": "EMP001",
                                "leave_date": "2025-11-15",
                                "intent": "remove",
                            },
                            {
                                "name": "李四",
                                "employee_id": "EMP002",
                                "leave_date": "",  # Empty leave_date
                                "intent": "remove",
                            },
                        ],
                        "metadata": {},
                        "warnings": [],
                    },
                },
            ],
            "merged": {"data": []},
        }
        
        template_schema = _build_template_schema()
        llm = MockLLMClient()
        
        fill_plan = fp._plan_insurance_template_with_llm(
            template_schema,
            extracted_json,
            llm,
            template_filename="减员模板.xlsx",
            template_intent="remove"
        )
        
        # Get rows
        row_writes = []
        if hasattr(fill_plan, 'model_dump'):
            plan_dict = fill_plan.model_dump()
            row_writes = plan_dict.get("row_writes", [])
        
        all_rows = []
        for rw in row_writes:
            if isinstance(rw, dict):
                all_rows.extend(rw.get("rows", []))
        
        # Only 张三 should be included (李四 has no leave_date)
        assert len(all_rows) == 1
        assert all_rows[0].get("__name__") == "张三"
        
        # Should have warning about skipped records
        warnings = fill_plan.warnings or []
        has_skip_warning = any("skipped" in w.lower() or "date_parse" in w.lower() for w in warnings)
        assert has_skip_warning, f"Expected skip warning, got: {warnings}"
