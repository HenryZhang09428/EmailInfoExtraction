"""
Test fill_template error handling when plan_fill fails or returns None.
"""
import os
import tempfile
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock
from openpyxl import Workbook

from core.ir import IntermediateRepresentation, SourceDoc, SourceBlock, BlockType


def _create_minimal_template(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    wb.save(path)


def _create_minimal_ir() -> IntermediateRepresentation:
    return IntermediateRepresentation(
        sources=[
            SourceDoc(
                source_id="test_source",
                filename="test.xlsx",
                file_path="/tmp/test.xlsx",
                source_type="excel",
                blocks=[SourceBlock(order=1, type=BlockType.TEXT, content="", meta={})],
                extracted={"data": [{"name": "Test"}]},
                parent_source_id=None,
            )
        ],
        facts=[],
        target_schema=None,
        output=None,
        scores=None,
    )


class TestFillTemplateErrorHandling:

    def test_fill_template_handles_plan_fill_returning_none(self, tmp_path):
        template_path = tmp_path / "template.xlsx"
        _create_minimal_template(str(template_path))
        
        ir = _create_minimal_ir()
        
        with patch("core.pipeline.plan_fill") as mock_plan_fill, \
             patch("core.pipeline.get_llm_client") as mock_llm, \
             patch("core.pipeline.get_settings") as mock_settings:
            
            mock_plan_fill.return_value = None
            mock_llm.return_value = MagicMock()
            mock_settings.return_value = MagicMock()
            
            from core.pipeline import fill_template
            
            output_path, schema, fill_plan_dict = fill_template(
                ir, str(template_path), require_llm=False
            )
            
            assert output_path is not None
            assert schema is not None
            assert fill_plan_dict is not None
            assert isinstance(fill_plan_dict, dict)
            
            assert "warnings" in fill_plan_dict
            warnings = fill_plan_dict.get("warnings", [])
            assert any("None" in w or "failed" in w for w in warnings)
            
            assert "debug" in fill_plan_dict
            assert isinstance(fill_plan_dict["debug"], dict)

    def test_fill_template_handles_plan_fill_exception(self, tmp_path):
        template_path = tmp_path / "template.xlsx"
        _create_minimal_template(str(template_path))
        
        ir = _create_minimal_ir()
        
        with patch("core.pipeline.plan_fill") as mock_plan_fill, \
             patch("core.pipeline.get_llm_client") as mock_llm, \
             patch("core.pipeline.get_settings") as mock_settings:
            
            mock_plan_fill.side_effect = RuntimeError("Test error from plan_fill")
            mock_llm.return_value = MagicMock()
            mock_settings.return_value = MagicMock()
            
            from core.pipeline import fill_template
            
            output_path, schema, fill_plan_dict = fill_template(
                ir, str(template_path), require_llm=False
            )
            
            assert output_path is not None
            assert schema is not None
            assert fill_plan_dict is not None
            assert isinstance(fill_plan_dict, dict)
            
            warnings = fill_plan_dict.get("warnings", [])
            assert any("RuntimeError" in w or "Test error" in w for w in warnings)
            
            debug = fill_plan_dict.get("debug", {})
            assert "plan_fill_error" in debug

    def test_fill_template_handles_unexpected_return_type(self, tmp_path):
        template_path = tmp_path / "template.xlsx"
        _create_minimal_template(str(template_path))
        
        ir = _create_minimal_ir()
        
        with patch("core.pipeline.plan_fill") as mock_plan_fill, \
             patch("core.pipeline.get_llm_client") as mock_llm, \
             patch("core.pipeline.get_settings") as mock_settings:
            
            mock_plan_fill.return_value = "unexpected_string"
            mock_llm.return_value = MagicMock()
            mock_settings.return_value = MagicMock()
            
            from core.pipeline import fill_template
            
            output_path, schema, fill_plan_dict = fill_template(
                ir, str(template_path), require_llm=False
            )
            
            assert output_path is not None
            assert schema is not None
            assert fill_plan_dict is not None
            assert isinstance(fill_plan_dict, dict)
            
            warnings = fill_plan_dict.get("warnings", [])
            assert any("unexpected type" in w for w in warnings)

    def test_fill_template_returns_valid_structure_on_failure(self, tmp_path):
        template_path = tmp_path / "template.xlsx"
        _create_minimal_template(str(template_path))
        
        ir = _create_minimal_ir()
        
        with patch("core.pipeline.plan_fill") as mock_plan_fill, \
             patch("core.pipeline.get_llm_client") as mock_llm, \
             patch("core.pipeline.get_settings") as mock_settings:
            
            mock_plan_fill.return_value = None
            mock_llm.return_value = MagicMock()
            mock_settings.return_value = MagicMock()
            
            from core.pipeline import fill_template
            
            output_path, schema, fill_plan_dict = fill_template(
                ir, str(template_path), require_llm=False
            )
            
            assert "target" in fill_plan_dict
            assert "row_writes" in fill_plan_dict
            assert isinstance(fill_plan_dict.get("row_writes"), list)
            assert "writes" in fill_plan_dict
            assert isinstance(fill_plan_dict.get("writes"), list)
            assert "warnings" in fill_plan_dict
            assert isinstance(fill_plan_dict.get("warnings"), list)
            assert "llm_used" in fill_plan_dict
            assert "debug" in fill_plan_dict
            assert isinstance(fill_plan_dict.get("debug"), dict)

    def test_fill_template_output_file_exists_on_failure(self, tmp_path):
        template_path = tmp_path / "template.xlsx"
        _create_minimal_template(str(template_path))
        
        ir = _create_minimal_ir()
        
        with patch("core.pipeline.plan_fill") as mock_plan_fill, \
             patch("core.pipeline.get_llm_client") as mock_llm, \
             patch("core.pipeline.get_settings") as mock_settings:
            
            mock_plan_fill.return_value = None
            mock_llm.return_value = MagicMock()
            mock_settings.return_value = MagicMock()
            
            from core.pipeline import fill_template
            
            output_path, schema, fill_plan_dict = fill_template(
                ir, str(template_path), require_llm=False
            )
            
            assert os.path.exists(output_path)
