"""
Regression test for Excel header detection.
Ensures that the extractor correctly identifies header rows and extracts data
from files with multi-row headers and Chinese column names.
"""
import os
import tempfile
import pytest
from pathlib import Path
from openpyxl import Workbook

from core.extractors.excel_extractor import ExcelExtractor
from core.llm import LLMClient


class MockLLMClient(LLMClient):
    def __init__(self):
        pass

    def chat_json_once(self, prompt, system=None, temperature=0, step=None, timeout=None):
        if step == "excel_schema_infer":
            return {
                "semantic_key_by_header": {
                    "姓名": "name",
                    "员工工号": "employee_id",
                    "身份证号": "id_number",
                    "入职日期": "start_date",
                    "部门": "department",
                    "岗位": "position",
                },
                "row_filter": {
                    "min_nonempty_ratio": 0.2,
                    "exclude_if_contains_any": ["合计", "备注"],
                    "required_fields_any": ["name", "employee_id"]
                },
                "normalization": {
                    "date_fields": ["start_date"],
                    "id_fields": ["id_number", "employee_id"],
                    "phone_fields": []
                }
            }
        return {}

    def chat_json(self, prompt, system=None, temperature=0, step=None):
        return self.chat_json_once(prompt, system, temperature, step)


@pytest.fixture
def excel_with_multirow_header(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "入职合同"

    ws["A1"] = "国际客服入职合同花名册"
    ws.merge_cells("A1:H1")

    ws["A2"] = "序号"
    ws["B2"] = "姓名"
    ws["C2"] = "员工工号"
    ws["D2"] = "身份证号"
    ws["E2"] = "入职日期"
    ws["F2"] = "部门"
    ws["G2"] = "岗位"
    ws["H2"] = "合同期限"

    test_data = [
        (1, "张三", "EMP001", "110101199001011234", "2025-11-01", "客服部", "客服专员", "3年"),
        (2, "李四", "EMP002", "110101199002022345", "2025-11-05", "客服部", "客服专员", "3年"),
        (3, "王五", "EMP003", "110101199003033456", "2025-11-10", "技术部", "技术支持", "2年"),
        (4, "赵六", "EMP004", "110101199004044567", "2025-12-01", "客服部", "客服主管", "3年"),
        (5, "钱七", "EMP005", "110101199005055678", "2025-12-15", "运营部", "运营专员", "2年"),
    ]

    for row_idx, row_data in enumerate(test_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    file_path = tmp_path / "excel_header_detection_sample.xlsx"
    wb.save(file_path)
    return str(file_path)


class TestExcelHeaderDetectionRegression:

    def test_header_row_detected_correctly(self, excel_with_multirow_header):
        llm = MockLLMClient()
        prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
        extractor = ExcelExtractor(llm, prompts)

        source_doc = extractor.extract(excel_with_multirow_header)

        assert source_doc is not None
        assert source_doc.extracted is not None

    def test_extracted_data_not_empty(self, excel_with_multirow_header):
        llm = MockLLMClient()
        prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
        extractor = ExcelExtractor(llm, prompts)

        source_doc = extractor.extract(excel_with_multirow_header)
        extracted = source_doc.extracted

        assert isinstance(extracted, dict)
        data = extracted.get("data", [])
        assert len(data) > 0, "extracted['data'] should not be empty"

    def test_extracted_records_have_required_fields(self, excel_with_multirow_header):
        llm = MockLLMClient()
        prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
        extractor = ExcelExtractor(llm, prompts)

        source_doc = extractor.extract(excel_with_multirow_header)
        data = source_doc.extracted.get("data", [])

        assert len(data) >= 5, f"Expected at least 5 records, got {len(data)}"

        for record in data:
            assert isinstance(record, dict)
            has_name = any(k for k in record.keys() if "姓名" in k or k == "name")
            has_id = any(k for k in record.keys() if "工号" in k or k == "employee_id")
            has_date = any(k for k in record.keys() if "日期" in k or k == "start_date")
            assert has_name or has_id, f"Record missing name or employee_id: {record.keys()}"

    def test_expected_names_present(self, excel_with_multirow_header):
        llm = MockLLMClient()
        prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
        extractor = ExcelExtractor(llm, prompts)

        source_doc = extractor.extract(excel_with_multirow_header)
        data = source_doc.extracted.get("data", [])

        all_values = []
        for record in data:
            all_values.extend(str(v) for v in record.values() if v)

        all_text = " ".join(all_values)
        expected_names = ["张三", "李四", "王五", "赵六", "钱七"]
        found_names = [name for name in expected_names if name in all_text]
        assert len(found_names) >= 3, f"Expected at least 3 names, found: {found_names}"

    def test_dates_extracted_correctly(self, excel_with_multirow_header):
        llm = MockLLMClient()
        prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
        extractor = ExcelExtractor(llm, prompts)

        source_doc = extractor.extract(excel_with_multirow_header)
        data = source_doc.extracted.get("data", [])

        date_values = []
        for record in data:
            for key, value in record.items():
                if "日期" in key or "date" in key.lower():
                    if value:
                        date_values.append(str(value))

        assert len(date_values) > 0, "Should extract date values"
        date_text = " ".join(date_values)
        assert "2025" in date_text or "11" in date_text or "12" in date_text


class TestExcelHeaderEdgeCases:

    def test_title_row_not_treated_as_data(self, excel_with_multirow_header):
        llm = MockLLMClient()
        prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
        extractor = ExcelExtractor(llm, prompts)

        source_doc = extractor.extract(excel_with_multirow_header)
        data = source_doc.extracted.get("data", [])

        for record in data:
            values_str = " ".join(str(v) for v in record.values() if v)
            assert "花名册" not in values_str, "Title row should not be in data"
            assert "入职合同花名册" not in values_str

    def test_header_row_not_treated_as_data(self, excel_with_multirow_header):
        llm = MockLLMClient()
        prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
        extractor = ExcelExtractor(llm, prompts)

        source_doc = extractor.extract(excel_with_multirow_header)
        data = source_doc.extracted.get("data", [])

        for record in data:
            values_list = [str(v).strip() for v in record.values() if v]
            is_header_row = (
                "序号" in values_list and
                "姓名" in values_list and
                "员工工号" in values_list
            )
            assert not is_header_row, "Header row should not be in data records"
