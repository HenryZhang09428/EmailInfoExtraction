import pytest
from pathlib import Path
from openpyxl import Workbook

from core.ir import SourceDoc
from core.llm import LLMClient
from core.pipeline import _extract_single_doc


class MockLLMClient(LLMClient):
    def __init__(self):
        pass
    
    def chat_json_once(self, prompt, system=None, temperature=0, step=None, timeout=None):
        if step == "excel_schema_infer":
            return {
                "semantic_key_by_header": {
                    "姓名": "name",
                    "员工工号": "employee_id",
                    "离职日期": "leave_date",
                }
            }
        return {}
    
    def chat_json(self, prompt, system=None, temperature=0, step=None):
        return self.chat_json_once(prompt, system, temperature, step)


def _create_multisheet_workbook(path: Path) -> None:
    wb = Workbook()
    ws_active = wb.active
    ws_active.title = "在职"
    ws_active.append(["姓名", "员工工号", "离职日期"])
    ws_active.append(["张三", "EMP001", "2025-01-10"])
    ws_active.append(["李四", "EMP002", "2025-01-11"])
    
    ws_leave_1 = wb.create_sheet("离职_1")
    ws_leave_1.append(["姓名", "员工工号", "离职日期"])
    ws_leave_1.append(["王五", "EMP003", "2025-02-01"])
    ws_leave_1.append(["赵六", "EMP004", "2025-02-02"])
    
    ws_leave_2 = wb.create_sheet("离职_2")
    ws_leave_2.append(["姓名", "员工工号", "离职日期"])
    ws_leave_2.append(["钱七", "EMP005", "2025-03-01"])
    
    wb.save(path)


def test_email_attachment_multisheet_extracts_all_records(tmp_path):
    file_path = tmp_path / "multisheet_leave.xlsx"
    _create_multisheet_workbook(file_path)
    
    source_doc = SourceDoc(
        source_id="excel_source_1",
        filename=file_path.name,
        file_path=str(file_path),
        source_type="excel",
        blocks=[],
        extracted=None,
        parent_source_id="email_source_1"
    )
    llm = MockLLMClient()
    prompts = {"EXCEL_SCHEMA_INFER_PROMPT": "mock prompt"}
    
    _extract_single_doc(source_doc, llm, prompts)
    
    extracted = source_doc.extracted
    assert isinstance(extracted, dict)
    data = extracted.get("data", [])
    assert len(data) == 5
    assert all("__sheet_name__" in record for record in data)
    sheet_names = {record.get("__sheet_name__") for record in data}
    assert sheet_names == {"在职", "离职_1", "离职_2"}
    
    sheets_meta = extracted.get("metadata", {}).get("sheets", [])
    assert isinstance(sheets_meta, list)
    assert len(sheets_meta) == 3
