from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from core.template.writer import apply_fill_plan


def main() -> None:
    with TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        template_path = temp_path / "template.xlsx"
        output_path = temp_path / "output.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(template_path)
        wb.close()

        fill_plan = {
            "target": {
                "region_id": "region_1",
                "layout_type": "table",
                "clear_policy": "clear_values_keep_format",
            },
            "clear_ranges": [],
            "row_writes": [],
            "writes": [{"cell": "A1", "value": "ok"}],
        }

        cells_written = apply_fill_plan(str(template_path), fill_plan, str(output_path))

        out_wb = load_workbook(output_path)
        out_ws = out_wb.active
        value = out_ws["A1"].value
        sheet_title = out_ws.title
        out_wb.close()

        assert cells_written >= 1, f"Expected at least 1 cell written, got {cells_written}"
        assert value == "ok", f"Expected 'ok', got {value}"
        print(f"smoke ok: cells_written={cells_written}, sheet={sheet_title}, value={value}")


if __name__ == "__main__":
    main()
